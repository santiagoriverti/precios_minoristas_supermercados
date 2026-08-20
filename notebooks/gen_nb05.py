"""Script to generate 05_evolucion_productos_representativos.ipynb — Multi-producto version.

Igual que 02_evolucion_canasta_representativa.ipynb pero en vez de agrupar productos en
hasta 6 canastas ponderadas, sigue la evolución de precio de cada producto INDIVIDUAL
(un EAN = un producto). La lista de EANs se define en la CELDA 1 (EANS_INPUT), separados
por coma, sin límite de cantidad.
"""
import json, os, hashlib

def _cell_id(prefix, src):
    # id determinista (md5 del contenido) — evita 'drift' espurio al regenerar
    return prefix + hashlib.md5(src.encode('utf-8')).hexdigest()[:6]

def cell_md(src):
    lines = src.split('\n')
    source = [l + '\n' for l in lines[:-1]] + ([lines[-1]] if lines[-1] else [])
    return {'cell_type':'markdown','id':_cell_id('md', src),'metadata':{},'source':source}

def cell_code(src):
    return {'cell_type':'code','execution_count':None,'id':_cell_id('c', src),'metadata':{},'outputs':[],'source':[src]}

cells = []

# ── CELL 0 ─────────────────────────────────────────────────────────────────────
cells.append(cell_md("""# SEPA — Evolución de Productos Representativos (Multi-producto)

**Objetivo:** Igual que `02_evolucion_canasta_representativa`, pero en vez de agrupar
productos en canastas ponderadas, sigue la evolución de precio de **productos
individuales** (Coca-Cola, Fernet, etc.) por separado: precio nacional, mapas
provinciales, comparación con el IPC INDEC, rankings por cadena y barrio.

**Productos a analizar:** se definen en la **CELDA 1**, campo `EANS_INPUT`, como una
lista de códigos EAN separados por coma. **Sin límite de cantidad.** Ejemplo:

```python
EANS_INPUT = '7790895000782, 7790085000010, 7794000618213'
```

Cada EAN se cruza contra el Maestro de Productos del repo para mostrar su descripción;
si un EAN no aparece en el maestro, igual se analiza (se muestra por su código).

**Estructura:** Config (EANs) → Setup → Maestros → Productos desde EANs → ZIPs →
Mes actual → Precio por sucursal → Análisis provincial → Serie histórica → IPC →
Comparativa → Gráficos IPC → Cuadro por producto → Mapas coropléticos → Cobertura →
Rankings → Mapa Folium → Rankings CABA → Excel"""))

# ── CELL 1 — CONFIG ────────────────────────────────────────────────────────────
cells.append(cell_code("""\
# ===========================================================
# CONFIGURACIÓN — Modificar solo esta sección
# ===========================================================

# EANs de los productos a analizar, separados por coma. SIN LÍMITE de cantidad.
# Poné los códigos EAN (numéricos), no los nombres de los productos.
# Ejemplo: 'Coca-Cola 2.25L, Fernet Branca 750ml' -> EANS_INPUT = '7790895000782, 7790085000010'
EANS_INPUT = '7790895000782, 7790085000010, 7794000618213'

SEPA_SOURCE = 'mi_drive'   # 'mi_drive' | 'local'

SEPA_DIR   = '/content/drive/MyDrive/carga'
OUTPUT_DIR = '/content/drive/MyDrive/carga/output_productos'

USE_CACHE = True

# Período mínimo de la serie histórica
MES_INICIO_HISTORICO = '2024-01'

# Mes base para gráficos de índice (auto-adapta si no está en la serie)
MES_INICIO_GRAFICO = '2024-01'

# Mínimo sucursales por cadena para aparecer en rankings
MIN_SUCURSALES_RANKING = 10"""))

# ── CELL 2 — SETUP ─────────────────────────────────────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 2 — Montar Drive + dependencias + imports
# ============================================================
try:
    import google.colab
    from google.colab import drive
    drive.mount('/content/drive')
    print('Google Drive montado')
except ImportError:
    print('Entorno local')

import subprocess, sys
subprocess.run([sys.executable, '-m', 'pip', 'install',
                'folium', 'openpyxl', 'tqdm', 'pyarrow', '-q'], check=False)

import zipfile, gzip, re, shutil, warnings, gc, hashlib
import json as _json
from pathlib import Path
from tqdm.auto import tqdm
warnings.filterwarnings('ignore')

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import matplotlib.colors as mcolors
import matplotlib.dates as mdates
from matplotlib.colors import LinearSegmentedColormap, Normalize
import seaborn as sns
import folium
from branca.colormap import LinearColormap

plt.rcParams['figure.figsize'] = (13, 6)
plt.rcParams['font.size'] = 11
sns.set_theme(style='whitegrid')
pd.set_option('display.max_columns', 50)
pd.set_option('display.float_format', '{:,.2f}'.format)

SEPA_DIR   = Path(SEPA_DIR)
OUTPUT_DIR = Path(OUTPUT_DIR)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
CACHE_DIR  = OUTPUT_DIR / '_cache'
if USE_CACHE:
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
TMP_DIR = Path('/content/tmp_sepa_nb05')
TMP_DIR.mkdir(exist_ok=True)

_ipc_candidatos = [SEPA_DIR / n for n in ('IPC.xlsx', 'ipc.xlsx', 'IPC.XLSX')]
_ipc_encontrado = next((p for p in _ipc_candidatos if p.exists()), None)
IPC_PATH     = _ipc_encontrado if _ipc_encontrado else SEPA_DIR / 'IPC.xlsx'
GEOJSON_PATH = SEPA_DIR / 'ar.json'

def normalizar_ean(s):
    # EAN normalizado sin ceros a la izquierda; usado para cruzar EANs entre fuentes
    if pd.isna(s): return None
    s = str(s).strip().lstrip('0')
    return s if s else '0'

# EANs de referencia SOLO para detectar el factor centavos/pesos de cada mes
# (mismos que usa nb02). Acá hace falta anclarlo a EANs conocidos porque el
# usuario puede pedir 1 o 2 productos nada más: el precio de un único producto
# puede caer cerca del umbral de 10.000 por azar y el factor se detecta mal
# (bug real visto en producción: caídas ~100x en meses puntuales). Con 40-70
# productos (como en nb02) la mediana general ya es robusta y no hace falta.
REF_EANS_FACTOR = {'7790072002080'.lstrip('0'), '7790070320285'.lstrip('0'), '7790132098459'.lstrip('0')}

print(f'SEPA_DIR:   {SEPA_DIR}')
print(f'OUTPUT_DIR: {OUTPUT_DIR}')
print(f'  IPC.xlsx: {"OK — " + IPC_PATH.name if IPC_PATH.exists() else "NO ENCONTRADO"}')
print(f'  ar.json:  {"OK" if GEOJSON_PATH.exists() else "NO ENCONTRADO"}')"""))

# ── CELL 3 — MAESTROS (sucursales + productos) ─────────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 3 — Maestros de sucursales, cadenas, provincias y productos
# ============================================================
DATA_URL = 'https://raw.githubusercontent.com/santiagoriverti/precios_minoristas_supermercados/main/data'

def leer_maestro(nombre, **kwargs):
    local = Path('data') / nombre
    if local.exists():
        return pd.read_excel(local, **kwargs)
    import urllib.request, urllib.parse
    dl = Path('/content/data') / nombre
    Path('/content/data').mkdir(exist_ok=True)
    if not dl.exists():
        print(f'  Descargando {nombre}...')
        urllib.request.urlretrieve(f'{DATA_URL}/{urllib.parse.quote(nombre)}', dl)
    return pd.read_excel(dl, **kwargs)

print('Cargando maestros...')
maestro_suc = leer_maestro('maestro_sucursales_completo.xlsx')
for c in ['id_comercio','id_bandera','id_sucursal']:
    maestro_suc[c] = maestro_suc[c].astype(str)

suc_pais = maestro_suc[
    maestro_suc['sucursales_latitud'].notna() &
    maestro_suc['sucursales_longitud'].notna() &
    (maestro_suc['sucursales_latitud'].between(-55, -22)) &
    (maestro_suc['sucursales_longitud'].between(-73, -53))
].copy()

IDS_PAIS = set(zip(suc_pais['id_comercio'], suc_pais['id_bandera'], suc_pais['id_sucursal']))
print(f'  Sucursales validas: {len(suc_pais):,} | IDs unicos: {len(IDS_PAIS):,}')

NOMBRES_COMPUESTOS = {
    ('9','1'):'Vea',('9','2'):'Disco',('9','3'):'Jumbo',
    ('10','1'):'Carrefour',('10','2'):'Carrefour Market',('10','3'):'Carrefour Express',
    ('11','2'):'ChangoMas',('11','4'):'Hiper ChangoMas',('11','5'):'Mi ChangoMas',
    ('16','1'):'Hipermercado Libertad',('16','2'):'Mini Libertad',
}
NOMBRES_SIMPLES = {
    '2':'La Anonima','3':'Cadena 3','5':'Hipermercado Misiones',
    '8':'Cadena 8 (Cordoba)','12':'Coto','13':'Cooperativa Obrera',
    '15':'DIA','20':'LAR','21':'Toledo','23':'Cadena 23','47':'Pasamonte',
}

def asignar_cadena(row):
    k = (row['id_comercio'], row['id_bandera'])
    if k in NOMBRES_COMPUESTOS: return NOMBRES_COMPUESTOS[k]
    if row['id_comercio'] in NOMBRES_SIMPLES: return NOMBRES_SIMPLES[row['id_comercio']]
    return f"Cadena {row['id_comercio']}"

PROV_NORM = {
    'Ciudad Autonoma de Buenos Aires':'CABA',
    'Ciudad Autónoma de Buenos Aires':'CABA',
    'Provincia de Buenos Aires':'Buenos Aires',
    'Provincia de Catamarca':'Catamarca',
    'Provincia del Chaco':'Chaco','Provincia del Chubut':'Chubut',
    'Provincia de Cordoba':'Córdoba','Provincia de Córdoba':'Córdoba',
    'Provincia de Corrientes':'Corrientes',
    'Provincia de Entre Rios':'Entre Ríos','Provincia de Entre Ríos':'Entre Ríos',
    'Provincia de Formosa':'Formosa','Provincia de Jujuy':'Jujuy',
    'Provincia de La Pampa':'La Pampa','Provincia de La Rioja':'La Rioja',
    'Provincia de Mendoza':'Mendoza','Provincia de Misiones':'Misiones',
    'Provincia del Neuquen':'Neuquén','Provincia del Neuquén':'Neuquén','Neuquén':'Neuquén',
    'Provincia de Rio Negro':'Río Negro','Provincia de Río Negro':'Río Negro',
    'Provincia de Salta':'Salta','Provincia de San Juan':'San Juan',
    'Provincia de San Luis':'San Luis','Provincia de Santa Cruz':'Santa Cruz',
    'Provincia de Santa Fe':'Santa Fe',
    'Provincia de Santiago del Estero':'Santiago del Estero',
    'Provincia de Tierra del Fuego, Antartida e Islas del Atlantico Sur':'Tierra del Fuego',
    'Provincia de Tierra del Fuego, Antártida e Islas del Atlántico Sur':'Tierra del Fuego',
    'Tierra del Fuego':'Tierra del Fuego',
    'Provincia de Tucuman':'Tucumán','Provincia de Tucumán':'Tucumán',
    'Buenos Aires':'Buenos Aires','CABA':'CABA','Catamarca':'Catamarca',
    'Chaco':'Chaco','Chubut':'Chubut','Córdoba':'Córdoba','Corrientes':'Corrientes',
    'Entre Ríos':'Entre Ríos','Formosa':'Formosa','Jujuy':'Jujuy',
    'La Pampa':'La Pampa','La Rioja':'La Rioja','Mendoza':'Mendoza',
    'Misiones':'Misiones','Río Negro':'Río Negro','Salta':'Salta',
    'San Juan':'San Juan','San juan':'San Juan','SAN JUAN':'San Juan',
    'San Luis':'San Luis','Santa Cruz':'Santa Cruz',
    'Santa Fe':'Santa Fe','Santiago del Estero':'Santiago del Estero','Tucumán':'Tucumán',
}

PESOS_POBLACION = {
    'Buenos Aires':17709732,'CABA':3075646,'Catamarca':415438,
    'Chaco':1204541,'Chubut':618994,'Córdoba':3978984,
    'Corrientes':1120801,'Entre Ríos':1385961,'Formosa':605193,
    'Jujuy':770881,'La Pampa':368550,'La Rioja':393531,
    'Mendoza':2014533,'Misiones':1261294,'Neuquén':664057,
    'Río Negro':747610,'Salta':1441998,'San Juan':781217,
    'San Luis':531745,'Santa Cruz':333473,'Santa Fe':3556522,
    'Santiago del Estero':1019304,'Tierra del Fuego':190641,'Tucumán':1737127,
}

# ── Maestro de Productos: descripción, marca y rubro por EAN ─────────────────
# OJO: el EAN real está en 'producto_sepa_id' ('producto_ean' es un flag 0/1)
_mp_raw = leer_maestro('Maestro de Productos Interno.xlsx', dtype=str,
                       usecols=['producto_sepa_id','producto_descripcion','producto_marca','rubro'])
MP_META = _mp_raw.rename(columns={'producto_descripcion':'descripcion','producto_marca':'marca'})
MP_META['ean_norm'] = MP_META['producto_sepa_id'].map(normalizar_ean)
MP_META = (MP_META.dropna(subset=['ean_norm'])
           .drop_duplicates('ean_norm')
           .set_index('ean_norm')[['descripcion','marca','rubro']])
print(f'  Maestro de productos: {len(MP_META):,} EANs con metadata')
print('Maestros OK')"""))

# ── CELL 4 — PRODUCTOS DESDE EANS_INPUT ────────────────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 4 — Productos individuales desde EANS_INPUT (CELDA 1)
# ============================================================
_eans_raw = [e.strip() for e in EANS_INPUT.split(',')]
_eans_raw = [e for e in _eans_raw if e]   # descarta vacíos (comas de más, etc.)
if not _eans_raw:
    raise ValueError(
        "EANS_INPUT está vacío. Completá la CELDA 1 con al menos un código EAN, "
        "separados por coma. Ejemplo: EANS_INPUT = '7790895000782, 7790085000010'"
    )

_eans_norm  = []
_vistos     = set()
_invalidos  = []
for _e in _eans_raw:
    _limpio = re.sub(r'\\D', '', _e)   # solo dígitos (tolera espacios, guiones, puntos)
    if not _limpio:
        _invalidos.append(_e)
        continue
    _en = _limpio.lstrip('0') or '0'
    if _en in _vistos:
        continue   # EAN duplicado en la lista -> se ignora silenciosamente
    _vistos.add(_en)
    _eans_norm.append(_en)

if _invalidos:
    print(f'AVISO: se ignoraron {len(_invalidos)} valores sin dígitos en EANS_INPUT: {_invalidos}')
if not _eans_norm:
    raise ValueError('Ningún EAN válido encontrado en EANS_INPUT (deben ser numéricos).')

# ── Cruzar con el Maestro de Productos para descripción/marca/rubro ──────────
PRODUCTOS         = {}   # pid -> {ean_norm: (descripcion[:50], 1, rubro)}
PRODUCTO_NOMBRES  = {}   # pid -> nombre para mostrar (gráficos, hojas, prints)
PRODUCTO_SHORT    = {}   # pid -> slug único para nombres de archivo (basado en el EAN)
_sin_meta = []

for _i, _ean in enumerate(_eans_norm, start=1):
    _pid = f'p{_i:02d}'
    if _ean in MP_META.index:
        _row   = MP_META.loc[_ean]
        _desc  = str(_row['descripcion'])[:50] if pd.notna(_row['descripcion']) and str(_row['descripcion']).strip() else f'EAN {_ean}'
        _rubro = str(_row['rubro']) if pd.notna(_row['rubro']) and str(_row['rubro']).strip() else 'Sin rubro'
    else:
        _sin_meta.append(_ean)
        _desc, _rubro = f'EAN {_ean}', 'Sin rubro'
    PRODUCTOS[_pid]        = {_ean: (_desc, 1, _rubro)}
    PRODUCTO_NOMBRES[_pid] = _desc
    PRODUCTO_SHORT[_pid]   = f'ean{_ean}'

if _sin_meta:
    print(f'AVISO: {len(_sin_meta)} EAN sin metadata en el Maestro de Productos '
          f'(se muestran por su código): {_sin_meta}')

PRODUCTOS_ACTIVOS  = list(PRODUCTOS.keys())
PRODUCTOS_EANS_NORM = set(_eans_norm)
N_PRODUCTOS = len(PRODUCTOS_ACTIVOS)

# ── Colores, estilos y marcadores dinámicos (uno por producto, sin límite) ───
_TAB20    = plt.cm.tab20.colors   # 20 colores discretos, se ciclan si hay más productos
_LS_CICLO = ['-', '--', '-.', ':']
_MK_CICLO = ['o', 's', '^', 'D', 'v', 'P', 'X', '*', 'h', '<', '>', '8']
PRODUCTO_COLORS     = {p: mcolors.to_hex(_TAB20[i % 20]) for i, p in enumerate(PRODUCTOS_ACTIVOS)}
PRODUCTO_LINESTYLES = {p: _LS_CICLO[(i // 20) % len(_LS_CICLO)] for i, p in enumerate(PRODUCTOS_ACTIVOS)}
PRODUCTO_MARKERS    = {p: _MK_CICLO[i % len(_MK_CICLO)] for i, p in enumerate(PRODUCTOS_ACTIVOS)}

print(f'\\nProductos a analizar: {N_PRODUCTOS}')
for _pid in PRODUCTOS_ACTIVOS:
    _ean = next(iter(PRODUCTOS[_pid]))
    print(f'  [{_pid}] {PRODUCTO_NOMBRES[_pid]}  (EAN {_ean})')"""))

# ── CELL 5 — ZIP FUNCTIONS ─────────────────────────────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 5 — Funciones de lectura de ZIPs SEPA
# ============================================================
_PAT_SEM = re.compile(r'^(\\d{4})(A|B)$', re.IGNORECASE)
_PAT_ARC = re.compile(r'^(\\d{2})(\\d{4})_pais_parte.*COMPLETO.*\\.csv\\.gz$', re.IGNORECASE)

def detectar_semestres():
    result = []
    for z in sorted(SEPA_DIR.glob('*.zip')):
        m = _PAT_SEM.match(z.stem)
        if m:
            result.append((z, int(m.group(1)), m.group(2).upper()))
    return result

def archivos_por_mes(zip_path):
    meses = {}
    with zipfile.ZipFile(zip_path) as zf:
        for nombre in zf.namelist():
            m = _PAT_ARC.match(Path(nombre).name)
            if m:
                mes_n, anio = int(m.group(1)), int(m.group(2))
                meses.setdefault((anio, mes_n), []).append(nombre)
    return meses

def detectar_ultimo_mes():
    sems = detectar_semestres()
    if not sems:
        raise RuntimeError(f'No hay ZIPs semestrales en {SEPA_DIR}')
    for zip_path, anio, sem in reversed(sems):
        meses = archivos_por_mes(zip_path)
        if meses:
            (a, m) = max(meses.keys())
            return zip_path, a, m, meses[(a, m)]
    raise RuntimeError('No se pudo detectar el ultimo mes')

sems = detectar_semestres()
print(f'Semestres encontrados: {len(sems)}')
for z, a, s in sems:
    meses = archivos_por_mes(z)
    meses_str = [f'{m:02d}/{an}' for (an,m) in sorted(meses.keys())]
    print(f'  {a}{s}: {meses_str}')"""))

# ── CELL 6 — CURRENT MONTH ─────────────────────────────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 6 — Mes actual: carga per-sucursal desde ZIP
# ============================================================
zip_actual, ANIO_ACTUAL, MES_NUM_ACTUAL, archivos_mes = detectar_ultimo_mes()

MES          = f'{MES_NUM_ACTUAL:02d}{ANIO_ACTUAL}'
PERIODO      = f'{ANIO_ACTUAL}-{MES_NUM_ACTUAL:02d}'
ULTIMO_MES   = PERIODO

_NOM = {'01':'enero','02':'febrero','03':'marzo','04':'abril','05':'mayo','06':'junio',
        '07':'julio','08':'agosto','09':'septiembre','10':'octubre','11':'noviembre','12':'diciembre'}
NOMBRE_MES       = f"{_NOM[f'{MES_NUM_ACTUAL:02d}']} {ANIO_ACTUAL}"
NOMBRE_MES_TITLE = NOMBRE_MES.title()

CADENAS_FILTRAR = {'19','2013','3001','4'}
PAT_FECHA = re.compile(r'^precio_(\\d{8})$')

print(f'Procesando: {NOMBRE_MES_TITLE}  ({zip_actual.name})')
print(f'Archivos: {archivos_mes}\\n')

# ── Detección de MES PARCIAL (menos días cargados que los del calendario) ─────
# El último mes en curso (ej. junio al día 23) tiene menos días que el mes
# completo. Su NIVEL es un snapshot válido, pero su VARIACIÓN mensual queda
# subestimada al compararse contra meses completos. Se marca para no confundir.
import calendar as _cal
_dias_cargados = set()
with zipfile.ZipFile(zip_actual) as _zfh:
    for _arch in archivos_mes:
        with _zfh.open(_arch) as _s:
            _hdr0 = gzip.open(_s, 'rt', encoding='utf-8', errors='replace').readline()
        for _c in _hdr0.strip().split(','):
            _mm = PAT_FECHA.match(_c)
            if _mm: _dias_cargados.add(_mm.group(1))
DIAS_CARGADOS  = len(_dias_cargados)
DIAS_MES       = _cal.monthrange(ANIO_ACTUAL, MES_NUM_ACTUAL)[1]
MES_PARCIAL    = 0 < DIAS_CARGADOS < DIAS_MES
SUFIJO_PARCIAL = f' (parcial: {DIAS_CARGADOS}/{DIAS_MES} días)' if MES_PARCIAL else ''
if MES_PARCIAL:
    print(f'⚠️  {NOMBRE_MES_TITLE} es un MES PARCIAL: {DIAS_CARGADOS}/{DIAS_MES} días cargados.')
    print(f'    El NIVEL es un snapshot al día {DIAS_CARGADOS}; la VARIACIÓN mensual está SUBESTIMADA')
    print(f'    (se compara contra meses completos). Re-correr al cierre del mes para el dato definitivo.\\n')

_EANS_LECTURA = PRODUCTOS_EANS_NORM | REF_EANS_FACTOR   # + referencia, solo para el factor

acumulador = []
muestra_factor = []   # precios crudos de los EANs de referencia (para detectar el factor)
for archivo in sorted(archivos_mes):
    nombre = Path(archivo).name
    tmp_p  = TMP_DIR / nombre
    with zipfile.ZipFile(zip_actual) as zf:
        with zf.open(archivo) as src, open(tmp_p, 'wb') as dst:
            shutil.copyfileobj(src, dst, length=4*1024*1024)
    print(f'  {nombre}...')
    with gzip.open(tmp_p, 'rt', encoding='utf-8', errors='replace') as g:
        for chunk in pd.read_csv(g, dtype=str, chunksize=300_000, low_memory=False):
            chunk['ean_norm'] = chunk['id_producto'].apply(normalizar_ean)
            chunk = chunk[chunk['ean_norm'].isin(_EANS_LECTURA)].copy()
            if len(chunk) == 0: continue
            for c in ['id_comercio','id_bandera','id_sucursal']:
                chunk[c] = chunk[c].astype(str)
            chunk['_k'] = list(zip(chunk['id_comercio'],chunk['id_bandera'],chunk['id_sucursal']))
            chunk = chunk[chunk['_k'].isin(IDS_PAIS)].drop(columns=['_k']).copy()
            if len(chunk) == 0: continue
            cols_p = [c for c in chunk.columns if PAT_FECHA.match(c)]
            if not cols_p: continue
            df_long = chunk.melt(
                id_vars=['id_comercio','id_bandera','id_sucursal','ean_norm'],
                value_vars=cols_p, var_name='_col', value_name='precio_raw')
            df_long['precio'] = pd.to_numeric(
                df_long['precio_raw'].replace('NA', np.nan), errors='coerce')
            df_long = df_long[df_long['precio'].notna() & (df_long['precio'] > 0)].copy()
            df_long.drop(columns=['_col','precio_raw'], inplace=True)
            _es_ref = df_long['ean_norm'].isin(REF_EANS_FACTOR)
            if _es_ref.any():
                muestra_factor.extend(df_long.loc[_es_ref, 'precio'].tolist())
            df_long = df_long[df_long['ean_norm'].isin(PRODUCTOS_EANS_NORM)].copy()
            if len(df_long) > 0:
                acumulador.append(df_long)
    tmp_p.unlink(missing_ok=True)
    del chunk, df_long; gc.collect()

if not acumulador:
    raise RuntimeError('Sin datos de precios para el mes actual y los EANs pedidos. Verificá los códigos EAN.')

datos = pd.concat(acumulador, ignore_index=True)
del acumulador; gc.collect()
datos = datos.drop_duplicates(
    subset=['id_comercio','id_bandera','id_sucursal','ean_norm'], keep='first')

# Factor centavos->pesos: se ancla a los EANs de referencia (muestra robusta,
# independiente de cuántos productos pidió el usuario). Si por algún motivo
# la referencia no aparece ese mes, cae a la mediana de los productos propios.
med_r = (pd.Series(muestra_factor).median() if muestra_factor
         else datos['precio'].median())
FACTOR = 100 if med_r > 10_000 else 1
if FACTOR == 100:
    datos['precio'] /= 100
    print(f'Factor: {FACTOR} (centavos -> pesos)')
else:
    print(f'Factor: {FACTOR} (ya en pesos)')

print(f'Datos: {len(datos):,} obs | {datos.groupby(["id_comercio","id_bandera","id_sucursal"]).ngroups:,} sucursales')
print(f'EANs con datos: {datos["ean_norm"].nunique()} / {N_PRODUCTOS} (productos consultados)')"""))

# ── CELL 7 — PER-SUCURSAL PRICES (multi-producto) ─────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 7 — Precio por sucursal para cada producto activo
# ============================================================
# A diferencia de una canasta (varios productos ponderados, con imputación por
# mediana nacional si falta un producto), acá cada 'producto' es un único EAN:
# una sucursal aparece para un producto SOLO si lo vende (sin imputación).
precio_mes = (datos.groupby(['id_comercio','id_bandera','id_sucursal','ean_norm'])
              ['precio'].mean().reset_index())
precio_mes = precio_mes[~precio_mes['id_comercio'].isin(CADENAS_FILTRAR)].copy()

# ── Bboxes y función de geocodificación ─────────────────────────────────────
_PROV_BBOX = {
    'CABA':                (-34.72,-34.52,-58.54,-58.33),
    'Tucumán':             (-28.0, -26.0, -66.5, -64.5),
    'Jujuy':               (-24.5, -21.5, -67.5, -63.5),
    'Misiones':            (-28.5, -25.5, -56.5, -53.0),
    'Chaco':               (-27.5, -24.0, -63.0, -57.5),
    'Formosa':             (-26.5, -22.0, -62.5, -58.0),
    'Corrientes':          (-30.5, -27.0, -60.0, -55.5),
    'Entre Ríos':          (-34.0, -30.0, -60.5, -57.5),
    'San Luis':            (-36.0, -32.5, -68.5, -65.0),
    'San Juan':            (-34.5, -27.5, -71.0, -65.0),
    'La Rioja':            (-32.5, -27.0, -70.0, -65.0),
    'Catamarca':           (-29.5, -25.0, -70.5, -64.5),
    'Salta':               (-26.5, -21.5, -68.5, -62.5),
    'Santiago del Estero': (-30.0, -25.5, -65.5, -61.5),
    'Mendoza':             (-37.5, -32.0, -70.5, -66.5),
    'Neuquén':             (-40.5, -36.0, -71.5, -68.5),
    'La Pampa':            (-40.0, -35.0, -68.5, -63.5),
    'Santa Fe':            (-34.5, -28.5, -62.5, -59.0),
    'Córdoba':             (-39.0, -29.5, -67.0, -62.0),
    'Río Negro':           (-42.5, -38.5, -71.5, -62.5),
    'Chubut':              (-46.5, -41.0, -72.5, -63.0),
    'Buenos Aires':        (-42.5, -33.5, -63.5, -56.5),
    'Santa Cruz':          (-52.5, -46.0, -72.5, -65.5),
    'Tierra del Fuego':    (-55.5, -51.0, -70.5, -63.5),
}

def _geocodif(lat, lon):
    # Primera provincia cuyo bbox contiene (lat, lon)
    if pd.isna(lat) or pd.isna(lon): return None
    for p, (la0, la1, lo0, lo1) in _PROV_BBOX.items():
        if la0 <= lat <= la1 and lo0 <= lon <= lo1: return p
    return None

# ── Pre-limpiar geografía (ONCE para todos los productos) ────────────────────
_cols_suc = ['id_comercio','id_bandera','id_sucursal',
             'sucursales_nombre','sucursales_latitud','sucursales_longitud',
             'sucursales_barrio','sucursales_localidad','PROVINCIA']
if 'sucursales_tipo' in suc_pais.columns:
    _cols_suc.append('sucursales_tipo')

suc_geo_clean = suc_pais[_cols_suc].copy()
if 'sucursales_tipo' not in suc_geo_clean.columns:
    suc_geo_clean['sucursales_tipo'] = 'N/D'
suc_geo_clean['cadena'] = suc_geo_clean.apply(asignar_cadena, axis=1)
suc_geo_clean['PROVINCIA_NORM'] = suc_geo_clean['PROVINCIA'].map(PROV_NORM).fillna(suc_geo_clean['PROVINCIA'])
suc_geo_clean = suc_geo_clean[suc_geo_clean['sucursales_tipo'] != 'Web'].copy()
_mask_caba = (
    suc_geo_clean['PROVINCIA_NORM'].eq('CABA') & (
        (suc_geo_clean['sucursales_latitud'] < -34.71) |
        (suc_geo_clean['sucursales_latitud'] > -34.53) |
        (suc_geo_clean['sucursales_longitud'] < -58.53) |
        (suc_geo_clean['sucursales_longitud'] > -58.34)))
suc_geo_clean = suc_geo_clean[~_mask_caba].copy()
_n_reclasif = 0
for _idx, _row in suc_geo_clean.iterrows():
    _p = _row['PROVINCIA_NORM']
    if _p not in _PROV_BBOX: continue
    _la0,_la1,_lo0,_lo1 = _PROV_BBOX[_p]
    _lat,_lon = _row['sucursales_latitud'],_row['sucursales_longitud']
    if _la0<=_lat<=_la1 and _lo0<=_lon<=_lo1: continue
    _nueva = _geocodif(_lat,_lon)
    if _nueva and _nueva != _p:
        suc_geo_clean.at[_idx,'PROVINCIA_NORM'] = _nueva
        _n_reclasif += 1
if _n_reclasif:
    print(f'  Reclasificadas {_n_reclasif} sucursales por coordenadas')

# ── Precio por sucursal para CADA producto activo (solo donde hay dato propio) ─
producto_geo_dict = {}

for _pid, _prod in PRODUCTOS.items():
    _name = PRODUCTO_NOMBRES[_pid]
    _ean  = next(iter(_prod))   # único EAN de este producto

    _suc  = (precio_mes[precio_mes['ean_norm'] == _ean]
             [['id_comercio','id_bandera','id_sucursal','precio']]
             .rename(columns={'precio': 'precio_producto'}))
    _pgeo = _suc.merge(suc_geo_clean, on=['id_comercio','id_bandera','id_sucursal'], how='inner')
    producto_geo_dict[_pid] = _pgeo.copy()
    if len(_pgeo) > 0:
        print(f'  [{_name}] {len(_pgeo):,} sucursales | '
              f'${_pgeo["precio_producto"].min():,.2f} – ${_pgeo["precio_producto"].max():,.2f}')
    else:
        print(f'  [{_name}] AVISO: sin sucursales con precio este mes (EAN {_ean})')

_prods_con_datos = [p for p in PRODUCTOS_ACTIVOS if len(producto_geo_dict[p]) > 0]
if not _prods_con_datos:
    raise RuntimeError('Ningún producto de EANS_INPUT tiene precios este mes. Revisá los códigos EAN.')

print()
print(f'Productos con datos este mes: {len(_prods_con_datos)} / {N_PRODUCTOS}')
print('Cadenas (primer producto con datos):')
print(producto_geo_dict[_prods_con_datos[0]]['cadena'].value_counts().to_string())"""))

# ── CELL 8 — PROVINCE ANALYSIS (multi-producto) ────────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 8 — Análisis provincial para cada producto activo
# ============================================================
def fmt_ar(x, dec=0):
    s = f'{x:,.{dec}f}'
    return s.replace(',','X').replace('.',',').replace('X','.')

# _tmean: media RECORTADA al 1% (descarta el 1% inferior y superior antes de
# promediar). Robusta a los outliers de carga del SEPA; se usa para TODAS las
# columnas 'media' nuevas. Se define aca porque la CELDA 8 es el primer consumidor
# (tambien la usa _leer_mes_hist en la CELDA 9, que corre despues).
def _tmean(_s, _p=0.01):
    _s = pd.to_numeric(_s, errors='coerce').dropna().sort_values()
    _n = len(_s)
    if _n == 0: return float('nan')
    _k = int(_n * _p)   # cuantos recortar de CADA extremo (0 si la muestra es chica)
    if _k > 0: _s = _s.iloc[_k:_n - _k]
    return _s.mean() if len(_s) else float('nan')

serie_prov_dict     = {}   # pid -> serie_provincia_valida (mediana + media por provincia)
prom_nac_dict       = {}   # pid -> precio nacional ponderado MEDIANA (NaN si sin datos)
prom_nac_media_dict = {}   # pid -> precio nacional ponderado MEDIA recortada (NaN si sin datos)

for _pid in PRODUCTOS_ACTIVOS:
    _pgeo = producto_geo_dict[_pid]
    if len(_pgeo) == 0:
        serie_prov_dict[_pid]     = pd.DataFrame(columns=['mes','provincia','precio_producto','precio_producto_media'])
        prom_nac_dict[_pid]       = float('nan')
        prom_nac_media_dict[_pid] = float('nan')
        continue
    _cpp  = (_pgeo.groupby('PROVINCIA_NORM')['precio_producto']
             .agg(precio_producto='median', precio_producto_media=_tmean).reset_index()
             .rename(columns={'PROVINCIA_NORM':'provincia'}))
    _cpp['mes']  = PERIODO
    _cpp['peso'] = _cpp['provincia'].map(PESOS_POBLACION).fillna(0)
    _pob = _cpp[_cpp['peso'] > 0]['peso'].sum()
    _prom = ((_cpp['precio_producto'] * _cpp['peso']).sum() / _pob
             if _pob > 0 else _cpp['precio_producto'].mean())
    _prom_media = ((_cpp['precio_producto_media'] * _cpp['peso']).sum() / _pob
                   if _pob > 0 else _cpp['precio_producto_media'].mean())
    serie_prov_dict[_pid]     = _cpp[['mes','provincia','precio_producto','precio_producto_media']].copy()
    prom_nac_dict[_pid]       = _prom
    prom_nac_media_dict[_pid] = _prom_media

print(f'=== CUADRO: Precio por provincia — {NOMBRE_MES_TITLE} ===')
for _pid in _prods_con_datos:
    _name = PRODUCTO_NOMBRES[_pid]
    _spv  = serie_prov_dict[_pid].sort_values('precio_producto')
    _prom = prom_nac_dict[_pid]
    print(f'\\n  ── {_name} ──')
    for _, r in _spv.iterrows():
        _vs = ((r['precio_producto'] / _prom) - 1) * 100
        print(f'  {r["provincia"]:<25} ${r["precio_producto"]:>10,.2f}  {_vs:+.2f}%')
    print(f'  {"Promedio (ponderado)":<25} ${_prom:>10,.2f}   0.00%')
print(f'\\nProductos con datos provinciales: {len(_prods_con_datos)} / {N_PRODUCTOS}')"""))

# ── CELL 9 — HISTORICAL SERIES (one raw cache, per-producto aggregation) ──────
cells.append(cell_code("""\
# ============================================================
# CELDA 9 — Serie histórica (cache de meses cerrados + mes en curso fresco)
# ============================================================
# Cache identificado por la unión de EANs activos (agregar/quitar EANs invalida
# el cache). IMPORTANTE: el cache guarda SOLO meses CERRADOS. El último mes
# disponible (en curso, crece día a día) se RELEE SIEMPRE fresco.
_cache_key  = hashlib.md5('|'.join(sorted(PRODUCTOS_EANS_NORM)).encode()).hexdigest()[:8]
# _v2m: esquema con precio_medio (media recortada 1%) ademas de precio_mediano.
# Al cambiar el nombre, la primera corrida reconstruye el cache con ambas medidas
# (los .parquet viejos sin precio_medio quedan huerfanos; se pueden borrar).
_cache_path = CACHE_DIR / f'hist_union_{_cache_key}_v2m.parquet'
_EANS_LECTURA = PRODUCTOS_EANS_NORM | REF_EANS_FACTOR   # + referencia, solo para el factor

# Mapa de meses disponibles -> (zip, archivos del mes)
_mapa_mes = {}
for _zip_path, _anio, _sem in detectar_semestres():
    for (_anio_m, _mes_m), _archs in archivos_por_mes(_zip_path).items():
        _lbl = f'{_anio_m}-{_mes_m:02d}'
        if _lbl >= MES_INICIO_HISTORICO:
            _mapa_mes[_lbl] = (_zip_path, _archs)
_meses_disp = sorted(_mapa_mes)
if not _meses_disp:
    raise RuntimeError(f'No hay meses disponibles >= {MES_INICIO_HISTORICO}')
_mes_actual = _meses_disp[-1]   # mes en curso -> NUNCA se cachea

def _leer_mes_hist(_lbl):
    # Lee los archivos del mes _lbl, filtra a los EANs pedidos (+ referencia
    # para el factor) y devuelve [ean_norm, precio_mediano, anio_mes] (mediana
    # sobre todos los días cargados de ese mes). Devuelve None si no hay datos.
    _zip_path, _archs = _mapa_mes[_lbl]
    _all_rows = []
    _muestra_ref = []
    for _archivo in sorted(_archs):
        _tmp_p = TMP_DIR / Path(_archivo).name
        with zipfile.ZipFile(_zip_path) as _zf:
            with _zf.open(_archivo) as _s, open(_tmp_p,'wb') as _d:
                shutil.copyfileobj(_s, _d, length=4*1024*1024)
        with gzip.open(_tmp_p,'rt',encoding='utf-8',errors='replace') as _g:
            for _chunk in pd.read_csv(_g, dtype=str, chunksize=300_000, low_memory=False):
                _chunk['ean_norm'] = _chunk['id_producto'].apply(normalizar_ean)
                _chunk = _chunk[_chunk['ean_norm'].isin(_EANS_LECTURA)].copy()
                if len(_chunk) == 0: continue
                _cols_p = [c for c in _chunk.columns if re.match(r'^precio_\\d{8}$', c)]
                if not _cols_p: continue
                _sub = _chunk[['ean_norm']+_cols_p].copy()
                for _cp in _cols_p:
                    _sub[_cp] = pd.to_numeric(_sub[_cp].replace('NA',np.nan), errors='coerce')
                _mlt = _sub.melt(id_vars='ean_norm', value_vars=_cols_p,
                                 var_name='_c', value_name='precio')
                _mlt = _mlt[_mlt['precio'].notna() & (_mlt['precio']>0)]
                _es_ref = _mlt['ean_norm'].isin(REF_EANS_FACTOR)
                if _es_ref.any():
                    _muestra_ref.extend(_mlt.loc[_es_ref, 'precio'].tolist())
                _mlt = _mlt[_mlt['ean_norm'].isin(PRODUCTOS_EANS_NORM)]
                if len(_mlt) > 0:
                    _all_rows.append(_mlt[['ean_norm','precio']])
        _tmp_p.unlink(missing_ok=True)
    if not _all_rows:
        return None
    _df_m = pd.concat(_all_rows, ignore_index=True)
    # Factor anclado a la referencia (robusto aunque se pidan 1-2 productos);
    # si la referencia no aparece ese mes, cae a la mediana de lo leído.
    _med_ref = (pd.Series(_muestra_ref).median() if _muestra_ref
                else _df_m['precio'].median())
    _fac = 100 if _med_ref > 10_000 else 1
    if _fac == 100: _df_m['precio'] /= 100
    _agg = (_df_m.groupby('ean_norm')['precio']
            .agg(precio_mediano='median', precio_medio=_tmean).reset_index())
    _agg['anio_mes'] = _lbl
    del _df_m, _all_rows, _muestra_ref; gc.collect()
    return _agg

# ── 1) Meses CERRADOS (todos menos el último): usar/actualizar cache ──────────
if USE_CACHE and _cache_path.exists():
    df_cache = pd.read_parquet(_cache_path)
    df_cache = df_cache[df_cache['anio_mes'] < _mes_actual].copy()
else:
    df_cache = pd.DataFrame(columns=['ean_norm','precio_mediano','precio_medio','anio_mes'])

_meses_en_cache = set(df_cache['anio_mes'].unique())
_faltantes = [m for m in _meses_disp if m < _mes_actual and m not in _meses_en_cache]
_nuevos = []
for _lbl in _faltantes:
    _agg = _leer_mes_hist(_lbl)
    if _agg is not None:
        _nuevos.append(_agg)
        print(f'  {_lbl}: {_agg["ean_norm"].nunique()} EANs (agregado al cache)')
if _nuevos:
    df_cache = pd.concat([df_cache] + _nuevos, ignore_index=True)
    if USE_CACHE:
        df_cache.to_parquet(_cache_path, compression='snappy', index=False)
        print(f'Cache actualizado: {_cache_path.name} ({df_cache["anio_mes"].nunique()} meses cerrados)')
elif USE_CACHE and _cache_path.exists():
    print(f'Cache al día: {_cache_path.name} ({len(_meses_en_cache)} meses cerrados)')

# ── 2) Mes en curso: SIEMPRE fresco (usa los días cargados a la fecha) ────────
_actual = _leer_mes_hist(_mes_actual)
if _actual is not None:
    print(f'  {_mes_actual}: {_actual["ean_norm"].nunique()} EANs (mes en curso — recalculado fresco)')

df_hist_raw = (pd.concat([df_cache] + ([_actual] if _actual is not None else []),
                         ignore_index=True)
               if (len(df_cache) > 0 or _actual is not None)
               else pd.DataFrame(columns=['ean_norm','precio_mediano','precio_medio','anio_mes']))
df_hist_raw = df_hist_raw.sort_values('anio_mes').reset_index(drop=True)

# ── Serie histórica por producto (un único EAN cada uno) ─────────────────────
serie_nac_dict = {}   # pid -> serie_nacional_valida (mes, precio_nacional_ponderado, ...)

for _pid, _prod in PRODUCTOS.items():
    _name = PRODUCTO_NOMBRES[_pid]
    _eans = set(_prod.keys())
    _dh   = df_hist_raw[df_hist_raw['ean_norm'].isin(_eans)].copy()
    if len(_dh) == 0:
        # Sin NINGÚN dato histórico para este EAN (típico si está mal tipeado o
        # nunca se vendió en el período). Se arma vacío con dtypes explícitos:
        # dejar que .map()/.round() corran sobre un DataFrame vacío de origen
        # produce columnas dtype=object y .round() falla más abajo.
        serie_nac_dict[_pid] = pd.DataFrame(columns=[
            'mes','precio_nacional_ponderado','precio_nacional_ponderado_media','n_eans',
            'variacion_mensual_%','variacion_mensual_media_%','indice_precio_base100'])
        print(f'  [{_name}] 0 meses | sin datos históricos (revisar EAN)')
        continue
    _dh['qty']              = _dh['ean_norm'].map(lambda e, c=_prod: c.get(e,('?',0,'?'))[1]).astype(float)
    _dh['costo_item']       = _dh['precio_mediano'].astype(float) * _dh['qty']
    _dh['costo_item_media'] = _dh['precio_medio'].astype(float)   * _dh['qty']
    _sn = (_dh.groupby('anio_mes')
           .agg(precio_nacional_ponderado=('costo_item','sum'),
                precio_nacional_ponderado_media=('costo_item_media','sum'),
                n_eans=('ean_norm','nunique'))
           .reset_index().rename(columns={'anio_mes':'mes'})
           .sort_values('mes').reset_index(drop=True))
    _sn = _sn[_sn['mes'] >= MES_INICIO_HISTORICO].copy()
    _sn['variacion_mensual_%']       = _sn['precio_nacional_ponderado'].pct_change() * 100
    _sn['variacion_mensual_media_%'] = _sn['precio_nacional_ponderado_media'].pct_change() * 100
    _bv = _sn['precio_nacional_ponderado'].iloc[0] if len(_sn) > 0 else 1
    _sn['indice_precio_base100'] = (_sn['precio_nacional_ponderado'] / _bv * 100).round(2)
    serie_nac_dict[_pid] = _sn.copy()
    _rng = f'{_sn["mes"].min()} -> {_sn["mes"].max()}' if len(_sn) > 0 else 'sin datos'
    print(f'  [{_name}] {len(_sn)} meses | {_rng}')"""))

# ── CELL 10 — IPC ──────────────────────────────────────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 10 — IPC INDEC desde carga/IPC.xlsx
# ============================================================
if not IPC_PATH.exists():
    raise FileNotFoundError(
        f'IPC.xlsx no encontrado en {SEPA_DIR}\\n'
        'Asegurate de tener el archivo IPC.xlsx en la carpeta carga/'
    )

ipc_raw = pd.read_excel(IPC_PATH)
print(f'IPC cargado: {len(ipc_raw)} filas, columnas: {list(ipc_raw.columns[:4])} ...')

fecha_col = next((c for c in ipc_raw.columns
                  if str(c).lower().strip() in ('date','fecha','mes','period')),
                 ipc_raw.columns[0])

if pd.api.types.is_datetime64_any_dtype(ipc_raw[fecha_col]):
    ipc_raw['mes'] = ipc_raw[fecha_col].dt.strftime('%Y-%m')
else:
    _MES_ESP = {'ene':1,'feb':2,'mar':3,'abr':4,'may':5,'jun':6,
                'jul':7,'ago':8,'sep':9,'oct':10,'nov':11,'dic':12}
    def _parse_ipc_fecha(val):
        if pd.isna(val): return pd.NaT
        if isinstance(val, pd.Timestamp): return val
        s = str(val).strip().lower()
        try:
            partes = s.split('-')
            if len(partes) == 2 and partes[0] in _MES_ESP:
                return pd.Timestamp(year=int(partes[1]), month=_MES_ESP[partes[0]], day=1)
        except Exception:
            pass
        return pd.to_datetime(val, errors='coerce')
    ipc_raw['mes'] = ipc_raw[fecha_col].apply(_parse_ipc_fecha).dt.strftime('%Y-%m')

rename_map = {}
for c in ipc_raw.columns:
    cs = str(c).strip()
    if 'nivel general' in cs.lower():               rename_map[c] = 'ipc_general'
    elif 'alimentos y bebidas no alc' in cs.lower(): rename_map[c] = 'ipc_alimentos'
ipc_raw = ipc_raw.rename(columns=rename_map)

for c in ['ipc_general','ipc_alimentos']:
    if c in ipc_raw.columns:
        ipc_raw[c] = pd.to_numeric(
            ipc_raw[c].astype(str).str.replace(',', '.', regex=False), errors='coerce')

_ipc_cols = ['mes','ipc_general'] + (['ipc_alimentos'] if 'ipc_alimentos' in ipc_raw.columns else [])
ipc = (ipc_raw[_ipc_cols]
       .dropna(subset=['ipc_general']).sort_values('mes').reset_index(drop=True))
if 'ipc_alimentos' not in ipc.columns:
    ipc['ipc_alimentos'] = np.nan
ipc['ipc_general_var_%']   = ipc['ipc_general'].pct_change(fill_method=None) * 100
ipc['ipc_alimentos_var_%'] = ipc['ipc_alimentos'].pct_change(fill_method=None) * 100

print(f'IPC procesado: {len(ipc)} meses | {ipc["mes"].min()} -> {ipc["mes"].max()}')
print(ipc[['mes','ipc_general','ipc_general_var_%','ipc_alimentos','ipc_alimentos_var_%']].tail(6).to_string(index=False))"""))

# ── CELL 11 — COMPARATIVA (multi-producto) ─────────────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 11 — Comparativa SEPA vs IPC para cada producto
# ============================================================
comparativa_dict  = {}   # pid -> comparativa DataFrame
df_g_dict         = {}   # pid -> df_g (desde MES_INICIO_GRAFICO)
_lbl_base_dict    = {}   # pid -> label base ej '03-24'
_serie_vacia_dict = {}   # pid -> bool

for _pid in PRODUCTOS_ACTIVOS:
    _sn   = serie_nac_dict[_pid]
    _name = PRODUCTO_NOMBRES[_pid]

    if len(_sn) == 0:
        print(f'AVISO [{_name}]: Serie histórica vacía — gráficos de índices no disponibles.')
        _serie_vacia_dict[_pid] = True
        comparativa_dict[_pid]  = pd.DataFrame()
        df_g_dict[_pid]         = pd.DataFrame()
        _lbl_base_dict[_pid]    = PERIODO[5:7] + '-' + PERIODO[2:4]
        continue

    _comp = _sn.merge(
        ipc[['mes','ipc_general','ipc_general_var_%','ipc_alimentos','ipc_alimentos_var_%']],
        on='mes', how='left')
    _ipc_b = _comp['ipc_general'].dropna().iloc[0] if _comp['ipc_general'].notna().any() else 1
    _ialb  = _comp['ipc_alimentos'].dropna().iloc[0] if _comp['ipc_alimentos'].notna().any() else 1
    _comp['indice_ipc_general_base100']   = (_comp['ipc_general']   / _ipc_b * 100).round(2)
    _comp['indice_ipc_alimentos_base100'] = (_comp['ipc_alimentos'] / _ialb  * 100).round(2)

    _mg = MES_INICIO_GRAFICO
    if _mg not in _comp['mes'].values:
        _mg = _comp['mes'].min()
    _dg = _comp[_comp['mes'] >= _mg].copy().reset_index(drop=True)
    _bg  = _dg['precio_nacional_ponderado'].iloc[0]
    _big = _dg['ipc_general'].dropna().iloc[0] if _dg['ipc_general'].notna().any() else 1
    _bia = _dg['ipc_alimentos'].dropna().iloc[0] if _dg['ipc_alimentos'].notna().any() else 1
    _lbl = _mg[5:7] + '-' + _mg[2:4]
    _dg['idx_producto_base']      = (_dg['precio_nacional_ponderado'] / _bg  * 100).round(2)
    _dg['idx_ipc_general_base']   = (_dg['ipc_general']               / _big * 100).round(2)
    _dg['idx_ipc_alimentos_base'] = (_dg['ipc_alimentos']             / _bia * 100).round(2)
    _dg['fecha'] = pd.to_datetime(_dg['mes'] + '-01')

    comparativa_dict[_pid]  = _comp
    df_g_dict[_pid]         = _dg
    _lbl_base_dict[_pid]    = _lbl
    _serie_vacia_dict[_pid] = False

    print(f'  [{_name}] {len(_comp)} meses | desde {_mg} ({len(_dg)} pts) | '
          f'último: ${_dg["precio_nacional_ponderado"].iloc[-1]:,.2f} '
          f'({_dg["variacion_mensual_%"].iloc[-1]:+.1f}%)')"""))

# ── CELL 12 — CHARTS (multi-producto) ──────────────────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 12 — Gráficos: índices y variaciones (todos los productos)
# ============================================================
COLOR_IPC_GEN = '#D62728'
COLOR_IPC_ALI = '#FF7F0E'
out1 = out2 = None

_MESES_ES = {1:'ene',2:'feb',3:'mar',4:'abr',5:'may',6:'jun',
             7:'jul',8:'ago',9:'sep',10:'oct',11:'nov',12:'dic'}
def _fmt_mes_es(x, pos):
    try:
        ts = mdates.num2date(x)
        return f'{_MESES_ES[ts.month]}-{str(ts.year)[2:]}'
    except Exception:
        return ''

# Productos con datos
_activos_con_datos = [p for p in PRODUCTOS_ACTIVOS if not _serie_vacia_dict[p] and len(df_g_dict[p]) > 0]
if not _activos_con_datos:
    print('AVISO: Sin serie histórica para ningún producto. Saltando gráficos de índices.')
else:
    # Usar el lbl_base del primer producto activo con datos
    _lbl_base = _lbl_base_dict[_activos_con_datos[0]]

    # ── GRAFICO 1: Índices base ─────────────────────────────────────────────
    fig1, ax1 = plt.subplots(figsize=(13, 6))
    for _pid in _activos_con_datos:
        _dg   = df_g_dict[_pid]
        _name = PRODUCTO_NOMBRES[_pid]
        ax1.plot(_dg['fecha'], _dg['idx_producto_base'],
                 color=PRODUCTO_COLORS[_pid], linewidth=2.5,
                 linestyle=PRODUCTO_LINESTYLES[_pid],
                 marker=PRODUCTO_MARKERS[_pid], markersize=5,
                 label=_name)
        _ult = _dg.iloc[-1]
        ax1.annotate(f"{_ult['idx_producto_base']:.0f}",
                     xy=(_ult['fecha'], _ult['idx_producto_base']),
                     xytext=(8, 0), textcoords='offset points',
                     color=PRODUCTO_COLORS[_pid], fontweight='bold', fontsize=9)
    _dg0 = df_g_dict[_activos_con_datos[0]]
    if _dg0['idx_ipc_general_base'].notna().any():
        ax1.plot(_dg0['fecha'], _dg0['idx_ipc_general_base'],
                 color=COLOR_IPC_GEN, linewidth=1.8, linestyle='--', marker='s', markersize=4,
                 label='IPC INDEC - Nivel general')
        _iu = _dg0.dropna(subset=['idx_ipc_general_base']).iloc[-1]
        ax1.annotate(f"{_iu['idx_ipc_general_base']:.0f}",
                     xy=(_iu['fecha'], _iu['idx_ipc_general_base']),
                     xytext=(8,-3), textcoords='offset points',
                     color=COLOR_IPC_GEN, fontweight='bold', fontsize=9)
    if _dg0['idx_ipc_alimentos_base'].notna().any():
        ax1.plot(_dg0['fecha'], _dg0['idx_ipc_alimentos_base'],
                 color=COLOR_IPC_ALI, linewidth=1.8, linestyle=':', marker='^', markersize=4,
                 label='IPC INDEC - Alimentos y bebidas')
        _ia = _dg0.dropna(subset=['idx_ipc_alimentos_base']).iloc[-1]
        ax1.annotate(f"{_ia['idx_ipc_alimentos_base']:.0f}",
                     xy=(_ia['fecha'], _ia['idx_ipc_alimentos_base']),
                     xytext=(8,3), textcoords='offset points',
                     color=COLOR_IPC_ALI, fontweight='bold', fontsize=9)
    ax1.set_ylabel(f'Índice ({_lbl_base} = 100)', fontsize=11)
    ax1.legend(loc='upper left', fontsize=8, framealpha=0.95, ncol=max(1, len(_activos_con_datos)//10+1))
    ax1.grid(True, alpha=0.3)
    ax1.xaxis.set_major_locator(mdates.MonthLocator(interval=1))
    ax1.xaxis.set_major_formatter(mticker.FuncFormatter(_fmt_mes_es))
    plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45, ha='right')
    if MES_PARCIAL:
        ax1.set_title(f'Último mes ({NOMBRE_MES_TITLE}) PARCIAL — {DIAS_CARGADOS}/{DIAS_MES} días · última variación preliminar',
                      fontsize=10, color='#C00000', style='italic', pad=8)
    plt.tight_layout()
    out1 = OUTPUT_DIR / f'indices_productos_vs_ipc_{MES}.png'
    plt.savefig(out1, dpi=600, bbox_inches='tight', facecolor='white')
    plt.show()
    print(f'Gráfico 1 guardado: {out1}')

    # ── GRAFICO 2: Variaciones mensuales (solo barras verticales) ───────────────
    # Productos activos + IPC General + IPC Alimentos como barras agrupadas
    _series_bar = (
        [(PRODUCTO_COLORS[p], PRODUCTO_NOMBRES[p], df_g_dict[p]['variacion_mensual_%'])
         for p in _activos_con_datos] +
        [(COLOR_IPC_GEN, 'IPC INDEC - Nivel general', _dg0['ipc_general_var_%']),
         (COLOR_IPC_ALI, 'IPC INDEC - Alimentos y bebidas', _dg0['ipc_alimentos_var_%'])]
    )
    _n_b    = len(_series_bar)
    # Ancho y figura adaptados al número de series: más series → más ancho, barras más anchas
    _fig_w  = max(20, _n_b * 2 + 10)
    _bw2    = pd.Timedelta(days=max(1, int(22 / max(_n_b, 1))))
    _offs2  = [(_i - (_n_b - 1) / 2) * _bw2 for _i in range(_n_b)]
    _tick_i = 2 if _n_b > 5 else 1   # ticks cada 2 meses con muchas series
    fig2, ax2 = plt.subplots(figsize=(_fig_w, 8))
    for _i, (_col, _lbl, _vals) in enumerate(_series_bar):
        if _vals.notna().any():
            _alpha = 0.88 if _i < len(_activos_con_datos) else 0.72
            ax2.bar(_dg0['fecha'] + _offs2[_i], _vals,
                    width=_bw2, color=_col, alpha=_alpha, label=_lbl, edgecolor='none')
    ax2.axhline(0, color='#444444', linewidth=0.8)
    ax2.set_ylabel('Variación mensual (%)', fontsize=11)
    ax2.legend(loc='upper right', fontsize=8, framealpha=0.95,
               ncol=max(1, _n_b // 4 + 1))
    ax2.grid(True, alpha=0.2, axis='y', linewidth=0.7)
    ax2.xaxis.set_major_locator(mdates.MonthLocator(interval=_tick_i))
    ax2.xaxis.set_major_formatter(mticker.FuncFormatter(_fmt_mes_es))
    plt.setp(ax2.xaxis.get_majorticklabels(), rotation=45, ha='right', fontsize=9)
    for sp in ['top', 'right']: ax2.spines[sp].set_visible(False)
    ax2.spines['bottom'].set_color('#cccccc')
    ax2.spines['left'].set_color('#cccccc')
    _vmax_l2 = [s.dropna().max() for _,_,s in _series_bar if s.notna().any()]
    if _vmax_l2: ax2.set_ylim(top=max(_vmax_l2) * 1.35)
    if MES_PARCIAL:
        ax2.set_title(f'Último mes ({NOMBRE_MES_TITLE}) PARCIAL — {DIAS_CARGADOS}/{DIAS_MES} días · variación preliminar (subestimada)',
                      fontsize=10, color='#C00000', style='italic', pad=8)
    plt.tight_layout()
    out2 = OUTPUT_DIR / f'variaciones_productos_vs_ipc_{MES}.png'
    plt.savefig(out2, dpi=600, bbox_inches='tight', facecolor='white')
    plt.show()
    print(f'Gráfico 2 guardado: {out2}')

    # ── GRAFICO 3: Ranking de precios absolutos por producto ────────────────────
    _abs_data = sorted(
        [(PRODUCTO_NOMBRES[p], prom_nac_dict[p], PRODUCTO_COLORS[p])
         for p in PRODUCTOS_ACTIVOS if not pd.isna(prom_nac_dict[p])],
        key=lambda x: x[1])
    _pnames  = [d[0] for d in _abs_data]
    _pvals   = [d[1] for d in _abs_data]
    _pcolors = [d[2] for d in _abs_data]
    _base_v  = _pvals[0] if _pvals else 1   # producto más barato como referencia
    _n_bars  = len(_pvals)
    fig3, ax3 = plt.subplots(figsize=(12, max(5, _n_bars * 0.5 + 2)))
    bars3 = ax3.barh(_pnames, _pvals, color=_pcolors,
                     edgecolor='none', height=0.55, zorder=2)
    ax3.barh(_pnames, _pvals, color='black', alpha=0.06,
             height=0.60, zorder=1)
    ax3.axvline(_base_v, color='#aaaaaa', linewidth=1.2, linestyle='--', zorder=3)
    for bar, val, name in zip(bars3, _pvals, _pnames):
        _ratio = val / _base_v if _base_v > 0 else 1
        _ratio_str = f'  ×{_ratio:.1f}' if _ratio > 1.05 else '  base'
        ax3.text(val + max(_pvals) * 0.008,
                 bar.get_y() + bar.get_height() / 2,
                 f'${val:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.') + _ratio_str,
                 va='center', fontsize=10, fontweight='bold',
                 color='#2c3e50')
    ax3.set_yticklabels(_pnames, fontsize=10, fontweight='bold')
    ax3.set_xlabel('Precio promedio nacional (ARS)', fontsize=11, color='#444')
    ax3.xaxis.set_major_formatter(mticker.FuncFormatter(
        lambda x, _: f'${x:,.0f}'.replace(',', '.')))
    ax3.set_xlim(0, max(_pvals) * 1.25)
    ax3.tick_params(axis='x', labelsize=10, colors='#555')
    ax3.grid(True, alpha=0.25, axis='x', zorder=0); ax3.set_axisbelow(True)
    for sp in ['top', 'right', 'left']: ax3.spines[sp].set_visible(False)
    ax3.spines['bottom'].set_color('#cccccc')
    fig3.text(0.98, 0.01, f'{NOMBRE_MES_TITLE}{SUFIJO_PARCIAL}',
              ha='right', va='bottom', fontsize=9, color=('#C00000' if MES_PARCIAL else '#999'))
    plt.tight_layout()
    out3 = OUTPUT_DIR / f'ranking_productos_{MES}.png'
    plt.savefig(out3, dpi=600, bbox_inches='tight', facecolor='white')
    plt.show()
    print(f'Gráfico 3 guardado: {out3}')"""))

# ── CELL 13 — CUADRO 1 + LaTeX (per-producto) ──────────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 13 — Cuadro 1 provincial + LaTeX para cada producto
# ============================================================
def fmt_ar(x, dec=0):
    s = f'{x:,.{dec}f}'
    return s.replace(',','X').replace('.',',').replace('X','.')

nom_mes = {'01':'enero','02':'febrero','03':'marzo','04':'abril','05':'mayo','06':'junio',
           '07':'julio','08':'agosto','09':'septiembre','10':'octubre','11':'noviembre','12':'diciembre'}
_mes_s  = nom_mes[ULTIMO_MES[5:7]]
_anio_s = ULTIMO_MES[:4]

for _pid in _prods_con_datos:
    _name  = PRODUCTO_NOMBRES[_pid]
    _short = PRODUCTO_SHORT[_pid]
    _spv   = serie_prov_dict[_pid].copy()
    _prom  = prom_nac_dict[_pid]
    if len(_spv) == 0 or pd.isna(_prom):
        print(f'  [{_name}] Sin datos provinciales — saltear cuadro/LaTeX'); continue
    _spv['vs_%'] = ((_spv['precio_producto'] / _prom) - 1) * 100
    _spv = _spv.sort_values('precio_producto').reset_index(drop=True)

    print(f'\\n=== CUADRO 1: {_name.upper()} — {NOMBRE_MES_TITLE} ===\\n')
    print(f'{"Provincia":<25} {"Precio":>14} {"Vs. promedio":>14}')
    print('-'*55)
    for _, r in _spv.iterrows():
        _c = fmt_ar(r['precio_producto'], dec=2)
        _v = f"{r['vs_%']:+.2f}%".replace('.',',')
        print(f"{r['provincia']:<25} {_c:>14} {_v:>14}")
    print('-'*55)
    print(f'{"Promedio nacional":<25} {fmt_ar(_prom, dec=2):>14} {"0,00%":>14}')

    # Todas las filas se arman como f-strings (mismo nivel de escape en todo el
    # bloque, incluido el encabezado) para no mezclar convenciones de backslash.
    ltx = [
        f'\\\\begin{{table}}[H]',
        f'\\\\centering',
        f'\\\\renewcommand{{\\\\arraystretch}}{{1.15}}',
        f'\\\\caption{{{_name} por provincia ({_mes_s} {_anio_s})}}',
        f'\\\\begin{{tabular}}{{@{{}}l r r@{{}}}}',
        f'\\\\toprule',
        f'\\\\textbf{{Provincia}} & \\\\textbf{{Precio}} & \\\\shortstack{{\\\\textbf{{Vs. promedio}}\\\\\\\\\\\\textbf{{pais (\\\\%)}}}} \\\\\\\\',
        f'\\\\midrule',
    ]
    for _, r in _spv.iterrows():
        _c = fmt_ar(r['precio_producto'], dec=2)
        _v = f"{r['vs_%']:+.2f}".replace('.',',')
        ltx.append(f"{r['provincia']:<22} & {_c} & {_v}\\\\% \\\\\\\\")
    ltx += [
        f'\\\\midrule',
        f'\\\\textbf{{Promedio}} & {fmt_ar(_prom, dec=2)} & 0,00\\\\% \\\\\\\\',
        f'\\\\bottomrule',
        f'\\\\end{{tabular}}\\\\\\\\[0.2cm]',
        f'\\\\caption*{{Fuente: Elaboracion propia en base a SEPA}}',
        f'\\\\label{{tab:producto_{_short}_{ULTIMO_MES}}}',
        f'\\\\end{{table}}',
    ]
    _latex_out = '\\n'.join(ltx)
    _out_tex = OUTPUT_DIR / f'tabla_producto_{_short}_{ULTIMO_MES}.tex'
    _out_tex.write_text(_latex_out, encoding='utf-8')
    print(f'  LaTeX guardado: {_out_tex.name}')"""))

# ── CELL 14 — CHOROPLETH MAPS (one per producto) ────────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 14 — Mapa coroplético por producto
# ============================================================
if not GEOJSON_PATH.exists():
    print(f'GeoJSON no encontrado en {GEOJSON_PATH} — saltear celda')
else:
    with open(GEOJSON_PATH, 'r', encoding='utf-8') as f:
        geo = _json.load(f)

    NORM_GEO = {'Ciudad de Buenos Aires': 'CABA'}
    AJUST = {
        'Salta':(0,-1),'Tucuman':(0.3,0),'Tucumán':(0.3,0),'Chaco':(0,-1),
        'Tierra del Fuego':(-1,-0.2),'Santa Fe':(0,1),'Santiago del Estero':(0.7,0),
    }

    def centroide(coords):
        xs,ys = [],[]
        if isinstance(coords[0][0][0],(int,float)):
            for p in coords[0]: xs.append(p[0]); ys.append(p[1])
        else:
            poly = max(coords, key=lambda p: len(p[0]))
            for p in poly[0]: xs.append(p[0]); ys.append(p[1])
        return sum(xs)/len(xs), sum(ys)/len(ys)

    def draw(ax, coords, color):
        if isinstance(coords[0][0][0],(int,float)):
            ax.fill([c[0] for c in coords[0]], [c[1] for c in coords[0]],
                    facecolor=color, edgecolor='white', linewidth=0.6)
        else:
            for poly in coords:
                ax.fill([c[0] for c in poly[0]], [c[1] for c in poly[0]],
                        facecolor=color, edgecolor='white', linewidth=0.6)

    cmap_m = LinearSegmentedColormap.from_list('c',
        ['#1a9850','#66bd63','#a6d96a','#d9ef8b','#fee08b','#fdae61','#f46d43','#d73027'], N=256)

    for _pid in _prods_con_datos:
        _name   = PRODUCTO_NOMBRES[_pid]
        _short  = PRODUCTO_SHORT[_pid]
        _spv    = serie_prov_dict[_pid]
        if len(_spv) == 0:
            print(f'  [{_name}] Sin datos provinciales — saltear mapa')
            continue
        _prod_prov = dict(zip(_spv['provincia'], _spv['precio_producto']))
        _vals = list(_prod_prov.values())
        norm_c = Normalize(vmin=min(_vals), vmax=max(_vals))

        fig, ax = plt.subplots(figsize=(12, 16))
        caba_c = None
        for feat in geo['features']:
            ng  = feat['properties']['name']
            nom = NORM_GEO.get(ng, ng)
            val = _prod_prov.get(nom)
            col = cmap_m(norm_c(val)) if val is not None else '#dddddd'
            gt  = feat['geometry']['type']
            co  = feat['geometry']['coordinates']
            draw(ax, [co] if gt=='Polygon' else co, col)
            cx, cy = centroide([co] if gt=='Polygon' else co)
            if nom == 'CABA':
                caba_c = (cx, cy); continue
            dx, dy = AJUST.get(nom, (0,0))
            if val is not None:
                ax.text(cx+dx, cy+dy, f'{nom}\\n${val:,.0f}',
                        ha='center', va='center', fontsize=7.5, fontweight='bold',
                        bbox=dict(boxstyle='round,pad=0.25', facecolor='white', alpha=0.75, edgecolor='none'))
        if caba_c and 'CABA' in _prod_prov:
            vc = _prod_prov['CABA']
            cc = cmap_m(norm_c(vc))
            lx, ly = caba_c[0]+2.2, caba_c[1]+0.8
            ax.annotate('', xy=caba_c, xytext=(lx,ly),
                        arrowprops=dict(arrowstyle='-', color='black', linewidth=1.0))
            ax.text(lx, ly, f'CABA\\n${vc:,.0f}',
                    ha='center', va='center', fontsize=9, fontweight='bold',
                    bbox=dict(boxstyle='round,pad=0.5', facecolor=cc, alpha=0.95,
                              edgecolor='black', linewidth=1.0))
            ax.plot(*caba_c, marker='o', markersize=10, markerfacecolor=cc,
                    markeredgecolor='black', markeredgewidth=1.2, zorder=5)
        ax.set_aspect('equal'); ax.axis('off')
        plt.tight_layout()
        _out_m = OUTPUT_DIR / f'mapa_producto_{_short}_{ULTIMO_MES}.png'
        plt.savefig(_out_m, dpi=600, bbox_inches='tight', facecolor='white')
        plt.show()
        print(f'  Mapa [{_name}] guardado: {_out_m.name}')"""))

# ── CELL 15 — COVERAGE ─────────────────────────────────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 15 — Gráficos de cobertura (todos los productos combinados)
# ============================================================
# A diferencia de nb02 (que usa la primera canasta como referencia), acá cada
# producto es independiente, así que la cobertura se calcula sobre la UNIÓN de
# todos los productos consultados con datos este mes.
_pm_info = (suc_geo_clean[['id_comercio','id_bandera','id_sucursal',
                           'PROVINCIA_NORM','cadena']].drop_duplicates())
_pm_geo  = precio_mes.merge(_pm_info, on=['id_comercio','id_bandera','id_sucursal'], how='inner')
_pm_geo  = _pm_geo[_pm_geo['ean_norm'].isin(PRODUCTOS_EANS_NORM)].copy()
_pm_geo['suc_key'] = (_pm_geo['id_comercio'] + '_' +
                      _pm_geo['id_bandera']   + '_' + _pm_geo['id_sucursal'])

cob_provincia = (_pm_geo.groupby('PROVINCIA_NORM')
    .agg(n_productos_unicos=('ean_norm','nunique'),
         n_cadenas=('cadena','nunique'),
         n_sucursales=('suc_key','nunique'))
    .reset_index().rename(columns={'PROVINCIA_NORM':'provincia'}))

cob_cadena = (_pm_geo.groupby('cadena')
    .agg(n_productos_unicos=('ean_norm','nunique'),
         n_provincias=('PROVINCIA_NORM','nunique'),
         n_sucursales=('suc_key','nunique'))
    .reset_index())

matriz_cad_prov = (_pm_geo.groupby(['cadena','PROVINCIA_NORM'])['ean_norm']
    .nunique().unstack(fill_value=0))

df_p = cob_provincia.sort_values('n_productos_unicos', ascending=True).reset_index(drop=True)
fig, axes = plt.subplots(1, 3, figsize=(15, max(9, len(df_p)*0.38+2)), sharey=True)
for ax, (col, titulo, color) in zip(axes, [
        ('n_productos_unicos','Productos únicos','#0055A4'),
        ('n_cadenas','Cadenas presentes','#27ae60'),
        ('n_sucursales','Sucursales','#c0392b')]):
    ax.barh(df_p['provincia'], df_p[col], color=color, edgecolor='white')
    for i, v in enumerate(df_p[col]):
        ax.text(v+max(df_p[col])*0.01, i, f'{int(v):,}', va='center', fontsize=8, color='#2c3e50')
    ax.set_title(titulo, fontsize=11, fontweight='bold', color=color)
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_: f'{int(x):,}'))
    ax.set_xlim(0, max(df_p[col])*1.15)
    for sp in ['top','right']: ax.spines[sp].set_visible(False)
plt.tight_layout()
out_cp = OUTPUT_DIR / f'cobertura_provincia_{MES}.png'
plt.savefig(out_cp, dpi=600, bbox_inches='tight'); plt.show()

df_c = cob_cadena.sort_values('n_productos_unicos', ascending=True).reset_index(drop=True)
fig, axes = plt.subplots(1, 3, figsize=(15, max(8, len(df_c)*0.5+2)), sharey=True)
for ax, (col, titulo, color) in zip(axes, [
        ('n_productos_unicos','Productos únicos','#0055A4'),
        ('n_provincias','Provincias presentes','#27ae60'),
        ('n_sucursales','Sucursales','#c0392b')]):
    ax.barh(df_c['cadena'], df_c[col], color=color, edgecolor='white')
    for i, v in enumerate(df_c[col]):
        ax.text(v+max(df_c[col])*0.01, i, f'{int(v):,}', va='center', fontsize=8, color='#2c3e50')
    ax.set_title(titulo, fontsize=11, fontweight='bold', color=color)
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_: f'{int(x):,}'))
    ax.set_xlim(0, max(df_c[col])*1.15)
    for sp in ['top','right']: ax.spines[sp].set_visible(False)
axes[0].tick_params(axis='y', labelsize=8)
plt.tight_layout()
out_cc = OUTPUT_DIR / f'cobertura_cadena_{MES}.png'
plt.savefig(out_cc, dpi=600, bbox_inches='tight'); plt.show()

df_m = matriz_cad_prov.copy()
df_m['_total'] = df_m.sum(axis=1)
df_m = df_m.sort_values('_total', ascending=False).drop(columns='_total')
orden_prov = (df_m > 0).sum(axis=0).sort_values(ascending=False).index
df_m = df_m[orden_prov]

fig, ax = plt.subplots(figsize=(13, max(6, len(df_m)*0.55+2)))
binaria = (df_m > 0).astype(int)
sns.heatmap(binaria, cmap='Blues', cbar=False,
            linewidths=0.5, linecolor='white', xticklabels=True, yticklabels=True, ax=ax)
for i in range(binaria.shape[0]):
    for j in range(binaria.shape[1]):
        if binaria.iat[i, j]:
            ax.text(j+0.5, i+0.5, '●', ha='center', va='center', color='white', fontsize=9)
ax.set_xlabel(''); ax.set_ylabel('')
plt.xticks(rotation=45, ha='right', fontsize=9); plt.yticks(rotation=0, fontsize=9)
plt.tight_layout()
out_mp = OUTPUT_DIR / f'matriz_presencia_{MES}.png'
plt.savefig(out_mp, dpi=600, bbox_inches='tight'); plt.show()

fig, ax = plt.subplots(figsize=(13, max(6, len(df_m)*0.55+2)))
data_log = np.log10(df_m.replace(0, np.nan))
sns.heatmap(data_log, cmap='YlOrRd', linewidths=0.5, linecolor='white',
            cbar_kws={'label': 'log10(productos unicos)'},
            xticklabels=True, yticklabels=True, ax=ax)
ax.set_xlabel(''); ax.set_ylabel('')
plt.xticks(rotation=45, ha='right', fontsize=9); plt.yticks(rotation=0, fontsize=9)
plt.tight_layout()
out_mi = OUTPUT_DIR / f'matriz_intensidad_{MES}.png'
plt.savefig(out_mi, dpi=600, bbox_inches='tight'); plt.show()
print('Gráficos cobertura guardados')"""))

# ── CELL 16 — RANKINGS (per-producto) ──────────────────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 16 — Rankings de cadenas: una por producto activo
# ============================================================
def fmtn(x): return f'{x:,.2f}'.replace(',','X').replace('.',',').replace('X','.')

CADENAS_EXCLUIR_RANKING = {'Cadena 23'}   # se saca de estos rankings (y de su promedio)

for _pid in _prods_con_datos:
    _name  = PRODUCTO_NOMBRES[_pid]
    _short = PRODUCTO_SHORT[_pid]
    _pgeo  = producto_geo_dict[_pid]
    _pgeo  = _pgeo[~_pgeo['cadena'].isin(CADENAS_EXCLUIR_RANKING)]

    _rk_nac = (_pgeo.groupby('cadena')
               .agg(n_sucursales=('precio_producto','count'),
                    precio_promedio=('precio_producto','mean'))
               .round(2).reset_index())
    _rk_nac = _rk_nac[_rk_nac['n_sucursales'] >= MIN_SUCURSALES_RANKING].sort_values('precio_promedio')
    _prom_nac_rk = _pgeo['precio_producto'].mean()

    _amba = _pgeo[_pgeo['PROVINCIA_NORM'].isin(['Buenos Aires','CABA'])]
    _rk_amba = (_amba.groupby('cadena')
                .agg(n_sucursales=('precio_producto','count'),
                     precio_promedio=('precio_producto','mean'))
                .round(2).reset_index())
    _rk_amba = _rk_amba[_rk_amba['n_sucursales'] >= MIN_SUCURSALES_RANKING].sort_values('precio_promedio')
    _prom_amba_rk = _amba['precio_producto'].mean() if len(_amba) > 0 else 0

    for (_rk, _prom_r, _titulo, _out_name) in [
        (_rk_nac,  _prom_nac_rk,  f'Ranking nacional [{_name}]',  f'ranking_cadenas_nacional_{MES}_{_short}'),
        (_rk_amba, _prom_amba_rk, f'Ranking AMBA [{_name}]',      f'ranking_cadenas_amba_{MES}_{_short}'),
    ]:
        if len(_rk) == 0:
            print(f'  Sin datos para {_titulo}'); continue
        fig, ax = plt.subplots(figsize=(11, max(5, len(_rk)*0.5+2)))
        labs   = [f"{r.cadena}  ({int(r.n_sucursales)})" for r in _rk.itertuples()]
        _n_c   = len(_rk)
        _cols  = plt.cm.RdYlGn_r(np.linspace(0.1, 0.9, _n_c)) if _n_c > 1 else [PRODUCTO_COLORS[_pid]]
        bars   = ax.barh(labs, _rk['precio_promedio'], color=_cols, edgecolor='black', linewidth=0.4)
        for bar, val in zip(bars, _rk['precio_promedio']):
            ax.text(bar.get_width() + _rk['precio_promedio'].max()*0.005,
                    bar.get_y()+bar.get_height()/2,
                    f'${fmtn(val)}', va='center', fontsize=9, fontweight='bold')
        ax.axvline(_prom_r, color='#666', linestyle='--', linewidth=1.5,
                   label=f'Promedio: ${fmtn(_prom_r)}')
        ax.set_xlabel('Precio promedio (ARS)', fontsize=11)
        ax.set_xlim(_rk['precio_promedio'].min()*0.95, _rk['precio_promedio'].max()*1.07)
        ax.xaxis.set_major_formatter(mticker.FuncFormatter(
            lambda x,_: f'${x:,.0f}'.replace(',','.')))
        ax.legend(loc='lower right', fontsize=10)
        ax.grid(True, alpha=0.3, axis='x'); ax.set_axisbelow(True)
        for sp in ['top','right']: ax.spines[sp].set_visible(False)
        plt.tight_layout()
        _out_r = OUTPUT_DIR / f'{_out_name}.png'
        plt.savefig(_out_r, dpi=600, bbox_inches='tight'); plt.show()
        print(f'  Ranking [{_name}] guardado: {_out_r.name}')

    print(f'\\n  === RANKING NACIONAL [{_name}] ===')
    for i, r in enumerate(_rk_nac.sort_values('precio_promedio', ascending=False).itertuples(), 1):
        print(f'    {i:>2}. {r.cadena:<25} ${fmtn(r.precio_promedio):>12}  ({int(r.n_sucursales)} sucs)')"""))

# ── CELL 17 — FOLIUM MAP (selector de producto, lazy popup) ───────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 17 — Mapa Folium: lazy popup (datos JSON on-demand)
# Arquitectura: datos almacenados una vez como JSON compacto;
# popup HTML construido por JS al hacer click → archivo ~80% más liviano
# ============================================================
def fmtm(x): return f'{x:,.2f}'.replace(',','X').replace('.',',').replace('X','.')

# ── Construir datos compactos para popups (almacenados una vez) ──────────────
_pgeo_ref = producto_geo_dict[_prods_con_datos[0]]
provs_u   = sorted(_pgeo_ref['PROVINCIA_NORM'].unique())

# Popup compacto: solo precio por sucursal/producto (sin detalle adicional)
_popup_data = {}
for _pid in _prods_con_datos:
    for _, _r in producto_geo_dict[_pid].iterrows():
        _sk = f"{_r['id_comercio']}_{_r['id_bandera']}_{_r['id_sucursal']}"
        if _sk not in _popup_data:
            _popup_data[_sk] = {
                'nom': str(_r['sucursales_nombre'])[:40],
                'bar': str(_r.get('sucursales_barrio') or _r.get('sucursales_localidad') or '')[:30],
                'prv': _r['PROVINCIA_NORM'],
                'cad': _r['cadena'],
                'tip': str(_r.get('sucursales_tipo') or 'N/D'),
                'can': {}
            }
        _popup_data[_sk]['can'][_pid] = {'t': float(_r['precio_producto'])}

_popup_json = _json.dumps(_popup_data, ensure_ascii=False, separators=(',',':'))
print(f'Datos popup: {len(_popup_data):,} sucursales | {len(_popup_json)/1024/1024:.1f} MB JSON compacto')

# ── Mapa Folium ───────────────────────────────────────────────────────────────
m = folium.Map(location=[-38.0,-63.5], zoom_start=5,
               tiles='cartodbpositron', control_scale=True)
folium.map.Marker(
    location=[-51.7963,-59.5236],
    icon=folium.DivIcon(icon_size=(140,28), icon_anchor=(70,14),
        html='<div style="background:rgba(255,255,255,.95);border:1px solid #777;border-radius:3px;padding:3px 7px;font-family:Arial;font-size:11px;font-weight:600;text-align:center;white-space:nowrap;">Islas Malvinas (ARG)</div>')
).add_to(m)

_producto_fg_ids = {}
for _pid in _prods_con_datos:
    _name  = PRODUCTO_NOMBRES[_pid]
    _short = PRODUCTO_SHORT[_pid]
    _pgeo  = producto_geo_dict[_pid]
    _is_default = (_pid == _prods_con_datos[0])
    _vmin = _pgeo['precio_producto'].quantile(0.05)
    _vmax = _pgeo['precio_producto'].quantile(0.95)
    if _vmin == _vmax: _vmin, _vmax = _pgeo['precio_producto'].min(), _pgeo['precio_producto'].max()
    _cm = LinearColormap(
        colors=['#1a9850','#66bd63','#a6d96a','#fee08b','#fdae61','#f46d43','#d73027'],
        vmin=_vmin, vmax=_vmax, caption=f'{_name} — {NOMBRE_MES_TITLE} (ARS)')
    if _is_default: _cm.add_to(m)
    _fg = folium.FeatureGroup(name=_short, show=_is_default)
    _producto_fg_ids[_pid] = _fg.get_name()
    for _, _r in _pgeo.iterrows():
        val  = _r['precio_producto']
        col  = _cm(max(_vmin, min(_vmax, val)))
        cad  = _r['cadena']
        prv  = _r['PROVINCIA_NORM']
        _sk  = f"{_r['id_comercio']}_{_r['id_bandera']}_{_r['id_sucursal']}"
        cl   = (f'sucursal-marker producto-{_short}'
                f' cadena-{cad.replace(" ","_").replace("(","").replace(")","").replace("/","")}'
                f' prov-{prv.replace(" ","_").replace("(","").replace(")","").replace("/","")}')
        # Popup mínimo: placeholder que JS rellena on-demand al hacer click
        _ph = (f'<div class="lz-pop" data-key="{_sk}" data-can="{_pid}"'
               f' style="font-family:Arial;min-width:200px;text-align:center;padding:15px">'
               f'<span style="color:#aaa;font-size:12px">Cargando detalle...</span></div>')
        folium.CircleMarker(
            location=[_r['sucursales_latitud'], _r['sucursales_longitud']],
            radius=5, color=col, fill=True, fillColor=col, fillOpacity=0.8, weight=1,
            tooltip=f'<b>{cad}</b><br>{prv}<br><b>${fmtm(val)}</b>',
            popup=folium.Popup(_ph, max_width=450), className=cl
        ).add_to(_fg)
    _fg.add_to(m)

_map_var    = m.get_name()
_fg_ids_str = '{' + ','.join(f'"{k}":"{v}"' for k,v in _producto_fg_ids.items()) + '}'
_names_str  = '{' + ','.join(f'"{k}":"{PRODUCTO_NOMBRES[k]}"' for k in _prods_con_datos) + '}'
_avgs_str   = '{' + ','.join(f'"{k}":{producto_geo_dict[k]["precio_producto"].mean():.2f}' for k in _prods_con_datos) + '}'

# Embeber JSON en script tag de tipo application/json (sin escape JS)
m.get_root().html.add_child(folium.Element(
    f'<script type="application/json" id="_pd_json">{_popup_json}</script>'))

prov_opts  = ''.join([f'<option value="prov-{p.replace(" ","_")}">{p}</option>' for p in provs_u])
_can_opts  = ''.join([f'<option value="{k}">{PRODUCTO_NOMBRES[k]}</option>' for k in _prods_con_datos])
_cadenas_u = sorted(_pgeo_ref['cadena'].unique())
_cad_opts  = ''.join([f'<option value="cadena-{c.replace(" ","_").replace("(","").replace(")","").replace("/","")}">{c}</option>' for c in _cadenas_u])

info_h = (f'<div style="position:fixed;top:10px;left:50px;width:340px;background:white;border:2px solid #0055A4;'
          f'border-radius:8px;padding:12px 15px;font-family:Arial;z-index:9999;box-shadow:0 2px 8px rgba(0,0,0,.15);">'
          f'<div style="color:#0055A4;font-size:15px;font-weight:bold;margin-bottom:5px;">Productos — {NOMBRE_MES_TITLE}</div>'
          f'<div style="font-size:11px;color:#555;line-height:1.5;">'
          f'<b>{len(_pgeo_ref):,}</b> sucursales · <b>{len(_prods_con_datos)}</b> productos<br>'
          f'Promedio: <span id="info_avg" style="font-weight:bold;"></span></div></div>')
m.get_root().html.add_child(folium.Element(info_h))

filtros_h = (
    f'<div id="pf" style="position:fixed;bottom:25px;left:50px;width:280px;background:white;'
    f'border:2px solid #0055A4;border-radius:8px;padding:12px 15px;font-family:Arial;z-index:9999;">'
    f'<div style="color:#0055A4;font-size:13px;font-weight:bold;margin-bottom:8px;">🔍 Filtros</div>'
    f'<label style="font-size:11px;color:#555;display:block;margin-top:4px;">Producto:'
    f'<select id="fcan" style="width:100%;padding:4px;font-size:11px;margin-top:3px;">{_can_opts}</select></label>'
    f'<label style="font-size:11px;color:#555;display:block;margin-top:6px;">Cadena:'
    f'<select id="fca" style="width:100%;padding:4px;font-size:11px;margin-top:3px;">'
    f'<option value="all">Todas</option>{_cad_opts}</select></label>'
    f'<label style="font-size:11px;color:#555;display:block;margin-top:6px;">Provincia:'
    f'<select id="fp" style="width:100%;padding:4px;font-size:11px;margin-top:3px;">'
    f'<option value="all">Todas</option>{prov_opts}</select></label>'
    f'<button id="fr" style="width:100%;margin-top:10px;padding:6px;background:#f0f0f0;'
    f'border:1px solid #ccc;border-radius:4px;font-size:11px;cursor:pointer;">Restablecer</button></div>'
    f'<style>'
    f'.lz-w{{font-family:Arial;font-size:12px;width:420px;max-height:500px;overflow-y:auto}}'
    f'.lz-h4{{margin:0;color:#0055A4}}'
    f'.lz-nfo{{font-size:11px;color:#555;margin-bottom:5px}}'
    f'.lz-bg{{background:#e6eef7;padding:2px 6px;border-radius:3px;font-size:10px}}'
    f'.lz-cx{{text-align:center;margin:8px 0}}'
    f'.lz-lbl{{font-size:11px;color:#555;margin-bottom:2px}}'
    f'.lz-tot{{color:#0055A4;font-size:22px;font-weight:bold}}'
    f'.lz-sub{{font-size:11px;color:#888;text-align:center;margin-top:3px}}'
    f'.lz-ft{{font-size:9px;color:#666;margin-top:4px}}'
    f'.lz-hr{{margin:5px 0}}'
    f'</style>'
    f'<script>'
    f'var _fg_ids={_fg_ids_str};var _names={_names_str};var _avgs={_avgs_str};var _pd=null;'
    f'function _gPD(){{if(!_pd){{var el=document.getElementById("_pd_json");if(el)_pd=JSON.parse(el.textContent);}}return _pd;}}'
    # _bPop usa template literals JS (backtick) + clases CSS → sin single-quote CSS = sin conflicto Python
    f'function _bPop(key,cid){{'
    f'var pd=_gPD();if(!pd||!pd[key]||!pd[key].can[cid])return "<div>Sin datos.</div>";'
    f'var d=pd[key];var c=d.can[cid];var nm=_names[cid];'
    f'var fmt=function(x){{return "$"+x.toLocaleString("es-AR",{{minimumFractionDigits:2,maximumFractionDigits:2}});}};'
    f'return `<div class=lz-w>`'
    f'+`<h4 class=lz-h4>${{d.cad}}</h4>`'
    f'+`<div class=lz-nfo><b>${{d.nom}}</b><br>${{d.bar?d.bar+" — ":""}}${{d.prv}}<br><span class=lz-bg>${{d.tip}}</span></div>`'
    f'+"<hr class=lz-hr>"'
    f'+`<div class=lz-cx><div class=lz-lbl>${{nm}}</div><span class=lz-tot>${{fmt(c.t)}}</span></div>`'
    f'+"</div>";}}'
    f'function _initEvt(){{var mp=window["{_map_var}"];if(!mp)return;'
    f'mp.on("popupopen",function(e){{'
    f'var el=e.popup.getElement().querySelector(".lz-pop");'
    f'if(el&&el.getAttribute("data-built")!=="1"){{'
    f'el.innerHTML=_bPop(el.getAttribute("data-key"),el.getAttribute("data-can"));'
    f'el.setAttribute("data-built","1");e.popup.update();}}}});}}'
    f'function switchProducto(sel){{var mp=window["{_map_var}"];if(!mp)return;'
    f'Object.keys(_fg_ids).forEach(function(k){{var fg=window[_fg_ids[k]];if(!fg)return;'
    f'if(k===sel){{mp.addLayer(fg);}}else{{mp.removeLayer(fg);}}}});'
    f'var avgEl=document.getElementById("info_avg");'
    f'if(avgEl)avgEl.innerHTML="$"+_avgs[sel].toLocaleString("es-AR",{{minimumFractionDigits:2,maximumFractionDigits:2}})+" ("+_names[sel]+")";apl();}}'
    f'function apl(){{var p=document.getElementById("fp").value;var ca=document.getElementById("fca").value;'
    f'document.querySelectorAll(".sucursal-marker").forEach(function(el){{'
    f'var c=el.className.baseVal||el.className||"";'
    f'var mp=(p==="all")||c.indexOf(p)>=0;'
    f'var mc=(ca==="all")||c.indexOf(ca)>=0;'
    f'el.style.display=(mp&&mc)?"":"none";}});}}'
    f'setTimeout(function(){{'
    f'var fc=document.getElementById("fcan"),sp=document.getElementById("fp"),fca=document.getElementById("fca"),btn=document.getElementById("fr");'
    f'var def=Object.keys(_fg_ids)[0];'
    f'_initEvt();switchProducto(def);'
    f'if(fc)fc.addEventListener("change",function(){{switchProducto(this.value);}});'
    f'if(sp)sp.addEventListener("change",apl);'
    f'if(fca)fca.addEventListener("change",apl);'
    f'if(btn)btn.addEventListener("click",function(){{'
    f'if(fc){{fc.value=Object.keys(_fg_ids)[0];switchProducto(fc.value);}}'
    f'if(sp)sp.value="all";if(fca)fca.value="all";'
    f'document.querySelectorAll(".sucursal-marker").forEach(e=>e.style.display="");}});'
    f'}},1200);</script>'
)
m.get_root().html.add_child(folium.Element(filtros_h))

out_map = OUTPUT_DIR / f'mapa_interactivo_{MES}.html'
m.save(str(out_map))
print(f'Mapa guardado: {out_map.name} ({len(_prods_con_datos)} productos · {len(_pgeo_ref):,} sucs)')"""))

# ── CELL 18 — CABA RANKINGS (per-producto) ─────────────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 18 — Ranking de barrios CABA para cada producto activo
# ============================================================
BARRIOS_BBOX = {
    'Agronomia':           (-34.604,-34.587,-58.498,-58.476),
    'Almagro':             (-34.622,-34.598,-58.435,-58.405),
    'Balvanera':           (-34.617,-34.598,-58.418,-58.388),
    'Barracas':            (-34.661,-34.628,-58.395,-58.366),
    'Belgrano':            (-34.575,-34.547,-58.471,-58.434),
    'Boedo':               (-34.638,-34.620,-58.426,-58.408),
    'Caballito':           (-34.628,-34.602,-58.460,-58.421),
    'Chacarita':           (-34.595,-34.575,-58.470,-58.443),
    'Coghlan':             (-34.572,-34.555,-58.485,-58.469),
    'Colegiales':          (-34.580,-34.563,-58.460,-58.439),
    'Constitucion':        (-34.631,-34.620,-58.395,-58.378),
    'Flores':              (-34.642,-34.615,-58.480,-58.435),
    'Floresta':            (-34.633,-34.615,-58.500,-58.479),
    'La Boca':             (-34.643,-34.620,-58.371,-58.350),
    'La Paternal':         (-34.605,-34.585,-58.475,-58.456),
    'Liniers':             (-34.652,-34.628,-58.534,-58.506),
    'Mataderos':           (-34.665,-34.641,-58.522,-58.488),
    'Monte Castro':        (-34.628,-34.610,-58.520,-58.500),
    'Monserrat':           (-34.625,-34.605,-58.391,-58.371),
    'Nueva Pompeya':       (-34.658,-34.638,-58.418,-58.396),
    'Nunez':               (-34.553,-34.532,-58.475,-58.443),
    'Palermo':             (-34.595,-34.560,-58.435,-58.398),
    'Parque Avellaneda':   (-34.660,-34.638,-58.495,-58.470),
    'Parque Chacabuco':    (-34.645,-34.625,-58.448,-58.422),
    'Parque Chas':         (-34.591,-34.578,-58.487,-58.475),
    'Parque Patricios':    (-34.652,-34.628,-58.418,-58.395),
    'Puerto Madero':       (-34.625,-34.587,-58.371,-58.349),
    'Recoleta':            (-34.598,-34.575,-58.405,-58.378),
    'Retiro':              (-34.595,-34.578,-58.388,-58.365),
    'Saavedra':            (-34.560,-34.540,-58.495,-58.467),
    'San Cristobal':       (-34.625,-34.612,-58.408,-58.391),
    'San Nicolas':         (-34.610,-34.595,-58.395,-58.371),
    'San Telmo':           (-34.625,-34.610,-58.378,-58.365),
    'Velez Sarsfield':     (-34.642,-34.624,-58.510,-58.493),
    'Versalles':           (-34.640,-34.621,-58.525,-58.508),
    'Villa Crespo':        (-34.605,-34.585,-58.452,-58.428),
    'Villa del Parque':    (-34.615,-34.595,-58.498,-58.472),
    'Villa Devoto':        (-34.612,-34.585,-58.518,-58.490),
    'Villa General Mitre': (-34.615,-34.600,-58.475,-58.458),
    'Villa Lugano':        (-34.690,-34.660,-58.475,-58.435),
    'Villa Luro':          (-34.645,-34.628,-58.510,-58.491),
    'Villa Ortuzar':       (-34.590,-34.575,-58.475,-58.456),
    'Villa Pueyrredon':    (-34.585,-34.565,-58.510,-58.485),
    'Villa Real':          (-34.628,-34.615,-58.530,-58.512),
    'Villa Riachuelo':     (-34.695,-34.680,-58.470,-58.450),
    'Villa Santa Rita':    (-34.622,-34.605,-58.488,-58.470),
    'Villa Soldati':       (-34.682,-34.655,-58.460,-58.420),
    'Villa Urquiza':       (-34.590,-34.565,-58.495,-58.470),
}

def det_barrio(lat, lon):
    if pd.isna(lat) or pd.isna(lon): return 'Sin clasificar'
    for b,(lmin,lmax,lonmin,lonmax) in BARRIOS_BBOX.items():
        if lmin <= lat <= lmax and lonmin <= lon <= lonmax: return b
    return 'Sin clasificar'

def fmtb(x): return f'${x:,.2f}'.replace(',','X').replace('.',',').replace('X','.')

for _pid in _prods_con_datos:
    _name  = PRODUCTO_NOMBRES[_pid]
    _pgeo  = producto_geo_dict[_pid]
    _caba  = _pgeo[_pgeo['PROVINCIA_NORM'] == 'CABA'].copy()
    if len(_caba) == 0:
        print(f'  [{_name}] Sin sucursales en CABA'); continue

    _caba['barrio'] = _caba.apply(
        lambda r: det_barrio(r['sucursales_latitud'], r['sucursales_longitud']), axis=1)
    _rk_b = (_caba[_caba['barrio'] != 'Sin clasificar']
             .groupby('barrio')
             .agg(n_sucs=('precio_producto','count'),
                  promedio=('precio_producto','mean'),
                  mediana=('precio_producto','median'))
             .round(2).sort_values('promedio'))
    _rk_b_fil = _rk_b[_rk_b['n_sucs'] >= 2]
    _pc = _caba['precio_producto'].mean()
    _pp = _pgeo['precio_producto'].mean()

    print(f'\\n{"="*78}')
    print(f'  RANKING BARRIOS CABA [{_name.upper()}] — {NOMBRE_MES_TITLE}')
    print(f'{"="*78}')
    print(f'  {"#":<3} {"Barrio":<22} {"Sucs.":<7} {"Promedio":<13} {"vs CABA":<10} {"vs Pais"}')
    print(f'  {"-"*70}')
    for i, (b, r) in enumerate(_rk_b_fil.iterrows(), 1):
        vc = (r["promedio"]/max(_pc,1)-1)*100
        vp = (r["promedio"]/max(_pp,1)-1)*100
        print(f'  {i:<3} {b:<22} {int(r["n_sucs"]):<7} {fmtb(r["promedio"]):<13} {vc:+.2f}%  {vp:+.2f}%')
    print(f'  {"-"*70}')
    print(f'  Promedio CABA: {fmtb(_pc)}   Promedio pais: {fmtb(_pp)}')
    _sin = set(BARRIOS_BBOX.keys()) - set(_rk_b.index)
    if _sin:
        print(f'  Sin sucursales ({len(_sin)}): {", ".join(sorted(_sin))}')"""))

# ── CELL 19 — EXCEL EXPORT (multi-producto) ────────────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 19 — Exportación Excel multi-producto
# ============================================================
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HDR_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HDR_FONT = Font(bold=True, color='FFFFFF', size=10)
HDR_ALIG = Alignment(horizontal='center', wrap_text=True, vertical='center')

def fmt_ws(ws):
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'
    for cell in ws[1]:
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = HDR_ALIG

def auto_widths(ws):
    for ci in range(1, ws.max_column+1):
        cl  = get_column_letter(ci)
        hdr = str(ws.cell(1,ci).value or '').lower()
        w   = 28 if any(x in hdr for x in ('nombre','provincia','cadena','barrio','producto')) else (42 if 'desc' in hdr else 14)
        ws.column_dimensions[cl].width = w
        if any(x in hdr for x in ('precio','ipc','promedio','mediana')):
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=ci, max_col=ci):
                for cell in row: cell.number_format = '#,##0.00'
        elif '%' in hdr:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=ci, max_col=ci):
                for cell in row: cell.number_format = '+0.00"%"'

def _safe_sheet(name, used, maxlen=28):
    base = re.sub(r'[\\\\/*?:\\[\\]]', ' ', name)
    base = re.sub(r'\\s+', ' ', base).strip()[:maxlen]
    if not base: base = 'Producto'
    out, k = base, 1
    while out.lower() in used:
        suf = f' {k}'; out = base[:maxlen-len(suf)] + suf; k += 1
    used.add(out.lower())
    return out

out_xls = OUTPUT_DIR / f'productos_analisis_{ULTIMO_MES}.xlsx'
with pd.ExcelWriter(out_xls, engine='openpyxl') as writer:

    # ── Hoja Evolucion_IPC: todos los productos + IPC en una tabla ─────────
    _ipc_base = ipc[['mes','ipc_general','ipc_general_var_%',
                      'ipc_alimentos','ipc_alimentos_var_%']].copy()
    _evo = _ipc_base
    for _pid in _prods_con_datos:
        if len(serie_nac_dict[_pid]) == 0: continue
        _sn = serie_nac_dict[_pid][['mes','precio_nacional_ponderado','variacion_mensual_%',
                                    'precio_nacional_ponderado_media','variacion_mensual_media_%']].copy()
        _n  = PRODUCTO_SHORT[_pid]
        _sn = _sn.rename(columns={
            'precio_nacional_ponderado':       f'precio_{_n}',
            'variacion_mensual_%':             f'var_{_n}_%',
            'precio_nacional_ponderado_media': f'precio_{_n}_media',
            'variacion_mensual_media_%':       f'var_{_n}_media_%'
        })
        _evo = _evo.merge(_sn, on='mes', how='outer')
    _evo = _evo.sort_values('mes').reset_index(drop=True)
    _evo.to_excel(writer, sheet_name='Evolucion_IPC', index=False)

    # ── Hoja índice: un producto por fila (EAN, descripción, rubro, hojas) ──
    _usados_sh = set()
    _idx_rows = []
    _sheet_by_pid = {}
    for _pid in PRODUCTOS_ACTIVOS:
        _ean = next(iter(PRODUCTOS[_pid]))
        _sh  = _safe_sheet(PRODUCTO_NOMBRES[_pid], _usados_sh)
        _sheet_by_pid[_pid] = _sh
        _idx_rows.append({
            'producto': PRODUCTO_NOMBRES[_pid], 'EAN': _ean,
            'rubro': next(iter(PRODUCTOS[_pid].values()))[2],
            'con_datos_este_mes': _pid in _prods_con_datos,
            'precio_promedio': round(prom_nac_dict.get(_pid, float('nan')), 2)
                if not pd.isna(prom_nac_dict.get(_pid, float('nan'))) else None,
            'precio_promedio_media': round(prom_nac_media_dict.get(_pid, float('nan')), 2)
                if not pd.isna(prom_nac_media_dict.get(_pid, float('nan'))) else None,
            'hoja': _sh,
        })
    pd.DataFrame(_idx_rows).to_excel(writer, sheet_name='Productos', index=False)

    # ── Hoja por producto: Provincias y Ranking ─────────────────────────────
    # Nombres de hoja finales (con prefijo Prov_/Rank_/Sucs_) deduplicados en un
    # único set: truncar a 31 chars (límite de Excel) DESPUÉS de anteponer el
    # prefijo podría, en casos raros, hacer coincidir dos nombres que sin el
    # prefijo eran distintos. _usados_full evita esa colisión.
    _usados_full = set()
    for _pid in _prods_con_datos:
        _sh  = _sheet_by_pid[_pid]
        _spv = serie_prov_dict[_pid].copy()
        _prom = prom_nac_dict[_pid]
        _prom_media = prom_nac_media_dict[_pid]
        if len(_spv) > 0 and not pd.isna(_prom):
            _spv['vs_promedio_%']       = ((_spv['precio_producto'] / _prom) - 1) * 100
            _spv['vs_promedio_media_%'] = ((_spv['precio_producto_media'] / _prom_media) - 1) * 100
            _spv = _spv[['mes','provincia','precio_producto','vs_promedio_%',
                         'precio_producto_media','vs_promedio_media_%']]
            _spv.sort_values('precio_producto').to_excel(
                writer, sheet_name=_safe_sheet(f'Prov_{_sh}', _usados_full, maxlen=31), index=False)

        _pgeo = producto_geo_dict[_pid]
        _rk  = (_pgeo.groupby('cadena')
                .agg(n_sucursales=('precio_producto','count'),
                     precio_promedio=('precio_producto','mean'),
                     precio_mediana=('precio_producto','median'))
                .round(2).reset_index()
                .sort_values('precio_promedio', ascending=False))
        if len(_pgeo) > 0:
            _rk['vs_promedio_%'] = ((_rk['precio_promedio'] / _pgeo['precio_producto'].mean()) - 1) * 100
            _rk['vs_mediana_%']  = ((_rk['precio_mediana'] / _pgeo['precio_producto'].median()) - 1) * 100
            _rk = _rk[['cadena','n_sucursales','precio_promedio','vs_promedio_%',
                       'precio_mediana','vs_mediana_%']]
        _rk.to_excel(writer, sheet_name=_safe_sheet(f'Rank_{_sh}', _usados_full, maxlen=31), index=False)

        _suc_exp = _pgeo[[
            'id_comercio','id_bandera','id_sucursal','cadena','PROVINCIA_NORM',
            'sucursales_nombre','sucursales_localidad','sucursales_barrio',
            'sucursales_latitud','sucursales_longitud','sucursales_tipo',
            'precio_producto'
        ]].sort_values(['PROVINCIA_NORM','cadena','precio_producto']).copy()
        _suc_exp.to_excel(writer, sheet_name=_safe_sheet(f'Sucs_{_sh}', _usados_full, maxlen=31), index=False)

    # ── Hoja Serie_precios: precio mediano por producto x mes ───────────────
    _sp_rows = []
    for _pid, _prod in PRODUCTOS.items():
        _eans = set(_prod.keys())
        _dh   = df_hist_raw[df_hist_raw['ean_norm'].isin(_eans)].copy()
        if len(_dh) == 0: continue
        _dh['id_producto']  = _dh['ean_norm'].apply(lambda e: e.zfill(13))
        _dh['descripcion']  = _dh['ean_norm'].map(lambda e, c=_prod: c.get(e,('',0,''))[0])
        _dh['rubro']        = _dh['ean_norm'].map(lambda e, c=_prod: c.get(e,('',0,'?'))[2])
        _dh['producto_id']  = _pid
        _sp_rows.append(_dh[['producto_id','descripcion','anio_mes',
                              'id_producto','rubro','precio_mediano','precio_medio']])
    if _sp_rows:
        _sp_all = pd.concat(_sp_rows, ignore_index=True).sort_values(
            ['producto_id','anio_mes']).reset_index(drop=True)
        _sp_all = _sp_all.rename(columns={'anio_mes':'mes'})
        _sp_all.to_excel(writer, sheet_name='Serie_precios', index=False)

    # ── Formato ──────────────────────────────────────────────────────────────
    for sn in writer.sheets:
        ws = writer.sheets[sn]
        fmt_ws(ws)
        auto_widths(ws)

print(f'Excel guardado: {out_xls}')
print()
print('='*65)
print(f'  RESUMEN — {NOMBRE_MES_TITLE.upper()}')
print('='*65)
for _pid in _prods_con_datos:
    _name = PRODUCTO_NOMBRES[_pid]
    _pgeo = producto_geo_dict[_pid]
    _prom = prom_nac_dict[_pid]
    _sn   = serie_nac_dict[_pid]
    _rng  = f'{_sn["mes"].min()} -> {_sn["mes"].max()}' if len(_sn) > 0 else 'sin historia'
    print(f'  [{_name}] {len(_pgeo):,} sucs | Promedio: ${_prom:,.2f} | Serie: {_rng}')
print('='*65)
if MES_PARCIAL:
    print(f'\\n⚠️  {NOMBRE_MES_TITLE}: MES PARCIAL ({DIAS_CARGADOS}/{DIAS_MES} días).')
    print('    Los promedios son un snapshot al día; la VARIACIÓN mensual es PRELIMINAR')
    print('    (subestimada vs meses completos). Re-correr al cierre del mes para el dato final.')"""))

# ── CELL 20 — VALORES RESUMEN (genérico, sin IDs de producto hardcodeados) ────
cells.append(cell_code("""\
# ============================================================
# CELDA 20 — Valores resumen para reporte (genérico, por producto)
# ============================================================
# A diferencia de nb02 (que asume canastas fijas como 'Media'), acá NO hay
# ningún producto especial hardcodeado: todo se calcula iterando sobre los
# productos que efectivamente tuvieron datos este mes.

def _ar(x): return f"${x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
def _pp(x, dec=2):
    s = "+" if x >= 0 else ""; return f"{s}{x:.{dec}f}%".replace(".", ",")

_vals = {p: prom_nac_dict[p] for p in _prods_con_datos if not pd.isna(prom_nac_dict.get(p))}
_vals_media = {p: prom_nac_media_dict[p] for p in _prods_con_datos if not pd.isna(prom_nac_media_dict.get(p))}
_vars = {}
_vars_media = {}
for p in _prods_con_datos:
    _sn = serie_nac_dict.get(p)
    if _sn is not None and len(_sn) > 1:
        _vars[p] = _sn['variacion_mensual_%'].iloc[-1]
        _vars_media[p] = _sn['variacion_mensual_media_%'].iloc[-1]

if not _vals:
    print('AVISO: ningún producto tiene precio válido este mes — no hay valores resumen para calcular.')
else:
    _sorted_ids = sorted(_vals, key=lambda p: _vals[p])
    _min_id, _max_id = _sorted_ids[0], _sorted_ids[-1]
    _brecha_abs = _vals[_max_id] - _vals[_min_id]
    _brecha_pct = (_brecha_abs / _vals[_min_id] * 100) if _vals[_min_id] else float('nan')

    _rows_doc = [['Seccion', 'Variable', 'Valor_LaTeX', 'Valor_numero']]
    _rows_doc.append(['Portada', 'Mes', NOMBRE_MES_TITLE, NOMBRE_MES_TITLE])
    _rows_doc.append(['Portada', 'Producto mas caro', PRODUCTO_NOMBRES[_max_id], PRODUCTO_NOMBRES[_max_id]])
    _rows_doc.append(['Portada', 'Producto mas barato', PRODUCTO_NOMBRES[_min_id], PRODUCTO_NOMBRES[_min_id]])

    print("=" * 72)
    print(f"  VALORES RESUMEN — {NOMBRE_MES_TITLE.upper()}")
    print("=" * 72)

    print("\\n[Precio nacional y variación mensual, por producto]")
    for p in _sorted_ids:
        _nm = PRODUCTO_NOMBRES[p]
        _line = f"{_pp(_vars.get(p, float('nan')))}"
        print(f"  {_nm:<40} {_ar(_vals[p]):>14}   var.mensual: {_line}")
        _rows_doc.append(['Productos', f'{_nm} — precio', _ar(_vals[p]), round(_vals[p], 2)])
        if p in _vals_media:
            _rows_doc.append(['Productos', f'{_nm} — precio (media)', _ar(_vals_media[p]), round(_vals_media[p], 2)])
        if p in _vars:
            _rows_doc.append(['Productos', f'{_nm} — var.mensual', _pp(_vars[p]), round(_vars[p], 2)])
        if p in _vars_media:
            _rows_doc.append(['Productos', f'{_nm} — var.mensual (media)', _pp(_vars_media[p]), round(_vars_media[p], 2)])
    print(f"  Producto mas caro:    {PRODUCTO_NOMBRES[_max_id]:<30} {_ar(_vals[_max_id])}")
    print(f"  Producto mas barato:  {PRODUCTO_NOMBRES[_min_id]:<30} {_ar(_vals[_min_id])}")
    print(f"  Brecha:               {_ar(_brecha_abs)}  ({_brecha_pct:.1f}%)")
    _rows_doc.append(['Productos', 'Brecha absoluta (mas caro - mas barato)', _ar(_brecha_abs), round(_brecha_abs, 2)])
    _rows_doc.append(['Productos', 'Brecha relativa (%)', f'{_brecha_pct:.1f}%', round(_brecha_pct, 1)])

    print("\\n[Dispersión provincial y cadenas, por producto]")
    for p in _sorted_ids:
        _nm  = PRODUCTO_NOMBRES[p]
        _spv = serie_prov_dict[p]
        if len(_spv) < 2:
            continue
        _prov_min = _spv.loc[_spv['precio_producto'].idxmin()]
        _prov_max = _spv.loc[_spv['precio_producto'].idxmax()]
        _disp_prov = (_prov_max['precio_producto'] - _prov_min['precio_producto']) / _prov_min['precio_producto'] * 100
        print(f"  {_nm}:")
        print(f"    Prov. mas barata: {_prov_min['provincia']:<22} {_ar(_prov_min['precio_producto'])}")
        print(f"    Prov. mas cara:   {_prov_max['provincia']:<22} {_ar(_prov_max['precio_producto'])}")
        print(f"    Dispersion prov.: {_disp_prov:.1f}%")
        _rows_doc.append(['Provincias', f'{_nm} — prov. mas barata', _prov_min['provincia'], _prov_min['provincia']])
        _rows_doc.append(['Provincias', f'{_nm} — prov. mas cara', _prov_max['provincia'], _prov_max['provincia']])
        _rows_doc.append(['Provincias', f'{_nm} — dispersion (%)', f'{_disp_prov:.1f}%', round(_disp_prov, 1)])

        _pgeo = producto_geo_dict[p]
        _rk = (_pgeo.groupby('cadena')['precio_producto']
               .agg(n='count', prom='mean').reset_index())
        _rk = _rk[_rk['n'] >= MIN_SUCURSALES_RANKING].sort_values('prom')
        if len(_rk) >= 2:
            _bar = _rk.iloc[0]; _car = _rk.iloc[-1]
            print(f"    Cadena mas barata: {_bar['cadena']:<22} {_ar(_bar['prom'])}")
            print(f"    Cadena mas cara:   {_car['cadena']:<22} {_ar(_car['prom'])}")
            _rows_doc.append(['Cadenas', f'{_nm} — cadena mas barata', _bar['cadena'], _bar['cadena']])
            _rows_doc.append(['Cadenas', f'{_nm} — cadena mas cara', _car['cadena'], _car['cadena']])

    print(f"\\n[Acumulados desde {MES_INICIO_GRAFICO}]")
    _ipc_b = ipc[ipc['mes'] == MES_INICIO_GRAFICO] if MES_INICIO_GRAFICO in ipc['mes'].values else ipc.iloc[[0]]
    _ipc_l = ipc.iloc[-1]
    _acum_ipc_gen = (_ipc_l['ipc_general'] / _ipc_b['ipc_general'].values[0] - 1) * 100
    for p in _sorted_ids:
        _sn = serie_nac_dict[p]
        _mg = MES_INICIO_GRAFICO if MES_INICIO_GRAFICO in _sn['mes'].values else (_sn['mes'].min() if len(_sn) > 0 else None)
        if _mg and len(_sn) >= 2:
            _dgp = _sn[_sn['mes'] >= _mg].reset_index(drop=True)
            if len(_dgp) >= 2:
                _acum = (_dgp['precio_nacional_ponderado'].iloc[-1] / _dgp['precio_nacional_ponderado'].iloc[0] - 1) * 100
                print(f"  {PRODUCTO_NOMBRES[p]:<40} {_pp(_acum)}")
                _rows_doc.append(['Acumulados', f'{PRODUCTO_NOMBRES[p]} desde {MES_INICIO_GRAFICO}', _pp(_acum), round(_acum, 1)])
    print(f"  {'IPC Nivel General':<40} {_pp(_acum_ipc_gen)}")
    _rows_doc.append(['Acumulados', f'IPC Nivel General desde {MES_INICIO_GRAFICO}', _pp(_acum_ipc_gen), round(_acum_ipc_gen, 1)])
    print("=" * 72)

    # ── Exportar a Excel (hoja Valores_Documento) ─────────────────────────────
    import openpyxl as _opxl
    _xls_path = OUTPUT_DIR / f'productos_analisis_{PERIODO}.xlsx'
    try:
        _wb = _opxl.load_workbook(_xls_path)
        _ws = _wb.create_sheet('Valores_Documento')
        _hdr = _ws
        for row in _rows_doc:
            _ws.append(row)
        for _cell in _ws[1]:
            _cell.font = _opxl.styles.Font(bold=True)
        _ws.column_dimensions['A'].width = 20
        _ws.column_dimensions['B'].width = 42
        _ws.column_dimensions['C'].width = 20
        _ws.column_dimensions['D'].width = 16
        _wb.save(_xls_path)
        print(f'Hoja Valores_Documento agregada a: {_xls_path.name}')
    except Exception as _e:
        print(f'AVISO: No se pudo agregar hoja al Excel: {_e}')"""))

# ── Write notebook ──────────────────────────────────────────────────────────────
nb = {
    'cells': cells,
    'metadata': {
        'kernelspec': {'display_name':'Python 3','language':'python','name':'python3'},
        'language_info': {'codemirror_mode':{'name':'ipython','version':3},
                          'file_extension':'.py','mimetype':'text/x-python',
                          'name':'python','nbformat':4,'pygments_lexer':'ipython3','version':'3.10.0'}
    },
    'nbformat': 4,
    'nbformat_minor': 5
}

script_dir = os.path.dirname(os.path.abspath(__file__))
out_path = os.path.join(script_dir, '05_evolucion_productos_representativos.ipynb')
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump(nb, f, ensure_ascii=False, indent=1)

print(f'Written: {out_path}')
print(f'Cells: {len(cells)}')
