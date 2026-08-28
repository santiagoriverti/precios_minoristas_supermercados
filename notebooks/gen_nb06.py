"""Script to generate 06_evolucion_brecha_celiaca.ipynb — Brecha TACC vs sin-TACC.

Evalua la evolucion DIARIA, SEMANAL y MENSUAL de la BRECHA entre una canasta base
(productos con TACC) y su equivalente sin-TACC (canasta celiaca), usando SOLO tipos
de producto con dicotomia celiaca (sin diluir con limpieza/higiene/otros alimentos).

Metodologia (acordada con los investigadores):
- Cada TIPO de producto (fideos, galletitas, pan rallado, etc.) tiene 2-3 EANs TACC y
  2-3 EANs sin-TACC representativos. El precio del tipo en una sucursal/dia = PROMEDIO
  de los representativos PRESENTES (robusto a faltantes y a la eleccion del representativo).
- La brecha se calcula POOLED por grupo (nacional/provincia/cadena/localidad): para cada
  tipo se toma el precio TACC y el sin-TACC sobre las sucursales del grupo que ofrecen cada
  lado, y se arma la canasta base y celiaca. Esto es robusto a la BAJA COBERTURA sin-TACC
  (los productos sin-TACC estan en pocas gondolas, asi que exigir ambos lados en la MISMA
  sucursal el mismo dia da 0 obs). Se calcula ademas una brecha intra-sucursal por mes
  (best-effort) para las sucursales que si ofrecen ambos lados.
- Series por partida doble: MEDIANA (robusta) y PROMEDIO (outliers fuera), como nb02/nb05.
"""
import json, os, hashlib

def _cell_id(prefix, src):
    return prefix + hashlib.md5(src.encode('utf-8')).hexdigest()[:6]

def cell_md(src):
    lines = src.split('\n')
    source = [l + '\n' for l in lines[:-1]] + ([lines[-1]] if lines[-1] else [])
    return {'cell_type':'markdown','id':_cell_id('md', src),'metadata':{},'source':source}

def cell_code(src):
    return {'cell_type':'code','execution_count':None,'id':_cell_id('c', src),'metadata':{},'outputs':[],'source':[src]}

cells = []

# ── CELL 0 ─────────────────────────────────────────────────────────────────────
cells.append(cell_md("""# SEPA — Brecha Celíaca (TACC vs sin-TACC)

**Objetivo:** medir la **brecha** entre una canasta **base** (productos con TACC) y su
equivalente **sin-TACC** (canasta celíaca), y su evolución **diaria, semanal y mensual**,
desagregada por provincia, cadena, concentración de comercios y localidad.

**Metodología** (acordada con los investigadores):
- Solo **tipos de producto con dicotomía celíaca** (fideos, galletitas, pan rallado,
  harina/premezcla, rebozador, caldo, cerveza→sidra…). Nada de limpieza/higiene/otros
  alimentos: la brecha se reporta sobre esa canasta acotada, sin maquillar.
- Cada tipo usa **2–3 EANs TACC y 2–3 sin-TACC representativos**; el precio del tipo en
  una sucursal/día = **promedio de los representativos presentes** (robusto a faltantes).
- La brecha se calcula **pooled por grupo** (nacional/provincia/cadena/localidad): por tipo se
  toma el precio TACC y el sin-TACC sobre las sucursales del grupo que ofrecen cada lado. Es
  robusto a la **baja cobertura** de los sin-TACC (exigir ambos lados en la misma sucursal el
  mismo día da 0 obs). Hay además una brecha **intra-sucursal por mes** (best-effort).
- Series por partida doble: **mediana** (robusta) y **promedio** (outliers fuera).

**Config en la CELDA 1**: el diccionario `TIPOS` (tipo → EANs TACC / sin-TACC / cantidad).

**Estructura:** Config → Setup → Maestros → Parseo de tipos → ZIPs → Lectura diaria →
Brecha por sucursal×día → Series (diaria/semanal/mensual) → Gráficos → Mapa → Excel"""))

# ── CELL 1 — CONFIG ────────────────────────────────────────────────────────────
cells.append(cell_code("""\
# ===========================================================
# CONFIGURACIÓN — Modificar solo esta sección
# ===========================================================
# TIPOS: un diccionario tipo -> {qty, tacc:[EANs], sin_tacc:[EANs]}.
#   - Poné VARIOS EANs candidatos por lado (cuantos más, más cobertura por sucursal).
#   - En cada sucursal/día se usa el precio de LOS candidatos que esa sucursal tenga
#     (mediana robusta, $/100g). Un tipo cuenta EN ESA SUCURSAL si tiene al menos 1
#     candidato TACC y 1 sin-TACC → la brecha del tipo se calcula INTRA-SUCURSAL.
#   - 'qty' = peso del tipo en la canasta (para la brecha agregada por sucursal).
# LISTAS AMPLIAS curadas del Maestro (marcas mainstream + varias presentaciones).
# Ajustá con la hoja Cobertura del primer run (n_tacc/n_sin por tipo).
TIPOS = {
    'Fideos secos': {
        'qty': 4,
        # TACC: Lucchetti · Matarazzo · Favorita (todos 500 g) — alta cobertura (canasta_repr. 07/2026, ~1900-2500 suc.)
        'tacc':     ['7790070336385','7790070336118','7790070336149','7790070336316',
                     '7790070336293','7790070320285','7790070320308','7790070320292'],
        # sin-TACC (maíz/arroz/legumbres, mayormente 500 g) — AMPLIADA [2026-08-27] con marcas
        # mainstream (Matarazzo/Gallo GF, Blue Patna, Grandiet, Yuka) para maximizar cobertura; §3.8.
        'sin_tacc': ['7730114000780','7730114000797','7730114100015','7730114100046','7730114100077',
                     '7797330105590','7797330105606','7797330105613',
                     '7790070321800','7790070321794','7790070321855','7790070335999','7790070335982','7790070336224',
                     '7790070321817','7790070321824',
                     '7798031470024','7794903232240','7794903232257','7794903232264','7794903232271'],
    },
    'Galletitas dulces': {
        'qty': 3,
        # TACC: 9deOro · Don Satur · Chocolinas · Bagley Rumba · Sonrisas · Maná (galletita dulce base) — alta cobertura
        'tacc':     ['7792200000128','7795735000335','7790040143234','7790040143524',
                     '7790040133471','7790040137844'],
        # sin-TACC — AMPLIADA [2026-08-27]: Santa María (línea completa 200 g), Natuzen, Smams,
        # Arrozen, Nina. Se quitó 7798079230062 (Vainillas 12Un, sin gramos). §3.8.
        'sin_tacc': ['7798079230017','7798079230147','7798079230109','7798079230116','7798079230598',
                     '7798079230628','7798079230161','7798079230659','7798079230673',
                     '7798082000317','7798082000331','7798082000393','7798082000492',
                     '7798181510120','7798181510199','7798181510441','7798181510236','7798181510243',
                     '7798082001017','7798082001024','0655257736631','7798294150435','7798308250205'],
    },
    'Galletitas saladas / crackers': {
        'qty': 2,
        # TACC (base trigo): 9deOro · Don Satur · Traviata · Tosti · Hogareñas — alta cobertura
        'tacc':     ['7792200000159','7795735000328','7790040144095','7794529041608',
                     '7790040136069'],
        # sin-TACC (crackers/tostadas de arroz) — AMPLIADA [2026-08-27]: Tía Maruca, Carrefour,
        # Granix, Crisppino, Shiva, Olienka, Viavita, Yuka, Apóstoles. §3.8.
        'sin_tacc': ['7798199770035','7798199770042','7798199770103','7798199770110','7798199770134','7798199770141',
                     '7798289620080','7798289620097','0617308824087','0617308824094','0617308824100','0617308824117',
                     '7798195940173','7798082743726','7798082743849','7791720019054','7791720019061',
                     '7790045826859','7794903232226','7791120098857'],
    },
    'Pan rallado / rebozador': {
        'qty': 2,
        # TACC: Preferido · Mamá Cocina · Lucchetti · Pureza · Favorita (pan rallado) — alta cobertura
        'tacc':     ['7790070433169','7792180004741','7790070433275','7792180136480','7790070433312'],
        # sin-TACC (pan rallado/rebozador de arroz) — AMPLIADA [2026-08-27]: se sumaron marca
        # blanca de cadena (Carrefour) y mainstream (Preferido, Santa María, Natuzen) porque las
        # marcas de nicho previas (Bio/La Delfina/Maizena/Marvese) tenían ~0 cobertura. §3.8.
        'sin_tacc': ['7791720018453','7791720018460','7791720018477','7791720003411',
                     '7790070100283','7798079230086','7798082000294','7798082000553',
                     '7798221641845','7798221641944','7798131130200','7798131130231',
                     '7794000005303','7794000007291','7794000008212',
                     '7798306830164','7798306830171','7798306830201','7798075280559','7798239780123'],
    },
    # ⚠️ SUSTITUCIÓN (harina de trigo vs premezcla) — NO es el mismo producto sin gluten
    #    y su brecha es enorme (~+700%). Descomentá SOLO si la querés incluir; para una
    #    brecha "like-for-like" pura dejala comentada.
    # 'Harina / premezcla': {
    #     'qty': 2,
    #     'tacc':     ['7792180004512','7790070506924','7792180001528'],
    #     'sin_tacc': ['7790070508010','7794000005273','7798239780178'],
    # },
}

# EANs a EXCLUIR a mano (escape hatch documentado, §3.8). Se filtran en la CELDA 7 DESPUÉS
# de leer → NO cambian el hash del caché. Vacío por diseño: la limpieza de errores de carga
# la hace el FILTRO DE PLAUSIBILIDAD por regla (banda ×FACTOR_PLAUS, CELDA 7), que es simétrico
# y reproducible (preferible a sacar productos a mano en un paper). Usar este set solo para
# casos de NO-comparabilidad documentados que la regla no capture (hoy: ninguno).
EANS_EXCLUIR = set()
FACTOR_PLAUS = 4   # factor de la banda de plausibilidad a nivel EAN (data-quality, §3.8)

# ── Selección sin-TACC POR NOMBRE (regla reproducible) ─────────────────────────
# Además de las listas curadas de arriba, se agregan como candidatos sin-TACC TODOS los
# productos del maestro cuya descripción diga "sin tacc/gluten" y matcheen las keywords del
# tipo (menos las exclusiones). Con el maestro SEPA COMPLETO esto captura los celíacos que
# no están en el Maestro interno (p.ej. los que vende DIA). (kw_incluir, kw_excluir) por tipo.
SIN_TACC_REGLA = {
    'Fideos secos': (
        r'fideo|spaghetti|spagueti|tallarin|mostachol|tirabuz|penne|fusilli|codito|cintas|caracol|rigat|letras|moñito|farfalle|dedalitos|cornetti|risoni|canestri|spirali',
        r'premezcla|ñoqui|noqui|harina|frola|salsa|rallado|rebozador|tapa|tarta|sopa|salvado|arroz integral|yamani'),
    # dulces: REQUIERE sustantivo de galletita/bizcocho (no solo el sabor) → evita que
    # 'naranja'/'frutilla'/'chocolate' sueltos metan yerba, mermelada, caramelos, tabletas, etc.
    'Galletitas dulces': (
        r'galletit|galleta|biscuit|bizcoch|oblea|scon|pepa|magdalena|masita|vainilla|marmolad|nevadit|rueditas',
        r'salad|queso|jamon|cracker|tostada|arroz|pizza|premezcla|bizcochuelo|flan|postre|leche|grisin|palito|semilla|rallado|ensalada|hamburg|pochoclo|pan dulce|cereal|copitas|barrita|turron|budin|budín|yerba|mate|mermelada|caramelo|tableta|bocadito|bombon|gomita|chupetin|jugo|bebida|agua|infusion|pimenton|cacao|dulce de leche|polvo|helado|alfajor'),
    'Galletitas saladas / crackers': (
        r'cracker|tostada|tortita|talita|chalita|grisin|bizcochito|tostadita|(?:galletit|galleta).*(?:salad|queso|jamon|de arroz|pizza)|snack.*queso|palito.*(?:maiz|queso)',
        r'dulce|vainilla|chocolate|coco|limon|pepas|marmolad|premezcla|budin|oblea|alfajor|flan|leche|manteca|fideo|ensalada|hamburg|dambo|yamani|arroz integral|bizcochuelo|scon|rallado'),
    'Pan rallado / rebozador': (
        r'pan rallado|rebozador|rebozar|apanad',
        r'premezcla|coco|queso rallado'),
}

SEPA_SOURCE = 'mi_drive'   # 'mi_drive' | 'local'
SEPA_DIR    = '/content/drive/MyDrive/carga'
OUTPUT_DIR  = '/content/drive/MyDrive/carga/output_brecha'

USE_CACHE = True

# Mínimo de tipos (con ambos lados) para que una sucursal cuente en la brecha de CANASTA.
# La brecha POR TIPO no lo usa (basta ese tipo presente en la sucursal). 1 = permisivo.
MIN_TIPOS = 1
# Mínimo de sucursales para reportar una desagregación (provincia/cadena/localidad) como
# CONFIABLE. Las de menos quedan marcadas confiable=False y NO se usan en titulares ni plots
# (evita brechas basadas en 3-10 sucursales).
MIN_SUC_AGG = 30
# Panel temporal BALANCEADO: una sucursal entra en la serie balanceada si aparece en al menos
# esta fracción de los meses disponibles → trayectoria comparable (mitiga que la serie completa
# se confunda con la composición cuando entran/salen sucursales).
FRAC_MESES_PANEL = 0.7

# Período mínimo de la serie histórica (semanal/mensual)
MES_INICIO_HISTORICO = '2024-01'

# Ventana de la serie DIARIA (meses hacia atrás desde el último mes). La semanal y
# mensual usan TODO el histórico; la diaria solo esta ventana (evita gráficos gigantes).
VENTANA_DIARIA_MESES = 3

# Excluir estaciones de servicio / comercios no minoristas de referencia
CADENAS_FILTRAR = {'19', '2013', '3001', '4'}"""))

# ── CELL 2 — SETUP (idéntica a nb05, TMP/OUTPUT propios) ───────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 2 — Montar Drive + dependencias + imports
# ============================================================
try:
    import google.colab
    from google.colab import drive
    drive.mount('/content/drive')
    print('Google Drive montado')
except ImportError:
    print('Entorno local')

import subprocess, sys
subprocess.run([sys.executable, '-m', 'pip', 'install',
                'openpyxl', 'tqdm', 'pyarrow', '-q'], check=False)

import zipfile, gzip, re, shutil, warnings, gc, hashlib, calendar as _cal
import json as _json
from pathlib import Path
from tqdm.auto import tqdm
warnings.filterwarnings('ignore')

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import matplotlib.dates as mdates
from matplotlib.colors import LinearSegmentedColormap, Normalize
import seaborn as sns

plt.rcParams['figure.figsize'] = (13, 6)
plt.rcParams['font.size'] = 11
sns.set_theme(style='whitegrid')
pd.set_option('display.max_columns', 50)
pd.set_option('display.float_format', '{:,.2f}'.format)

SEPA_DIR   = Path(SEPA_DIR)
OUTPUT_DIR = Path(OUTPUT_DIR)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
CACHE_DIR  = OUTPUT_DIR / '_cache'
if USE_CACHE:
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
TMP_DIR = Path('/content/tmp_sepa_nb06')
TMP_DIR.mkdir(exist_ok=True)

GEOJSON_PATH = SEPA_DIR / 'ar.json'

def normalizar_ean(s):
    if pd.isna(s): return None
    s = str(s).strip().lstrip('0')
    return s if s else '0'

# EANs de referencia SOLO para detectar el factor centavos/pesos de cada mes.
REF_EANS_FACTOR = {'7790072002080'.lstrip('0'), '7790070320285'.lstrip('0'), '7790132098459'.lstrip('0')}

# ── Promedio con outliers fuera (banda relativa a la mediana) ────────────────
# Descarta valores fuera de [mediana/4, mediana x4] antes de promediar. Robusto a
# errores gruesos del SEPA a cualquier tamaño de muestra. (Igual que nb02/nb05.)
def _pmean(_s):
    _s = pd.to_numeric(_s, errors='coerce').dropna()
    if len(_s) == 0: return float('nan')
    _m = _s.median()
    if _m and _m > 0:
        _f = _s[(_s >= _m/4) & (_s <= _m*4)]
        if len(_f) > 0: _s = _f
    return _s.mean()

print(f'SEPA_DIR:   {SEPA_DIR}')
print(f'OUTPUT_DIR: {OUTPUT_DIR}')
print(f'  ar.json:  {"OK" if GEOJSON_PATH.exists() else "NO ENCONTRADO (mapa se saltea)"}')"""))

# ── CELL 3 — MAESTROS (idéntica a nb05) ────────────────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 3 — Maestros de sucursales, cadenas, provincias y productos
# ============================================================
DATA_URL = 'https://raw.githubusercontent.com/santiagoriverti/precios_minoristas_supermercados/main/data'

def leer_maestro(nombre, **kwargs):
    local = Path('data') / nombre
    if local.exists():
        return pd.read_excel(local, **kwargs)
    import urllib.request, urllib.parse
    dl = Path('/content/data') / nombre
    Path('/content/data').mkdir(exist_ok=True)
    if not dl.exists():
        print(f'  Descargando {nombre}...')
        urllib.request.urlretrieve(f'{DATA_URL}/{urllib.parse.quote(nombre)}', dl)
    return pd.read_excel(dl, **kwargs)

print('Cargando maestros...')
maestro_suc = leer_maestro('maestro_sucursales_completo.xlsx')
for c in ['id_comercio','id_bandera','id_sucursal']:
    maestro_suc[c] = maestro_suc[c].astype(str)

suc_pais = maestro_suc[
    maestro_suc['sucursales_latitud'].notna() &
    maestro_suc['sucursales_longitud'].notna() &
    (maestro_suc['sucursales_latitud'].between(-55, -22)) &
    (maestro_suc['sucursales_longitud'].between(-73, -53))
].copy()
IDS_PAIS = set(zip(suc_pais['id_comercio'], suc_pais['id_bandera'], suc_pais['id_sucursal']))
print(f'  Sucursales validas: {len(suc_pais):,} | IDs unicos: {len(IDS_PAIS):,}')

NOMBRES_COMPUESTOS = {
    ('9','1'):'Vea',('9','2'):'Disco',('9','3'):'Jumbo',
    ('10','1'):'Carrefour',('10','2'):'Carrefour Market',('10','3'):'Carrefour Express',
    ('11','2'):'ChangoMas',('11','4'):'Hiper ChangoMas',('11','5'):'Mi ChangoMas',
    ('16','1'):'Hipermercado Libertad',('16','2'):'Mini Libertad',
}
NOMBRES_SIMPLES = {
    '2':'La Anonima','3':'Cadena 3','5':'Hipermercado Misiones',
    '8':'Mariano Max','12':'Coto','13':'Cooperativa Obrera',
    '15':'DIA','20':'LAR','21':'Toledo','23':'Cadena 23','47':'Pasamonte',
}

def asignar_cadena(row):
    k = (row['id_comercio'], row['id_bandera'])
    if k in NOMBRES_COMPUESTOS: return NOMBRES_COMPUESTOS[k]
    if row['id_comercio'] in NOMBRES_SIMPLES: return NOMBRES_SIMPLES[row['id_comercio']]
    return f"Cadena {row['id_comercio']}"

PROV_NORM = {
    'Ciudad Autonoma de Buenos Aires':'CABA',
    'Ciudad Autónoma de Buenos Aires':'CABA',
    'Provincia de Buenos Aires':'Buenos Aires',
    'Provincia de Catamarca':'Catamarca',
    'Provincia del Chaco':'Chaco','Provincia del Chubut':'Chubut',
    'Provincia de Cordoba':'Córdoba','Provincia de Córdoba':'Córdoba',
    'Provincia de Corrientes':'Corrientes',
    'Provincia de Entre Rios':'Entre Ríos','Provincia de Entre Ríos':'Entre Ríos',
    'Provincia de Formosa':'Formosa','Provincia de Jujuy':'Jujuy',
    'Provincia de La Pampa':'La Pampa','Provincia de La Rioja':'La Rioja',
    'Provincia de Mendoza':'Mendoza','Provincia de Misiones':'Misiones',
    'Provincia del Neuquen':'Neuquén','Provincia del Neuquén':'Neuquén','Neuquén':'Neuquén',
    'Provincia de Rio Negro':'Río Negro','Provincia de Río Negro':'Río Negro',
    'Provincia de Salta':'Salta','Provincia de San Juan':'San Juan',
    'Provincia de San Luis':'San Luis','Provincia de Santa Cruz':'Santa Cruz',
    'Provincia de Santa Fe':'Santa Fe',
    'Provincia de Santiago del Estero':'Santiago del Estero',
    'Provincia de Tierra del Fuego, Antartida e Islas del Atlantico Sur':'Tierra del Fuego',
    'Provincia de Tierra del Fuego, Antártida e Islas del Atlántico Sur':'Tierra del Fuego',
    'Tierra del Fuego':'Tierra del Fuego',
    'Provincia de Tucuman':'Tucumán','Provincia de Tucumán':'Tucumán',
    'Buenos Aires':'Buenos Aires','CABA':'CABA','Catamarca':'Catamarca',
    'Chaco':'Chaco','Chubut':'Chubut','Córdoba':'Córdoba','Corrientes':'Corrientes',
    'Entre Ríos':'Entre Ríos','Formosa':'Formosa','Jujuy':'Jujuy',
    'La Pampa':'La Pampa','La Rioja':'La Rioja','Mendoza':'Mendoza',
    'Misiones':'Misiones','Río Negro':'Río Negro','Salta':'Salta',
    'San Juan':'San Juan','San juan':'San Juan','SAN JUAN':'San Juan',
    'San Luis':'San Luis','Santa Cruz':'Santa Cruz',
    'Santa Fe':'Santa Fe','Santiago del Estero':'Santiago del Estero','Tucumán':'Tucumán',
}
PESOS_POBLACION = {
    'Buenos Aires':17709732,'CABA':3075646,'Catamarca':415438,
    'Chaco':1204541,'Chubut':618994,'Córdoba':3978984,
    'Corrientes':1120801,'Entre Ríos':1385961,'Formosa':605193,
    'Jujuy':770881,'La Pampa':368550,'La Rioja':393531,
    'Mendoza':2014533,'Misiones':1261294,'Neuquén':664057,
    'Río Negro':747610,'Salta':1441998,'San Juan':781217,
    'San Luis':531745,'Santa Cruz':333473,'Santa Fe':3556522,
    'Santiago del Estero':1019304,'Tierra del Fuego':190641,'Tucumán':1737127,
}

_mp_raw = leer_maestro('Maestro de Productos Interno.xlsx', dtype=str,
                       usecols=['producto_sepa_id','producto_descripcion','producto_marca','rubro',
                                'producto_cantidad_presentacion','producto_unidad_medida_presentac'])
# Maestro SEPA COMPLETO (opcional, del Drive carga/): trae TODOS los productos que se venden
# con nombre/presentación → captura celíacos que NO están en el interno (p.ej. los de DIA).
_msepa = None
try:
    _msp = Path(SEPA_DIR) / 'maestro_sepa_completo.csv.gz'
    if _msp.exists():
        _msepa = pd.read_csv(_msp, dtype=str)
        print(f'  Maestro SEPA completo (Drive): {len(_msepa):,} EANs')
except Exception as _e:
    print(f'  (aviso: no se pudo leer maestro_sepa_completo.csv.gz: {_e})')

def _prep_master(_df):
    _df = _df.rename(columns={'producto_descripcion':'descripcion','producto_marca':'marca'})
    if 'rubro' not in _df.columns: _df['rubro'] = ''
    _df['ean_norm'] = _df['producto_sepa_id'].map(normalizar_ean)
    return _df[['ean_norm','descripcion','marca','rubro',
                'producto_cantidad_presentacion','producto_unidad_medida_presentac']]

_base = _prep_master(_mp_raw)
if _msepa is not None:
    _extra = _prep_master(_msepa)
    _extra = _extra[~_extra['ean_norm'].isin(set(_base['ean_norm']))]   # solo EANs que faltan
    MP_META = pd.concat([_base, _extra], ignore_index=True)
    print(f'  Fusión de maestros: interno {len(_base):,} + {len(_extra):,} nuevos del SEPA = {len(MP_META):,} EANs')
else:
    MP_META = _base

# gramos/ml de la presentación → para normalizar el precio a $/100g. Se toma del CAMPO
# presentación y, si no da (muchos productos —p.ej. de DIA— reportan '1 UNI' y los gramos van
# en el TEXTO 'X 400 GR'), se PARSEA de la descripción.
def _to_gramos(_q, _u):
    _q = pd.to_numeric(str(_q).replace(',', '.'), errors='coerce'); _u = str(_u).strip().lower()
    if pd.isna(_q) or _q <= 0: return np.nan
    if _u in ('kg','kgm','l','lt','litro','litros','kilogramo','kilogramos'): return _q * 1000
    if _u in ('gr','g','grs','gramo','gramos','ml','cc','mililitro'): return _q
    return np.nan
_RE_GR = re.compile(r'(\\d+(?:[.,]\\d+)?)\\s*(kg|kgm|kilo|grs?|gramos?|ml|cc|lts?|litros?|g)\\b')
def _gramos_desc(_s):
    _m = _RE_GR.search(str(_s).lower())
    if not _m: return np.nan
    _v = float(_m.group(1).replace(',', '.')); _u = _m.group(2)
    if _u.startswith('k') or _u in ('lt','l','lts','litro','litros'): return _v * 1000
    return _v
_g1 = [_to_gramos(a, b) for a, b in zip(MP_META['producto_cantidad_presentacion'],
                                        MP_META['producto_unidad_medida_presentac'])]
_g2 = [_gramos_desc(x) for x in MP_META['descripcion']]
MP_META['grams'] = [(_a if (_a == _a) else _b) for _a, _b in zip(_g1, _g2)]
MP_META = (MP_META.dropna(subset=['ean_norm']).drop_duplicates('ean_norm')
           .set_index('ean_norm')[['descripcion','marca','rubro','grams']])
print(f'  Maestro de productos (total): {len(MP_META):,} EANs con metadata')
print('Maestros OK')"""))

# ── CELL 4 — PARSEO DE TIPOS (config) ──────────────────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 4 — Parseo de tipos TACC / sin-TACC (CELDA 1)
# ============================================================
def _norm_lista(_lst):
    _out = []
    for _e in _lst:
        _limpio = re.sub(r'\\D', '', str(_e))
        if _limpio:
            _out.append(_limpio.lstrip('0') or '0')
    return _out

TIPO_TACC = {}      # tipo -> set(ean_norm) TACC
TIPO_SIN  = {}      # tipo -> set(ean_norm) sin-TACC
TIPO_QTY  = {}      # tipo -> cantidad
EAN_TIPO  = {}      # ean_norm -> tipo
EAN_ROL   = {}      # ean_norm -> 'tacc' | 'sin'
EAN_DESC  = {}      # ean_norm -> descripcion (del maestro o el propio EAN)

# Expansión sin-TACC POR NOMBRE: candidatos = lista curada ∪ {EANs del maestro cuya
# descripción diga "sin tacc/gluten" y matcheen las keywords del tipo (§ SIN_TACC_REGLA)}.
_desc_all = MP_META['descripcion'].fillna('').astype(str).str.lower()
_gf_all = _desc_all.str.contains(r'sin tacc|sin gluten|libre de gluten|s/tacc', regex=True)
_REGLAS_SIN = globals().get('SIN_TACC_REGLA', {})
def _sin_por_regla(_tp):
    _r = _REGLAS_SIN.get(_tp)
    if not _r: return set()
    _kw, _ex = _r
    _m = _gf_all & _desc_all.str.contains(_kw, regex=True) & ~_desc_all.str.contains(_ex, regex=True)
    return set(MP_META.index[_m])

for _tipo, _cfg in TIPOS.items():
    _tacc = set(_norm_lista(_cfg.get('tacc', [])))
    _sin  = set(_norm_lista(_cfg.get('sin_tacc', []))) | _sin_por_regla(_tipo)
    if not _tacc or not _sin:
        print(f'AVISO: el tipo \"{_tipo}\" no tiene EANs en ambos lados — se ignora.')
        continue
    TIPO_TACC[_tipo] = _tacc
    TIPO_SIN[_tipo]  = _sin
    TIPO_QTY[_tipo]  = float(_cfg.get('qty', 1))
    for _e in _tacc: EAN_TIPO[_e] = _tipo; EAN_ROL[_e] = 'tacc'
    for _e in _sin:  EAN_TIPO[_e] = _tipo; EAN_ROL[_e] = 'sin'

if not TIPO_TACC:
    raise ValueError('Ningún tipo válido en TIPOS (cada tipo necesita EANs TACC y sin-TACC).')

EANS_CONFIG = set(EAN_TIPO.keys())
EAN_GRAMS = {}      # ean_norm -> gramos/ml de la presentación (para normalizar a $/100g)
_sin_gramos = []
for _e in EANS_CONFIG:
    if _e in MP_META.index and pd.notna(MP_META.loc[_e, 'descripcion']):
        EAN_DESC[_e] = str(MP_META.loc[_e, 'descripcion'])[:50]
    else:
        EAN_DESC[_e] = f'EAN {_e}'
    _g = MP_META.loc[_e, 'grams'] if _e in MP_META.index else np.nan
    if pd.notna(_g) and _g > 0:
        EAN_GRAMS[_e] = float(_g)
    else:
        _sin_gramos.append(_e)
if _sin_gramos:
    print(f'AVISO: {len(_sin_gramos)} EAN sin presentación en el maestro — usan precio por paquete '
          f'(NO normalizado a $/100g): {_sin_gramos}')

# ── Subtipo / EQUIVALENCIA dentro de cada tipo ───────────────────────────────
# Para comparar producto-vs-producto equivalente, la GRANULARIDAD depende del tipo, porque
# solo en algunos tipos la sub-variedad alinea entre TACC y sin-TACC:
#  · Fideos: por FORMA (Spaghetti↔Spaghetti, Tirabuzón↔Tirabuzón…), que sí es equivalencia real.
#  · Pan rallado, Galletitas dulces, Galletitas saladas: los surtidos TACC y sin-TACC son
#    variedades DISTINTAS (bizcocho agridulce vs vainilla; galleta de trigo vs de arroz), así
#    que la equivalencia honesta es a NIVEL TIPO → un solo grupo (evita filas espurias/vacías).
_SUBTIPO_KW = {   # solo tipos con equivalencia fina (por forma)
    'Fideos secos': [
        ('Spaghetti', ['spaghetti','spagueti','spaghetti','spaguetti','spaghettini','cabello','vermicelli','fideo fino','fideos finos']),
        ('Tallarín/Cintas', ['tallarin','tallarín','fettuccini','fetuccini','fettuccine','cintas','nido','canestri']),
        ('Mostachol', ['mostachol','rigati','rigato']),
        ('Tirabuzón/Fusilli', ['tirabuzon','tirabuzón','fusilli','fusili','caracol','spirali']),
        ('Penne', ['penne','pluma','plumita','rigate']),
        ('Codito/Corto', ['codito','dedalitos','tornillo','moñito','moñitos','farfalle','risoni','cornetti','coditos']),
        ('Ñoquis', ['ñoqui','noqui']),
    ],
}
_SUBTIPO_FIJO = {   # tipos donde la equivalencia es a nivel tipo (un solo grupo)
    'Pan rallado / rebozador': 'Pan rallado / rebozador',
    'Galletitas dulces': 'Galletita dulce',
    'Galletitas saladas / crackers': 'Galleta / cracker salado',
}
def _subtipo(_desc, _tipo):
    if _tipo in _SUBTIPO_FIJO:
        return _SUBTIPO_FIJO[_tipo]
    _d = str(_desc).lower()
    for _lab, _kws in _SUBTIPO_KW.get(_tipo, []):
        if any(_k in _d for _k in _kws):
            return _lab
    return 'Otros'
EAN_SUBTIPO = {_e: _subtipo(EAN_DESC.get(_e, ''), EAN_TIPO[_e]) for _e in EANS_CONFIG}

print(f'Tipos activos: {len(TIPO_TACC)} | EANs de config: {len(EANS_CONFIG)}')
for _t in TIPO_TACC:
    print(f'  [{_t}] qty={TIPO_QTY[_t]:.0f} | TACC: {len(TIPO_TACC[_t])} · sin-TACC: {len(TIPO_SIN[_t])}')"""))

# ── CELL 5 — ZIP FUNCTIONS (idéntica a nb05) ───────────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 5 — Funciones de lectura de ZIPs SEPA
# ============================================================
_PAT_SEM  = re.compile(r'^(\\d{4})(A|B)$', re.IGNORECASE)
_PAT_ARC  = re.compile(r'^(\\d{2})(\\d{4})_pais_parte.*COMPLETO.*\\.csv\\.gz$', re.IGNORECASE)
PAT_FECHA = re.compile(r'^precio_(\\d{8})$')

def detectar_semestres():
    result = []
    for z in sorted(SEPA_DIR.glob('*.zip')):
        m = _PAT_SEM.match(z.stem)
        if m:
            result.append((z, int(m.group(1)), m.group(2).upper()))
    return result

def archivos_por_mes(zip_path):
    meses = {}
    with zipfile.ZipFile(zip_path) as zf:
        for nombre in zf.namelist():
            m = _PAT_ARC.match(Path(nombre).name)
            if m:
                mes_n, anio = int(m.group(1)), int(m.group(2))
                meses.setdefault((anio, mes_n), []).append(nombre)
    return meses

# Mapa de meses disponibles -> (zip, archivos)
_mapa_mes = {}
for _zip_path, _anio, _sem in detectar_semestres():
    for (_anio_m, _mes_m), _archs in archivos_por_mes(_zip_path).items():
        _lbl = f'{_anio_m}-{_mes_m:02d}'
        if _lbl >= MES_INICIO_HISTORICO:
            _mapa_mes[_lbl] = (_zip_path, _archs)
_meses_disp = sorted(_mapa_mes)
if not _meses_disp:
    raise RuntimeError(f'No hay meses disponibles >= {MES_INICIO_HISTORICO}')
_mes_actual = _meses_disp[-1]   # mes en curso -> NUNCA se cachea
print(f'Meses disponibles: {len(_meses_disp)}  ({_meses_disp[0]} → {_meses_disp[-1]})')"""))

# ── CELL 6 — LECTURA DIARIA (todos los meses, dimensión fecha) ─────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 6 — Lectura MENSUAL de precios (una fila por sucursal×EAN×mes)
# ============================================================
# Lee los EANs de config en todos los meses >= MES_INICIO_HISTORICO. Dentro de cada
# mes COLAPSA los días a la MEDIANA mensual por (sucursal, EAN) → footprint ~30× menor
# (clave para no reventar la RAM de Colab cuando los EANs tienen mucha cobertura). El
# método vigente agrega por MES, así que no se pierde nada. Cache de meses cerrados.
_cache_key    = hashlib.md5('|'.join(sorted(EANS_CONFIG)).encode()).hexdigest()[:8]
_cache_path   = CACHE_DIR / f'brecha_dia_{_cache_key}_v2.parquet'
_EANS_LECTURA = EANS_CONFIG | REF_EANS_FACTOR

def _leer_mes_dia(_lbl):
    _zip_path, _archs = _mapa_mes[_lbl]
    _rows = []; _muestra_ref = []
    for _archivo in sorted(_archs):
        _tmp_p = TMP_DIR / Path(_archivo).name
        with zipfile.ZipFile(_zip_path) as _zf:
            with _zf.open(_archivo) as _s, open(_tmp_p, 'wb') as _d:
                shutil.copyfileobj(_s, _d, length=4*1024*1024)
        with gzip.open(_tmp_p, 'rt', encoding='utf-8', errors='replace') as _g:
            for _chunk in pd.read_csv(_g, dtype=str, chunksize=300_000, low_memory=False):
                _chunk['ean_norm'] = _chunk['id_producto'].apply(normalizar_ean)
                _chunk = _chunk[_chunk['ean_norm'].isin(_EANS_LECTURA)].copy()
                if len(_chunk) == 0: continue
                for _c in ['id_comercio','id_bandera','id_sucursal']:
                    _chunk[_c] = _chunk[_c].astype(str)
                _chunk['_k'] = list(zip(_chunk['id_comercio'],_chunk['id_bandera'],_chunk['id_sucursal']))
                _chunk = _chunk[_chunk['_k'].isin(IDS_PAIS)].drop(columns=['_k']).copy()
                if len(_chunk) == 0: continue
                _cols_p = [c for c in _chunk.columns if PAT_FECHA.match(c)]
                if not _cols_p: continue
                _mlt = _chunk.melt(
                    id_vars=['id_comercio','id_bandera','id_sucursal','ean_norm'],
                    value_vars=_cols_p, var_name='_col', value_name='precio_raw')
                _mlt['precio'] = pd.to_numeric(_mlt['precio_raw'].replace('NA', np.nan), errors='coerce')
                _mlt = _mlt[_mlt['precio'].notna() & (_mlt['precio'] > 0)].copy()
                _mlt['fecha'] = pd.to_datetime(_mlt['_col'].str[-8:], format='%Y%m%d', errors='coerce')
                _mlt = _mlt[_mlt['fecha'].notna()]
                _es_ref = _mlt['ean_norm'].isin(REF_EANS_FACTOR)
                if _es_ref.any():
                    _muestra_ref.extend(_mlt.loc[_es_ref, 'precio'].tolist())
                _mlt = _mlt[_mlt['ean_norm'].isin(EANS_CONFIG)]
                if len(_mlt) > 0:
                    # Colapso parcial por chunk: mediana del mes por (suc, EAN) sobre los días
                    # del chunk. Reduce el pico de RAM antes de concatenar (los días de un
                    # mismo mes pueden venir en varios archivos/chunks → se re-agrega al final).
                    _rows.append(_mlt.groupby(
                        ['id_comercio','id_bandera','id_sucursal','ean_norm'],
                        as_index=False)['precio'].median())
        _tmp_p.unlink(missing_ok=True)
    if not _rows:
        return None
    _df = pd.concat(_rows, ignore_index=True)
    _med_ref = (pd.Series(_muestra_ref).median() if _muestra_ref else _df['precio'].median())
    _fac = 100 if _med_ref > 10_000 else 1
    if _fac == 100: _df['precio'] /= 100
    # Mediana mensual definitiva por (sucursal, EAN) — colapsa los distintos chunks/archivos
    _df = _df.groupby(['id_comercio','id_bandera','id_sucursal','ean_norm'],
                      as_index=False)['precio'].median()
    _df['mes'] = _lbl
    del _rows, _muestra_ref; gc.collect()
    return _df

# Meses cerrados: cache incremental
if USE_CACHE and _cache_path.exists():
    df_cache = pd.read_parquet(_cache_path)
    df_cache = df_cache[df_cache['mes'] < _mes_actual].copy()
else:
    df_cache = pd.DataFrame(columns=['id_comercio','id_bandera','id_sucursal','ean_norm','precio','mes'])

_en_cache = set(df_cache['mes'].unique())
_faltantes = [m for m in _meses_disp if m < _mes_actual and m not in _en_cache]
_nuevos = []
for _lbl in tqdm(_faltantes, desc='Meses cerrados'):
    _d = _leer_mes_dia(_lbl)
    if _d is not None: _nuevos.append(_d)
if _nuevos:
    df_cache = pd.concat([df_cache] + _nuevos, ignore_index=True)
    if USE_CACHE:
        df_cache.to_parquet(_cache_path, compression='snappy', index=False)
        print(f'Cache actualizado: {_cache_path.name} ({df_cache["mes"].nunique()} meses cerrados)')

# Mes en curso: siempre fresco
_actual = _leer_mes_dia(_mes_actual)
datos_dia = (pd.concat([df_cache] + ([_actual] if _actual is not None else []), ignore_index=True)
             if (len(df_cache) > 0 or _actual is not None)
             else pd.DataFrame(columns=df_cache.columns))
if len(datos_dia) == 0:
    raise RuntimeError('Sin datos para los EANs de config. Revisá los EANs de TIPOS (CELDA 1).')
datos_dia = datos_dia.sort_values('mes').reset_index(drop=True)

ULTIMO_MES = _mes_actual
_NOM = {'01':'enero','02':'febrero','03':'marzo','04':'abril','05':'mayo','06':'junio',
        '07':'julio','08':'agosto','09':'septiembre','10':'octubre','11':'noviembre','12':'diciembre'}
NOMBRE_MES_TITLE = f"{_NOM[_mes_actual[5:7]]} {_mes_actual[:4]}".title()
print(f'Observaciones (sucursal×EAN×mes): {len(datos_dia):,} | EANs con datos: {datos_dia["ean_norm"].nunique()}/{len(EANS_CONFIG)}')
print(f'Meses: {datos_dia["mes"].min()} → {datos_dia["mes"].max()} | Sucursales: {datos_dia.groupby(["id_comercio","id_bandera","id_sucursal"]).ngroups:,}')"""))

# ── CELL 7 — BRECHA POR SUCURSAL × DÍA ─────────────────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 7 — Brecha INTRA-SUPERMERCADO por tipo (por sucursal, sobre el mes)
# ============================================================
# Para cada SUCURSAL y MES, por tipo, el precio de cada lado = MEDIANA $/100g (§3.3) de
# los candidatos que esa sucursal tuvo ese mes. Un tipo cuenta en esa sucursal-mes si tuvo
# ≥1 candidato TACC y ≥1 sin-TACC → la brecha del tipo se calcula DENTRO del mismo super
# (sin exigir el MISMO día → mucho más robusto). Luego se promedia entre supers (y por
# zonas/provincia/cadena). Se guarda en paralelo el "más barato" (mínimo) como referencia
# ilustrativa (sesgada por el nº desigual de candidatos por lado; §3.7).
datos_dia['tipo'] = datos_dia['ean_norm'].map(EAN_TIPO)
datos_dia['rol']  = datos_dia['ean_norm'].map(EAN_ROL)
datos_dia = datos_dia[~datos_dia['id_comercio'].isin(CADENAS_FILTRAR)].copy()
# Exclusión de EANs por curación de comparabilidad (§3.8) — cache-preserving
_ex = EANS_EXCLUIR if 'EANS_EXCLUIR' in dir() else set()
if _ex:
    _n0e = len(datos_dia)
    datos_dia = datos_dia[~datos_dia['ean_norm'].isin(_ex)].copy()
    print(f'  Excluidos {len(_ex)} EANs por curación (§3.8): {_n0e - len(datos_dia):,} obs')

# Precio normalizado a $/100g. Los EANs SIN presentación en el maestro se EXCLUYEN
# (definición §3.8 de docs/BRECHA_CELIACA.md: la brecha va siempre en $/100g; no se
# mezcla precio por paquete). Filtrar acá NO cambia el hash del caché (cache-preserving).
datos_dia['grams'] = datos_dia['ean_norm'].map(EAN_GRAMS)
_n0 = len(datos_dia)
datos_dia = datos_dia[datos_dia['grams'].notna() & (datos_dia['grams'] > 0)].copy()
_n_excl = _n0 - len(datos_dia)
if _n_excl:
    print(f'  Excluidas {_n_excl:,} obs de EANs sin presentación (no normalizables a $/100g)')
datos_dia['precio_100'] = datos_dia['precio'] / datos_dia['grams'] * 100

# ── Filtro de plausibilidad a nivel EAN (data-quality, §3.8) ──────────────────
# Regla SIMÉTRICA, reproducible e INFLACIÓN-ROBUSTA (no curación a mano): cada EAN se mide
# RELATIVO a sus pares CONTEMPORÁNEOS. Referencia = mediana de los precios-EAN dentro de
# (tipo, lado, MES), ponderada POR PRODUCTO (no por sucursal → la cobertura no sesga la
# referencia). Un EAN se descarta si su precio relativo mediano (en el panel) cae fuera de
# [1/FACTOR_PLAUS, FACTOR_PLAUS]. Medir DENTRO del mes evita que la inflación 2024-2026
# confunda la banda; así se quitan solo errores de carga groseros, sin elegir a mano ni
# favorecer la dirección de la brecha. FACTOR_PLAUS = parámetro de robustez documentado.
_fp = FACTOR_PLAUS if 'FACTOR_PLAUS' in dir() else 4
_eanm = datos_dia.groupby(['tipo','rol','mes','ean_norm'])['precio_100'].median().reset_index(name='_p')
_eanm['_ref'] = _eanm.groupby(['tipo','rol','mes'])['_p'].transform('median')
_eanm['_rel'] = _eanm['_p'] / _eanm['_ref']
_emr = _eanm.groupby('ean_norm')['_rel'].median()
_bad = set(_emr[(_emr < 1/_fp) | (_emr > _fp)].index)
if _bad:
    _n0p = len(datos_dia)
    _desc = MP_META['descripcion'] if 'MP_META' in dir() else pd.Series(dtype=str)
    datos_dia = datos_dia[~datos_dia['ean_norm'].isin(_bad)].copy()
    print(f'  Filtro plausibilidad (banda ×{_fp}, inflación-robusta): {len(_bad)} EAN(s) fuera → '
          f'{_n0p - len(datos_dia):,} obs quitadas (data-quality):')
    for _e in sorted(_bad):
        print(f'      - {_e}  rel={_emr[_e]:.2f}  {str(_desc.get(_e, "?"))[:50]}')

# Precio del tipo por (sucursal, mes, lado) — definición §3.3: MEDIANA $/100g de los
# candidatos presentes (estimador PRIMARIO: estable a la asimetría en el nº de candidatos
# por lado). En paralelo se calcula el "más barato" (mínimo robusto) SOLO como referencia
# ilustrativa — está sesgado al alza porque el lado TACC tiene muchos más candidatos que
# el sin-TACC (min de más tiros baja más); ver §3.7.
def _min_robusto(_s):
    _s = pd.to_numeric(_s, errors='coerce').dropna()
    if len(_s) == 0: return float('nan')
    _m = _s.median()
    if _m and _m > 0:
        _f = _s[(_s >= _m/4) & (_s <= _m*4)]
        if len(_f) > 0: _s = _f
    return _s.min()
_sk = ['id_comercio','id_bandera','id_sucursal']
tp = (datos_dia.groupby(_sk + ['mes','tipo','rol'])['precio_100']
      .agg(precio_rep='median', precio_min=_min_robusto).reset_index())

# ── Geo (una vez) ────────────────────────────────────────────────────────────
_PROV_BBOX = {
    'CABA':(-34.72,-34.52,-58.54,-58.33),'Tucumán':(-28.0,-26.0,-66.5,-64.5),
    'Jujuy':(-24.5,-21.5,-67.5,-63.5),'Misiones':(-28.5,-25.5,-56.5,-53.0),
    'Chaco':(-27.5,-24.0,-63.0,-57.5),'Formosa':(-26.5,-22.0,-62.5,-58.0),
    'Corrientes':(-30.5,-27.0,-60.0,-55.5),'Entre Ríos':(-34.0,-30.0,-60.5,-57.5),
    'San Luis':(-36.0,-32.5,-68.5,-65.0),'San Juan':(-34.5,-27.5,-71.0,-65.0),
    'La Rioja':(-32.5,-27.0,-70.0,-65.0),'Catamarca':(-29.5,-25.0,-70.5,-64.5),
    'Salta':(-26.5,-21.5,-68.5,-62.5),'Santiago del Estero':(-30.0,-25.5,-65.5,-61.5),
    'Mendoza':(-37.5,-32.0,-70.5,-66.5),'Neuquén':(-40.5,-36.0,-71.5,-68.5),
    'La Pampa':(-40.0,-35.0,-68.5,-63.5),'Santa Fe':(-34.5,-28.5,-62.5,-59.0),
    'Córdoba':(-39.0,-29.5,-67.0,-62.0),'Río Negro':(-42.5,-38.5,-71.5,-62.5),
    'Chubut':(-46.5,-41.0,-72.5,-63.0),'Buenos Aires':(-42.5,-33.5,-63.5,-56.5),
    'Santa Cruz':(-52.5,-46.0,-72.5,-65.5),'Tierra del Fuego':(-55.5,-51.0,-70.5,-63.5),
}
def _geocodif(lat, lon):
    if pd.isna(lat) or pd.isna(lon): return None
    for p,(la0,la1,lo0,lo1) in _PROV_BBOX.items():
        if la0 <= lat <= la1 and lo0 <= lon <= lo1: return p
    return None
_cols_suc = ['id_comercio','id_bandera','id_sucursal','sucursales_nombre',
             'sucursales_latitud','sucursales_longitud','sucursales_localidad','PROVINCIA']
suc_geo = suc_pais[_cols_suc].copy()
suc_geo['cadena'] = suc_geo.apply(asignar_cadena, axis=1)
suc_geo['PROVINCIA_NORM'] = suc_geo['PROVINCIA'].map(PROV_NORM).fillna(suc_geo['PROVINCIA'])
_n_recl = 0
for _idx,_row in suc_geo.iterrows():
    _p = _row['PROVINCIA_NORM']
    if _p not in _PROV_BBOX: continue
    _la0,_la1,_lo0,_lo1 = _PROV_BBOX[_p]
    if _la0 <= _row['sucursales_latitud'] <= _la1 and _lo0 <= _row['sucursales_longitud'] <= _lo1: continue
    _nueva = _geocodif(_row['sucursales_latitud'], _row['sucursales_longitud'])
    if _nueva and _nueva != _p:
        suc_geo.at[_idx,'PROVINCIA_NORM'] = _nueva; _n_recl += 1
if _n_recl: print(f'  Reclasificadas {_n_recl} sucursales por coordenadas')

tp = tp.merge(suc_geo[['id_comercio','id_bandera','id_sucursal','cadena','PROVINCIA_NORM',
                       'sucursales_localidad','sucursales_nombre']],
              on=['id_comercio','id_bandera','id_sucursal'], how='inner')
tp['localidad'] = (tp['sucursales_localidad'].astype(str).str.strip()
                   .replace({'':'N/D','nan':'N/D','None':'N/D'}))

# ── DIAGNÓSTICO de cobertura por tipo ────────────────────────────────────────
_cob_rows = []
print('=== Cobertura por tipo (sucursales que ofrecen cada lado, todo el período) ===')
for _t in TIPO_TACC:
    _st = tp[tp['tipo'] == _t]
    _set_t = set(map(tuple, _st[_st['rol']=='tacc'][_sk].drop_duplicates().to_numpy()))
    _set_s = set(map(tuple, _st[_st['rol']=='sin'][_sk].drop_duplicates().to_numpy()))
    _both  = _set_t & _set_s
    print(f'  [{_t}] TACC: {len(_set_t):>5} · sin-TACC: {len(_set_s):>5} · con AMBOS: {len(_both):>5}')
    _cob_rows.append({'tipo':_t,'sucursales_tacc':len(_set_t),'sucursales_sin_tacc':len(_set_s),'sucursales_ambos':len(_both)})
cobertura_tipo = pd.DataFrame(_cob_rows)

# ── ATÓMICO: brecha por (sucursal, mes, tipo) con AMBOS lados en ese super-mes ─
_idx = _sk + ['mes','PROVINCIA_NORM','cadena','localidad','sucursales_nombre','tipo']
_tacc = tp[tp['rol']=='tacc'][_idx + ['precio_rep','precio_min']].rename(columns={'precio_rep':'tacc','precio_min':'tacc_min'})
_sin  = tp[tp['rol']=='sin'][_idx + ['precio_rep','precio_min']].rename(columns={'precio_rep':'sin','precio_min':'sin_min'})
bt_sm = _tacc.merge(_sin, on=_idx, how='inner')   # inner = ambos lados en el mismo super-mes
bt_sm['brecha_pct'] = (bt_sm['sin'] / bt_sm['tacc'] - 1) * 100          # PRIMARIO: mediana de candidatos (§3.4)
bt_sm['brecha_min'] = (bt_sm['sin_min'] / bt_sm['tacc_min'] - 1) * 100  # ilustrativo "más barato" (sesgo por nº de candidatos, §3.7)
bt_sm['qty'] = bt_sm['tipo'].map(TIPO_QTY)
bt_sm['base_c'] = bt_sm['tacc'] * bt_sm['qty']
bt_sm['cel_c']  = bt_sm['sin']  * bt_sm['qty']

# ── Canasta por (sucursal, mes): suma sobre los tipos con ambos lados ────────
if len(bt_sm):
    brecha_suc_mes = (bt_sm.groupby(_sk + ['mes','PROVINCIA_NORM','cadena','localidad'])
                      .agg(base=('base_c','sum'), celiaca=('cel_c','sum'), n_tipos=('tipo','nunique'))
                      .reset_index())
    brecha_suc_mes = brecha_suc_mes[brecha_suc_mes['n_tipos'] >= MIN_TIPOS].copy()
    brecha_suc_mes['brecha_pct'] = (brecha_suc_mes['celiaca'] / brecha_suc_mes['base'] - 1) * 100
else:
    brecha_suc_mes = pd.DataFrame(columns=_sk + ['mes','PROVINCIA_NORM','cadena','localidad',
                                                 'base','celiaca','n_tipos','brecha_pct'])

_nsuc_bt = bt_sm[_sk].drop_duplicates().shape[0] if len(bt_sm) else 0
print(f'\\nObs (sucursal×mes×tipo) con ambos lados: {len(bt_sm):,} | sucursales con ≥1 par: {_nsuc_bt:,}')
print(f'Canasta por (sucursal×mes, ≥{MIN_TIPOS} tipos): {len(brecha_suc_mes):,} obs'
      + (f' | brecha mediana {brecha_suc_mes["brecha_pct"].median():+.1f}%' if len(brecha_suc_mes) else ''))"""))

# ── CELL 8 — SERIES (diaria / semanal / mensual) + desagregaciones ─────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 8 — Brecha POR TIPO (intra-super) + evolución + desagregaciones
# ============================================================
# Todo se deriva de bt_sm (brecha por sucursal×mes×tipo) y brecha_suc_mes (canasta
# por sucursal×mes). Se promedia ENTRE sucursales (mediana + promedio robusto).
# brecha_mediana/_prom = agregación ENTRE sucursales (mediana/promedio) de la brecha
# intra-super, que usa la MEDIANA de candidatos (§3.4). brecha_min_ilustr = misma
# agregación pero con el "más barato" → referencia ilustrativa, sesgada al alza (§3.7).
def _agg_tipo(df, keys):
    _cols = keys + ['tipo','tacc_100','sin_100','brecha_mediana','brecha_prom','brecha_min_ilustr','n_sucursales']
    if len(df) == 0:
        return pd.DataFrame(columns=_cols)
    return (df.groupby(keys + ['tipo'])
            .agg(tacc_100=('tacc','median'), sin_100=('sin','median'),
                 brecha_mediana=('brecha_pct','median'), brecha_prom=('brecha_pct', _pmean),
                 brecha_min_ilustr=('brecha_min','median'),
                 n_sucursales=('id_sucursal','nunique')).reset_index())

def _agg_suc(df, keys):
    _cols = keys + ['brecha_mediana','brecha_prom','n_sucursales']
    if len(df) == 0:
        return pd.DataFrame(columns=_cols)
    return (df.groupby(keys)
            .agg(brecha_mediana=('brecha_pct','median'), brecha_prom=('brecha_pct', _pmean),
                 n_sucursales=('id_sucursal','nunique')).reset_index())

# Brecha POR TIPO (el número clave): mediana de las brechas intra-super
brecha_tipo         = _agg_tipo(bt_sm, [])
if len(brecha_tipo): brecha_tipo = brecha_tipo.sort_values('brecha_mediana')
brecha_tipo_mensual = _agg_tipo(bt_sm, ['mes'])
brecha_tipo_prov    = _agg_tipo(bt_sm, ['PROVINCIA_NORM']).rename(columns={'PROVINCIA_NORM':'provincia'})

# Series de CANASTA (brecha por sucursal, promediada entre supers). Granularidad MENSUAL
# (intra-super por mes; diaria/semanal no aplican a este método y quedan vacías).
serie_mensual = _agg_suc(brecha_suc_mes, ['mes'])
if len(serie_mensual): serie_mensual = serie_mensual.sort_values('mes')
serie_semanal = pd.DataFrame(columns=['semana','brecha_mediana','brecha_prom','n_sucursales'])
serie_diaria  = pd.DataFrame(columns=['fecha','brecha_mediana','brecha_prom','n_sucursales'])

# ── Cobertura temporal: PANEL BALANCEADO + diagnóstico de meses con baja cobertura ──
# La serie completa mezcla sucursales que entran/salen (n_sucursales inestable) → la
# comparación entre meses se confunde con la composición. La serie BALANCEADA usa solo
# las sucursales presentes en ≥ FRAC_MESES_PANEL de los meses → trayectoria comparable.
_sk3 = ['id_comercio','id_bandera','id_sucursal']
if len(brecha_suc_mes):
    _nmes  = brecha_suc_mes['mes'].nunique()
    _minm  = max(1, int(round(FRAC_MESES_PANEL * _nmes)))
    _pres  = brecha_suc_mes.groupby(_sk3)['mes'].nunique()
    _panel = set(_pres[_pres >= _minm].index)
    _bsm_bal = brecha_suc_mes[brecha_suc_mes.set_index(_sk3).index.isin(_panel)].copy()
    serie_mensual_bal = _agg_suc(_bsm_bal, ['mes'])
    if len(serie_mensual_bal): serie_mensual_bal = serie_mensual_bal.sort_values('mes')
    if len(serie_mensual):
        _nmed = serie_mensual['n_sucursales'].median()
        _bajos = serie_mensual[serie_mensual['n_sucursales'] < 0.6 * _nmed]
        if len(_bajos):
            print(f'  ⚠️ Cobertura temporal BAJA en {len(_bajos)} mes(es) (n < 60% de la mediana {_nmed:.0f}): '
                  + ', '.join(f'{_m}(n={int(_n)})' for _m, _n in zip(_bajos['mes'], _bajos['n_sucursales'])))
    print(f'  Panel balanceado: {len(_panel):,} sucursales presentes en ≥{_minm}/{_nmes} meses '
          f'→ serie_mensual_bal ({len(serie_mensual_bal)} meses, n≈{int(serie_mensual_bal["n_sucursales"].median()) if len(serie_mensual_bal) else 0})')
else:
    serie_mensual_bal = pd.DataFrame(columns=['mes','brecha_mediana','brecha_prom','n_sucursales'])

# Desagregaciones (promedio entre supers, por zona). Se marca confiable = n >= MIN_SUC_AGG.
def _marcar_conf(_df):
    _df['confiable'] = (_df['n_sucursales'] >= MIN_SUC_AGG) if len(_df) else pd.Series(dtype=bool)
    return _df
brecha_prov   = _marcar_conf(_agg_suc(brecha_suc_mes, ['PROVINCIA_NORM']).rename(columns={'PROVINCIA_NORM':'provincia'}))
if len(brecha_prov): brecha_prov = brecha_prov.sort_values('brecha_mediana')
brecha_cadena = _marcar_conf(_agg_suc(brecha_suc_mes, ['cadena']))
if len(brecha_cadena): brecha_cadena = brecha_cadena.sort_values('brecha_mediana')
concentracion = _marcar_conf(_agg_suc(brecha_suc_mes, ['localidad']))
if len(concentracion):
    concentracion = concentracion[concentracion['localidad'] != 'N/D'].sort_values('n_sucursales', ascending=False)

# ── Resumen ──────────────────────────────────────────────────────────────────
print('=== Brecha POR TIPO (intra-super, $/100g) — el número clave ===')
print('    brecha_mediana    = estimador PRIMARIO (mediana de candidatos, §3.4)')
print('    brecha_min_ilustr = referencia "más barato" (sesgada al alza por nº de candidatos, §3.7)')
if len(brecha_tipo):
    print(brecha_tipo[['tipo','tacc_100','sin_100','brecha_mediana','brecha_min_ilustr','n_sucursales']].round(1).to_string(index=False))
    print('  ⚠️ Revisá n_sucursales por tipo (y la hoja Cobertura): pocas = poco confiable.')
else:
    print('  (sin datos — revisá la Cobertura por tipo y los EANs de TIPOS)')

print('\\n=== Brecha de CANASTA mensual (intra-super) — últimos 6 ===')
if len(serie_mensual):
    print(serie_mensual.tail(6).round(1).to_string(index=False))
else:
    print('  (sin datos)')
_prov_conf = brecha_prov[brecha_prov['confiable']] if len(brecha_prov) else brecha_prov
if len(_prov_conf):
    print(f'\\nProvincia MENOR brecha (n≥{MIN_SUC_AGG}): {_prov_conf.iloc[0]["provincia"]} ({_prov_conf.iloc[0]["brecha_mediana"]:+.1f}%, n={int(_prov_conf.iloc[0]["n_sucursales"])})')
    print(f'Provincia MAYOR brecha (n≥{MIN_SUC_AGG}): {_prov_conf.iloc[-1]["provincia"]} ({_prov_conf.iloc[-1]["brecha_mediana"]:+.1f}%, n={int(_prov_conf.iloc[-1]["n_sucursales"])})')

# ── Resumen ──────────────────────────────────────────────────────────────────
print('=== Brecha mensual (nacional) — últimos 6 ===')
if len(serie_mensual):
    print(serie_mensual.tail(6)[['mes','brecha_mediana','brecha_prom','n_sucursales']].to_string(index=False))
    if len(serie_mensual) >= 2:
        _b0, _b1 = serie_mensual['brecha_mediana'].iloc[0], serie_mensual['brecha_mediana'].iloc[-1]
        print(f'Serie COMPLETA (n variable): {_b0:+.1f}% ({serie_mensual["mes"].iloc[0]}) -> {_b1:+.1f}% ({serie_mensual["mes"].iloc[-1]}) | cambio {(_b1-_b0):+.1f} pp')
    if len(serie_mensual_bal) >= 2:
        _c0, _c1 = serie_mensual_bal['brecha_mediana'].iloc[0], serie_mensual_bal['brecha_mediana'].iloc[-1]
        print(f'Panel BALANCEADO (n≈estable): {_c0:+.1f}% ({serie_mensual_bal["mes"].iloc[0]}) -> {_c1:+.1f}% ({serie_mensual_bal["mes"].iloc[-1]}) | cambio {(_c1-_c0):+.1f} pp  ← usar este para la tendencia temporal')
else:
    print('  (sin datos — revisá la cobertura por tipo arriba y los EANs de TIPOS)')
if len(_prov_conf):
    print(f'Provincia MENOR brecha (n≥{MIN_SUC_AGG}): {_prov_conf.iloc[0]["provincia"]} ({_prov_conf.iloc[0]["brecha_mediana"]:+.2f}%)')
    print(f'Provincia MAYOR brecha (n≥{MIN_SUC_AGG}): {_prov_conf.iloc[-1]["provincia"]} ({_prov_conf.iloc[-1]["brecha_mediana"]:+.2f}%)')"""))

# ── CELL 9 — GRÁFICOS ──────────────────────────────────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 9 — Gráficos de la brecha
# ============================================================
MES = f'{ULTIMO_MES[5:7]}{ULTIMO_MES[:4]}'
_C_MED, _C_PRO = '#0055A4', '#D62728'

# 1) Serie temporal mensual: serie COMPLETA (n variable) vs PANEL BALANCEADO (n≈estable),
#    con la cobertura (n_sucursales) en un eje secundario para leer la (in)estabilidad.
fig, ax = plt.subplots(figsize=(13, 6))
_sm = serie_mensual.copy(); _sm['fecha'] = pd.to_datetime(_sm['mes'] + '-01')
ax.plot(_sm['fecha'], _sm['brecha_mediana'], color=_C_MED, lw=2.5, marker='o', label='Serie completa (n variable)')
if len(serie_mensual_bal):
    _smb = serie_mensual_bal.copy(); _smb['fecha'] = pd.to_datetime(_smb['mes'] + '-01')
    ax.plot(_smb['fecha'], _smb['brecha_mediana'], color='#2ca02c', lw=2.5, marker='D',
            label=f'Panel balanceado (n≈{int(_smb["n_sucursales"].median())} estable)')
ax.set_ylabel('Brecha celíaca (%)'); ax.set_title('Evolución de la brecha celíaca — serie completa vs panel balanceado')
ax.grid(True, alpha=0.3); ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_: f'{x:+.0f}%'))
_ax2 = ax.twinx()
_ax2.bar(_sm['fecha'], _sm['n_sucursales'], width=20, color='#cccccc', alpha=0.35, label='n sucursales (der.)')
_ax2.set_ylabel('n sucursales', color='#999'); _ax2.set_ylim(0, _sm['n_sucursales'].max()*3)
_l1,_lb1 = ax.get_legend_handles_labels(); _l2,_lb2 = _ax2.get_legend_handles_labels()
ax.legend(_l1+_l2, _lb1+_lb2, loc='upper left', fontsize=9)
plt.tight_layout(); _o = OUTPUT_DIR / f'brecha_mensual_{MES}.png'
plt.savefig(_o, dpi=200, bbox_inches='tight', facecolor='white'); plt.show()
print(f'Guardado: {_o.name}')

# 2) Serie DIARIA (ventana)
if len(serie_diaria) > 1:
    fig, ax = plt.subplots(figsize=(13, 5))
    ax.plot(serie_diaria['fecha'], serie_diaria['brecha_mediana'], color=_C_MED, lw=1.5, label='Diaria (mediana)')
    ax.plot(serie_diaria['fecha'], serie_diaria['brecha_prom'], color=_C_PRO, lw=1.2, alpha=0.8, label='Diaria (promedio)')
    ax.set_ylabel('Brecha (%)'); ax.set_title(f'Brecha diaria — últimos {VENTANA_DIARIA_MESES} meses')
    ax.legend(); ax.grid(True, alpha=0.3)
    plt.tight_layout(); _o = OUTPUT_DIR / f'brecha_diaria_{MES}.png'
    plt.savefig(_o, dpi=200, bbox_inches='tight', facecolor='white'); plt.show()
    print(f'Guardado: {_o.name}')

# 3) Brecha por provincia (barras) — solo provincias CONFIABLES (n >= MIN_SUC_AGG)
_bp = brecha_prov[brecha_prov['confiable']] if len(brecha_prov) else brecha_prov
fig, ax = plt.subplots(figsize=(11, max(5, len(_bp)*0.35+2)))
_cols = plt.cm.RdYlGn_r(np.linspace(0.15, 0.9, max(1, len(_bp))))
ax.barh(_bp['provincia'], _bp['brecha_mediana'], color=_cols, edgecolor='black', lw=0.4)
for _i,(_,_r) in enumerate(_bp.iterrows()):
    ax.text(_r['brecha_mediana'], _i, f' {_r["brecha_mediana"]:+.1f}% (n={int(_r["n_sucursales"])})', va='center', fontsize=8)
ax.set_xlabel('Brecha mediana (%)'); ax.set_title(f'Brecha celíaca por provincia (n≥{MIN_SUC_AGG} sucursales)')
ax.grid(True, alpha=0.3, axis='x')
plt.tight_layout(); _o = OUTPUT_DIR / f'brecha_provincia_{MES}.png'
plt.savefig(_o, dpi=200, bbox_inches='tight', facecolor='white'); plt.show()
print(f'Guardado: {_o.name}')

# 4) Brecha por cadena (barras) — solo cadenas CONFIABLES (n >= MIN_SUC_AGG)
_bc = brecha_cadena[brecha_cadena['n_sucursales'] >= MIN_SUC_AGG]
if len(_bc) > 0:
    fig, ax = plt.subplots(figsize=(11, max(4, len(_bc)*0.4+2)))
    _cols = plt.cm.RdYlGn_r(np.linspace(0.15, 0.9, len(_bc)))
    ax.barh(_bc['cadena'], _bc['brecha_mediana'], color=_cols, edgecolor='black', lw=0.4)
    for _i,(_,_r) in enumerate(_bc.iterrows()):
        ax.text(_r['brecha_mediana'], _i, f' {_r["brecha_mediana"]:+.1f}% (n={int(_r["n_sucursales"])})', va='center', fontsize=8)
    ax.set_xlabel('Brecha mediana (%)'); ax.set_title(f'Brecha celíaca por cadena (n≥{MIN_SUC_AGG} sucursales)')
    ax.grid(True, alpha=0.3, axis='x')
    plt.tight_layout(); _o = OUTPUT_DIR / f'brecha_cadena_{MES}.png'
    plt.savefig(_o, dpi=200, bbox_inches='tight', facecolor='white'); plt.show()
    print(f'Guardado: {_o.name}')

# 5) Brecha vs concentración de comercios (scatter)
_cc = concentracion[concentracion['n_sucursales'] >= 2]
if len(_cc) >= 3:
    fig, ax = plt.subplots(figsize=(9, 6))
    ax.scatter(_cc['n_sucursales'], _cc['brecha_mediana'], s=28, alpha=0.6, color=_C_MED, edgecolor='white')
    _corr = _cc['n_sucursales'].corr(_cc['brecha_mediana'])
    _corr_str = f'{_corr:+.2f}' if pd.notna(_corr) else 'n/d'
    ax.set_xscale('log')
    ax.set_xlabel('Nº de sucursales en la localidad (log)'); ax.set_ylabel('Brecha mediana (%)')
    ax.set_title(f'Brecha vs concentración de comercios (corr={_corr_str})')
    ax.grid(True, alpha=0.3)
    plt.tight_layout(); _o = OUTPUT_DIR / f'brecha_concentracion_{MES}.png'
    plt.savefig(_o, dpi=200, bbox_inches='tight', facecolor='white'); plt.show()
    print(f'Guardado: {_o.name} | correlación brecha–concentración: {_corr_str}')

# 6) Brecha POR TIPO (barras) — el gráfico clave
if len(brecha_tipo):
    _bt = brecha_tipo.sort_values('brecha_mediana')
    fig, ax = plt.subplots(figsize=(11, max(4, len(_bt)*0.7+2)))
    _cols = plt.cm.RdYlGn_r(np.linspace(0.2, 0.9, len(_bt)))
    ax.barh(_bt['tipo'], _bt['brecha_mediana'], color=_cols, edgecolor='black', lw=0.4)
    for _i,(_,_r) in enumerate(_bt.iterrows()):
        ax.text(_r['brecha_mediana'], _i, f'  {_r["brecha_mediana"]:+.0f}%  ({int(_r["n_sucursales"])} sucs con ambos)',
                va='center', fontsize=8)
    ax.set_xlabel('Brecha mediana (%) — precio sin-TACC vs TACC, $/100g')
    ax.set_title('Brecha celíaca POR TIPO de producto')
    ax.grid(True, alpha=0.3, axis='x')
    plt.tight_layout(); _o = OUTPUT_DIR / f'brecha_por_tipo_{MES}.png'
    plt.savefig(_o, dpi=200, bbox_inches='tight', facecolor='white'); plt.show()
    print(f'Guardado: {_o.name}')

# 7) Evolución mensual de la brecha por tipo
if len(brecha_tipo_mensual):
    fig, ax = plt.subplots(figsize=(13, 6))
    for _t, _g in brecha_tipo_mensual.groupby('tipo'):
        _g = _g.sort_values('mes'); _x = pd.to_datetime(_g['mes'] + '-01')
        ax.plot(_x, _g['brecha_mediana'], marker='o', ms=4, lw=1.8, label=_t)
    ax.set_ylabel('Brecha (%)'); ax.set_title('Evolución mensual de la brecha por tipo')
    ax.legend(fontsize=8, ncol=2); ax.grid(True, alpha=0.3)
    plt.tight_layout(); _o = OUTPUT_DIR / f'brecha_tipo_mensual_{MES}.png'
    plt.savefig(_o, dpi=200, bbox_inches='tight', facecolor='white'); plt.show()
    print(f'Guardado: {_o.name}')"""))

# ── CELL 10 — MAPA COROPLÉTICO de la brecha por provincia ──────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 10 — Mapa coroplético de la brecha por provincia
# ============================================================
if not GEOJSON_PATH.exists():
    print(f'GeoJSON no encontrado en {GEOJSON_PATH} — saltear mapa')
else:
    with open(GEOJSON_PATH, 'r', encoding='utf-8') as f:
        geo = _json.load(f)
    NORM_GEO = {'Ciudad de Buenos Aires':'CABA'}
    AJUST = {'Salta':(0,-1),'Tucumán':(0.3,0),'Chaco':(0,-1),'Tierra del Fuego':(-1,-0.2),
             'Santa Fe':(0,1),'Santiago del Estero':(0.7,0)}
    def centroide(coords):
        xs,ys=[],[]
        if isinstance(coords[0][0][0],(int,float)):
            for p in coords[0]: xs.append(p[0]); ys.append(p[1])
        else:
            poly=max(coords,key=lambda p:len(p[0]))
            for p in poly[0]: xs.append(p[0]); ys.append(p[1])
        return sum(xs)/len(xs), sum(ys)/len(ys)
    def draw(ax,coords,color):
        if isinstance(coords[0][0][0],(int,float)):
            ax.fill([c[0] for c in coords[0]],[c[1] for c in coords[0]],facecolor=color,edgecolor='white',linewidth=0.6)
        else:
            for poly in coords:
                ax.fill([c[0] for c in poly[0]],[c[1] for c in poly[0]],facecolor=color,edgecolor='white',linewidth=0.6)
    cmap_m = LinearSegmentedColormap.from_list('c',
        ['#1a9850','#66bd63','#a6d96a','#fee08b','#fdae61','#f46d43','#d73027'], N=256)

    _prov_val = dict(zip(brecha_prov['provincia'], brecha_prov['brecha_mediana']))
    _vals = list(_prov_val.values())
    if _vals:
        norm_c = Normalize(vmin=min(_vals), vmax=max(_vals))
        fig, ax = plt.subplots(figsize=(12, 16)); caba_c = None
        for feat in geo['features']:
            nom = NORM_GEO.get(feat['properties']['name'], feat['properties']['name'])
            val = _prov_val.get(nom)
            col = cmap_m(norm_c(val)) if val is not None else '#dddddd'
            gt, co = feat['geometry']['type'], feat['geometry']['coordinates']
            draw(ax, [co] if gt=='Polygon' else co, col)
            cx, cy = centroide([co] if gt=='Polygon' else co)
            if nom == 'CABA': caba_c=(cx,cy); continue
            dx,dy = AJUST.get(nom,(0,0))
            if val is not None:
                ax.text(cx+dx, cy+dy, f'{nom}\\n{val:+.1f}%', ha='center', va='center', fontsize=7.5,
                        fontweight='bold', bbox=dict(boxstyle='round,pad=0.25', facecolor='white', alpha=0.75, edgecolor='none'))
        if caba_c and 'CABA' in _prov_val:
            vc = _prov_val['CABA']; cc = cmap_m(norm_c(vc)); lx,ly = caba_c[0]+2.2, caba_c[1]+0.8
            ax.annotate('', xy=caba_c, xytext=(lx,ly), arrowprops=dict(arrowstyle='-', color='black', linewidth=1.0))
            ax.text(lx, ly, f'CABA\\n{vc:+.1f}%', ha='center', va='center', fontsize=9, fontweight='bold',
                    bbox=dict(boxstyle='round,pad=0.5', facecolor=cc, alpha=0.95, edgecolor='black', linewidth=1.0))
            ax.plot(*caba_c, marker='o', markersize=10, markerfacecolor=cc, markeredgecolor='black', markeredgewidth=1.2, zorder=5)
        ax.set_title('Brecha celíaca por provincia (mediana %)', fontsize=13, fontweight='bold', pad=6)
        ax.set_aspect('equal'); ax.axis('off'); plt.tight_layout()
        _o = OUTPUT_DIR / f'mapa_brecha_{MES}.png'
        plt.savefig(_o, dpi=300, bbox_inches='tight', facecolor='white'); plt.show()
        print(f'Guardado: {_o.name}')"""))

# ── CELL 11 — EXCEL EXPORT ─────────────────────────────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 11 — Exportación Excel
# ============================================================
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
HDR_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HDR_FONT = Font(bold=True, color='FFFFFF', size=10)
HDR_ALIG = Alignment(horizontal='center', wrap_text=True, vertical='center')
def fmt_ws(ws):
    ws.row_dimensions[1].height = 28; ws.freeze_panes = 'A2'
    for cell in ws[1]: cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = HDR_ALIG
def auto_widths(ws):
    for ci in range(1, ws.max_column+1):
        cl = get_column_letter(ci); hdr = str(ws.cell(1,ci).value or '').lower()
        ws.column_dimensions[cl].width = 30 if any(x in hdr for x in ('provincia','cadena','localidad','tipo','desc','producto','sucursal','nombre')) else 14
        if any(x in hdr for x in ('brecha','base','celiaca','precio','n_','100','tacc','grams')):
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=ci, max_col=ci):
                for cell in row:
                    cell.number_format = ('+0.00"%"' if 'brecha' in hdr else '#,##0.00')

# ── Detalle intra-sucursal por producto (último mes): precio por sucursal×EAN ──
_ult = datos_dia[datos_dia['mes'] == ULTIMO_MES]
_det = (_ult.groupby(['id_comercio','id_bandera','id_sucursal','ean_norm','tipo','rol'])['precio']
        .median().reset_index().rename(columns={'precio':'precio_mediano_mes'}))
_det['descripcion'] = _det['ean_norm'].map(EAN_DESC)
_det['grams'] = _det['ean_norm'].map(EAN_GRAMS)
_det['precio_100g'] = np.where(_det['grams'].notna() & (_det['grams'] > 0),
                               _det['precio_mediano_mes'] / _det['grams'] * 100, np.nan)
_det = _det.merge(suc_geo[['id_comercio','id_bandera','id_sucursal','cadena','PROVINCIA_NORM',
                           'sucursales_localidad','sucursales_nombre']],
                  on=['id_comercio','id_bandera','id_sucursal'], how='left')
_det = _det[['cadena','PROVINCIA_NORM','sucursales_localidad','sucursales_nombre',
             'id_comercio','id_bandera','id_sucursal','tipo','rol','ean_norm','descripcion',
             'grams','precio_mediano_mes','precio_100g']]
_det = _det.sort_values(['PROVINCIA_NORM','cadena','sucursales_nombre','tipo','rol'])

# ── Brecha por SUCURSAL (canasta intra-super) — último mes ────────────────────
_bsm = brecha_suc_mes[brecha_suc_mes['mes'] == ULTIMO_MES].copy()
if len(_bsm):
    _brecha_suc = (_bsm.merge(suc_geo[['id_comercio','id_bandera','id_sucursal','sucursales_nombre']],
                              on=['id_comercio','id_bandera','id_sucursal'], how='left')
                   .rename(columns={'base':'base_mediana','celiaca':'celiaca_mediana','brecha_pct':'brecha_mediana'})
                   [['cadena','PROVINCIA_NORM','localidad','sucursales_nombre','id_comercio','id_bandera',
                     'id_sucursal','base_mediana','celiaca_mediana','n_tipos','brecha_mediana']]
                   .sort_values('brecha_mediana', ascending=False))
else:
    _brecha_suc = pd.DataFrame(columns=['cadena','PROVINCIA_NORM','localidad','sucursales_nombre',
                                        'base_mediana','celiaca_mediana','n_tipos','brecha_mediana'])

# ── Comparación de productos EQUIVALENTES por sucursal (último mes) ────────────
# Para cada sucursal, tipo y SUBTIPO (equivalencia: Spaghetti vs Spaghetti, Vainilla vs
# Vainilla, Pan rallado vs Pan rallado…): qué producto(s) con TACC se comparan con qué
# producto(s) sin-TACC, con presentación y $/100g, y la brecha de ese subtipo.
_cu = _ult.copy()
_cu['subtipo'] = _cu['ean_norm'].map(EAN_SUBTIPO)
_cu['descripcion'] = _cu['ean_norm'].map(EAN_DESC)
_cu['grams'] = _cu['ean_norm'].map(EAN_GRAMS)
def _items_lado(_g):
    _g = _g.sort_values('precio_100')
    _p = []
    for _t in _g.itertuples(index=False):
        _gt = f'{int(_t.grams)}g' if (_t.grams == _t.grams) else 's/pres'
        _pt = f'${_t.precio_100:,.0f}/100g' if (_t.precio_100 == _t.precio_100) else 's/precio'
        _p.append(f'{_t.descripcion} ({_gt}, {_pt})')
    return '; '.join(_p)
_rows_c = []
for _kk, _g in _cu.groupby(['id_comercio','id_bandera','id_sucursal','tipo','subtipo','rol']):
    _rows_c.append((*_kk, _items_lado(_g), float(_g['precio_100'].median()), int(_g['ean_norm'].nunique())))
_cg = pd.DataFrame(_rows_c, columns=['id_comercio','id_bandera','id_sucursal','tipo','subtipo','rol',
                                     'productos','precio_100g_med','n'])
if len(_cg):
    _tc = _cg[_cg['rol']=='tacc'].drop(columns='rol').rename(
        columns={'productos':'productos_tacc','precio_100g_med':'precio_tacc_100g','n':'n_tacc'})
    _sc = _cg[_cg['rol']=='sin'].drop(columns='rol').rename(
        columns={'productos':'productos_sin','precio_100g_med':'precio_sin_100g','n':'n_sin'})
    _comp = _tc.merge(_sc, on=['id_comercio','id_bandera','id_sucursal','tipo','subtipo'], how='outer')
    _comp['n_tacc'] = _comp['n_tacc'].fillna(0).astype(int)
    _comp['n_sin']  = _comp['n_sin'].fillna(0).astype(int)
    _comp['equivalencia'] = np.where((_comp['n_tacc']>0) & (_comp['n_sin']>0), 'ambos',
                              np.where(_comp['n_tacc']>0, 'solo con-TACC', 'solo sin-TACC'))
    _comp['brecha_subtipo_pct'] = np.where((_comp['precio_tacc_100g']>0) & (_comp['precio_sin_100g']>0),
                                           (_comp['precio_sin_100g']/_comp['precio_tacc_100g']-1)*100, np.nan)
    _comp = _comp.merge(suc_geo[['id_comercio','id_bandera','id_sucursal','cadena','PROVINCIA_NORM',
                                 'sucursales_localidad','sucursales_nombre']],
                        on=['id_comercio','id_bandera','id_sucursal'], how='left')
    _comp = _comp[['cadena','PROVINCIA_NORM','sucursales_localidad','sucursales_nombre',
                   'id_comercio','id_bandera','id_sucursal','tipo','subtipo','equivalencia',
                   'n_tacc','productos_tacc','precio_tacc_100g','n_sin','productos_sin','precio_sin_100g',
                   'brecha_subtipo_pct']].sort_values(['PROVINCIA_NORM','cadena','sucursales_nombre','tipo','subtipo'])
else:
    _comp = pd.DataFrame(columns=['cadena','PROVINCIA_NORM','sucursales_localidad','sucursales_nombre',
                                  'id_comercio','id_bandera','id_sucursal','tipo','subtipo','equivalencia',
                                  'n_tacc','productos_tacc','precio_tacc_100g','n_sin','productos_sin',
                                  'precio_sin_100g','brecha_subtipo_pct'])

out_xls = OUTPUT_DIR / f'brecha_celiaca_{ULTIMO_MES}.xlsx'
with pd.ExcelWriter(out_xls, engine='openpyxl') as writer:
    cobertura_tipo.to_excel(writer, sheet_name='Cobertura', index=False)
    brecha_tipo.to_excel(writer, sheet_name='Brecha_tipo', index=False)
    brecha_tipo_mensual.to_excel(writer, sheet_name='Brecha_tipo_mes', index=False)
    brecha_tipo_prov.to_excel(writer, sheet_name='Brecha_tipo_prov', index=False)
    serie_mensual.to_excel(writer, sheet_name='Serie_mensual', index=False)
    serie_mensual_bal.to_excel(writer, sheet_name='Serie_mensual_balanceada', index=False)
    # Serie_diaria/Serie_semanal solo si tienen datos (el método es mensual → normalmente vacías)
    if len(serie_diaria):  serie_diaria.to_excel(writer, sheet_name='Serie_diaria', index=False)
    if len(serie_semanal): serie_semanal.to_excel(writer, sheet_name='Serie_semanal', index=False)
    brecha_prov.to_excel(writer, sheet_name='Brecha_provincia', index=False)
    brecha_cadena.to_excel(writer, sheet_name='Brecha_cadena', index=False)
    concentracion.to_excel(writer, sheet_name='Concentracion', index=False)
    _brecha_suc.to_excel(writer, sheet_name='Brecha_sucursal', index=False)
    _comp.to_excel(writer, sheet_name='Comparacion_productos', index=False)
    _det.to_excel(writer, sheet_name='Detalle_producto', index=False)
    for sn in writer.sheets:
        ws = writer.sheets[sn]; fmt_ws(ws); auto_widths(ws)
print(f'Excel guardado: {out_xls}')
print(f'  Hojas: Cobertura · Brecha_tipo/_mes/_prov · Serie_mensual(+balanceada) · Brecha_provincia/cadena (col confiable) · Concentracion · Brecha_sucursal · Comparacion_productos · Detalle_producto')
_np = int(brecha_prov['confiable'].sum()) if len(brecha_prov) else 0
_nc = int(brecha_cadena['confiable'].sum()) if len(brecha_cadena) else 0
print(f'  Confiables (n≥{MIN_SUC_AGG}): {_np}/{len(brecha_prov)} provincias · {_nc}/{len(brecha_cadena)} cadenas | Panel balanceado: {len(serie_mensual_bal)} meses')
print(f'  Detalle_producto: {len(_det):,} filas · Comparacion_productos: {len(_comp):,} filas (sucursal×tipo×subtipo) · Brecha_sucursal: {len(_brecha_suc):,}')
print(f'  ⚠️ El número clave está en Brecha_tipo (brecha por tipo en $/100g). La canasta pooled mezcla tipos de brecha muy distinta.')"""))

# ── CELL 12 — MAPA FOLIUM POR SUCURSAL (lazy-load) ─────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 12 — Mapa Folium interactivo POR SUCURSAL (brecha celíaca, último mes)
# ============================================================
# Un punto por supermercado; color VERDE→ROJO graduado por la brecha de canasta.
# LIVIANO: los popups se arman al hacer clic (lazy-load desde un JSON compacto), así el
# HTML no pesa (antes ~15 MB con todos los popups embebidos → carga al instante). Popup:
# (1) valor de cada canasta y % de brecha; (2) productos por EQUIVALENCIA (subtipo) —
# Con TACC vs Sin TACC, con presentación y $/100g; (3) cadena + localidad/provincia. La
# brecha AUTORITATIVA es la del tipo (§3). Filtro por cadena (LayerControl) + pantalla
# completa. Tiles OpenStreetMap (sin API key).
import folium
from folium.plugins import Fullscreen
from branca.colormap import LinearColormap
import json as _json

# Snapshot = ÚLTIMA brecha disponible POR SUCURSAL dentro de una ventana de meses. Así el
# mapa incluye cadenas que no reportaron el último mes (p.ej. DIA sin datos en agosto): cada
# sucursal muestra su mes más reciente con dato (el popup indica de qué mes es).
_sk3m = ['id_comercio','id_bandera','id_sucursal']
VENTANA_MAPA_MESES = 6
_meses_ord = sorted(brecha_suc_mes['mes'].unique())
_meses_win = set(_meses_ord[-VENTANA_MAPA_MESES:]) if _meses_ord else set()
_bsm_w = brecha_suc_mes[brecha_suc_mes['mes'].isin(_meses_win)]
_bk = (_bsm_w.sort_values('mes').groupby(_sk3m, as_index=False).tail(1)   # última fila = mes más reciente por sucursal
       .merge(suc_geo[['id_comercio','id_bandera','id_sucursal','sucursales_nombre',
                       'sucursales_latitud','sucursales_longitud']],
              on=_sk3m, how='left')
       .dropna(subset=['sucursales_latitud','sucursales_longitud']))
_um = _meses_ord[-1] if _meses_ord else ULTIMO_MES
_sel_mes = _bk[_sk3m + ['mes']]   # mes elegido por sucursal → alinea bt_sm y datos_dia a ESE mes

_btm = bt_sm.merge(_sel_mes, on=_sk3m + ['mes'], how='inner')[_sk3m + ['tipo','tacc','sin','brecha_pct']]
_bt_por_suc = {}
for _r in _btm.itertuples(index=False):
    _bt_por_suc.setdefault((_r.id_comercio, _r.id_bandera, _r.id_sucursal), {})[_r.tipo] = (_r.tacc, _r.sin, _r.brecha_pct)

_pu = datos_dia.merge(_sel_mes, on=_sk3m + ['mes'], how='inner').copy()
_pu['descripcion'] = _pu['ean_norm'].map(EAN_DESC)
_pu['subtipo'] = _pu['ean_norm'].map(EAN_SUBTIPO)
_prod = {}
for _r in _pu.itertuples(index=False):
    _d = (_prod.setdefault((_r.id_comercio, _r.id_bandera, _r.id_sucursal), {})
              .setdefault(_r.tipo, {}).setdefault(_r.subtipo, {'A': [], 'B': []}))
    _d['A' if _r.rol == 'tacc' else 'B'].append((str(_r.descripcion)[:38], _r.grams, _r.precio_100))

def _iround(_x):
    try: return int(round(float(_x)))
    except Exception: return None
def _p_arr(_lst):
    _lst = sorted(_lst, key=lambda x: (x[2] if x[2] == x[2] else 9e9))[:8]
    return [[_d, (int(_g) if (_g == _g and _g > 0) else 0), (_iround(_p) if _p == _p else None)] for _d, _g, _p in _lst]

# JSON compacto por sucursal (lo consume el JS al abrir el popup)
_pd = {}
for _r in _bk.itertuples(index=False):
    _sk = (_r.id_comercio, _r.id_bandera, _r.id_sucursal)
    _key = f'{_r.id_comercio}_{_r.id_bandera}_{_r.id_sucursal}'
    _tipos = _bt_por_suc.get(_sk, {})
    _T = []
    for _tp in sorted(_tipos, key=lambda t: -_tipos[t][2]):
        _tacc_p, _sin_p, _brk = _tipos[_tp]
        _subs = _prod.get(_sk, {}).get(_tp, {})
        def _ord(_s):
            _d = _subs[_s]; return (0 if (_d['A'] and _d['B']) else 1, _s)
        _S = [{'s': _s, 'A': _p_arr(_subs[_s]['A']), 'B': _p_arr(_subs[_s]['B'])} for _s in sorted(_subs, key=_ord)]
        _T.append({'t': _tp, 'ta': _iround(_tacc_p), 'si': _iround(_sin_p), 'g': _iround(_brk), 'S': _S})
    _pd[_key] = {'c': str(_r.cadena), 'n': str(_r.sucursales_nombre)[:46], 'p': str(_r.PROVINCIA_NORM),
                 'l': str(_r.localidad), 'm': str(_r.mes), 'b': _iround(_r.brecha_pct), 'bs': _iround(_r.base),
                 'cl': _iround(_r.celiaca), 'nt': int(_r.n_tipos), 'T': _T}
_json_str = _json.dumps(_pd, ensure_ascii=False, separators=(',', ':'))

_vmin = float(_bk['brecha_pct'].quantile(0.05)); _vmax = float(_bk['brecha_pct'].quantile(0.95))
if _vmin == _vmax: _vmin, _vmax = float(_bk['brecha_pct'].min()), float(_bk['brecha_pct'].max())
_cm = LinearColormap(colors=['#1a9850','#66bd63','#a6d96a','#fee08b','#fdae61','#f46d43','#d73027'],
                     vmin=_vmin, vmax=_vmax, caption=f'Brecha de canasta celíaca (%) — {NOMBRE_MES_TITLE}')

m = folium.Map(location=[-38.0, -63.5], zoom_start=5, tiles='cartodbpositron', control_scale=True)
# Límites PROVINCIALES desde el GeoJSON local (ar.json) → se ven SIEMPRE, aun si la red
# bloquea los tiles externos (el fondo con nombres es un extra si los tiles cargan).
try:
    if 'GEOJSON_PATH' in dir() and Path(GEOJSON_PATH).exists():
        with open(GEOJSON_PATH, 'r', encoding='utf-8') as _gf:
            _geo_base = _json.load(_gf)
        folium.GeoJson(_geo_base, name='Límites provinciales', control=False,
                       style_function=lambda _f: {'color': '#8a8f98', 'weight': 1.0,
                                                  'fillColor': '#eef0f2', 'fillOpacity': 0.35}
                       ).add_to(m)
except Exception as _e:
    print(f'  (aviso: no se pudieron dibujar los límites provinciales: {_e})')
_cm.add_to(m)
Fullscreen(position='topright', title='Pantalla completa', title_cancel='Salir').add_to(m)

_fgs = {}
def _fg(_cad):
    if _cad not in _fgs:
        _fgs[_cad] = folium.FeatureGroup(name=str(_cad), show=True)
    return _fgs[_cad]

for _r in _bk.itertuples(index=False):
    _key = f'{_r.id_comercio}_{_r.id_bandera}_{_r.id_sucursal}'
    _b = float(_r.brecha_pct); _col = _cm(max(_vmin, min(_vmax, _b)))
    _ph = (f'<div class="lz-pop" data-key="{_key}" style="min-width:320px">'
           f'<span style="color:#aaa;font-family:Arial;font-size:12px">Cargando detalle…</span></div>')
    folium.CircleMarker(
        location=[_r.sucursales_latitud, _r.sucursales_longitud],
        radius=6, color='#333', weight=0.6, fill=True, fillColor=_col, fillOpacity=0.9,
        tooltip=f'<b>{_r.cadena}</b><br>{_r.PROVINCIA_NORM}<br>Brecha <b>+{_b:.0f}%</b>',
        popup=folium.Popup(_ph, max_width=500)
    ).add_to(_fg(_r.cadena))

for _c in sorted(_fgs): _fgs[_c].add_to(m)
folium.LayerControl(collapsed=True, position='topright').add_to(m)

_med_nac = _bk['brecha_pct'].median()
_info = (f'<div style="position:fixed;top:10px;left:50px;width:335px;background:white;border:2px solid #0055A4;'
         f'border-radius:8px;padding:12px 15px;font-family:Arial;z-index:9999;box-shadow:0 2px 8px rgba(0,0,0,.15)">'
         f'<div style="color:#0055A4;font-size:15px;font-weight:bold;margin-bottom:4px">Brecha celíaca por sucursal</div>'
         f'<div style="font-size:11px;color:#555;line-height:1.5">Última brecha disponible por sucursal '
         f'(últimos {VENTANA_MAPA_MESES} meses) · <b>{len(_bk):,}</b> sucursales · <b>{len(_fgs)}</b> cadenas<br>'
         f'Brecha mediana: <b>+{_med_nac:.0f}%</b> · el popup indica de qué mes es cada dato<br>'
         f'<span style="color:#1a9850">■</span> menor &nbsp; <span style="color:#d73027">■</span> mayor brecha &nbsp;·&nbsp; '
         f'filtrá cadenas con el control de capas (arriba der.)</div></div>')
m.get_root().html.add_child(folium.Element(_info))
m.get_root().html.add_child(folium.Element(f'<script type="application/json" id="_pd_json">{_json_str}</script>'))

_MV = m.get_name()
_JS = '''<script>
var _MV="__MAPVAR__", _MES="__MES__", _PD=null;
function _gpd(){ if(!_PD){ var el=document.getElementById("_pd_json"); if(el) _PD=JSON.parse(el.textContent); } return _PD; }
function _fmt(x){ return "$"+Number(x).toLocaleString("es-AR"); }
function _plist(a){ if(!a||!a.length) return '<span style="color:#bbb">—</span>'; var s=""; for(var i=0;i<a.length;i++){ var g=a[i][1]?a[i][1]+" g":"s/pres"; var p=(a[i][2]!=null)?_fmt(a[i][2])+"/100g":"s/precio"; s+='<div style="margin:1px 0">• '+a[i][0]+' <span style="color:#777">('+g+", "+p+')</span></div>'; } return s; }
function _bpop(k){ var pd=_gpd(); if(!pd||!pd[k]) return "<div>Sin datos.</div>"; var d=pd[k];
  var h='<div style="font-family:Arial;font-size:12px;min-width:360px;max-width:480px;max-height:520px;overflow-y:auto">';
  h+='<div style="font-size:14px;font-weight:bold">'+d.n+'</div>';
  h+='<div style="color:#555;margin-bottom:5px">'+d.c+" · "+d.l+", "+d.p+'</div>';
  h+='<div style="padding:6px;background:#f5f5f5;border-radius:4px;margin-bottom:4px">Brecha de canasta celíaca: <b style="font-size:15px">+'+d.b+'%</b><br>Canasta convencional: <b>'+_fmt(d.bs)+'</b> · celíaca: <b>'+_fmt(d.cl)+'</b> <span style="color:#888">(índice $/100g ponderado)</span><br><span style="color:#888;font-size:11px">'+d.nt+" tipo(s) · dato de "+d.m+'</span></div>';
  for(var i=0;i<d.T.length;i++){ var t=d.T[i];
    h+='<div style="margin-top:7px;font-weight:bold;color:#1F4E79">'+t.t+" · brecha del tipo +"+t.g+'% <span style="color:#888;font-weight:normal">(mediana con-TACC '+_fmt(t.ta)+" vs sin "+_fmt(t.si)+'/100g)</span></div>';
    h+='<table style="border-collapse:collapse;width:100%;font-size:11px"><tr style="background:#1F4E79;color:#fff"><th style="padding:2px">Equivalente</th><th style="padding:2px">Con TACC</th><th style="padding:2px">Sin TACC</th></tr>';
    for(var j=0;j<t.S.length;j++){ var s=t.S[j];
      h+='<tr><td style="border:1px solid #ddd;padding:3px;vertical-align:top;font-weight:600">'+s.s+'</td><td style="border:1px solid #ddd;padding:3px;vertical-align:top">'+_plist(s.A)+'</td><td style="border:1px solid #ddd;padding:3px;vertical-align:top">'+_plist(s.B)+'</td></tr>';
    }
    h+="</table>";
  }
  h+='<div style="color:#999;font-size:10px;margin-top:5px">Filas = productos equivalentes. La brecha autoritativa es la del tipo (§3); acá se listan los productos comparados de cada lado.</div></div>';
  return h;
}
function _init(){ var mp=window[_MV]; if(!mp){ setTimeout(_init,300); return; } mp.on("popupopen",function(e){ var el=e.popup.getElement().querySelector(".lz-pop"); if(el&&el.getAttribute("data-b")!=="1"){ el.innerHTML=_bpop(el.getAttribute("data-key")); el.setAttribute("data-b","1"); e.popup.update(); } }); }
setTimeout(_init,500);
</script>'''
_JS = _JS.replace('__MAPVAR__', _MV).replace('__MES__', NOMBRE_MES_TITLE)
m.get_root().html.add_child(folium.Element(_JS))

_out_map = OUTPUT_DIR / f'mapa_sucursales_brecha_{MES}.html'
m.save(str(_out_map))
print(f'Mapa Folium por sucursal guardado: {_out_map}')
print(f'  {len(_bk):,} sucursales · {len(_fgs)} cadenas (últimos {VENTANA_MAPA_MESES} meses; incluye cadenas sin dato el último mes)'
      f' · popup lazy-load · límites del GeoJSON local · JSON {len(_json_str)/1024/1024:.1f} MB')
_mmix = _bk['mes'].value_counts().to_dict()
print(f'  Meses del snapshot (sucursal→mes más reciente): ' + ', '.join(f'{_m}:{_n}' for _m,_n in sorted(_mmix.items(), reverse=True)))"""))

# ── Write notebook ──────────────────────────────────────────────────────────────
nb = {
    'cells': cells,
    'metadata': {
        'kernelspec': {'display_name':'Python 3','language':'python','name':'python3'},
        'language_info': {'codemirror_mode':{'name':'ipython','version':3},
                          'file_extension':'.py','mimetype':'text/x-python',
                          'name':'python','nbformat':4,'pygments_lexer':'ipython3','version':'3.10.0'}
    },
    'nbformat': 4,
    'nbformat_minor': 5
}
script_dir = os.path.dirname(os.path.abspath(__file__))
out_path = os.path.join(script_dir, '06_evolucion_brecha_celiaca.ipynb')
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump(nb, f, ensure_ascii=False, indent=1)
print(f'Written: {out_path}')
print(f'Cells: {len(cells)}')
