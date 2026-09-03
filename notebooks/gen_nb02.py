"""Script to generate 02_evolucion_canasta_representativa.ipynb — Multi-canasta version"""
import json, os, hashlib

def _cell_id(prefix, src):
    # id determinista (md5 del contenido) — evita 'drift' espurio al regenerar
    return prefix + hashlib.md5(src.encode('utf-8')).hexdigest()[:6]

def cell_md(src):
    lines = src.split('\n')
    source = [l + '\n' for l in lines[:-1]] + ([lines[-1]] if lines[-1] else [])
    return {'cell_type':'markdown','id':_cell_id('md', src),'metadata':{},'source':source}

def cell_code(src):
    return {'cell_type':'code','execution_count':None,'id':_cell_id('c', src),'metadata':{},'outputs':[],'source':[src]}

cells = []

# ── CELL 0 ─────────────────────────────────────────────────────────────────────
cells.append(cell_md("""# SEPA — Evolución de Canastas Representativas (Multi-canasta)

**Objetivo:** Calcular el costo mensual de hasta 6 canastas definidas por el economista,
mapearlas por provincia, compararlas con el IPC INDEC, y generar rankings por cadena y barrio.

**Canastas:** leídas desde las columnas `cantidad_01`...`cantidad_06` de la hoja `Selección`
del Excel `canasta_representativa_YYYY-MM.xlsx`. Solo se procesan las columnas con al menos
un producto con cantidad > 0.

**Canastas predefinidas:** Vulnerable · Popular · Media · Media Alta · Canasta 05 · Canasta 06

**Estructura:** Config → Setup → Canastas desde Excel → Maestros → ZIPs → Mes actual →
Canasta por sucursal → Análisis provincial → Serie histórica → IPC → Comparativa →
Gráficos IPC → Cuadro por canasta → Mapas coropléticos → Cobertura → Rankings →
Mapas Folium → Rankings CABA → Excel"""))

# ── CELL 1 — CONFIG ────────────────────────────────────────────────────────────
cells.append(cell_code("""\
# ===========================================================
# CONFIGURACIÓN — Modificar solo esta sección
# ===========================================================

SEPA_SOURCE = 'mi_drive'   # 'mi_drive' | 'local'

SEPA_DIR   = '/content/drive/MyDrive/carga'
OUTPUT_DIR = '/content/drive/MyDrive/carga/output_canasta'

USE_CACHE = True

# Período mínimo de la serie histórica
MES_INICIO_HISTORICO = '2024-01'

# Mes base para gráficos de índice (auto-adapta si no está en la serie)
MES_INICIO_GRAFICO = '2024-01'

# Mínimo productos propios para incluir sucursal en análisis
MIN_PRODUCTOS_PROPIOS = 15

# Mínimo sucursales por cadena para aparecer en rankings
MIN_SUCURSALES_RANKING = 10"""))

# ── CELL 2 — SETUP ─────────────────────────────────────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 2 — Montar Drive + dependencias + imports
# ============================================================
try:
    import google.colab
    from google.colab import drive
    drive.mount('/content/drive')
    print('Google Drive montado')
except ImportError:
    print('Entorno local')

import subprocess, sys
subprocess.run([sys.executable, '-m', 'pip', 'install',
                'folium', 'openpyxl', 'tqdm', 'pyarrow', '-q'], check=False)

import zipfile, gzip, re, shutil, warnings, gc, hashlib
import json as _json
from pathlib import Path
from tqdm.auto import tqdm
warnings.filterwarnings('ignore')

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import matplotlib.colors as mcolors
import matplotlib.dates as mdates
from matplotlib.colors import LinearSegmentedColormap, Normalize
import seaborn as sns
import folium
from branca.colormap import LinearColormap

plt.rcParams['figure.figsize'] = (13, 6)
plt.rcParams['font.size'] = 11
sns.set_theme(style='whitegrid')
pd.set_option('display.max_columns', 50)
pd.set_option('display.float_format', '{:,.2f}'.format)

SEPA_DIR   = Path(SEPA_DIR)
OUTPUT_DIR = Path(OUTPUT_DIR)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
CACHE_DIR  = OUTPUT_DIR / '_cache'
if USE_CACHE:
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
TMP_DIR = Path('/content/tmp_sepa_nb02')
TMP_DIR.mkdir(exist_ok=True)

_ipc_candidatos = [SEPA_DIR / n for n in ('IPC.xlsx', 'ipc.xlsx', 'IPC.XLSX')]
_ipc_encontrado = next((p for p in _ipc_candidatos if p.exists()), None)
IPC_PATH     = _ipc_encontrado if _ipc_encontrado else SEPA_DIR / 'IPC.xlsx'
GEOJSON_PATH = SEPA_DIR / 'ar.json'

print(f'SEPA_DIR:   {SEPA_DIR}')
print(f'OUTPUT_DIR: {OUTPUT_DIR}')
print(f'  IPC.xlsx: {"OK — " + IPC_PATH.name if IPC_PATH.exists() else "NO ENCONTRADO"}')
print(f'  ar.json:  {"OK" if GEOJSON_PATH.exists() else "NO ENCONTRADO"}')"""))

# ── CELL 3 — MULTI-CANASTA FROM EXCEL ──────────────────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 3 — Canastas desde Excel (columnas cantidad_01..06)
# ============================================================
import glob as _glob

patrones = sorted(_glob.glob(str(OUTPUT_DIR / 'canasta_representativa_*.xlsx')), reverse=True)
if not patrones:
    raise FileNotFoundError(
        f'No se encontro canasta_representativa_*.xlsx en {OUTPUT_DIR}\\n'
        'Ejecuta primero 01_exploracion_productos.ipynb y completa las columnas cantidad_01..06.'
    )

CANASTA_EXCEL = Path(patrones[0])
print(f'Excel de canasta: {CANASTA_EXCEL.name}')

for _sn in ['Selección', 'Seleccion']:
    try:
        sel = pd.read_excel(CANASTA_EXCEL, sheet_name=_sn, dtype={'id_producto': str})
        break
    except Exception:
        pass
else:
    raise ValueError(f'No se encontró la hoja Seleccion/Selección en {CANASTA_EXCEL.name}')

sel['id_producto'] = sel['id_producto'].str.strip().str.zfill(13)
sel['ean_norm']    = sel['id_producto'].str.lstrip('0')

desc_col = next((c for c in ['descripcion','descripcion_producto','nombre']
                 if c in sel.columns), sel.columns[3])
cat_col  = next((c for c in ['categoria','rubro'] if c in sel.columns), 'categoria')

# ── Nombres, colores y estilos por canasta ───────────────────────────────────
CANT_COLS = ['cantidad_01','cantidad_02','cantidad_03','cantidad_04','cantidad_05','cantidad_06']
CANASTA_NAMES = {
    'cantidad_01': 'Vulnerable',
    'cantidad_02': 'Popular',
    'cantidad_03': 'Media',
    'cantidad_04': 'Media Alta',
    'cantidad_05': 'Celíaca Media',
    'cantidad_06': 'Vegana Básica',
}
CANASTA_SHORT = {
    'cantidad_01': 'vulnerable',
    'cantidad_02': 'popular',
    'cantidad_03': 'media',
    'cantidad_04': 'media_alta',
    'cantidad_05': 'celiaca_media',
    'cantidad_06': 'vegana_basica',
}
CANASTA_COLORS = {
    'cantidad_01': '#0055A4',
    'cantidad_02': '#e74c3c',
    'cantidad_03': '#27ae60',
    'cantidad_04': '#f39c12',
    'cantidad_05': '#9b59b6',
    'cantidad_06': '#1abc9c',
}
CANASTA_LINESTYLES = {
    'cantidad_01': '-',
    'cantidad_02': '--',
    'cantidad_03': '-.',
    'cantidad_04': ':',
    'cantidad_05': (0,(5,1)),
    'cantidad_06': (0,(3,1,1,1)),
}
CANASTA_MARKERS = {
    'cantidad_01': 'o', 'cantidad_02': 's', 'cantidad_03': '^',
    'cantidad_04': 'D', 'cantidad_05': 'v', 'cantidad_06': 'P',
}

# ── Construir CANASTAS (solo las activas) ────────────────────────────────────
CANASTAS = {}
for _col in CANT_COLS:
    if _col not in sel.columns:
        continue
    _qty_s = pd.to_numeric(sel[_col], errors='coerce').fillna(0).astype(int)
    _activos = sel[_qty_s > 0].copy()
    _activos['_qty'] = _qty_s[_qty_s > 0].values
    if len(_activos) == 0:
        continue
    CANASTAS[_col] = {
        row['ean_norm']: (str(row[desc_col])[:50], int(row['_qty']), str(row[cat_col]))
        for _, row in _activos.iterrows()
    }

if not CANASTAS:
    raise ValueError(
        'Ninguna columna cantidad_01..06 tiene productos con cantidad > 0.\\n'
        'Completa al menos una columna cantidad en la hoja Seleccion.'
    )

CANASTAS_ACTIVAS   = list(CANASTAS.keys())
CANASTA_EANS_NORM  = set().union(*[set(c.keys()) for c in CANASTAS.values()])
CANASTA_EANS_ZFILL = {e.zfill(13) for e in CANASTA_EANS_NORM}
N_EANS_TOTAL       = len(CANASTA_EANS_NORM)
N_CANASTAS         = len(CANASTAS)

# Safeguard MIN_PRODUCTOS_PROPIOS vs tamaño de canasta
_max_n = max(len(c) for c in CANASTAS.values())
if MIN_PRODUCTOS_PROPIOS >= _max_n:
    MIN_PRODUCTOS_PROPIOS = max(1, _max_n // 2)
    print(f'AVISO: MIN_PRODUCTOS_PROPIOS ajustado a {MIN_PRODUCTOS_PROPIOS} (mayor canasta: {_max_n} prods)')

print(f'\\nCanastas activas: {N_CANASTAS}')
for _col in CANASTAS_ACTIVAS:
    _c = CANASTAS[_col]
    _u = sum(v[1] for v in _c.values())
    print(f'  [{CANASTA_NAMES[_col]}] {len(_c)} productos, {_u} unidades/mes')
print(f'EANs únicos (unión): {N_EANS_TOTAL}')"""))

# ── CELL 4 — MAESTROS ──────────────────────────────────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 4 — Maestros de sucursales, cadenas, provincias
# ============================================================
DATA_URL = 'https://raw.githubusercontent.com/santiagoriverti/precios_minoristas_supermercados/main/data'

def leer_maestro(nombre):
    local = Path('data') / nombre
    if local.exists():
        return pd.read_excel(local)
    import urllib.request
    dl = Path('/content/data') / nombre
    Path('/content/data').mkdir(exist_ok=True)
    if not dl.exists():
        print(f'  Descargando {nombre}...')
        urllib.request.urlretrieve(f'{DATA_URL}/{nombre}', dl)
    return pd.read_excel(dl)

print('Cargando maestros...')
maestro_suc = leer_maestro('maestro_sucursales_completo.xlsx')
for c in ['id_comercio','id_bandera','id_sucursal']:
    maestro_suc[c] = maestro_suc[c].astype(str)

suc_pais = maestro_suc[
    maestro_suc['sucursales_latitud'].notna() &
    maestro_suc['sucursales_longitud'].notna() &
    (maestro_suc['sucursales_latitud'].between(-55, -22)) &
    (maestro_suc['sucursales_longitud'].between(-73, -53))
].copy()

IDS_PAIS = set(zip(suc_pais['id_comercio'], suc_pais['id_bandera'], suc_pais['id_sucursal']))
print(f'  Sucursales validas: {len(suc_pais):,} | IDs unicos: {len(IDS_PAIS):,}')

NOMBRES_COMPUESTOS = {
    ('9','1'):'Vea',('9','2'):'Disco',('9','3'):'Jumbo',
    ('10','1'):'Carrefour',('10','2'):'Carrefour Market',('10','3'):'Carrefour Express',
    ('11','2'):'ChangoMas',('11','4'):'Hiper ChangoMas',('11','5'):'Mi ChangoMas',
    ('16','1'):'Hipermercado Libertad',('16','2'):'Mini Libertad',
}
NOMBRES_SIMPLES = {
    '2':'La Anonima','3':'Cadena 3','5':'Hipermercado Misiones',
    '8':'Cadena 8 (Cordoba)','12':'Coto','13':'Cooperativa Obrera',
    '15':'DIA','20':'LAR','21':'Toledo','23':'Cadena 23','47':'Pasamonte',
}

def asignar_cadena(row):
    k = (row['id_comercio'], row['id_bandera'])
    if k in NOMBRES_COMPUESTOS: return NOMBRES_COMPUESTOS[k]
    if row['id_comercio'] in NOMBRES_SIMPLES: return NOMBRES_SIMPLES[row['id_comercio']]
    return f"Cadena {row['id_comercio']}"

PROV_NORM = {
    'Ciudad Autonoma de Buenos Aires':'CABA',
    'Ciudad Autónoma de Buenos Aires':'CABA',
    'Provincia de Buenos Aires':'Buenos Aires',
    'Provincia de Catamarca':'Catamarca',
    'Provincia del Chaco':'Chaco','Provincia del Chubut':'Chubut',
    'Provincia de Cordoba':'Córdoba','Provincia de Córdoba':'Córdoba',
    'Provincia de Corrientes':'Corrientes',
    'Provincia de Entre Rios':'Entre Ríos','Provincia de Entre Ríos':'Entre Ríos',
    'Provincia de Formosa':'Formosa','Provincia de Jujuy':'Jujuy',
    'Provincia de La Pampa':'La Pampa','Provincia de La Rioja':'La Rioja',
    'Provincia de Mendoza':'Mendoza','Provincia de Misiones':'Misiones',
    'Provincia del Neuquen':'Neuquén','Provincia del Neuquén':'Neuquén','Neuquén':'Neuquén',
    'Provincia de Rio Negro':'Río Negro','Provincia de Río Negro':'Río Negro',
    'Provincia de Salta':'Salta','Provincia de San Juan':'San Juan',
    'Provincia de San Luis':'San Luis','Provincia de Santa Cruz':'Santa Cruz',
    'Provincia de Santa Fe':'Santa Fe',
    'Provincia de Santiago del Estero':'Santiago del Estero',
    'Provincia de Tierra del Fuego, Antartida e Islas del Atlantico Sur':'Tierra del Fuego',
    'Provincia de Tierra del Fuego, Antártida e Islas del Atlántico Sur':'Tierra del Fuego',
    'Tierra del Fuego':'Tierra del Fuego',
    'Provincia de Tucuman':'Tucumán','Provincia de Tucumán':'Tucumán',
    'Buenos Aires':'Buenos Aires','CABA':'CABA','Catamarca':'Catamarca',
    'Chaco':'Chaco','Chubut':'Chubut','Córdoba':'Córdoba','Corrientes':'Corrientes',
    'Entre Ríos':'Entre Ríos','Formosa':'Formosa','Jujuy':'Jujuy',
    'La Pampa':'La Pampa','La Rioja':'La Rioja','Mendoza':'Mendoza',
    'Misiones':'Misiones','Río Negro':'Río Negro','Salta':'Salta',
    'San Juan':'San Juan','San juan':'San Juan','SAN JUAN':'San Juan',
    'San Luis':'San Luis','Santa Cruz':'Santa Cruz',
    'Santa Fe':'Santa Fe','Santiago del Estero':'Santiago del Estero','Tucumán':'Tucumán',
}

PESOS_POBLACION = {
    'Buenos Aires':17709732,'CABA':3075646,'Catamarca':415438,
    'Chaco':1204541,'Chubut':618994,'Córdoba':3978984,
    'Corrientes':1120801,'Entre Ríos':1385961,'Formosa':605193,
    'Jujuy':770881,'La Pampa':368550,'La Rioja':393531,
    'Mendoza':2014533,'Misiones':1261294,'Neuquén':664057,
    'Río Negro':747610,'Salta':1441998,'San Juan':781217,
    'San Luis':531745,'Santa Cruz':333473,'Santa Fe':3556522,
    'Santiago del Estero':1019304,'Tierra del Fuego':190641,'Tucumán':1737127,
}
print('Maestros OK')"""))

# ── CELL 5 — ZIP FUNCTIONS ─────────────────────────────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 5 — Funciones de lectura de ZIPs SEPA
# ============================================================
_PAT_SEM = re.compile(r'^(\\d{4})(A|B)$', re.IGNORECASE)
_PAT_ARC = re.compile(r'^(\\d{2})(\\d{4})_pais_parte.*COMPLETO.*\\.csv\\.gz$', re.IGNORECASE)

def detectar_semestres():
    result = []
    for z in sorted(SEPA_DIR.glob('*.zip')):
        m = _PAT_SEM.match(z.stem)
        if m:
            result.append((z, int(m.group(1)), m.group(2).upper()))
    return result

def archivos_por_mes(zip_path):
    meses = {}
    with zipfile.ZipFile(zip_path) as zf:
        for nombre in zf.namelist():
            m = _PAT_ARC.match(Path(nombre).name)
            if m:
                mes_n, anio = int(m.group(1)), int(m.group(2))
                meses.setdefault((anio, mes_n), []).append(nombre)
    return meses

def detectar_ultimo_mes():
    sems = detectar_semestres()
    if not sems:
        raise RuntimeError(f'No hay ZIPs semestrales en {SEPA_DIR}')
    for zip_path, anio, sem in reversed(sems):
        meses = archivos_por_mes(zip_path)
        if meses:
            (a, m) = max(meses.keys())
            return zip_path, a, m, meses[(a, m)]
    raise RuntimeError('No se pudo detectar el ultimo mes')

def normalizar_ean(s):
    if pd.isna(s): return None
    s = str(s).strip().lstrip('0')
    return s if s else '0'

sems = detectar_semestres()
print(f'Semestres encontrados: {len(sems)}')
for z, a, s in sems:
    meses = archivos_por_mes(z)
    meses_str = [f'{m:02d}/{an}' for (an,m) in sorted(meses.keys())]
    print(f'  {a}{s}: {meses_str}')"""))

# ── CELL 6 — CURRENT MONTH ─────────────────────────────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 6 — Mes actual: carga per-sucursal desde ZIP
# ============================================================
zip_actual, ANIO_ACTUAL, MES_NUM_ACTUAL, archivos_mes = detectar_ultimo_mes()

MES          = f'{MES_NUM_ACTUAL:02d}{ANIO_ACTUAL}'
PERIODO      = f'{ANIO_ACTUAL}-{MES_NUM_ACTUAL:02d}'
ULTIMO_MES   = PERIODO

_NOM = {'01':'enero','02':'febrero','03':'marzo','04':'abril','05':'mayo','06':'junio',
        '07':'julio','08':'agosto','09':'septiembre','10':'octubre','11':'noviembre','12':'diciembre'}
NOMBRE_MES       = f"{_NOM[f'{MES_NUM_ACTUAL:02d}']} {ANIO_ACTUAL}"
NOMBRE_MES_TITLE = NOMBRE_MES.title()

CADENAS_FILTRAR = {'19','2013','3001','4'}
PAT_FECHA = re.compile(r'^precio_(\\d{8})$')

print(f'Procesando: {NOMBRE_MES_TITLE}  ({zip_actual.name})')
print(f'Archivos: {archivos_mes}\\n')

# ── Detección de MES PARCIAL (menos días cargados que los del calendario) ─────
# El último mes en curso (ej. junio al día 23) tiene menos días que el mes
# completo. Su NIVEL es un snapshot válido, pero su VARIACIÓN mensual queda
# subestimada al compararse contra meses completos. Se marca para no confundir.
import calendar as _cal
_dias_cargados = set()
with zipfile.ZipFile(zip_actual) as _zfh:
    for _arch in archivos_mes:
        with _zfh.open(_arch) as _s:
            _hdr0 = gzip.open(_s, 'rt', encoding='utf-8', errors='replace').readline()
        for _c in _hdr0.strip().split(','):
            _mm = PAT_FECHA.match(_c)
            if _mm: _dias_cargados.add(_mm.group(1))
DIAS_CARGADOS  = len(_dias_cargados)
DIAS_MES       = _cal.monthrange(ANIO_ACTUAL, MES_NUM_ACTUAL)[1]
MES_PARCIAL    = 0 < DIAS_CARGADOS < DIAS_MES
SUFIJO_PARCIAL = f' (parcial: {DIAS_CARGADOS}/{DIAS_MES} días)' if MES_PARCIAL else ''
if MES_PARCIAL:
    print(f'⚠️  {NOMBRE_MES_TITLE} es un MES PARCIAL: {DIAS_CARGADOS}/{DIAS_MES} días cargados.')
    print(f'    El NIVEL es un snapshot al día {DIAS_CARGADOS}; la VARIACIÓN mensual está SUBESTIMADA')
    print(f'    (se compara contra meses completos). Re-correr al cierre del mes para el dato definitivo.\\n')

acumulador = []
for archivo in sorted(archivos_mes):
    nombre = Path(archivo).name
    tmp_p  = TMP_DIR / nombre
    with zipfile.ZipFile(zip_actual) as zf:
        with zf.open(archivo) as src, open(tmp_p, 'wb') as dst:
            shutil.copyfileobj(src, dst, length=4*1024*1024)
    print(f'  {nombre}...')
    with gzip.open(tmp_p, 'rt', encoding='utf-8', errors='replace') as g:
        for chunk in pd.read_csv(g, dtype=str, chunksize=300_000, low_memory=False):
            chunk['ean_norm'] = chunk['id_producto'].apply(normalizar_ean)
            chunk = chunk[chunk['ean_norm'].isin(CANASTA_EANS_NORM)].copy()
            if len(chunk) == 0: continue
            for c in ['id_comercio','id_bandera','id_sucursal']:
                chunk[c] = chunk[c].astype(str)
            chunk['_k'] = list(zip(chunk['id_comercio'],chunk['id_bandera'],chunk['id_sucursal']))
            chunk = chunk[chunk['_k'].isin(IDS_PAIS)].drop(columns=['_k']).copy()
            if len(chunk) == 0: continue
            cols_p = [c for c in chunk.columns if PAT_FECHA.match(c)]
            if not cols_p: continue
            df_long = chunk.melt(
                id_vars=['id_comercio','id_bandera','id_sucursal','ean_norm'],
                value_vars=cols_p, var_name='_col', value_name='precio_raw')
            df_long['precio'] = pd.to_numeric(
                df_long['precio_raw'].replace('NA', np.nan), errors='coerce')
            df_long = df_long[df_long['precio'].notna() & (df_long['precio'] > 0)].copy()
            df_long.drop(columns=['_col','precio_raw'], inplace=True)
            acumulador.append(df_long)
    tmp_p.unlink(missing_ok=True)
    del chunk, df_long; gc.collect()

if not acumulador:
    raise RuntimeError('Sin datos de canasta para el mes actual')

datos = pd.concat(acumulador, ignore_index=True)
del acumulador; gc.collect()
# NO se deduplica: se conservan TODAS las observaciones (sucursal x dia) del mes.
# El precio por sucursal (mediana Y promedio sobre los dias) se calcula en CELDA 7.

# ── Promedio con outliers fuera (banda relativa a la mediana) ────────────────
# Para el analisis PROMEDIO: en cada grupo se descartan los valores fuera de
# [mediana/4, mediana x4] (errores gruesos del SEPA: 100x, centavos sueltos) y
# recien despues se promedia. Robusto a cualquier tamano de muestra. Se usa en
# TODAS las agregaciones chicas del analisis promedio (provincias, cadenas,
# barrios). Para el groupby grande (sucursal x EAN sobre dias) se usa una
# version VECTORIZADA equivalente en la CELDA 7 (mismo criterio, mas rapida).
def _pmean(_s):
    _s = pd.to_numeric(_s, errors='coerce').dropna()
    if len(_s) == 0: return float('nan')
    _m = _s.median()
    if _m and _m > 0:
        _f = _s[(_s >= _m/4) & (_s <= _m*4)]
        if len(_f) > 0: _s = _f
    return _s.mean()

ref_e = {'7790072002080'.lstrip('0'), '7790070320285'.lstrip('0'), '7790132098459'.lstrip('0')}
ref_d = datos[datos['ean_norm'].isin(ref_e)]
med_r = ref_d['precio'].median() if len(ref_d) > 0 else datos['precio'].median()
FACTOR = 100 if med_r > 10_000 else 1
if FACTOR == 100:
    datos['precio'] /= 100
    print(f'Factor: {FACTOR} (centavos -> pesos)')
else:
    print(f'Factor: {FACTOR} (ya en pesos)')

_n_suc = datos.groupby(['id_comercio','id_bandera','id_sucursal']).ngroups
print(f'Observaciones (sucursal x dia): {len(datos):,} | Sucursales: {_n_suc:,}')
print(f'EANs únicos con datos: {datos["ean_norm"].nunique()} / {N_EANS_TOTAL} (unión de canastas activas)')"""))

# ── CELL 7 — PER-SUCURSAL BASKETS (multi-canasta) ─────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 7 — Canasta por sucursal para cada canasta activa
# ============================================================
# Precio por sucursal-EAN agregando sobre TODOS los dias del mes:
#   precio_mediana = mediana de los dias (analisis MEDIANA)
#   precio_prom    = media con outliers fuera [med/4, med x4] (analisis PROMEDIO)
# La limpieza de outliers del promedio es VECTORIZADA (equivalente a _pmean).
_keys = ['id_comercio','id_bandera','id_sucursal','ean_norm']
_med_dia = datos.groupby(_keys)['precio'].transform('median')
_mask_ok = (datos['precio'] >= _med_dia/4) & (datos['precio'] <= _med_dia*4)
_gm = datos.groupby(_keys)['precio'].median().rename('precio_mediana')
_gp = datos[_mask_ok].groupby(_keys)['precio'].mean().rename('precio_prom')
precio_mes = pd.concat([_gm, _gp], axis=1).reset_index()
precio_mes['precio_prom'] = precio_mes['precio_prom'].fillna(precio_mes['precio_mediana'])
precio_mes = precio_mes[~precio_mes['id_comercio'].isin(CADENAS_FILTRAR)].copy()
# Referencias nacionales por EAN para imputar productos faltantes (una por analisis):
precio_med_nac  = precio_mes.groupby('ean_norm')['precio_mediana'].median().to_dict()
precio_prom_nac = precio_mes.groupby('ean_norm')['precio_prom'].agg(_pmean).to_dict()

# ── Bboxes y función de geocodificación ─────────────────────────────────────
_PROV_BBOX = {
    'CABA':                (-34.72,-34.52,-58.54,-58.33),
    'Tucumán':             (-28.0, -26.0, -66.5, -64.5),
    'Jujuy':               (-24.5, -21.5, -67.5, -63.5),
    'Misiones':            (-28.5, -25.5, -56.5, -53.0),
    'Chaco':               (-27.5, -24.0, -63.0, -57.5),
    'Formosa':             (-26.5, -22.0, -62.5, -58.0),
    'Corrientes':          (-30.5, -27.0, -60.0, -55.5),
    'Entre Ríos':          (-34.0, -30.0, -60.5, -57.5),
    'San Luis':            (-36.0, -32.5, -68.5, -65.0),
    'San Juan':            (-34.5, -27.5, -71.0, -65.0),
    'La Rioja':            (-32.5, -27.0, -70.0, -65.0),
    'Catamarca':           (-29.5, -25.0, -70.5, -64.5),
    'Salta':               (-26.5, -21.5, -68.5, -62.5),
    'Santiago del Estero': (-30.0, -25.5, -65.5, -61.5),
    'Mendoza':             (-37.5, -32.0, -70.5, -66.5),
    'Neuquén':             (-40.5, -36.0, -71.5, -68.5),
    'La Pampa':            (-40.0, -35.0, -68.5, -63.5),
    'Santa Fe':            (-34.5, -28.5, -62.5, -59.0),
    'Córdoba':             (-39.0, -29.5, -67.0, -62.0),
    'Río Negro':           (-42.5, -38.5, -71.5, -62.5),
    'Chubut':              (-46.5, -41.0, -72.5, -63.0),
    'Buenos Aires':        (-42.5, -33.5, -63.5, -56.5),
    'Santa Cruz':          (-52.5, -46.0, -72.5, -65.5),
    'Tierra del Fuego':    (-55.5, -51.0, -70.5, -63.5),
}

def _geocodif(lat, lon):
    # Primera provincia cuyo bbox contiene (lat, lon)
    if pd.isna(lat) or pd.isna(lon): return None
    for p, (la0, la1, lo0, lo1) in _PROV_BBOX.items():
        if la0 <= lat <= la1 and lo0 <= lon <= lo1: return p
    return None

# ── Pre-limpiar geografía (ONCE para todas las canastas) ─────────────────────
_cols_suc = ['id_comercio','id_bandera','id_sucursal',
             'sucursales_nombre','sucursales_latitud','sucursales_longitud',
             'sucursales_barrio','sucursales_localidad','PROVINCIA']
if 'sucursales_tipo' in suc_pais.columns:
    _cols_suc.append('sucursales_tipo')

suc_geo_clean = suc_pais[_cols_suc].copy()
if 'sucursales_tipo' not in suc_geo_clean.columns:
    suc_geo_clean['sucursales_tipo'] = 'N/D'
suc_geo_clean['cadena'] = suc_geo_clean.apply(asignar_cadena, axis=1)
suc_geo_clean['PROVINCIA_NORM'] = suc_geo_clean['PROVINCIA'].map(PROV_NORM).fillna(suc_geo_clean['PROVINCIA'])
suc_geo_clean = suc_geo_clean[suc_geo_clean['sucursales_tipo'] != 'Web'].copy()
_mask_caba = (
    suc_geo_clean['PROVINCIA_NORM'].eq('CABA') & (
        (suc_geo_clean['sucursales_latitud'] < -34.71) |
        (suc_geo_clean['sucursales_latitud'] > -34.53) |
        (suc_geo_clean['sucursales_longitud'] < -58.53) |
        (suc_geo_clean['sucursales_longitud'] > -58.34)))
suc_geo_clean = suc_geo_clean[~_mask_caba].copy()
_n_reclasif = 0
for _idx, _row in suc_geo_clean.iterrows():
    _p = _row['PROVINCIA_NORM']
    if _p not in _PROV_BBOX: continue
    _la0,_la1,_lo0,_lo1 = _PROV_BBOX[_p]
    _lat,_lon = _row['sucursales_latitud'],_row['sucursales_longitud']
    if _la0<=_lat<=_la1 and _lo0<=_lon<=_lo1: continue
    _nueva = _geocodif(_lat,_lon)
    if _nueva and _nueva != _p:
        suc_geo_clean.at[_idx,'PROVINCIA_NORM'] = _nueva
        _n_reclasif += 1
if _n_reclasif:
    print(f'  Reclasificadas {_n_reclasif} sucursales por coordenadas')

# ── Calcular canasta por sucursal para CADA canasta activa ───────────────────
canasta_geo_dict = {}

for _col_id, _canasta in CANASTAS.items():
    _name    = CANASTA_NAMES[_col_id]
    _min_p   = min(MIN_PRODUCTOS_PROPIOS, len(_canasta))

    # Cada sucursal aporta DOS costos de canasta: uno con precios medianos
    # (canasta_total) y otro con precios promedio-sin-outliers (canasta_total_prom).
    # Productos faltantes se imputan con la referencia nacional del mismo analisis.
    def _calc(grupo, _c=_canasta, _pm=precio_med_nac, _pp=precio_prom_nac):
        loc_m = dict(zip(grupo['ean_norm'], grupo['precio_mediana']))
        loc_p = dict(zip(grupo['ean_norm'], grupo['precio_prom']))
        tot_m = 0.0; tot_p = 0.0; propios = 0
        for ean, (nom, qty, cat) in _c.items():
            if ean in loc_m:
                prm = loc_m[ean]; prp = loc_p[ean]; propios += 1
            else:
                prm = _pm.get(ean, 0); prp = _pp.get(ean, 0)
            tot_m += prm * qty; tot_p += prp * qty
        return pd.Series({'canasta_total': tot_m, 'canasta_total_prom': tot_p,
                          'productos_propios': propios})

    _suc = (precio_mes.groupby(['id_comercio','id_bandera','id_sucursal'])
            .apply(_calc, include_groups=False).reset_index())
    _suc = _suc[_suc['productos_propios'] >= _min_p].copy()
    _cgeo = _suc.merge(suc_geo_clean, on=['id_comercio','id_bandera','id_sucursal'], how='inner')
    canasta_geo_dict[_col_id] = _cgeo.copy()
    print(f'  [{_name}] {len(_cgeo):,} sucursales | '
          f'mediana ${_cgeo["canasta_total"].min():,.0f}–${_cgeo["canasta_total"].max():,.0f}')

print()
print('Cadenas (primera canasta activa):')
print(canasta_geo_dict[CANASTAS_ACTIVAS[0]]['cadena'].value_counts().to_string())"""))

# ── CELL 8 — PROVINCE ANALYSIS (multi-canasta) ────────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 8 — Análisis provincial para cada canasta activa
# ============================================================
def fmt_ar(x, dec=0):
    s = f'{x:,.{dec}f}'
    return s.replace(',','X').replace('.',',').replace('X','.')

serie_prov_dict     = {}   # col_id -> provincia: canasta_total (mediana) + canasta_total_prom (promedio)
prom_nac_dict       = {}   # col_id -> nacional ponderado por poblacion — MEDIANA
prom_nac_prom_dict  = {}   # col_id -> nacional ponderado por poblacion — PROMEDIO (outliers fuera)

for _col_id in CANASTAS_ACTIVAS:
    _cgf  = canasta_geo_dict[_col_id]
    # MEDIANA : mediana provincial de las canastas medianas por sucursal.
    # PROMEDIO: media provincial (outliers fuera, _pmean) de las canastas promedio por sucursal.
    _cpp  = (_cgf.groupby('PROVINCIA_NORM')
             .agg(canasta_total=('canasta_total','median'),
                  canasta_total_prom=('canasta_total_prom', _pmean)).reset_index()
             .rename(columns={'PROVINCIA_NORM':'provincia'}))
    _cpp['mes']  = PERIODO
    _cpp['peso'] = _cpp['provincia'].map(PESOS_POBLACION).fillna(0)
    _pob = _cpp[_cpp['peso'] > 0]['peso'].sum()
    _prom      = ((_cpp['canasta_total'] * _cpp['peso']).sum() / _pob
                  if _pob > 0 else _cpp['canasta_total'].mean())
    _prom_prom = ((_cpp['canasta_total_prom'] * _cpp['peso']).sum() / _pob
                  if _pob > 0 else _cpp['canasta_total_prom'].mean())
    serie_prov_dict[_col_id]    = _cpp[['mes','provincia','canasta_total','canasta_total_prom']].copy()
    prom_nac_dict[_col_id]      = _prom
    prom_nac_prom_dict[_col_id] = _prom_prom

print(f'=== CUADRO: Canastas por provincia — {NOMBRE_MES_TITLE} ===')
for _col_id in CANASTAS_ACTIVAS:
    _name = CANASTA_NAMES[_col_id]
    _spv  = serie_prov_dict[_col_id].sort_values('canasta_total')
    _prom = prom_nac_dict[_col_id]
    print(f'\\n  ── {_name} ──')
    for _, r in _spv.iterrows():
        _vs = ((r['canasta_total'] / _prom) - 1) * 100
        print(f'  {r["provincia"]:<25} ${r["canasta_total"]:>10,.0f}  {_vs:+.2f}%')
    print(f'  {"Promedio (ponderado)":<25} ${_prom:>10,.0f}   0.00%')
print(f'\\nProvincias con datos (primera canasta): {len(serie_prov_dict[CANASTAS_ACTIVAS[0]])}')"""))

# ── CELL 9 — HISTORICAL SERIES (one raw cache, per-canasta aggregation) ───────
cells.append(cell_code("""\
# ============================================================
# CELDA 9 — Serie histórica (cache de meses cerrados + mes en curso fresco)
# ============================================================
# Cache identificado por la unión de EANs activos (cambiar cantidades reusa;
# agregar/quitar EANs invalida). IMPORTANTE: el cache guarda SOLO meses
# CERRADOS. El último mes disponible (en curso, crece día a día) se RELEE
# SIEMPRE fresco, así sus promedios usan los días efectivamente cargados.
_cache_key  = hashlib.md5('|'.join(sorted(CANASTA_EANS_NORM)).encode()).hexdigest()[:8]
# _v2m: esquema con precio_medio (media recortada 1%) ademas de precio_mediano.
# Al cambiar el nombre, la primera corrida reconstruye el cache con ambas medidas
# (los .parquet viejos sin precio_medio quedan huerfanos; se pueden borrar).
_cache_path = CACHE_DIR / f'hist_union_{_cache_key}_v2m.parquet'

# Mapa de meses disponibles -> (zip, archivos del mes)
_mapa_mes = {}
for _zip_path, _anio, _sem in detectar_semestres():
    for (_anio_m, _mes_m), _archs in archivos_por_mes(_zip_path).items():
        _lbl = f'{_anio_m}-{_mes_m:02d}'
        if _lbl >= MES_INICIO_HISTORICO:
            _mapa_mes[_lbl] = (_zip_path, _archs)
_meses_disp = sorted(_mapa_mes)
if not _meses_disp:
    raise RuntimeError(f'No hay meses disponibles >= {MES_INICIO_HISTORICO}')
_mes_actual = _meses_disp[-1]   # mes en curso -> NUNCA se cachea

def _leer_mes_hist(_lbl):
    # Lee los archivos del mes _lbl, filtra a los EANs de las canastas y
    # devuelve [ean_norm, precio_mediano, anio_mes] (mediana sobre todos los
    # días cargados de ese mes). Devuelve None si no hay datos.
    _zip_path, _archs = _mapa_mes[_lbl]
    _all_rows = []
    for _archivo in sorted(_archs):
        _tmp_p = TMP_DIR / Path(_archivo).name
        with zipfile.ZipFile(_zip_path) as _zf:
            with _zf.open(_archivo) as _s, open(_tmp_p,'wb') as _d:
                shutil.copyfileobj(_s, _d, length=4*1024*1024)
        with gzip.open(_tmp_p,'rt',encoding='utf-8',errors='replace') as _g:
            for _chunk in pd.read_csv(_g, dtype=str, chunksize=300_000, low_memory=False):
                _chunk['ean_norm'] = _chunk['id_producto'].apply(normalizar_ean)
                _chunk = _chunk[_chunk['ean_norm'].isin(CANASTA_EANS_NORM)].copy()
                if len(_chunk) == 0: continue
                _cols_p = [c for c in _chunk.columns if re.match(r'^precio_\\d{8}$', c)]
                if not _cols_p: continue
                _sub = _chunk[['ean_norm']+_cols_p].copy()
                for _cp in _cols_p:
                    _sub[_cp] = pd.to_numeric(_sub[_cp].replace('NA',np.nan), errors='coerce')
                _mlt = _sub.melt(id_vars='ean_norm', value_vars=_cols_p,
                                 var_name='_c', value_name='precio')
                _mlt = _mlt[_mlt['precio'].notna() & (_mlt['precio']>0)]
                _all_rows.append(_mlt[['ean_norm','precio']])
        _tmp_p.unlink(missing_ok=True)
    if not _all_rows:
        return None
    _df_m = pd.concat(_all_rows, ignore_index=True)
    _fac = 100 if _df_m['precio'].median() > 10_000 else 1
    if _fac == 100: _df_m['precio'] /= 100
    _agg = (_df_m.groupby('ean_norm')['precio']
            .agg(precio_mediano='median', precio_medio=_pmean).reset_index())
    _agg['anio_mes'] = _lbl
    del _df_m, _all_rows; gc.collect()
    return _agg

# ── 1) Meses CERRADOS (todos menos el último): usar/actualizar cache ──────────
if USE_CACHE and _cache_path.exists():
    df_cache = pd.read_parquet(_cache_path)
    df_cache = df_cache[df_cache['anio_mes'] < _mes_actual].copy()
else:
    df_cache = pd.DataFrame(columns=['ean_norm','precio_mediano','precio_medio','anio_mes'])

_meses_en_cache = set(df_cache['anio_mes'].unique())
_faltantes = [m for m in _meses_disp if m < _mes_actual and m not in _meses_en_cache]
_nuevos = []
for _lbl in _faltantes:
    _agg = _leer_mes_hist(_lbl)
    if _agg is not None:
        _nuevos.append(_agg)
        print(f'  {_lbl}: {_agg["ean_norm"].nunique()} EANs (agregado al cache)')
if _nuevos:
    df_cache = pd.concat([df_cache] + _nuevos, ignore_index=True)
    if USE_CACHE:
        df_cache.to_parquet(_cache_path, compression='snappy', index=False)
        print(f'Cache actualizado: {_cache_path.name} ({df_cache["anio_mes"].nunique()} meses cerrados)')
elif USE_CACHE and _cache_path.exists():
    print(f'Cache al día: {_cache_path.name} ({len(_meses_en_cache)} meses cerrados)')

# ── 2) Mes en curso: SIEMPRE fresco (usa los días cargados a la fecha) ────────
_actual = _leer_mes_hist(_mes_actual)
if _actual is not None:
    print(f'  {_mes_actual}: {_actual["ean_norm"].nunique()} EANs (mes en curso — recalculado fresco)')

df_hist_raw = (pd.concat([df_cache] + ([_actual] if _actual is not None else []),
                         ignore_index=True)
               if (len(df_cache) > 0 or _actual is not None)
               else pd.DataFrame(columns=['ean_norm','precio_mediano','precio_medio','anio_mes']))
df_hist_raw = df_hist_raw.sort_values('anio_mes').reset_index(drop=True)

# ── Agregar por canasta ──────────────────────────────────────────────────────
serie_nac_dict = {}   # col_id -> serie_nacional_valida

for _col_id, _canasta in CANASTAS.items():
    _name  = CANASTA_NAMES[_col_id]
    _eans  = set(_canasta.keys())
    _dh    = df_hist_raw[df_hist_raw['ean_norm'].isin(_eans)].copy()
    _dh['qty']              = _dh['ean_norm'].map(lambda e, c=_canasta: c.get(e,('?',0,'?'))[1])
    _dh['costo_item']      = _dh['precio_mediano'] * _dh['qty']
    _dh['costo_item_prom'] = _dh['precio_medio']   * _dh['qty']
    _sn = (_dh.groupby('anio_mes')
           .agg(canasta_nacional_ponderada=('costo_item','sum'),
                canasta_nacional_ponderada_prom=('costo_item_prom','sum'),
                n_eans=('ean_norm','nunique'))
           .reset_index().rename(columns={'anio_mes':'mes'})
           .sort_values('mes').reset_index(drop=True))
    _sn = _sn[_sn['mes'] >= MES_INICIO_HISTORICO].copy()
    _sn['variacion_mensual_%']      = _sn['canasta_nacional_ponderada'].pct_change() * 100
    _sn['variacion_mensual_prom_%'] = _sn['canasta_nacional_ponderada_prom'].pct_change() * 100
    _bv = _sn['canasta_nacional_ponderada'].iloc[0] if len(_sn) > 0 else 1
    _sn['indice_canasta_base100'] = (_sn['canasta_nacional_ponderada'] / _bv * 100).round(2)
    serie_nac_dict[_col_id] = _sn.copy()
    _rng = f'{_sn["mes"].min()} -> {_sn["mes"].max()}' if len(_sn) > 0 else 'sin datos'
    print(f'  [{_name}] {len(_sn)} meses | {_rng}')"""))

# ── CELL 10 — IPC ──────────────────────────────────────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 10 — IPC INDEC desde carga/IPC.xlsx
# ============================================================
if not IPC_PATH.exists():
    raise FileNotFoundError(
        f'IPC.xlsx no encontrado en {SEPA_DIR}\\n'
        'Asegurate de tener el archivo IPC.xlsx en la carpeta carga/'
    )

ipc_raw = pd.read_excel(IPC_PATH)
print(f'IPC cargado: {len(ipc_raw)} filas, columnas: {list(ipc_raw.columns[:4])} ...')

fecha_col = next((c for c in ipc_raw.columns
                  if str(c).lower().strip() in ('date','fecha','mes','period')),
                 ipc_raw.columns[0])

if pd.api.types.is_datetime64_any_dtype(ipc_raw[fecha_col]):
    ipc_raw['mes'] = ipc_raw[fecha_col].dt.strftime('%Y-%m')
else:
    _MES_ESP = {'ene':1,'feb':2,'mar':3,'abr':4,'may':5,'jun':6,
                'jul':7,'ago':8,'sep':9,'oct':10,'nov':11,'dic':12}
    def _parse_ipc_fecha(val):
        if pd.isna(val): return pd.NaT
        if isinstance(val, pd.Timestamp): return val
        s = str(val).strip().lower()
        try:
            partes = s.split('-')
            if len(partes) == 2 and partes[0] in _MES_ESP:
                return pd.Timestamp(year=int(partes[1]), month=_MES_ESP[partes[0]], day=1)
        except Exception:
            pass
        return pd.to_datetime(val, errors='coerce')
    ipc_raw['mes'] = ipc_raw[fecha_col].apply(_parse_ipc_fecha).dt.strftime('%Y-%m')

rename_map = {}
for c in ipc_raw.columns:
    cs = str(c).strip()
    if 'nivel general' in cs.lower():               rename_map[c] = 'ipc_general'
    elif 'alimentos y bebidas no alc' in cs.lower(): rename_map[c] = 'ipc_alimentos'
ipc_raw = ipc_raw.rename(columns=rename_map)

for c in ['ipc_general','ipc_alimentos']:
    if c in ipc_raw.columns:
        ipc_raw[c] = pd.to_numeric(
            ipc_raw[c].astype(str).str.replace(',', '.', regex=False), errors='coerce')

_ipc_cols = ['mes','ipc_general'] + (['ipc_alimentos'] if 'ipc_alimentos' in ipc_raw.columns else [])
ipc = (ipc_raw[_ipc_cols]
       .dropna(subset=['ipc_general']).sort_values('mes').reset_index(drop=True))
if 'ipc_alimentos' not in ipc.columns:
    ipc['ipc_alimentos'] = np.nan
ipc['ipc_general_var_%']   = ipc['ipc_general'].pct_change(fill_method=None) * 100
ipc['ipc_alimentos_var_%'] = ipc['ipc_alimentos'].pct_change(fill_method=None) * 100

print(f'IPC procesado: {len(ipc)} meses | {ipc["mes"].min()} -> {ipc["mes"].max()}')
print(ipc[['mes','ipc_general','ipc_general_var_%','ipc_alimentos','ipc_alimentos_var_%']].tail(6).to_string(index=False))"""))

# ── CELL 11 — COMPARATIVA (multi-canasta) ─────────────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 11 — Comparativa SEPA vs IPC para cada canasta
# ============================================================
comparativa_dict  = {}   # col_id -> comparativa DataFrame
df_g_dict         = {}   # col_id -> df_g (desde MES_INICIO_GRAFICO)
_lbl_base_dict    = {}   # col_id -> label base ej '03-24'
_serie_vacia_dict = {}   # col_id -> bool

for _col_id in CANASTAS_ACTIVAS:
    _sn   = serie_nac_dict[_col_id]
    _name = CANASTA_NAMES[_col_id]

    if len(_sn) == 0:
        print(f'AVISO [{_name}]: Serie histórica vacía — gráficos de índices no disponibles.')
        _serie_vacia_dict[_col_id] = True
        comparativa_dict[_col_id]  = pd.DataFrame()
        df_g_dict[_col_id]         = pd.DataFrame()
        _lbl_base_dict[_col_id]    = PERIODO[5:7] + '-' + PERIODO[2:4]
        continue

    _comp = _sn.merge(
        ipc[['mes','ipc_general','ipc_general_var_%','ipc_alimentos','ipc_alimentos_var_%']],
        on='mes', how='left')
    _ipc_b = _comp['ipc_general'].dropna().iloc[0] if _comp['ipc_general'].notna().any() else 1
    _ialb  = _comp['ipc_alimentos'].dropna().iloc[0] if _comp['ipc_alimentos'].notna().any() else 1
    _comp['indice_ipc_general_base100']   = (_comp['ipc_general']   / _ipc_b * 100).round(2)
    _comp['indice_ipc_alimentos_base100'] = (_comp['ipc_alimentos'] / _ialb  * 100).round(2)

    _mg = MES_INICIO_GRAFICO
    if _mg not in _comp['mes'].values:
        _mg = _comp['mes'].min()
    _dg = _comp[_comp['mes'] >= _mg].copy().reset_index(drop=True)
    _bg  = _dg['canasta_nacional_ponderada'].iloc[0]
    _bgp = _dg['canasta_nacional_ponderada_prom'].iloc[0]
    _big = _dg['ipc_general'].dropna().iloc[0] if _dg['ipc_general'].notna().any() else 1
    _bia = _dg['ipc_alimentos'].dropna().iloc[0] if _dg['ipc_alimentos'].notna().any() else 1
    _lbl = _mg[5:7] + '-' + _mg[2:4]
    _dg['idx_canasta_base']       = (_dg['canasta_nacional_ponderada']      / _bg  * 100).round(2)
    _dg['idx_canasta_base_prom']  = (_dg['canasta_nacional_ponderada_prom'] / _bgp * 100).round(2)
    _dg['idx_ipc_general_base']   = (_dg['ipc_general']                / _big * 100).round(2)
    _dg['idx_ipc_alimentos_base'] = (_dg['ipc_alimentos']              / _bia * 100).round(2)
    _dg['fecha'] = pd.to_datetime(_dg['mes'] + '-01')

    comparativa_dict[_col_id]  = _comp
    df_g_dict[_col_id]         = _dg
    _lbl_base_dict[_col_id]    = _lbl
    _serie_vacia_dict[_col_id] = False

    print(f'  [{_name}] {len(_comp)} meses | desde {_mg} ({len(_dg)} pts) | '
          f'último: ${_dg["canasta_nacional_ponderada"].iloc[-1]:,.0f} '
          f'({_dg["variacion_mensual_%"].iloc[-1]:+.1f}%)')"""))

# ── CELL 12 — CHARTS (multi-canasta) ──────────────────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 12 — Gráficos: índices y variaciones (MEDIANA y PROMEDIO, por duplicado)
# ============================================================
# Cada gráfico se genera DOS veces: análisis MEDIANA (archivos sin sufijo) y
# análisis PROMEDIO (sufijo _prom, precios con outliers fuera).
COLOR_IPC_GEN = '#D62728'
COLOR_IPC_ALI = '#FF7F0E'

_MESES_ES = {1:'ene',2:'feb',3:'mar',4:'abr',5:'may',6:'jun',
             7:'jul',8:'ago',9:'sep',10:'oct',11:'nov',12:'dic'}
def _fmt_mes_es(x, pos):
    try:
        ts = mdates.num2date(x)
        return f'{_MESES_ES[ts.month]}-{str(ts.year)[2:]}'
    except Exception:
        return ''

_activas_con_datos = [c for c in CANASTAS_ACTIVAS if not _serie_vacia_dict[c] and len(df_g_dict[c]) > 0]
if not _activas_con_datos:
    print('AVISO: Sin serie histórica para ninguna canasta. Saltando gráficos de índices.')
else:
    _lbl_base = _lbl_base_dict[_activas_con_datos[0]]
    _dg0 = df_g_dict[_activas_con_datos[0]]
    # (sufijo_archivo, titulo, col_indice_canasta, col_variacion, dict_prom_nacional)
    _MEDIDAS_G = [
        ('',      'mediana',  'idx_canasta_base',      'variacion_mensual_%',      prom_nac_dict),
        ('_prom', 'promedio', 'idx_canasta_base_prom', 'variacion_mensual_prom_%', prom_nac_prom_dict),
    ]
    for _SFX, _TIT, _IDXCOL, _VARCOL, _PROMD in _MEDIDAS_G:

        # ── GRAFICO 1: Índices base ─────────────────────────────────────────────
        fig1, ax1 = plt.subplots(figsize=(13, 6))
        for _col_id in _activas_con_datos:
            _dg   = df_g_dict[_col_id]
            _name = CANASTA_NAMES[_col_id]
            ax1.plot(_dg['fecha'], _dg[_IDXCOL],
                     color=CANASTA_COLORS[_col_id], linewidth=2.5,
                     linestyle=CANASTA_LINESTYLES[_col_id],
                     marker=CANASTA_MARKERS[_col_id], markersize=5,
                     label=f'ICR {_name}')
            _ult = _dg.iloc[-1]
            ax1.annotate(f"{_ult[_IDXCOL]:.0f}",
                         xy=(_ult['fecha'], _ult[_IDXCOL]),
                         xytext=(8, 0), textcoords='offset points',
                         color=CANASTA_COLORS[_col_id], fontweight='bold', fontsize=9)
        if _dg0['idx_ipc_general_base'].notna().any():
            ax1.plot(_dg0['fecha'], _dg0['idx_ipc_general_base'],
                     color=COLOR_IPC_GEN, linewidth=1.8, linestyle='--', marker='s', markersize=4,
                     label='IPC INDEC - Nivel general')
            _iu = _dg0.dropna(subset=['idx_ipc_general_base']).iloc[-1]
            ax1.annotate(f"{_iu['idx_ipc_general_base']:.0f}",
                         xy=(_iu['fecha'], _iu['idx_ipc_general_base']),
                         xytext=(8,-3), textcoords='offset points',
                         color=COLOR_IPC_GEN, fontweight='bold', fontsize=9)
        if _dg0['idx_ipc_alimentos_base'].notna().any():
            ax1.plot(_dg0['fecha'], _dg0['idx_ipc_alimentos_base'],
                     color=COLOR_IPC_ALI, linewidth=1.8, linestyle=':', marker='^', markersize=4,
                     label='IPC INDEC - Alimentos y bebidas')
            _ia = _dg0.dropna(subset=['idx_ipc_alimentos_base']).iloc[-1]
            ax1.annotate(f"{_ia['idx_ipc_alimentos_base']:.0f}",
                         xy=(_ia['fecha'], _ia['idx_ipc_alimentos_base']),
                         xytext=(8,3), textcoords='offset points',
                         color=COLOR_IPC_ALI, fontweight='bold', fontsize=9)
        ax1.set_ylabel(f'Índice ({_lbl_base} = 100)', fontsize=11)
        ax1.legend(loc='upper left', fontsize=9, framealpha=0.95)
        ax1.grid(True, alpha=0.3)
        ax1.xaxis.set_major_locator(mdates.MonthLocator(interval=1))
        ax1.xaxis.set_major_formatter(mticker.FuncFormatter(_fmt_mes_es))
        plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45, ha='right')
        _t1 = f'Índices de canastas vs IPC — análisis {_TIT}'
        if MES_PARCIAL:
            _t1 += f'  ·  {NOMBRE_MES_TITLE} PARCIAL {DIAS_CARGADOS}/{DIAS_MES} días (preliminar)'
        ax1.set_title(_t1, fontsize=10, color=('#C00000' if MES_PARCIAL else '#333'),
                      style=('italic' if MES_PARCIAL else 'normal'), pad=8)
        plt.tight_layout()
        out1 = OUTPUT_DIR / f'indices_canasta_vs_ipc_{MES}{_SFX}.png'
        plt.savefig(out1, dpi=600, bbox_inches='tight', facecolor='white')
        plt.show()
        print(f'Gráfico 1 [{_TIT}] guardado: {out1}')

        # ── GRAFICO 2: Variaciones mensuales (barras agrupadas) ─────────────────
        _series_bar = (
            [(CANASTA_COLORS[c], f'ICR {CANASTA_NAMES[c]}',
              df_g_dict[c][_VARCOL]) for c in _activas_con_datos] +
            [(COLOR_IPC_GEN, 'IPC INDEC - Nivel general', _dg0['ipc_general_var_%']),
             (COLOR_IPC_ALI, 'IPC INDEC - Alimentos y bebidas', _dg0['ipc_alimentos_var_%'])]
        )
        _n_b    = len(_series_bar)
        _fig_w  = max(20, _n_b * 2 + 10)
        _bw2    = pd.Timedelta(days=max(3, int(22 / max(_n_b, 1))))
        _offs2  = [(_i - (_n_b - 1) / 2) * _bw2 for _i in range(_n_b)]
        _tick_i = 2 if _n_b > 5 else 1
        fig2, ax2 = plt.subplots(figsize=(_fig_w, 8))
        for _i, (_col, _lbl, _vals) in enumerate(_series_bar):
            if _vals.notna().any():
                _alpha = 0.88 if _i < len(_activas_con_datos) else 0.72
                ax2.bar(_dg0['fecha'] + _offs2[_i], _vals,
                        width=_bw2, color=_col, alpha=_alpha, label=_lbl, edgecolor='none')
        ax2.axhline(0, color='#444444', linewidth=0.8)
        ax2.set_ylabel('Variación mensual (%)', fontsize=11)
        ax2.legend(loc='upper right', fontsize=9, framealpha=0.95,
                   ncol=max(1, _n_b // 4 + 1))
        ax2.grid(True, alpha=0.2, axis='y', linewidth=0.7)
        ax2.xaxis.set_major_locator(mdates.MonthLocator(interval=_tick_i))
        ax2.xaxis.set_major_formatter(mticker.FuncFormatter(_fmt_mes_es))
        plt.setp(ax2.xaxis.get_majorticklabels(), rotation=45, ha='right', fontsize=9)
        for sp in ['top', 'right']: ax2.spines[sp].set_visible(False)
        ax2.spines['bottom'].set_color('#cccccc')
        ax2.spines['left'].set_color('#cccccc')
        _vmax_l2 = [s.dropna().max() for _,_,s in _series_bar if s.notna().any()]
        if _vmax_l2: ax2.set_ylim(top=max(_vmax_l2) * 1.35)
        _t2 = f'Variación mensual — análisis {_TIT}'
        if MES_PARCIAL:
            _t2 += f'  ·  {NOMBRE_MES_TITLE} PARCIAL {DIAS_CARGADOS}/{DIAS_MES} días (preliminar)'
        ax2.set_title(_t2, fontsize=10, color=('#C00000' if MES_PARCIAL else '#333'),
                      style=('italic' if MES_PARCIAL else 'normal'), pad=8)
        plt.tight_layout()
        out2 = OUTPUT_DIR / f'variaciones_canasta_vs_ipc_{MES}{_SFX}.png'
        plt.savefig(out2, dpi=600, bbox_inches='tight', facecolor='white')
        plt.show()
        print(f'Gráfico 2 [{_TIT}] guardado: {out2}')

        # ── GRAFICO 3: Ranking de valores absolutos por canasta ─────────────────
        _abs_data = sorted(
            [(CANASTA_NAMES[c], _PROMD[c], CANASTA_COLORS[c]) for c in CANASTAS_ACTIVAS],
            key=lambda x: x[1])
        _pnames  = [d[0] for d in _abs_data]
        _pvals   = [d[1] for d in _abs_data]
        _pcolors = [d[2] for d in _abs_data]
        _base_v  = _pvals[0] if _pvals else 1
        _n_bars  = len(_pvals)
        fig3, ax3 = plt.subplots(figsize=(12, max(5, _n_bars * 1.6 + 2)))
        bars3 = ax3.barh(_pnames, _pvals, color=_pcolors,
                         edgecolor='none', height=0.55, zorder=2)
        ax3.barh(_pnames, _pvals, color='black', alpha=0.06,
                 height=0.60, zorder=1)
        ax3.axvline(_base_v, color='#aaaaaa', linewidth=1.2, linestyle='--', zorder=3)
        for bar, val, name in zip(bars3, _pvals, _pnames):
            _ratio = val / _base_v if _base_v > 0 else 1
            _ratio_str = f'  ×{_ratio:.1f}' if _ratio > 1.05 else '  base'
            ax3.text(val + max(_pvals) * 0.008,
                     bar.get_y() + bar.get_height() / 2,
                     f'${int(val):,}'.replace(',', '.') + _ratio_str,
                     va='center', fontsize=11, fontweight='bold',
                     color='#2c3e50')
        ax3.set_yticklabels(_pnames, fontsize=12, fontweight='bold')
        ax3.set_xlabel(f'Costo mensual nacional (ARS) — análisis {_TIT}', fontsize=11, color='#444')
        ax3.xaxis.set_major_formatter(mticker.FuncFormatter(
            lambda x, _: f'${int(x):,}'.replace(',', '.')))
        ax3.set_xlim(0, max(_pvals) * 1.22)
        ax3.tick_params(axis='x', labelsize=10, colors='#555')
        ax3.grid(True, alpha=0.25, axis='x', zorder=0); ax3.set_axisbelow(True)
        for sp in ['top', 'right', 'left']: ax3.spines[sp].set_visible(False)
        ax3.spines['bottom'].set_color('#cccccc')
        fig3.text(0.98, 0.01, f'{NOMBRE_MES_TITLE}{SUFIJO_PARCIAL}',
                  ha='right', va='bottom', fontsize=9, color=('#C00000' if MES_PARCIAL else '#999'))
        plt.tight_layout()
        out3 = OUTPUT_DIR / f'ranking_canastas_{MES}{_SFX}.png'
        plt.savefig(out3, dpi=600, bbox_inches='tight', facecolor='white')
        plt.show()
        print(f'Gráfico 3 [{_TIT}] guardado: {out3}')"""))

# ── CELL 13 — CUADRO 1 + LaTeX (per-canasta) ──────────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 13 — Cuadro 1 provincial + LaTeX (MEDIANA y PROMEDIO, por duplicado)
# ============================================================
def fmt_ar(x, dec=0):
    s = f'{x:,.{dec}f}'
    return s.replace(',','X').replace('.',',').replace('X','.')

nom_mes = {'01':'enero','02':'febrero','03':'marzo','04':'abril','05':'mayo','06':'junio',
           '07':'julio','08':'agosto','09':'septiembre','10':'octubre','11':'noviembre','12':'diciembre'}
_mes_s  = nom_mes[ULTIMO_MES[5:7]]
_anio_s = ULTIMO_MES[:4]

# (sufijo_archivo, titulo, col_valor, dict_prom_nacional)
_MEDIDAS13 = [
    ('',      'mediana',  'canasta_total',      prom_nac_dict),
    ('_prom', 'promedio', 'canasta_total_prom', prom_nac_prom_dict),
]
for _col_id in CANASTAS_ACTIVAS:
    _name  = CANASTA_NAMES[_col_id]
    _short = CANASTA_SHORT[_col_id]
    for _SFX, _TIT, _VCOL, _PROMD in _MEDIDAS13:
        _spv   = serie_prov_dict[_col_id].copy()
        _prom  = _PROMD[_col_id]
        _spv['vs_%'] = ((_spv[_VCOL] / _prom) - 1) * 100
        _spv = _spv.sort_values(_VCOL).reset_index(drop=True)

        print(f'\\n=== CUADRO 1: {_name.upper()} [{_TIT}] — {NOMBRE_MES_TITLE} ===\\n')
        print(f'{"Provincia":<25} {"Canasta":>14} {"Vs. promedio":>14}')
        print('-'*55)
        for _, r in _spv.iterrows():
            _c = fmt_ar(r[_VCOL])
            _v = f"{r['vs_%']:+.2f}%".replace('.',',')
            print(f"{r['provincia']:<25} {_c:>14} {_v:>14}")
        print('-'*55)
        print(f'{"Promedio nacional":<25} {fmt_ar(_prom):>14} {"0,00%":>14}')

        ltx = [
            r'\\begin{table}[H]',
            r'\\centering',
            r'\\renewcommand{\\arraystretch}{1.15}',
            f'\\\\caption{{Canasta {_name} por provincia ({_mes_s} {_anio_s}) --- analisis {_TIT}}}',
            r'\\begin{tabular}{@{}l r r@{}}',
            r'\\toprule',
            f'\\\\textbf{{Provincia}} & \\\\textbf{{Canasta}} & \\\\shortstack{{\\\\textbf{{Vs. promedio}}\\\\\\\\\\\\textbf{{pais (\\\\%)}}}} \\\\\\\\',
            r'\\midrule',
        ]
        for _, r in _spv.iterrows():
            _c = fmt_ar(r[_VCOL])
            _v = f"{r['vs_%']:+.2f}".replace('.',',')
            ltx.append(f"{r['provincia']:<22} & {_c} & {_v}\\\\% \\\\\\\\")
        ltx += [
            r'\\midrule',
            f'\\\\textbf{{Promedio}} & {fmt_ar(_prom)} & 0,00\\\\% \\\\\\\\',
            r'\\bottomrule',
            r'\\end{tabular}\\\\[0.2cm]',
            r'\\caption*{Fuente: Elaboracion propia en base a SEPA}',
            f'\\\\label{{tab:canasta_{_short}_{ULTIMO_MES}{_SFX}}}',
            r'\\end{table}',
        ]
        _latex_out = '\\n'.join(ltx)
        _out_tex = OUTPUT_DIR / f'tabla_canasta_{_short}_{ULTIMO_MES}{_SFX}.tex'
        _out_tex.write_text(_latex_out, encoding='utf-8')
        print(f'  LaTeX [{_TIT}] guardado: {_out_tex.name}')"""))

# ── CELL 14 — CHOROPLETH MAPS (one per canasta) ────────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 14 — Mapa coroplético por canasta (MEDIANA y PROMEDIO, por duplicado)
# ============================================================
if not GEOJSON_PATH.exists():
    print(f'GeoJSON no encontrado en {GEOJSON_PATH} — saltear celda')
else:
    with open(GEOJSON_PATH, 'r', encoding='utf-8') as f:
        geo = _json.load(f)

    NORM_GEO = {'Ciudad de Buenos Aires': 'CABA'}
    AJUST = {
        'Salta':(0,-1),'Tucuman':(0.3,0),'Tucumán':(0.3,0),'Chaco':(0,-1),
        'Tierra del Fuego':(-1,-0.2),'Santa Fe':(0,1),'Santiago del Estero':(0.7,0),
    }

    def centroide(coords):
        xs,ys = [],[]
        if isinstance(coords[0][0][0],(int,float)):
            for p in coords[0]: xs.append(p[0]); ys.append(p[1])
        else:
            poly = max(coords, key=lambda p: len(p[0]))
            for p in poly[0]: xs.append(p[0]); ys.append(p[1])
        return sum(xs)/len(xs), sum(ys)/len(ys)

    def draw(ax, coords, color):
        if isinstance(coords[0][0][0],(int,float)):
            ax.fill([c[0] for c in coords[0]], [c[1] for c in coords[0]],
                    facecolor=color, edgecolor='white', linewidth=0.6)
        else:
            for poly in coords:
                ax.fill([c[0] for c in poly[0]], [c[1] for c in poly[0]],
                        facecolor=color, edgecolor='white', linewidth=0.6)

    cmap_m = LinearSegmentedColormap.from_list('c',
        ['#1a9850','#66bd63','#a6d96a','#d9ef8b','#fee08b','#fdae61','#f46d43','#d73027'], N=256)

    for _col_id in CANASTAS_ACTIVAS:
        _name   = CANASTA_NAMES[_col_id]
        _short  = CANASTA_SHORT[_col_id]
        _spv    = serie_prov_dict[_col_id]
        if len(_spv) == 0:
            print(f'  [{_name}] Sin datos provinciales — saltear mapa')
            continue
        for _SFX, _TIT, _VCOL in [('','mediana','canasta_total'), ('_prom','promedio','canasta_total_prom')]:
            _can_prov = dict(zip(_spv['provincia'], _spv[_VCOL]))
            _vals = list(_can_prov.values())
            norm_c = Normalize(vmin=min(_vals), vmax=max(_vals))

            fig, ax = plt.subplots(figsize=(12, 16))
            caba_c = None
            for feat in geo['features']:
                ng  = feat['properties']['name']
                nom = NORM_GEO.get(ng, ng)
                val = _can_prov.get(nom)
                col = cmap_m(norm_c(val)) if val is not None else '#dddddd'
                gt  = feat['geometry']['type']
                co  = feat['geometry']['coordinates']
                draw(ax, [co] if gt=='Polygon' else co, col)
                cx, cy = centroide([co] if gt=='Polygon' else co)
                if nom == 'CABA':
                    caba_c = (cx, cy); continue
                dx, dy = AJUST.get(nom, (0,0))
                if val is not None:
                    ax.text(cx+dx, cy+dy, f'{nom}\\n${val/1000:.0f}k',
                            ha='center', va='center', fontsize=7.5, fontweight='bold',
                            bbox=dict(boxstyle='round,pad=0.25', facecolor='white', alpha=0.75, edgecolor='none'))
            if caba_c and 'CABA' in _can_prov:
                vc = _can_prov['CABA']
                cc = cmap_m(norm_c(vc))
                lx, ly = caba_c[0]+2.2, caba_c[1]+0.8
                ax.annotate('', xy=caba_c, xytext=(lx,ly),
                            arrowprops=dict(arrowstyle='-', color='black', linewidth=1.0))
                ax.text(lx, ly, f'CABA\\n${vc/1000:.0f}k',
                        ha='center', va='center', fontsize=9, fontweight='bold',
                        bbox=dict(boxstyle='round,pad=0.5', facecolor=cc, alpha=0.95,
                                  edgecolor='black', linewidth=1.0))
                ax.plot(*caba_c, marker='o', markersize=10, markerfacecolor=cc,
                        markeredgecolor='black', markeredgewidth=1.2, zorder=5)
            ax.set_title(f'Canasta {_name} por provincia — análisis {_TIT}',
                         fontsize=13, fontweight='bold', pad=6)
            ax.set_aspect('equal'); ax.axis('off')
            plt.tight_layout()
            _out_m = OUTPUT_DIR / f'mapa_canasta_{_short}_{ULTIMO_MES}{_SFX}.png'
            plt.savefig(_out_m, dpi=600, bbox_inches='tight', facecolor='white')
            plt.show()
            print(f'  Mapa [{_name}/{_TIT}] guardado: {_out_m.name}')"""))

# ── CELL 15 — COVERAGE ─────────────────────────────────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 15 — Gráficos de cobertura (primera canasta activa)
# ============================================================
_cgf0 = canasta_geo_dict[CANASTAS_ACTIVAS[0]]

_pm_info = (_cgf0[['id_comercio','id_bandera','id_sucursal',
                    'PROVINCIA_NORM','cadena']].drop_duplicates())
_pm_geo  = precio_mes.merge(_pm_info, on=['id_comercio','id_bandera','id_sucursal'], how='inner')
_pm_geo['suc_key'] = (_pm_geo['id_comercio'] + '_' +
                      _pm_geo['id_bandera']   + '_' + _pm_geo['id_sucursal'])

cob_provincia = (_pm_geo.groupby('PROVINCIA_NORM')
    .agg(n_productos_unicos=('ean_norm','nunique'),
         n_cadenas=('cadena','nunique'),
         n_sucursales=('suc_key','nunique'))
    .reset_index().rename(columns={'PROVINCIA_NORM':'provincia'}))

cob_cadena = (_pm_geo.groupby('cadena')
    .agg(n_productos_unicos=('ean_norm','nunique'),
         n_provincias=('PROVINCIA_NORM','nunique'),
         n_sucursales=('suc_key','nunique'))
    .reset_index())

matriz_cad_prov = (_pm_geo.groupby(['cadena','PROVINCIA_NORM'])['ean_norm']
    .nunique().unstack(fill_value=0))

df_p = cob_provincia.sort_values('n_productos_unicos', ascending=True).reset_index(drop=True)
fig, axes = plt.subplots(1, 3, figsize=(15, max(9, len(df_p)*0.38+2)), sharey=True)
for ax, (col, titulo, color) in zip(axes, [
        ('n_productos_unicos','Productos únicos','#0055A4'),
        ('n_cadenas','Cadenas presentes','#27ae60'),
        ('n_sucursales','Sucursales','#c0392b')]):
    ax.barh(df_p['provincia'], df_p[col], color=color, edgecolor='white')
    for i, v in enumerate(df_p[col]):
        ax.text(v+max(df_p[col])*0.01, i, f'{int(v):,}', va='center', fontsize=8, color='#2c3e50')
    ax.set_title(titulo, fontsize=11, fontweight='bold', color=color)
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_: f'{int(x):,}'))
    ax.set_xlim(0, max(df_p[col])*1.15)
    for sp in ['top','right']: ax.spines[sp].set_visible(False)
plt.tight_layout()
out_cp = OUTPUT_DIR / f'cobertura_provincia_{MES}.png'
plt.savefig(out_cp, dpi=600, bbox_inches='tight'); plt.show()

df_c = cob_cadena.sort_values('n_productos_unicos', ascending=True).reset_index(drop=True)
fig, axes = plt.subplots(1, 3, figsize=(15, max(8, len(df_c)*0.5+2)), sharey=True)
for ax, (col, titulo, color) in zip(axes, [
        ('n_productos_unicos','Productos únicos','#0055A4'),
        ('n_provincias','Provincias presentes','#27ae60'),
        ('n_sucursales','Sucursales','#c0392b')]):
    ax.barh(df_c['cadena'], df_c[col], color=color, edgecolor='white')
    for i, v in enumerate(df_c[col]):
        ax.text(v+max(df_c[col])*0.01, i, f'{int(v):,}', va='center', fontsize=8, color='#2c3e50')
    ax.set_title(titulo, fontsize=11, fontweight='bold', color=color)
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_: f'{int(x):,}'))
    ax.set_xlim(0, max(df_c[col])*1.15)
    for sp in ['top','right']: ax.spines[sp].set_visible(False)
axes[0].tick_params(axis='y', labelsize=8)
plt.tight_layout()
out_cc = OUTPUT_DIR / f'cobertura_cadena_{MES}.png'
plt.savefig(out_cc, dpi=600, bbox_inches='tight'); plt.show()

df_m = matriz_cad_prov.copy()
df_m['_total'] = df_m.sum(axis=1)
df_m = df_m.sort_values('_total', ascending=False).drop(columns='_total')
orden_prov = (df_m > 0).sum(axis=0).sort_values(ascending=False).index
df_m = df_m[orden_prov]

fig, ax = plt.subplots(figsize=(13, max(6, len(df_m)*0.55+2)))
binaria = (df_m > 0).astype(int)
sns.heatmap(binaria, cmap='Blues', cbar=False,
            linewidths=0.5, linecolor='white', xticklabels=True, yticklabels=True, ax=ax)
for i in range(binaria.shape[0]):
    for j in range(binaria.shape[1]):
        if binaria.iat[i, j]:
            ax.text(j+0.5, i+0.5, '●', ha='center', va='center', color='white', fontsize=9)
ax.set_xlabel(''); ax.set_ylabel('')
plt.xticks(rotation=45, ha='right', fontsize=9); plt.yticks(rotation=0, fontsize=9)
plt.tight_layout()
out_mp = OUTPUT_DIR / f'matriz_presencia_{MES}.png'
plt.savefig(out_mp, dpi=600, bbox_inches='tight'); plt.show()

fig, ax = plt.subplots(figsize=(13, max(6, len(df_m)*0.55+2)))
data_log = np.log10(df_m.replace(0, np.nan))
sns.heatmap(data_log, cmap='YlOrRd', linewidths=0.5, linecolor='white',
            cbar_kws={'label': 'log10(productos unicos)'},
            xticklabels=True, yticklabels=True, ax=ax)
ax.set_xlabel(''); ax.set_ylabel('')
plt.xticks(rotation=45, ha='right', fontsize=9); plt.yticks(rotation=0, fontsize=9)
plt.tight_layout()
out_mi = OUTPUT_DIR / f'matriz_intensidad_{MES}.png'
plt.savefig(out_mi, dpi=600, bbox_inches='tight'); plt.show()
print('Gráficos cobertura guardados')"""))

# ── CELL 16 — RANKINGS (per-canasta) ──────────────────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 16 — Rankings de cadenas (MEDIANA y PROMEDIO, por duplicado)
# ============================================================
def fmtn(x): return f'{x:,.0f}'.replace(',','.')

# (sufijo, titulo, col_valor, agregador_por_cadena)
_MEDIDAS16 = [('','mediana','canasta_total','median'), ('_prom','promedio','canasta_total_prom',_pmean)]
for _col_id in CANASTAS_ACTIVAS:
    _name  = CANASTA_NAMES[_col_id]
    _short = CANASTA_SHORT[_col_id]
    _cgf   = canasta_geo_dict[_col_id]
    for _SFX, _TIT, _VCOL, _AGG in _MEDIDAS16:

        _rk_nac = (_cgf.groupby('cadena')
                   .agg(n_sucursales=(_VCOL,'count'), valor=(_VCOL,_AGG))
                   .round(0).reset_index())
        _rk_nac = _rk_nac[_rk_nac['n_sucursales'] >= MIN_SUCURSALES_RANKING].sort_values('valor')
        _prom_nac_rk = (_cgf[_VCOL].median() if _SFX == '' else _pmean(_cgf[_VCOL]))

        _amba = _cgf[_cgf['PROVINCIA_NORM'].isin(['Buenos Aires','CABA'])]
        _rk_amba = (_amba.groupby('cadena')
                    .agg(n_sucursales=(_VCOL,'count'), valor=(_VCOL,_AGG))
                    .round(0).reset_index())
        _rk_amba = _rk_amba[_rk_amba['n_sucursales'] >= MIN_SUCURSALES_RANKING].sort_values('valor')
        _prom_amba_rk = ((_amba[_VCOL].median() if _SFX == '' else _pmean(_amba[_VCOL])) if len(_amba) > 0 else 0)

        for (_rk, _prom_r, _titulo, _out_name) in [
            (_rk_nac,  _prom_nac_rk,  f'Ranking nacional [{_name}/{_TIT}]',  f'ranking_cadenas_nacional_{MES}_{_short}{_SFX}'),
            (_rk_amba, _prom_amba_rk, f'Ranking AMBA [{_name}/{_TIT}]',      f'ranking_cadenas_amba_{MES}_{_short}{_SFX}'),
        ]:
            if len(_rk) == 0:
                print(f'  Sin datos para {_titulo}'); continue
            fig, ax = plt.subplots(figsize=(11, max(5, len(_rk)*0.5+2)))
            labs   = [f"{r.cadena}  ({int(r.n_sucursales)})" for r in _rk.itertuples()]
            _n_c   = len(_rk)
            _cols  = plt.cm.RdYlGn_r(np.linspace(0.1, 0.9, _n_c)) if _n_c > 1 else [CANASTA_COLORS[_col_id]]
            bars   = ax.barh(labs, _rk['valor'], color=_cols, edgecolor='black', linewidth=0.4)
            for bar, val in zip(bars, _rk['valor']):
                ax.text(bar.get_width() + _rk['valor'].max()*0.005,
                        bar.get_y()+bar.get_height()/2,
                        f'${fmtn(val)}', va='center', fontsize=9, fontweight='bold')
            ax.axvline(_prom_r, color='#666', linestyle='--', linewidth=1.5,
                       label=f'Nacional: ${fmtn(_prom_r)}')
            ax.set_xlabel(f'Canasta por cadena (ARS) — {_TIT}', fontsize=11)
            ax.set_title(_titulo, fontsize=11, fontweight='bold')
            ax.set_xlim(_rk['valor'].min()*0.95, _rk['valor'].max()*1.07)
            ax.xaxis.set_major_formatter(mticker.FuncFormatter(
                lambda x,_: f'${int(x):,}'.replace(',','.')))
            ax.legend(loc='lower right', fontsize=10)
            ax.grid(True, alpha=0.3, axis='x'); ax.set_axisbelow(True)
            for sp in ['top','right']: ax.spines[sp].set_visible(False)
            plt.tight_layout()
            _out_r = OUTPUT_DIR / f'{_out_name}.png'
            plt.savefig(_out_r, dpi=600, bbox_inches='tight'); plt.show()
            print(f'  Ranking [{_name}/{_TIT}] guardado: {_out_r.name}')

        print(f'\\n  === RANKING NACIONAL [{_name}/{_TIT}] ===')
        for i, r in enumerate(_rk_nac.sort_values('valor', ascending=False).itertuples(), 1):
            print(f'    {i:>2}. {r.cadena:<25} ${fmtn(r.valor):>12}  ({int(r.n_sucursales)} sucs)')"""))

# ── CELL 17 — FOLIUM MAP (único con selector de canasta, lazy popup) ──────────
cells.append(cell_code("""\
# ============================================================
# CELDA 17 — Mapa Folium: lazy popup (datos JSON on-demand)
# Arquitectura: datos almacenados una vez como JSON compacto;
# popup HTML construido por JS al hacer click → archivo ~80% más liviano
# ============================================================
def fmtm(x): return f'{x:,.0f}'.replace(',','.')

# Dos mapas: analisis MEDIANA (sin sufijo) y PROMEDIO (_prom)
for _SFX, _TIT, _VCOL in [('','mediana','canasta_total'), ('_prom','promedio','canasta_total_prom')]:
    # ── Construir datos compactos para popups (almacenados una vez) ──────────────
    _cgf_ref = canasta_geo_dict[CANASTAS_ACTIVAS[0]]
    provs_u  = sorted(_cgf_ref['PROVINCIA_NORM'].unique())

    # Popup compacto: solo totales por sucursal/canasta (sin detalle de productos)
    # → JSON ~300 KB en vez de ~60 MB
    _popup_data = {}
    for _col_id in CANASTAS_ACTIVAS:
        _nc = len(CANASTAS[_col_id])
        for _, _r in canasta_geo_dict[_col_id].iterrows():
            _sk = f"{_r['id_comercio']}_{_r['id_bandera']}_{_r['id_sucursal']}"
            if _sk not in _popup_data:
                _popup_data[_sk] = {
                    'nom': str(_r['sucursales_nombre'])[:40],
                    'bar': str(_r.get('sucursales_barrio') or _r.get('sucursales_localidad') or '')[:30],
                    'prv': _r['PROVINCIA_NORM'],
                    'cad': _r['cadena'],
                    'tip': str(_r.get('sucursales_tipo') or 'N/D'),
                    'can': {}
                }
            _popup_data[_sk]['can'][_col_id] = {
                't': int(_r[_VCOL]),
                'p': int(_r['productos_propios']),
                'n': _nc
            }

    _popup_json = _json.dumps(_popup_data, ensure_ascii=False, separators=(',',':'))
    print(f'Datos popup: {len(_popup_data):,} sucursales | {len(_popup_json)/1024/1024:.1f} MB JSON compacto')

    # ── Mapa Folium ───────────────────────────────────────────────────────────────
    m = folium.Map(location=[-38.0,-63.5], zoom_start=5,
                   tiles='cartodbpositron', control_scale=True)
    folium.map.Marker(
        location=[-51.7963,-59.5236],
        icon=folium.DivIcon(icon_size=(140,28), icon_anchor=(70,14),
            html='<div style="background:rgba(255,255,255,.95);border:1px solid #777;border-radius:3px;padding:3px 7px;font-family:Arial;font-size:11px;font-weight:600;text-align:center;white-space:nowrap;">Islas Malvinas (ARG)</div>')
    ).add_to(m)

    _canasta_fg_ids = {}
    for _col_id in CANASTAS_ACTIVAS:
        _name  = CANASTA_NAMES[_col_id]
        _short = CANASTA_SHORT[_col_id]
        _cgf   = canasta_geo_dict[_col_id]
        _is_default = (_col_id == CANASTAS_ACTIVAS[0])
        _vmin = _cgf[_VCOL].quantile(0.05)
        _vmax = _cgf[_VCOL].quantile(0.95)
        if _vmin == _vmax: _vmin, _vmax = _cgf[_VCOL].min(), _cgf[_VCOL].max()
        _cm = LinearColormap(
            colors=['#1a9850','#66bd63','#a6d96a','#fee08b','#fdae61','#f46d43','#d73027'],
            vmin=_vmin, vmax=_vmax, caption=f'ICR {_name} — {NOMBRE_MES_TITLE} ({_TIT}, ARS)')
        if _is_default: _cm.add_to(m)
        _fg = folium.FeatureGroup(name=_short, show=_is_default)
        _canasta_fg_ids[_col_id] = _fg.get_name()
        for _, _r in _cgf.iterrows():
            val  = _r[_VCOL]
            col  = _cm(max(_vmin, min(_vmax, val)))
            cad  = _r['cadena']
            prv  = _r['PROVINCIA_NORM']
            _sk  = f"{_r['id_comercio']}_{_r['id_bandera']}_{_r['id_sucursal']}"
            cl   = (f'sucursal-marker canasta-{_short}'
                    f' cadena-{cad.replace(" ","_").replace("(","").replace(")","").replace("/","")}'
                    f' prov-{prv.replace(" ","_").replace("(","").replace(")","").replace("/","")}')
            # Popup mínimo: placeholder que JS rellena on-demand al hacer click
            _ph = (f'<div class="lz-pop" data-key="{_sk}" data-can="{_col_id}"'
                   f' style="font-family:Arial;min-width:200px;text-align:center;padding:15px">'
                   f'<span style="color:#aaa;font-size:12px">Cargando detalle...</span></div>')
            folium.CircleMarker(
                location=[_r['sucursales_latitud'], _r['sucursales_longitud']],
                radius=5, color=col, fill=True, fillColor=col, fillOpacity=0.8, weight=1,
                tooltip=f'<b>{cad}</b><br>{prv}<br><b>${fmtm(val)}</b>',
                popup=folium.Popup(_ph, max_width=450), className=cl
            ).add_to(_fg)
        _fg.add_to(m)

    _map_var    = m.get_name()
    _fg_ids_str = '{' + ','.join(f'"{k}":"{v}"' for k,v in _canasta_fg_ids.items()) + '}'
    _names_str  = '{' + ','.join(f'"{k}":"{CANASTA_NAMES[k]}"' for k in CANASTAS_ACTIVAS) + '}'
    _avgs_str   = '{' + ','.join(f'"{k}":{int(canasta_geo_dict[k][_VCOL].mean())}' for k in CANASTAS_ACTIVAS) + '}'

    # Embeber JSON en script tag de tipo application/json (sin escape JS)
    m.get_root().html.add_child(folium.Element(
        f'<script type="application/json" id="_pd_json">{_popup_json}</script>'))

    prov_opts  = ''.join([f'<option value="prov-{p.replace(" ","_")}">{p}</option>' for p in provs_u])
    _can_opts  = ''.join([f'<option value="{k}">{CANASTA_NAMES[k]}</option>' for k in CANASTAS_ACTIVAS])
    _cadenas_u = sorted(_cgf_ref['cadena'].unique())
    _cad_opts  = ''.join([f'<option value="cadena-{c.replace(" ","_").replace("(","").replace(")","").replace("/","")}">{c}</option>' for c in _cadenas_u])

    info_h = (f'<div style="position:fixed;top:10px;left:50px;width:340px;background:white;border:2px solid #0055A4;'
              f'border-radius:8px;padding:12px 15px;font-family:Arial;z-index:9999;box-shadow:0 2px 8px rgba(0,0,0,.15);">'
              f'<div style="color:#0055A4;font-size:15px;font-weight:bold;margin-bottom:5px;">ICR — {NOMBRE_MES_TITLE} ({_TIT})</div>'
              f'<div style="font-size:11px;color:#555;line-height:1.5;">'
              f'<b>{len(_cgf_ref):,}</b> sucursales · <b>{len(CANASTAS_ACTIVAS)}</b> canastas<br>'
              f'Promedio: <span id="info_avg" style="font-weight:bold;"></span></div></div>')
    m.get_root().html.add_child(folium.Element(info_h))

    filtros_h = (
        f'<div id="pf" style="position:fixed;bottom:25px;left:50px;width:280px;background:white;'
        f'border:2px solid #0055A4;border-radius:8px;padding:12px 15px;font-family:Arial;z-index:9999;">'
        f'<div style="color:#0055A4;font-size:13px;font-weight:bold;margin-bottom:8px;">🔍 Filtros</div>'
        f'<label style="font-size:11px;color:#555;display:block;margin-top:4px;">Canasta:'
        f'<select id="fcan" style="width:100%;padding:4px;font-size:11px;margin-top:3px;">{_can_opts}</select></label>'
        f'<label style="font-size:11px;color:#555;display:block;margin-top:6px;">Cadena:'
        f'<select id="fca" style="width:100%;padding:4px;font-size:11px;margin-top:3px;">'
        f'<option value="all">Todas</option>{_cad_opts}</select></label>'
        f'<label style="font-size:11px;color:#555;display:block;margin-top:6px;">Provincia:'
        f'<select id="fp" style="width:100%;padding:4px;font-size:11px;margin-top:3px;">'
        f'<option value="all">Todas</option>{prov_opts}</select></label>'
        f'<button id="fr" style="width:100%;margin-top:10px;padding:6px;background:#f0f0f0;'
        f'border:1px solid #ccc;border-radius:4px;font-size:11px;cursor:pointer;">Restablecer</button></div>'
        f'<style>'
        f'.lz-w{{font-family:Arial;font-size:12px;width:420px;max-height:500px;overflow-y:auto}}'
        f'.lz-h4{{margin:0;color:#0055A4}}'
        f'.lz-nfo{{font-size:11px;color:#555;margin-bottom:5px}}'
        f'.lz-bg{{background:#e6eef7;padding:2px 6px;border-radius:3px;font-size:10px}}'
        f'.lz-cx{{text-align:center;margin:8px 0}}'
        f'.lz-lbl{{font-size:11px;color:#555;margin-bottom:2px}}'
        f'.lz-tot{{color:#0055A4;font-size:22px;font-weight:bold}}'
        f'.lz-sub{{font-size:11px;color:#888;text-align:center;margin-top:3px}}'
        f'.lz-tb{{width:100%;border-collapse:collapse;font-size:10px}}'
        f'.lz-hd{{background:#e6eef7;font-weight:bold}}'
        f'.lz-hd th{{padding:3px 5px;text-align:left}}'
        f'.lz-ch td{{background:#0055A4;color:white;padding:3px 5px;font-weight:bold}}'
        f'.lz-imp{{color:#888;font-style:italic}}'
        f'td{{padding:2px 5px}}'
        f'.lz-sb{{font-weight:600}}'
        f'.lz-ft{{font-size:9px;color:#666;margin-top:4px}}'
        f'.lz-hr{{margin:5px 0}}'
        f'</style>'
        f'<script>'
        f'var _fg_ids={_fg_ids_str};var _names={_names_str};var _avgs={_avgs_str};var _pd=null;'
        f'function _gPD(){{if(!_pd){{var el=document.getElementById("_pd_json");if(el)_pd=JSON.parse(el.textContent);}}return _pd;}}'
        # _bPop usa template literals JS (backtick) + clases CSS → sin single-quote CSS = sin conflicto Python
        f'function _bPop(key,cid){{'
        f'var pd=_gPD();if(!pd||!pd[key]||!pd[key].can[cid])return "<div>Sin datos.</div>";'
        f'var d=pd[key];var c=d.can[cid];var nm=_names[cid];'
        f'var fmt=function(x){{return "$"+Math.round(x).toLocaleString("es-AR");}};'
        f'var cov=Math.round(c.p/c.n*100);'
        f'return `<div class=lz-w>`'
        f'+`<h4 class=lz-h4>${{d.cad}}</h4>`'
        f'+`<div class=lz-nfo><b>${{d.nom}}</b><br>${{d.bar?d.bar+" — ":""}}${{d.prv}}<br><span class=lz-bg>${{d.tip}}</span></div>`'
        f'+"<hr class=lz-hr>"'
        f'+`<div class=lz-cx><div class=lz-lbl>${{nm}}</div><span class=lz-tot>${{fmt(c.t)}}</span></div>`'
        f'+`<div class=lz-sub>${{c.p}}/${{c.n}} productos propios (${{cov}}%)</div>`'
        f'+"</div>";}}'
        f'function _initEvt(){{var mp=window["{_map_var}"];if(!mp)return;'
        f'mp.on("popupopen",function(e){{'
        f'var el=e.popup.getElement().querySelector(".lz-pop");'
        f'if(el&&el.getAttribute("data-built")!=="1"){{'
        f'el.innerHTML=_bPop(el.getAttribute("data-key"),el.getAttribute("data-can"));'
        f'el.setAttribute("data-built","1");e.popup.update();}}}});}}'
        f'function switchCanasta(sel){{var mp=window["{_map_var}"];if(!mp)return;'
        f'Object.keys(_fg_ids).forEach(function(k){{var fg=window[_fg_ids[k]];if(!fg)return;'
        f'if(k===sel){{mp.addLayer(fg);}}else{{mp.removeLayer(fg);}}}});'
        f'var avgEl=document.getElementById("info_avg");'
        f'if(avgEl)avgEl.innerHTML="$"+_avgs[sel].toLocaleString("es-AR")+" ("+_names[sel]+")";apl();}}'
        f'function apl(){{var p=document.getElementById("fp").value;var ca=document.getElementById("fca").value;'
        f'document.querySelectorAll(".sucursal-marker").forEach(function(el){{'
        f'var c=el.className.baseVal||el.className||"";'
        f'var mp=(p==="all")||c.indexOf(p)>=0;'
        f'var mc=(ca==="all")||c.indexOf(ca)>=0;'
        f'el.style.display=(mp&&mc)?"":"none";}});}}'
        f'setTimeout(function(){{'
        f'var fc=document.getElementById("fcan"),sp=document.getElementById("fp"),fca=document.getElementById("fca"),btn=document.getElementById("fr");'
        f'var def=Object.keys(_fg_ids)[0];'
        f'_initEvt();switchCanasta(def);'
        f'if(fc)fc.addEventListener("change",function(){{switchCanasta(this.value);}});'
        f'if(sp)sp.addEventListener("change",apl);'
        f'if(fca)fca.addEventListener("change",apl);'
        f'if(btn)btn.addEventListener("click",function(){{'
        f'if(fc){{fc.value=Object.keys(_fg_ids)[0];switchCanasta(fc.value);}}'
        f'if(sp)sp.value="all";if(fca)fca.value="all";'
        f'document.querySelectorAll(".sucursal-marker").forEach(e=>e.style.display="");}});'
        f'}},1200);</script>'
    )
    m.get_root().html.add_child(folium.Element(filtros_h))

    out_map = OUTPUT_DIR / f'mapa_interactivo_{MES}{_SFX}.html'
    m.save(str(out_map))
    print(f'Mapa [{_TIT}] guardado: {out_map.name} ({len(CANASTAS_ACTIVAS)} canastas · {len(_cgf_ref):,} sucs)')"""))

# ── CELL 18 — CABA RANKINGS (per-canasta) ─────────────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 18 — Ranking de barrios CABA para cada canasta
# ============================================================
BARRIOS_BBOX = {
    'Agronomia':           (-34.604,-34.587,-58.498,-58.476),
    'Almagro':             (-34.622,-34.598,-58.435,-58.405),
    'Balvanera':           (-34.617,-34.598,-58.418,-58.388),
    'Barracas':            (-34.661,-34.628,-58.395,-58.366),
    'Belgrano':            (-34.575,-34.547,-58.471,-58.434),
    'Boedo':               (-34.638,-34.620,-58.426,-58.408),
    'Caballito':           (-34.628,-34.602,-58.460,-58.421),
    'Chacarita':           (-34.595,-34.575,-58.470,-58.443),
    'Coghlan':             (-34.572,-34.555,-58.485,-58.469),
    'Colegiales':          (-34.580,-34.563,-58.460,-58.439),
    'Constitucion':        (-34.631,-34.620,-58.395,-58.378),
    'Flores':              (-34.642,-34.615,-58.480,-58.435),
    'Floresta':            (-34.633,-34.615,-58.500,-58.479),
    'La Boca':             (-34.643,-34.620,-58.371,-58.350),
    'La Paternal':         (-34.605,-34.585,-58.475,-58.456),
    'Liniers':             (-34.652,-34.628,-58.534,-58.506),
    'Mataderos':           (-34.665,-34.641,-58.522,-58.488),
    'Monte Castro':        (-34.628,-34.610,-58.520,-58.500),
    'Monserrat':           (-34.625,-34.605,-58.391,-58.371),
    'Nueva Pompeya':       (-34.658,-34.638,-58.418,-58.396),
    'Nunez':               (-34.553,-34.532,-58.475,-58.443),
    'Palermo':             (-34.595,-34.560,-58.435,-58.398),
    'Parque Avellaneda':   (-34.660,-34.638,-58.495,-58.470),
    'Parque Chacabuco':    (-34.645,-34.625,-58.448,-58.422),
    'Parque Chas':         (-34.591,-34.578,-58.487,-58.475),
    'Parque Patricios':    (-34.652,-34.628,-58.418,-58.395),
    'Puerto Madero':       (-34.625,-34.587,-58.371,-58.349),
    'Recoleta':            (-34.598,-34.575,-58.405,-58.378),
    'Retiro':              (-34.595,-34.578,-58.388,-58.365),
    'Saavedra':            (-34.560,-34.540,-58.495,-58.467),
    'San Cristobal':       (-34.625,-34.612,-58.408,-58.391),
    'San Nicolas':         (-34.610,-34.595,-58.395,-58.371),
    'San Telmo':           (-34.625,-34.610,-58.378,-58.365),
    'Velez Sarsfield':     (-34.642,-34.624,-58.510,-58.493),
    'Versalles':           (-34.640,-34.621,-58.525,-58.508),
    'Villa Crespo':        (-34.605,-34.585,-58.452,-58.428),
    'Villa del Parque':    (-34.615,-34.595,-58.498,-58.472),
    'Villa Devoto':        (-34.612,-34.585,-58.518,-58.490),
    'Villa General Mitre': (-34.615,-34.600,-58.475,-58.458),
    'Villa Lugano':        (-34.690,-34.660,-58.475,-58.435),
    'Villa Luro':          (-34.645,-34.628,-58.510,-58.491),
    'Villa Ortuzar':       (-34.590,-34.575,-58.475,-58.456),
    'Villa Pueyrredon':    (-34.585,-34.565,-58.510,-58.485),
    'Villa Real':          (-34.628,-34.615,-58.530,-58.512),
    'Villa Riachuelo':     (-34.695,-34.680,-58.470,-58.450),
    'Villa Santa Rita':    (-34.622,-34.605,-58.488,-58.470),
    'Villa Soldati':       (-34.682,-34.655,-58.460,-58.420),
    'Villa Urquiza':       (-34.590,-34.565,-58.495,-58.470),
}

def det_barrio(lat, lon):
    if pd.isna(lat) or pd.isna(lon): return 'Sin clasificar'
    for b,(lmin,lmax,lonmin,lonmax) in BARRIOS_BBOX.items():
        if lmin <= lat <= lmax and lonmin <= lon <= lonmax: return b
    return 'Sin clasificar'

def fmtb(x): return f'${x:,.0f}'.replace(',','.')

# Dos rankings por canasta: analisis MEDIANA y PROMEDIO (outliers fuera).
_MEDIDAS18 = [('mediana','canasta_total','median'), ('promedio','canasta_total_prom',_pmean)]
for _col_id in CANASTAS_ACTIVAS:
    _name  = CANASTA_NAMES[_col_id]
    _cgf   = canasta_geo_dict[_col_id]
    _caba  = _cgf[_cgf['PROVINCIA_NORM'] == 'CABA'].copy()
    if len(_caba) == 0:
        print(f'  [{_name}] Sin sucursales en CABA'); continue
    _caba['barrio'] = _caba.apply(
        lambda r: det_barrio(r['sucursales_latitud'], r['sucursales_longitud']), axis=1)
    for _TIT, _VCOL, _AGG in _MEDIDAS18:
        _rk_b = (_caba[_caba['barrio'] != 'Sin clasificar']
                 .groupby('barrio')
                 .agg(n_sucs=(_VCOL,'count'), valor=(_VCOL,_AGG))
                 .round(0).sort_values('valor'))
        _rk_b_fil = _rk_b[_rk_b['n_sucs'] >= 2]
        _pc = (_caba[_VCOL].median() if _TIT == 'mediana' else _pmean(_caba[_VCOL]))
        _pp = (_cgf[_VCOL].median()  if _TIT == 'mediana' else _pmean(_cgf[_VCOL]))

        print(f'\\n{"="*78}')
        print(f'  RANKING BARRIOS CABA [{_name.upper()}/{_TIT}] — {NOMBRE_MES_TITLE}')
        print(f'{"="*78}')
        print(f'  {"#":<3} {"Barrio":<22} {"Sucs.":<7} {"Valor":<13} {"vs CABA":<10} {"vs Pais"}')
        print(f'  {"-"*70}')
        for i, (b, r) in enumerate(_rk_b_fil.iterrows(), 1):
            vc = (r["valor"]/max(_pc,1)-1)*100
            vp = (r["valor"]/max(_pp,1)-1)*100
            print(f'  {i:<3} {b:<22} {int(r["n_sucs"]):<7} {fmtb(r["valor"]):<13} {vc:+.2f}%  {vp:+.2f}%')
        print(f'  {"-"*70}')
        print(f'  [{_TIT}] Promedio CABA: {fmtb(_pc)}   Promedio pais: {fmtb(_pp)}')
        _sin = set(BARRIOS_BBOX.keys()) - set(_rk_b.index)
        if _sin:
            print(f'  Sin sucursales ({len(_sin)}): {", ".join(sorted(_sin))}')"""))

# ── CELL 19 — EXCEL EXPORT (multi-canasta) ────────────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 19 — Exportación Excel multi-canasta (MEDIANA y PROMEDIO, por duplicado)
# ============================================================
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HDR_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HDR_FONT = Font(bold=True, color='FFFFFF', size=10)
HDR_ALIG = Alignment(horizontal='center', wrap_text=True, vertical='center')

def fmt_ws(ws):
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'
    for cell in ws[1]:
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = HDR_ALIG

def auto_widths(ws):
    for ci in range(1, ws.max_column+1):
        cl  = get_column_letter(ci)
        hdr = str(ws.cell(1,ci).value or '').lower()
        w   = 28 if any(x in hdr for x in ('nombre','provincia','cadena','barrio','sucursal','localidad')) else (42 if 'desc' in hdr else 14)
        ws.column_dimensions[cl].width = w
        if any(x in hdr for x in ('canasta','precio','ipc','costo','promedio','mediana','prom')):
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=ci, max_col=ci):
                for cell in row: cell.number_format = '#,##0.00'
        elif '%' in hdr:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=ci, max_col=ci):
                for cell in row: cell.number_format = '+0.00"%"'

out_xls = OUTPUT_DIR / f'canasta_analisis_{ULTIMO_MES}.xlsx'
with pd.ExcelWriter(out_xls, engine='openpyxl') as writer:

    # ── Hoja Evolucion_IPC: todas las canastas + IPC (mediana y promedio) ──────
    _ipc_base = ipc[['mes','ipc_general','ipc_general_var_%',
                      'ipc_alimentos','ipc_alimentos_var_%']].copy()
    _evo = _ipc_base
    for _col_id in CANASTAS_ACTIVAS:
        if len(serie_nac_dict[_col_id]) == 0: continue
        _sn = serie_nac_dict[_col_id][['mes','canasta_nacional_ponderada','variacion_mensual_%',
                                       'canasta_nacional_ponderada_prom','variacion_mensual_prom_%']].copy()
        _n  = CANASTA_SHORT[_col_id]
        _sn = _sn.rename(columns={
            'canasta_nacional_ponderada':      f'canasta_{_n}',
            'variacion_mensual_%':             f'var_{_n}_%',
            'canasta_nacional_ponderada_prom': f'canasta_{_n}_prom',
            'variacion_mensual_prom_%':        f'var_{_n}_prom_%'
        })
        _evo = _evo.merge(_sn, on='mes', how='outer')
    _evo = _evo.sort_values('mes').reset_index(drop=True)
    _evo.to_excel(writer, sheet_name='Evolucion_IPC', index=False)

    # ── Hoja por canasta: Provincias, Ranking y Sucursales ──────────────────────
    for _col_id in CANASTAS_ACTIVAS:
        _name  = CANASTA_NAMES[_col_id]
        _short = CANASTA_SHORT[_col_id]
        _spv   = serie_prov_dict[_col_id].copy()
        _prom      = prom_nac_dict[_col_id]
        _prom_prom = prom_nac_prom_dict[_col_id]
        _spv['vs_promedio_%']      = ((_spv['canasta_total'] / _prom) - 1) * 100
        _spv['vs_promedio_prom_%'] = ((_spv['canasta_total_prom'] / _prom_prom) - 1) * 100
        _spv = _spv[['mes','provincia','canasta_total','vs_promedio_%',
                     'canasta_total_prom','vs_promedio_prom_%']]
        _spv.sort_values('canasta_total').to_excel(
            writer, sheet_name=f'Prov_{_short}', index=False)

        # Ranking por cadena: MEDIANA (mediana por cadena) y PROMEDIO (_pmean por cadena)
        _cgf = canasta_geo_dict[_col_id]
        _rk  = (_cgf.groupby('cadena')
                .agg(n_sucursales=('canasta_total','count'),
                     canasta_mediana=('canasta_total','median'),
                     canasta_prom=('canasta_total_prom', _pmean))
                .reset_index())
        _rk['canasta_mediana'] = _rk['canasta_mediana'].round(0)
        _rk['canasta_prom']    = _rk['canasta_prom'].round(0)
        _rk = _rk.sort_values('canasta_mediana', ascending=False)
        _rk['vs_mediana_%'] = ((_rk['canasta_mediana'] / _cgf['canasta_total'].median()) - 1) * 100
        _rk['vs_prom_%']    = ((_rk['canasta_prom'] / _pmean(_cgf['canasta_total_prom'])) - 1) * 100
        _rk = _rk[['cadena','n_sucursales','canasta_mediana','vs_mediana_%','canasta_prom','vs_prom_%']]
        _rk.to_excel(writer, sheet_name=f'Ranking_{_short}', index=False)

        _suc_exp = _cgf[[
            'id_comercio','id_bandera','id_sucursal','cadena','PROVINCIA_NORM',
            'sucursales_nombre','sucursales_localidad','sucursales_barrio',
            'sucursales_latitud','sucursales_longitud','sucursales_tipo',
            'canasta_total','canasta_total_prom','productos_propios'
        ]].sort_values(['PROVINCIA_NORM','cadena','canasta_total']).copy()
        _suc_exp.to_excel(writer, sheet_name=f'Sucs_{_short}', index=False)

    # ── Hoja Serie_precios: precio mediano y promedio por canasta x mes x producto ─
    _sp_rows = []
    for _col_id, _canasta in CANASTAS.items():
        _eans = set(_canasta.keys())
        _dh   = df_hist_raw[df_hist_raw['ean_norm'].isin(_eans)].copy()
        _dh['id_producto']  = _dh['ean_norm'].apply(lambda e: e.zfill(13))
        _dh['descripcion']  = _dh['ean_norm'].map(lambda e, c=_canasta: c.get(e,('',0,''))[0])
        _dh['categoria']    = _dh['ean_norm'].map(lambda e, c=_canasta: c.get(e,('',0,'?'))[2])
        _dh['qty']              = _dh['ean_norm'].map(lambda e, c=_canasta: c.get(e,('',0,''))[1])
        _dh['costo_item']       = _dh['precio_mediano'] * _dh['qty']
        _dh['costo_item_prom']  = _dh['precio_medio']   * _dh['qty']
        _dh['canasta_id']   = _col_id
        _dh['canasta_name'] = CANASTA_NAMES[_col_id]
        _sp_rows.append(_dh[['canasta_id','canasta_name','anio_mes',
                              'id_producto','descripcion','categoria',
                              'qty','precio_mediano','costo_item',
                              'precio_medio','costo_item_prom']])
    if _sp_rows:
        _sp_all = pd.concat(_sp_rows, ignore_index=True).sort_values(
            ['canasta_id','id_producto','anio_mes']).reset_index(drop=True)
        _sp_all = _sp_all.rename(columns={'anio_mes':'mes','precio_medio':'precio_prom'})
        _sp_all.to_excel(writer, sheet_name='Serie_precios', index=False)

    # ── Hojas de detalle geografico (canasta MEDIA) — belgrano, Pinamar, costa, CABA ──
    # MISMA fuente que Sucs_Media. Cada hoja trae canasta_mediana y canasta_prom
    # (promedio con outliers fuera), geolocalizando por lat/lon.
    _MID = 'cantidad_03'
    if _MID in canasta_geo_dict:
        _cgm    = canasta_geo_dict[_MID]
        _pais_m = prom_nac_dict[_MID]        # cuadro pais — mediana
        _pais_p = prom_nac_prom_dict[_MID]   # cuadro pais — promedio

        _cgm_caba = _cgm[_cgm['PROVINCIA_NORM'] == 'CABA'].copy()
        _cgm_caba['barrio'] = _cgm_caba.apply(
            lambda r: det_barrio(r['sucursales_latitud'], r['sucursales_longitud']), axis=1)

        # Hoja 'belgrano'
        _bel = _cgm_caba[_cgm_caba['barrio'] == 'Belgrano']
        if len(_bel):
            _bel_g = (_bel.groupby('cadena')
                      .agg(n_sucursales=('canasta_total','count'),
                           canasta_mediana=('canasta_total','median'),
                           canasta_prom=('canasta_total_prom', _pmean))
                      .reset_index().sort_values('canasta_mediana'))
            _bel_g['canasta_mediana'] = _bel_g['canasta_mediana'].round(0)
            _bel_g['canasta_prom']    = _bel_g['canasta_prom'].round(0)
            _bel_g = pd.concat([_bel_g, pd.DataFrame([{
                'cadena': 'Total Belgrano',
                'n_sucursales': int(_bel_g['n_sucursales'].sum()),
                'canasta_mediana': round(_bel['canasta_total'].median(), 0),
                'canasta_prom': round(_pmean(_bel['canasta_total_prom']), 0)}])],
                ignore_index=True)
            _bel_g.to_excel(writer, sheet_name='belgrano', index=False)

        # Hoja 'CABA'
        _caba_rk = (_cgm_caba[_cgm_caba['barrio'] != 'Sin clasificar']
                    .groupby('barrio')
                    .agg(n_sucursales=('canasta_total','count'),
                         canasta_mediana=('canasta_total','median'),
                         canasta_prom=('canasta_total_prom', _pmean))
                    .reset_index())
        _caba_rk = _caba_rk[_caba_rk['n_sucursales'] >= 2].sort_values(
            'canasta_mediana', ascending=False)
        _caba_rk['canasta_mediana'] = _caba_rk['canasta_mediana'].round(0)
        _caba_rk['canasta_prom']    = _caba_rk['canasta_prom'].round(0)
        if len(_caba_rk):
            _caba_rk = pd.concat([_caba_rk, pd.DataFrame([
                {'barrio': 'Promedio CABA',
                 'n_sucursales': int(_caba_rk['n_sucursales'].sum()),
                 'canasta_mediana': round(_cgm_caba['canasta_total'].median(), 0),
                 'canasta_prom': round(_pmean(_cgm_caba['canasta_total_prom']), 0)},
                {'barrio': 'Promedio pais', 'n_sucursales': '',
                 'canasta_mediana': round(_pais_m, 0),
                 'canasta_prom': round(_pais_p, 0)}])], ignore_index=True)
            _caba_rk.to_excel(writer, sheet_name='CABA', index=False)

        # ── Localidades de la Costa Atlantica bonaerense por lat/lon ──────────────
        COSTA_BBOX = {
            'San Clemente':  (-36.44,  -36.32,  -56.80, -56.64),
            'Mar del Tuyu':  (-36.62,  -36.50,  -56.76, -56.62),
            'San Bernardo':  (-36.712, -36.66,  -56.74, -56.61),
            'Mar de Ajo':    (-36.76,  -36.713, -56.74, -56.61),
            'Pinamar':       (-37.19,  -37.05,  -56.94, -56.81),
            'Gesell':        (-37.37,  -37.20,  -57.04, -56.91),
            'Madariaga':     (-37.07,  -36.93,  -57.24, -57.05),
            'Mar del Plata': (-38.13,  -37.87,  -57.74, -57.45),
            'Miramar':       (-38.35,  -38.19,  -57.94, -57.75),
            'Necochea':      (-38.63,  -38.47,  -58.84, -58.62),
        }
        def _det_costa(_lat, _lon):
            if pd.isna(_lat) or pd.isna(_lon): return 'Sin clasificar'
            for _loc, (_a0, _a1, _o0, _o1) in COSTA_BBOX.items():
                if _a0 <= _lat <= _a1 and _o0 <= _lon <= _o1: return _loc
            return 'Sin clasificar'

        _cgm_ba = _cgm[_cgm['PROVINCIA_NORM'] == 'Buenos Aires'].copy()
        _cgm_ba['loc_costa'] = _cgm_ba.apply(
            lambda r: _det_costa(r['sucursales_latitud'], r['sucursales_longitud']), axis=1)

        # Hoja 'Pinamar': una fila por sucursal
        _pin = _cgm_ba[_cgm_ba['loc_costa'] == 'Pinamar']
        if len(_pin):
            _pin_out = (_pin[['cadena','sucursales_nombre','canasta_total','canasta_total_prom']]
                        .rename(columns={'sucursales_nombre':'sucursal',
                                         'canasta_total':'canasta_mediana',
                                         'canasta_total_prom':'canasta_prom'})
                        .sort_values('canasta_mediana'))
            _pin_out['canasta_mediana'] = _pin_out['canasta_mediana'].round(0)
            _pin_out['canasta_prom']    = _pin_out['canasta_prom'].round(0)
            _pin_out = pd.concat([_pin_out, pd.DataFrame([{
                'cadena': 'Total Pinamar', 'sucursal': '',
                'canasta_mediana': round(_pin['canasta_total'].median(), 0),
                'canasta_prom': round(_pmean(_pin['canasta_total_prom']), 0)}])],
                ignore_index=True)
            _pin_out.to_excel(writer, sheet_name='Pinamar', index=False)

        # Hoja 'costa': ranking de localidades
        _costa = _cgm_ba[_cgm_ba['loc_costa'] != 'Sin clasificar']
        if len(_costa):
            _costa_g = (_costa.groupby('loc_costa')
                        .agg(n_sucursales=('canasta_total','count'),
                             canasta_mediana=('canasta_total','median'),
                             canasta_prom=('canasta_total_prom', _pmean))
                        .reset_index().rename(columns={'loc_costa':'localidad'}))
            _costa_g['canasta_mediana'] = _costa_g['canasta_mediana'].round(0)
            _costa_g['canasta_prom']    = _costa_g['canasta_prom'].round(0)
            _costa_g['vs_pais_%']      = (_costa_g['canasta_mediana'] / _pais_m - 1) * 100
            _costa_g['vs_pais_prom_%'] = (_costa_g['canasta_prom'] / _pais_p - 1) * 100
            _costa_g = _costa_g.sort_values('canasta_mediana')
            _costa_g = pd.concat([_costa_g, pd.DataFrame([{
                'localidad': 'Total Costa Atlantica',
                'n_sucursales': int(_costa_g['n_sucursales'].sum()),
                'canasta_mediana': round(_costa['canasta_total'].median(), 0),
                'canasta_prom': round(_pmean(_costa['canasta_total_prom']), 0),
                'vs_pais_%': (_costa['canasta_total'].median() / _pais_m - 1) * 100,
                'vs_pais_prom_%': (_pmean(_costa['canasta_total_prom']) / _pais_p - 1) * 100}])], ignore_index=True)
            _costa_g = _costa_g[['localidad','n_sucursales','canasta_mediana','vs_pais_%','canasta_prom','vs_pais_prom_%']]
            _costa_g.to_excel(writer, sheet_name='costa', index=False)

    # ── Formato ──────────────────────────────────────────────────────────────
    for sn in writer.sheets:
        ws = writer.sheets[sn]
        fmt_ws(ws)
        auto_widths(ws)

print(f'Excel guardado: {out_xls}')
print()
print('='*65)
print(f'  RESUMEN — {NOMBRE_MES_TITLE.upper()}')
print('='*65)
for _col_id in CANASTAS_ACTIVAS:
    _name = CANASTA_NAMES[_col_id]
    _cgf  = canasta_geo_dict[_col_id]
    _prom = prom_nac_dict[_col_id]
    _promp = prom_nac_prom_dict[_col_id]
    _sn   = serie_nac_dict[_col_id]
    _rng  = f'{_sn["mes"].min()} -> {_sn["mes"].max()}' if len(_sn) > 0 else 'sin historia'
    print(f'  [{_name}] {len(_cgf):,} sucs | Mediana: ${_prom:,.0f} | Promedio: ${_promp:,.0f} | Serie: {_rng}')
print('='*65)
if MES_PARCIAL:
    print(f'\\n⚠️  {NOMBRE_MES_TITLE}: MES PARCIAL ({DIAS_CARGADOS}/{DIAS_MES} días).')
    print('    Los promedios son un snapshot al día; la VARIACIÓN mensual es PRELIMINAR')
    print('    (subestimada vs meses completos). Re-correr al cierre del mes para el dato final.')"""))

# ── CELL 20 — VALORES PARA DOCUMENTO TÉCNICO ──────────────────────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 20 — Valores para actualizar el documento técnico
# ============================================================
# Calcula todos los números que aparecen en el LaTeX (portada,
# resumen ejecutivo, evolución, provincias, cadenas, barrios,
# canastas especiales, acumulados) y los exporta como:
#   - Print organizado por sección del documento
#   - Hoja 'Valores_Documento' en canasta_analisis_{MES}.xlsx
# ============================================================

def _ar(x): return f"${int(round(x)):,}".replace(",", ".")
def _pp(x, dec=2):
    s = "+" if x >= 0 else ""; return f"{s}{x:.{dec}f}%".replace(".", ",")

_MEDIA_ID = 'cantidad_03'
_POP_ID   = 'cantidad_02'

# ── 1. Valores por canasta ────────────────────────────────────────────────────
_vals = {c: prom_nac_dict[c] for c in CANASTAS_ACTIVAS}
_vals_prom = {c: prom_nac_prom_dict[c] for c in CANASTAS_ACTIVAS}
_vars = {}
_vars_prom = {}
for c in CANASTAS_ACTIVAS:
    _sn = serie_nac_dict[c]
    if len(_sn) > 1:
        _vars[c] = _sn['variacion_mensual_%'].iloc[-1]
        _vars_prom[c] = _sn['variacion_mensual_prom_%'].iloc[-1]

_sorted_ids = sorted(CANASTAS_ACTIVAS, key=lambda c: _vals[c])
_min_id, _max_id = _sorted_ids[0], _sorted_ids[-1]
_brecha_abs = _vals[_max_id] - _vals[_min_id]
_brecha_pct = _brecha_abs / _vals[_min_id] * 100

# ── 2. Análisis provincial (canasta media) ────────────────────────────────────
_v_media = _vals[_MEDIA_ID]
_v_media_prom = _vals_prom[_MEDIA_ID]
_spv = serie_prov_dict[_MEDIA_ID].copy()
_spv['vs_%'] = (_spv['canasta_total'] / _v_media - 1) * 100
_prov_min = _spv.loc[_spv['canasta_total'].idxmin()]
_prov_max = _spv.loc[_spv['canasta_total'].idxmax()]
_disp_prov = (_prov_max['canasta_total'] - _prov_min['canasta_total']) / _prov_min['canasta_total'] * 100

# ── 3. Rango sucursales (canasta media) ───────────────────────────────────────
_cgf_m = canasta_geo_dict[_MEDIA_ID]
_suc_min, _suc_max = _cgf_m['canasta_total'].min(), _cgf_m['canasta_total'].max()
_p25, _p75 = _cgf_m['canasta_total'].quantile(0.25), _cgf_m['canasta_total'].quantile(0.75)
_disp_suc = (_suc_max - _suc_min) / _suc_min * 100

# ── 4. Cadenas nacionales (canasta media) ─────────────────────────────────────
_rk_nac = (_cgf_m.groupby('cadena')
            .agg(n=('canasta_total','count'), prom=('canasta_total','mean'))
            .reset_index())
_rk_nac = _rk_nac[_rk_nac['n'] >= MIN_SUCURSALES_RANKING].sort_values('prom', ascending=False)
_disp_cad_nac = (_rk_nac['prom'].max() - _rk_nac['prom'].min()) / _rk_nac['prom'].min() * 100

# ── 5. Cadenas AMBA (canasta media) ───────────────────────────────────────────
_cgf_amba = _cgf_m[_cgf_m['PROVINCIA_NORM'].isin(['Buenos Aires','CABA'])].copy()
_rk_amba = (_cgf_amba.groupby('cadena')
             .agg(n=('canasta_total','count'), prom=('canasta_total','mean'))
             .reset_index())
_rk_amba = _rk_amba[_rk_amba['n'] >= MIN_SUCURSALES_RANKING].sort_values('prom', ascending=False)
_disp_cad_amba = (_rk_amba['prom'].max() - _rk_amba['prom'].min()) / _rk_amba['prom'].min() * 100 if len(_rk_amba) >= 2 else 0

# ── 6. Barrios CABA (canasta media) ───────────────────────────────────────────
_cgf_caba = _cgf_m[_cgf_m['PROVINCIA_NORM'] == 'CABA'].copy()
_cgf_caba['barrio'] = _cgf_caba.apply(
    lambda r: det_barrio(r['sucursales_latitud'], r['sucursales_longitud']), axis=1)
_brk = (_cgf_caba[_cgf_caba['barrio'] != 'Sin clasificar']
         .groupby('barrio')['canasta_total']
         .agg(['mean','count'])
         .reset_index()
         .rename(columns={'mean':'prom','count':'n'}))
_brk = _brk[_brk['n'] >= 2].sort_values('prom', ascending=False).reset_index(drop=True)
_prom_caba = _cgf_caba['canasta_total'].mean()
_disp_bar = (_brk['prom'].max() - _brk['prom'].min()) / _brk['prom'].min() * 100 if len(_brk) >= 2 else 0

# ── 7. Canastas especiales ────────────────────────────────────────────────────
_prima_cel  = ((_vals.get('cantidad_05', _v_media) / _v_media) - 1) * 100 if 'cantidad_05' in _vals else None
_ahorro_veg = ((_vals.get('cantidad_06', _vals.get(_POP_ID, 1)) / _vals.get(_POP_ID, 1)) - 1) * 100 if 'cantidad_06' in _vals and _POP_ID in _vals else None

# ── 8. Acumulados desde MES_INICIO_GRAFICO ────────────────────────────────────
_acum = {}
for c in CANASTAS_ACTIVAS:
    _sn = serie_nac_dict[c]
    _mg = MES_INICIO_GRAFICO if MES_INICIO_GRAFICO in _sn['mes'].values else (_sn['mes'].min() if len(_sn) > 0 else None)
    if _mg and len(_sn) >= 2:
        _dg = _sn[_sn['mes'] >= _mg].reset_index(drop=True)
        if len(_dg) >= 2:
            _acum[c] = (_dg['canasta_nacional_ponderada'].iloc[-1] / _dg['canasta_nacional_ponderada'].iloc[0] - 1) * 100

_ipc_b = ipc[ipc['mes'] == MES_INICIO_GRAFICO] if MES_INICIO_GRAFICO in ipc['mes'].values else ipc.iloc[[0]]
_ipc_l = ipc.iloc[-1]
_acum_ipc_gen = (_ipc_l['ipc_general'] / _ipc_b['ipc_general'].values[0] - 1) * 100
_acum_ipc_ali = (_ipc_l['ipc_alimentos'] / _ipc_b['ipc_alimentos'].values[0] - 1) * 100 if _ipc_b['ipc_alimentos'].notna().all() else None

# ── 9. Variaciones últimos 3 meses ────────────────────────────────────────────
_var3_canastas = {}
for c in CANASTAS_ACTIVAS:
    _sn = serie_nac_dict[c]
    if len(_sn) >= 3:
        _var3_canastas[c] = _sn[['mes','variacion_mensual_%']].tail(3).values.tolist()

_ipc_tail3 = ipc[['mes','ipc_general_var_%','ipc_alimentos_var_%']].tail(3)

# ═══════════════════════════════════════════════════════════════════════════════
# IMPRIMIR RESUMEN POR SECCIÓN DEL DOCUMENTO
# ═══════════════════════════════════════════════════════════════════════════════
print("=" * 72)
print(f"  VALORES PARA DOCUMENTO TÉCNICO — {NOMBRE_MES_TITLE.upper()}")
print("=" * 72)

print("\\n[PORTADA / TÍTULO]")
print(f"  Mes:                 {NOMBRE_MES_TITLE}{SUFIJO_PARCIAL}")
print(f"  Canasta media:       {_ar(_v_media)}")
print(f"  Var. mensual media:  {_pp(_vars.get(_MEDIA_ID, float('nan')))}" + ('  ⚠️ PRELIMINAR (mes parcial)' if MES_PARCIAL else ''))

print("\\n[RESUMEN EJECUTIVO — Item 1: valores y variaciones por canasta]")
for c in _sorted_ids:
    _nm = CANASTA_NAMES[c]
    print(f"  {_nm:<22} {_ar(_vals[c]):>12}   var.mensual: {_pp(_vars.get(c, float('nan')))}")
print(f"  Brecha absoluta:       {_ar(_brecha_abs)}")
print(f"  Brecha relativa:       {_brecha_pct:.1f}%  ({CANASTA_NAMES[_min_id]} a {CANASTA_NAMES[_max_id]})")

print("\\n[RESUMEN EJECUTIVO — Item 2: patrón provincial]")
print(f"  Prov. más barata:  {_prov_min['provincia']:<25} {_ar(_prov_min['canasta_total'])}  ({_pp(_prov_min['vs_%'])})")
print(f"  Prov. más cara:    {_prov_max['provincia']:<25} {_ar(_prov_max['canasta_total'])}  ({_pp(_prov_max['vs_%'])})")
print(f"  Dispersión prov.:  {_disp_prov:.1f}%")

print("\\n[RESUMEN EJECUTIVO — Item 3: barrios CABA (top 3 / bottom 3)]")
for _, r in _brk.head(3).iterrows():
    print(f"  CARO   {r['barrio']:<25} {_ar(r['prom'])}  ({_pp((r['prom']/_prom_caba-1)*100)})")
for _, r in _brk.tail(3).sort_values('prom').iterrows():
    print(f"  BARATO {r['barrio']:<25} {_ar(r['prom'])}  ({_pp((r['prom']/_prom_caba-1)*100)})")
print(f"  Dispersión CABA:   {_disp_bar:.2f}%")
print(f"  Promedio CABA:     {_ar(_prom_caba)}")

print("\\n[RESUMEN EJECUTIVO — Item 4: cadenas nacionales y AMBA]")
print("  Nacional:")
for _, r in _rk_nac.sort_values('prom', ascending=False).head(3).iterrows():
    print(f"    CARA   {r['cadena']:<28} {_ar(r['prom'])}")
for _, r in _rk_nac.sort_values('prom').head(1).iterrows():
    print(f"    BARATA {r['cadena']:<28} {_ar(r['prom'])}")
print(f"  Dispersión nacional: {_disp_cad_nac:.1f}%")
if len(_rk_amba) >= 2:
    print("  AMBA:")
    for _, r in _rk_amba.head(1).iterrows():
        print(f"    CARA   {r['cadena']:<28} {_ar(r['prom'])}")
    for _, r in _rk_amba.tail(1).iterrows():
        print(f"    BARATA {r['cadena']:<28} {_ar(r['prom'])}")
    print(f"  Dispersión AMBA:     {_disp_cad_amba:.1f}%")

print("\\n[RESUMEN EJECUTIVO — Item 5: canastas especiales]")
if _prima_cel is not None:
    print(f"  Prima celíaca:   +{abs(_prima_cel):.1f}%   ({_ar(_vals['cantidad_05'])} vs {_ar(_v_media)})")
if _ahorro_veg is not None:
    print(f"  Ahorro vegano:   {_pp(_ahorro_veg)}   ({_ar(_vals['cantidad_06'])} vs {_ar(_vals[_POP_ID])})")

print(f"\\n[RESUMEN EJECUTIVO — Item 6: acumulados desde {MES_INICIO_GRAFICO}]")
for c in _sorted_ids:
    if c in _acum:
        print(f"  {CANASTA_NAMES[c]:<22} {_pp(_acum[c])}")
print(f"  {'IPC Nivel General':<22} {_pp(_acum_ipc_gen)}")
if _acum_ipc_ali:
    print(f"  {'IPC Alimentos':<22} {_pp(_acum_ipc_ali)}")

print("\\n[SECCIÓN EVOLUCIÓN — Variaciones últimos 3 meses]")
print(f"  {'Mes':<10}  {'IPC Gral':>10}  {'IPC Ali':>10}")
for _, r in _ipc_tail3.iterrows():
    print(f"  {r['mes']:<10}  {_pp(r['ipc_general_var_%']):>10}  {_pp(r['ipc_alimentos_var_%']):>10}")
for c in CANASTAS_ACTIVAS:
    if c in _var3_canastas:
        _nm = CANASTA_NAMES[c]
        _vs = [_pp(v[1]) for v in _var3_canastas[c]]
        _ms = [v[0] for v in _var3_canastas[c]]
        print(f"  {_nm:<22}  " + "  ".join(f"{m}: {v}" for m, v in zip(_ms, _vs)))

print("\\n[SECCIÓN PROVINCIAL — Dispersión sucursales (canasta media)]")
print(f"  Sucursal mínima:  {_ar(_suc_min)}")
print(f"  Sucursal máxima:  {_ar(_suc_max)}")
print(f"  Dispersión:       {_disp_suc:.1f}%")
print(f"  P25:              {_ar(_p25)}")
print(f"  P75:              {_ar(_p75)}")

print("\\n[SECCIÓN CADENAS — Ranking nacional completo (canasta media)]")
for _, r in _rk_nac.sort_values('prom', ascending=False).iterrows():
    _vs_nac = (r['prom'] / _cgf_m['canasta_total'].mean() - 1) * 100
    print(f"  {r['cadena']:<28} {int(r['n']):>4} sucs   {_ar(r['prom'])}  ({_pp(_vs_nac)})")

if len(_rk_amba) >= 2:
    print("\\n[SECCIÓN CADENAS — Ranking AMBA completo (canasta media)]")
    _prom_amba = _cgf_amba['canasta_total'].mean()
    for _, r in _rk_amba.sort_values('prom', ascending=False).iterrows():
        _vs_amba = (r['prom'] / _prom_amba - 1) * 100
        print(f"  {r['cadena']:<28} {int(r['n']):>4} sucs   {_ar(r['prom'])}  ({_pp(_vs_amba)})")

print("\\n[SECCIÓN BARRIOS CABA — Ranking completo (canasta media)]")
for i, r in _brk.iterrows():
    _vs_c = (r['prom'] / _prom_caba - 1) * 100
    _vs_p = (r['prom'] / _v_media - 1) * 100
    print(f"  {str(i+1):>2}. {r['barrio']:<25} {int(r['n']):>3} sucs   {_ar(r['prom'])}  vs CABA: {_pp(_vs_c)}  vs país: {_pp(_vs_p)}")
print("=" * 72)

# ═══════════════════════════════════════════════════════════════════════════════
# EXPORTAR A EXCEL (nueva hoja en canasta_analisis_{MES}.xlsx)
# ═══════════════════════════════════════════════════════════════════════════════
import openpyxl as _opxl
_xls_path = OUTPUT_DIR / f'canasta_analisis_{PERIODO}.xlsx'
try:
    _wb = _opxl.load_workbook(_xls_path)
    _ws = _wb.create_sheet('Valores_Documento')

    # Cabecera
    _ws.append(['Seccion', 'Variable', 'Valor_LaTeX', 'Valor_numero'])
    _hdr = _ws[1]
    for _cell in _hdr:
        _cell.font = _opxl.styles.Font(bold=True)

    _rows_doc = []
    # Portada
    _rows_doc.append(['Portada', 'Mes', NOMBRE_MES_TITLE, NOMBRE_MES_TITLE])
    _rows_doc.append(['Portada', 'Canasta media valor', _ar(_v_media), round(_v_media, 0)])
    _rows_doc.append(['Portada', 'Canasta media valor (prom)', _ar(_v_media_prom), round(_v_media_prom, 0)])
    _rows_doc.append(['Portada', 'Canasta media var.mensual', _pp(_vars.get(_MEDIA_ID, 0)), round(_vars.get(_MEDIA_ID, 0), 2)])
    # Canastas (valor y var.mensual: MEDIANA y su gemela MEDIA recortada)
    for c in _sorted_ids:
        _nm = CANASTA_NAMES[c]
        _rows_doc.append(['Canastas', f'{_nm} — valor', _ar(_vals[c]), round(_vals[c], 0)])
        _rows_doc.append(['Canastas', f'{_nm} — valor (prom)', _ar(_vals_prom[c]), round(_vals_prom[c], 0)])
        _rows_doc.append(['Canastas', f'{_nm} — var.mensual', _pp(_vars.get(c, 0)), round(_vars.get(c, 0), 2)])
        _rows_doc.append(['Canastas', f'{_nm} — var.mensual (prom)', _pp(_vars_prom.get(c, 0)), round(_vars_prom.get(c, 0), 2)])
    _rows_doc.append(['Canastas', 'Brecha absoluta', _ar(_brecha_abs), round(_brecha_abs, 0)])
    _rows_doc.append(['Canastas', 'Brecha relativa (%)', f'{_brecha_pct:.1f}%', round(_brecha_pct, 1)])
    # Provincias
    _rows_doc.append(['Provincias', 'Prov. más barata (nombre)', _prov_min['provincia'], _prov_min['provincia']])
    _rows_doc.append(['Provincias', 'Prov. más barata (valor)', _ar(_prov_min['canasta_total']), round(_prov_min['canasta_total'], 0)])
    _rows_doc.append(['Provincias', 'Prov. más barata (vs. prom.)', _pp(_prov_min['vs_%']), round(_prov_min['vs_%'], 2)])
    _rows_doc.append(['Provincias', 'Prov. más cara (nombre)', _prov_max['provincia'], _prov_max['provincia']])
    _rows_doc.append(['Provincias', 'Prov. más cara (valor)', _ar(_prov_max['canasta_total']), round(_prov_max['canasta_total'], 0)])
    _rows_doc.append(['Provincias', 'Prov. más cara (vs. prom.)', _pp(_prov_max['vs_%']), round(_prov_max['vs_%'], 2)])
    _rows_doc.append(['Provincias', 'Dispersión interprovincial (%)', f'{_disp_prov:.1f}%', round(_disp_prov, 1)])
    _rows_doc.append(['Provincias', 'Sucursal más barata', _ar(_suc_min), round(_suc_min, 0)])
    _rows_doc.append(['Provincias', 'Sucursal más cara', _ar(_suc_max), round(_suc_max, 0)])
    _rows_doc.append(['Provincias', 'Dispersión inter-sucursal (%)', f'{_disp_suc:.1f}%', round(_disp_suc, 1)])
    _rows_doc.append(['Provincias', 'P25 sucursales', _ar(_p25), round(_p25, 0)])
    _rows_doc.append(['Provincias', 'P75 sucursales', _ar(_p75), round(_p75, 0)])
    # Cadenas nacionales
    for _, r in _rk_nac.sort_values('prom', ascending=False).iterrows():
        _rows_doc.append(['Cadenas_Nacional', r['cadena'], _ar(r['prom']), round(r['prom'], 0)])
    _rows_doc.append(['Cadenas_Nacional', 'Dispersión (%)', f'{_disp_cad_nac:.1f}%', round(_disp_cad_nac, 1)])
    # Cadenas AMBA
    for _, r in _rk_amba.sort_values('prom', ascending=False).iterrows():
        _rows_doc.append(['Cadenas_AMBA', r['cadena'], _ar(r['prom']), round(r['prom'], 0)])
    if len(_rk_amba) >= 2:
        _rows_doc.append(['Cadenas_AMBA', 'Dispersión (%)', f'{_disp_cad_amba:.1f}%', round(_disp_cad_amba, 1)])
    # Barrios CABA
    for i, r in _brk.iterrows():
        _rows_doc.append(['Barrios_CABA', r['barrio'], _ar(r['prom']), round(r['prom'], 0)])
    _rows_doc.append(['Barrios_CABA', 'Promedio CABA', _ar(_prom_caba), round(_prom_caba, 0)])
    _rows_doc.append(['Barrios_CABA', 'Dispersión barrios (%)', f'{_disp_bar:.2f}%', round(_disp_bar, 2)])
    # Canastas especiales
    if _prima_cel is not None:
        _rows_doc.append(['Especiales', 'Prima celíaca (%)', f'+{abs(_prima_cel):.1f}%', round(_prima_cel, 1)])
    if _ahorro_veg is not None:
        _rows_doc.append(['Especiales', 'Ahorro vegano (%)', _pp(_ahorro_veg), round(_ahorro_veg, 1)])
    # Acumulados
    for c in _sorted_ids:
        if c in _acum:
            _rows_doc.append(['Acumulados', f'{CANASTA_NAMES[c]} desde {MES_INICIO_GRAFICO}', _pp(_acum[c]), round(_acum[c], 1)])
    _rows_doc.append(['Acumulados', f'IPC Nivel General desde {MES_INICIO_GRAFICO}', _pp(_acum_ipc_gen), round(_acum_ipc_gen, 1)])
    if _acum_ipc_ali:
        _rows_doc.append(['Acumulados', f'IPC Alimentos desde {MES_INICIO_GRAFICO}', _pp(_acum_ipc_ali), round(_acum_ipc_ali, 1)])
    # Variaciones últimos 3 meses
    for _, r in _ipc_tail3.iterrows():
        _rows_doc.append(['Var_3meses', f'IPC General {r["mes"]}', _pp(r['ipc_general_var_%']), round(r['ipc_general_var_%'], 2)])
        _rows_doc.append(['Var_3meses', f'IPC Alimentos {r["mes"]}', _pp(r['ipc_alimentos_var_%']), round(r['ipc_alimentos_var_%'], 2)])
    for c in CANASTAS_ACTIVAS:
        if c in _var3_canastas:
            for _m, _v in _var3_canastas[c]:
                _rows_doc.append(['Var_3meses', f'{CANASTA_NAMES[c]} {_m}', _pp(_v), round(_v, 2)])

    for row in _rows_doc:
        _ws.append(row)

    # Ancho de columnas
    _ws.column_dimensions['A'].width = 20
    _ws.column_dimensions['B'].width = 40
    _ws.column_dimensions['C'].width = 18
    _ws.column_dimensions['D'].width = 16

    _wb.save(_xls_path)
    print(f'Hoja Valores_Documento agregada a: {_xls_path.name}')
except Exception as _e:
    print(f'AVISO: No se pudo agregar hoja al Excel: {_e}')
"""))

# ── CELL 21 — DIAGNOSTIC: trazabilidad temporal de Candidatos ──────────────────
cells.append(cell_code("""\
# ============================================================
# CELDA 20 — DIAGNÓSTICO: Trazabilidad temporal de Candidatos
# ============================================================
# Lee hoja Candidatos, escanea ZIPs con usecols=['id_producto'] solo
# (~15-20 min). Muestra qué candidatos son estables históricamente.
# No modifica ninguna canasta — es solo informativo.

try:
    df_cand = pd.read_excel(CANASTA_EXCEL, sheet_name='Candidatos',
                            dtype={'id_producto': str})
    df_cand['ean_norm'] = df_cand['id_producto'].str.lstrip('0')
    CAND_EANS = set(df_cand['ean_norm'])
    print(f'Candidatos cargados: {len(CAND_EANS):,} EANs únicos')
except Exception as _e:
    raise RuntimeError(f'No se pudo leer hoja Candidatos: {_e}')

_presencia = {}
_meses_vis = []
for _zip_path, _anio, _sem in detectar_semestres():
    _meses = archivos_por_mes(_zip_path)
    for (_anio_m, _mes_m), _archs in sorted(_meses.items()):
        _lbl = f'{_anio_m}-{_mes_m:02d}'
        if _lbl < MES_INICIO_HISTORICO: continue
        _meses_vis.append(_lbl)
        _found = set()
        for _archivo in sorted(_archs):
            _tmp_p = TMP_DIR / Path(_archivo).name
            with zipfile.ZipFile(_zip_path) as _zf:
                with _zf.open(_archivo) as _s, open(_tmp_p, 'wb') as _d:
                    shutil.copyfileobj(_s, _d, length=4*1024*1024)
            with gzip.open(_tmp_p, 'rt', encoding='utf-8', errors='replace') as _g:
                for _chunk in pd.read_csv(_g, dtype={'id_producto': str},
                                           usecols=['id_producto'],
                                           chunksize=500_000, low_memory=False):
                    _chunk['ean_norm'] = _chunk['id_producto'].apply(normalizar_ean)
                    _found |= set(_chunk[_chunk['ean_norm'].isin(CAND_EANS)]['ean_norm'])
            _tmp_p.unlink(missing_ok=True)
        for _ean in _found:
            _presencia.setdefault(_ean, set()).add(_lbl)
        print(f'  {_lbl}: {len(_found):,} candidatos encontrados')

_n_meses = len(set(_meses_vis))
_mes_min  = min(_meses_vis) if _meses_vis else MES_INICIO_HISTORICO
_mes_max  = max(_meses_vis) if _meses_vis else ULTIMO_MES

_rows = []
for _ean, _mp in _presencia.items():
    _rows.append({'ean_norm':_ean,'meses_presentes':len(_mp),
                  'pct_trazabilidad':len(_mp)/_n_meses*100,
                  'primer_mes':min(_mp),'ultimo_mes':max(_mp),
                  'en_canasta':_ean in CANASTA_EANS_NORM})
for _ean in CAND_EANS - set(_presencia.keys()):
    _rows.append({'ean_norm':_ean,'meses_presentes':0,'pct_trazabilidad':0.0,
                  'primer_mes':None,'ultimo_mes':None,'en_canasta':_ean in CANASTA_EANS_NORM})

df_traz = pd.DataFrame(_rows)
_mc = ['ean_norm'] + [c for c in ['descripcion','marca','categoria','subcategoria',
                                   'precio_mediano','score_cobertura'] if c in df_cand.columns]
df_traz = (df_traz.merge(df_cand[_mc], on='ean_norm', how='left')
           .sort_values('pct_trazabilidad', ascending=False).reset_index(drop=True))

print(f'\\n{"="*60}')
print(f'  TRAZABILIDAD TEMPORAL — {_mes_min} a {_mes_max} ({_n_meses} meses)')
print(f'{"="*60}')
print(f'  Candidatos con trazabilidad 100%  : {(df_traz["pct_trazabilidad"]==100).sum():>5,}')
print(f'  Candidatos con trazabilidad  >90% : {(df_traz["pct_trazabilidad"]> 90).sum():>5,}')
print(f'  Candidatos con trazabilidad  >75% : {(df_traz["pct_trazabilidad"]> 75).sum():>5,}')
print(f'  Candidatos con trazabilidad  <50% : {(df_traz["pct_trazabilidad"]< 50).sum():>5,}')
_can_traz = df_traz[df_traz['en_canasta']]['pct_trazabilidad']
if len(_can_traz) > 0:
    print(f'\\n  EANs en canastas activas ({len(_can_traz)} únicos):')
    print(f'    Trazabilidad promedio : {_can_traz.mean():.1f}%')
    print(f'    Trazabilidad mínima   : {_can_traz.min():.1f}%')
print(f'{"="*60}')

_cols_show = [c for c in ['descripcion','marca','categoria','meses_presentes',
                           'pct_trazabilidad','primer_mes','en_canasta'] if c in df_traz.columns]
display(df_traz[_cols_show].head(30).style
    .background_gradient(subset=['pct_trazabilidad'], cmap='RdYlGn', vmin=0, vmax=100)
    .format({'pct_trazabilidad': '{:.1f}%'})
    .set_caption(f'Top 30 candidatos por trazabilidad ({_mes_min} → {_mes_max})'))

_fig, _axes = plt.subplots(1, 2, figsize=(14, 5))
_axes[0].hist(df_traz['pct_trazabilidad'], bins=20, color='#0055A4', edgecolor='white', alpha=0.85)
if len(_can_traz) > 0:
    _axes[0].axvline(_can_traz.mean(), color='#D62728', linewidth=2,
                     linestyle='--', label=f'Promedio canastas activas ({_can_traz.mean():.0f}%)')
    _axes[0].legend(fontsize=9)
_axes[0].set_xlabel('Trazabilidad temporal (%)'); _axes[0].set_ylabel('Número de productos')
_axes[0].set_title('Distribución de trazabilidad (todos los candidatos)')
if 'categoria' in df_traz.columns:
    _top_cat = (df_traz.groupby('categoria')['pct_trazabilidad']
                .agg(_mean='mean', _n='count').query('_n >= 5')
                .sort_values('_mean', ascending=True).tail(15))
    _axes[1].barh(_top_cat.index, _top_cat['_mean'], color='#0055A4', alpha=0.85)
    _axes[1].axvline(90, color='gray', linestyle='--', linewidth=1, alpha=0.6)
    _axes[1].set_xlabel('Trazabilidad promedio (%)'); _axes[1].set_xlim(0, 105)
    _axes[1].set_title('Trazabilidad por categoría (>=5 candidatos)')
    for _i, (_idx, _row) in enumerate(_top_cat.iterrows()):
        _axes[1].text(_row['_mean']+0.5, _i, f'{_row["_mean"]:.0f}%', va='center', fontsize=8)
plt.tight_layout()
_fig.savefig(OUTPUT_DIR / f'trazabilidad_candidatos_{ULTIMO_MES}.png', dpi=600, bbox_inches='tight')
plt.show()

_out_traz = OUTPUT_DIR / f'trazabilidad_candidatos_{ULTIMO_MES}.xlsx'
df_traz.to_excel(_out_traz, index=False)
print(f'Tabla completa guardada: {_out_traz.name}')"""))

# ── CELL 22 — DATOS ECONOMETRIA (motor semanal + mensual, histórico completo) ────
cells.append(cell_code(r'''# ============================================================
# CELDA 22 — DATOS ECONOMETRIA: series SEMANAL + MENSUAL por nivel (histórico completo)
# ============================================================
# Exporta un Excel tidy (formato largo) para análisis de series de tiempo:
#   - Costo de CADA canasta activa (cantidad_01..06 de la hoja Selección)
#   - Precio de 10 PRODUCTOS representativos (config abajo)
# con frecuencia SEMANAL (ISO) y MENSUAL, a nivel NACIONAL (ponderado por población),
# por PROVINCIA y por CADENA. Historia completa (>= MES_INICIO_HISTORICO).
# Cachea meses cerrados; relee el mes en curso fresco.

# --- CONFIG: 10 productos representativos (EAN -> nombre). EDITAR con tus 10 EAN ---
PRODUCTOS_ECONOMETRIA = {
    '7790895000997': 'Coca Cola 2.25 L',
    '7790742335500': 'Leche Entera La Serenisima 1 L',
    '7790387013610': 'Yerba Mate Taragui 1 Kg',
    '7790070012050': 'Aceite Girasol Cocinero 900 Ml',
    '7790070336385': 'Fideos Spaghetti Lucchetti 500 Gr',
    '7791120031557': 'Arroz Molinos Ala 1 Kg',
    '7792540250450': 'Azucar Ledesma 1 Kg',
    '7790550000157': 'Cafe Molido Cabrales 250 Gr',
    '7790132098459': 'Lavandina Ayudin 1 L',
    '7791813888468': 'Gaseosa Cola Pepsi 2 L',
}
ECON_MIN_SUC = 5          # mínimo de sucursales para reportar una celda provincia/cadena
ECON_XLSX    = OUTPUT_DIR / f'datos_econometria_{ULTIMO_MES}.xlsx'   # se guarda en el Drive

import datetime as _dt
PRODUCTOS_ECONOMETRIA = {normalizar_ean(k): v for k, v in PRODUCTOS_ECONOMETRIA.items()}
PROD_ECON_EANS = set(PRODUCTOS_ECONOMETRIA)
EANS_ECON = set(CANASTA_EANS_NORM) | PROD_ECON_EANS
_PAT_P8 = re.compile(r'^precio_\d{8}$')

def _fecha_de_periodo(_freq, _periodo):
    if _freq == 'mensual':
        _y, _m = _periodo.split('-'); return f'{_y}-{_m}-01'
    _y, _w = _periodo.split('-W')
    return _dt.date.fromisocalendar(int(_y), int(_w), 1).isoformat()

def _leer_mes_dia_econ(_lbl):
    # Devuelve [id_comercio,id_bandera,id_sucursal,ean_norm,fecha,precio] del mes _lbl
    _zip_path, _archs = _mapa_mes[_lbl]
    _rows = []
    for _archivo in sorted(_archs):
        _tmp = TMP_DIR / Path(_archivo).name
        with zipfile.ZipFile(_zip_path) as _zf:
            with _zf.open(_archivo) as _s, open(_tmp, 'wb') as _d:
                shutil.copyfileobj(_s, _d, length=4*1024*1024)
        with gzip.open(_tmp, 'rt', encoding='utf-8', errors='replace') as _g:
            for _ch in pd.read_csv(_g, dtype=str, chunksize=300_000, low_memory=False):
                _ch['ean_norm'] = _ch['id_producto'].apply(normalizar_ean)
                _ch = _ch[_ch['ean_norm'].isin(EANS_ECON)]
                if len(_ch) == 0: continue
                for _c in ['id_comercio','id_bandera','id_sucursal']:
                    _ch[_c] = _ch[_c].astype(str)
                _ch['_k'] = list(zip(_ch['id_comercio'], _ch['id_bandera'], _ch['id_sucursal']))
                _ch = _ch[_ch['_k'].isin(IDS_PAIS)].drop(columns='_k')
                if len(_ch) == 0: continue
                _cp = [c for c in _ch.columns if _PAT_P8.match(c)]
                if not _cp: continue
                _m = _ch.melt(id_vars=['id_comercio','id_bandera','id_sucursal','ean_norm'],
                              value_vars=_cp, var_name='_c', value_name='precio_raw')
                _m['precio'] = pd.to_numeric(_m['precio_raw'].replace('NA', np.nan), errors='coerce')
                _m = _m[_m['precio'].notna() & (_m['precio'] > 0)]
                if len(_m) == 0: continue
                _m['fecha'] = pd.to_datetime(_m['_c'].str[-8:], format='%Y%m%d', errors='coerce')
                _rows.append(_m[['id_comercio','id_bandera','id_sucursal','ean_norm','fecha','precio']])
        _tmp.unlink(missing_ok=True)
    if not _rows: return None
    _dd = pd.concat(_rows, ignore_index=True)
    _dd = _dd[_dd['fecha'].notna()].copy()
    if len(_dd) and _dd['precio'].median() > 10_000:
        _dd['precio'] /= 100
    return _dd

def _agg_desde_diario(_d):
    # Toma el df diario y devuelve filas LARGAS (tidy) para canastas y productos,
    # frecuencia semanal + mensual, niveles nacional/provincia/cadena.
    if _d is None or len(_d) == 0: return None
    _d = _d.copy()
    _d['mensual'] = _d['fecha'].dt.strftime('%Y-%m')
    _d['semanal'] = _d['fecha'].dt.strftime('%G-W%V')
    _k3 = ['id_comercio','id_bandera','id_sucursal']
    _out = []
    for _freq in ['mensual','semanal']:
        _key = _k3 + ['ean_norm', _freq]
        _md = _d.groupby(_key)['precio'].transform('median')
        _ok = (_d['precio'] >= _md/4) & (_d['precio'] <= _md*4)
        _gm = _d.groupby(_key)['precio'].median().rename('precio_med')
        _gp = _d[_ok].groupby(_key)['precio'].mean().rename('precio_prom')
        _px = pd.concat([_gm, _gp], axis=1).reset_index()
        _px['precio_prom'] = _px['precio_prom'].fillna(_px['precio_med'])
        _px = _px[~_px['id_comercio'].isin(CADENAS_FILTRAR)]
        _px = _px.merge(suc_geo_clean[_k3 + ['PROVINCIA_NORM','cadena']], on=_k3, how='inner')
        if len(_px) == 0: continue
        _nac = (_px.groupby(['ean_norm', _freq])
                   .agg(nac_med=('precio_med','median'), nac_prom=('precio_prom', _pmean))
                   .reset_index())
        _px = _px.merge(_nac, on=['ean_norm', _freq], how='left')

        # ---------------- CANASTAS ----------------
        for _col_id, _canasta in CANASTAS.items():
            _nm_can = CANASTA_NAMES[_col_id]
            _rec = pd.DataFrame([(e, q) for e, (nm, q, ct) in _canasta.items()],
                                columns=['ean_norm','qty'])
            _min_p = min(MIN_PRODUCTOS_PROPIOS, len(_canasta))
            _pr = _px.merge(_rec, on='ean_norm', how='inner')
            if len(_pr) == 0: continue
            _pr['store_m'] = _pr['precio_med']  * _pr['qty']
            _pr['store_p'] = _pr['precio_prom'] * _pr['qty']
            _pr['nacv_m']  = _pr['nac_med']  * _pr['qty']
            _pr['nacv_p']  = _pr['nac_prom'] * _pr['qty']
            _g = (_pr.groupby(_k3 + [_freq, 'PROVINCIA_NORM', 'cadena'])
                     .agg(S_store_m=('store_m','sum'), S_store_p=('store_p','sum'),
                          S_nac_m=('nacv_m','sum'),   S_nac_p=('nacv_p','sum'),
                          n_prop=('qty','count')).reset_index())
            _rn = _nac.merge(_rec, on='ean_norm', how='inner')
            _rn['am'] = _rn['nac_med']  * _rn['qty']
            _rn['ap'] = _rn['nac_prom'] * _rn['qty']
            _allnac = _rn.groupby(_freq)[['am','ap']].sum().reset_index().rename(
                columns={'am':'all_m','ap':'all_p'})
            _g = _g.merge(_allnac, on=_freq, how='left')
            _g['valor_med']  = _g['S_store_m'] + (_g['all_m'] - _g['S_nac_m'])
            _g['valor_prom'] = _g['S_store_p'] + (_g['all_p'] - _g['S_nac_p'])
            _g = _g[_g['n_prop'] >= _min_p]
            if len(_g) == 0: continue

            # provincia
            _pv = (_g.groupby([_freq,'PROVINCIA_NORM'])
                     .agg(valor_mediana=('valor_med','median'),
                          valor_promedio=('valor_prom', _pmean),
                          n_sucursales=('valor_med','size')).reset_index())
            _pv = _pv[_pv['n_sucursales'] >= ECON_MIN_SUC]
            for _, r in _pv.iterrows():
                _out.append((_freq, r[_freq], 'canasta', _nm_can, 'provincia', r['PROVINCIA_NORM'],
                             int(r['n_sucursales']), r['valor_mediana'], r['valor_promedio']))
            # cadena
            _cd = (_g.groupby([_freq,'cadena'])
                     .agg(valor_mediana=('valor_med','median'),
                          valor_promedio=('valor_prom', _pmean),
                          n_sucursales=('valor_med','size')).reset_index())
            _cd = _cd[_cd['n_sucursales'] >= ECON_MIN_SUC]
            for _, r in _cd.iterrows():
                _out.append((_freq, r[_freq], 'canasta', _nm_can, 'cadena', r['cadena'],
                             int(r['n_sucursales']), r['valor_mediana'], r['valor_promedio']))
            # nacional ponderado por población (mediana provincial -> peso poblacional)
            _pvw = _pv.copy()
            _pvw['peso'] = _pvw['PROVINCIA_NORM'].map(PESOS_POBLACION).fillna(0)
            for _per, _sub in _pvw.groupby(_freq):
                _w = _sub[_sub['peso'] > 0]['peso'].sum()
                if _w > 0:
                    _vm = (_sub['valor_mediana'] * _sub['peso']).sum() / _w
                    _vp = (_sub['valor_promedio'] * _sub['peso']).sum() / _w
                else:
                    _vm = _sub['valor_mediana'].mean(); _vp = _sub['valor_promedio'].mean()
                _nsu = int(_g[_g[_freq] == _per]['id_sucursal'].nunique())
                _out.append((_freq, _per, 'canasta', _nm_can, 'nacional', 'Nacional',
                             _nsu, _vm, _vp))

        # ---------------- PRODUCTOS ----------------
        _pp = _px[_px['ean_norm'].isin(PROD_ECON_EANS)].copy()
        if len(_pp):
            _pp['prod'] = _pp['ean_norm'].map(PRODUCTOS_ECONOMETRIA)
            # provincia
            _ppv = (_pp.groupby([_freq,'prod','PROVINCIA_NORM'])
                      .agg(valor_mediana=('precio_med','median'),
                           valor_promedio=('precio_prom', _pmean),
                           n_sucursales=('precio_med','size')).reset_index())
            _ppv = _ppv[_ppv['n_sucursales'] >= ECON_MIN_SUC]
            for _, r in _ppv.iterrows():
                _out.append((_freq, r[_freq], 'producto', r['prod'], 'provincia', r['PROVINCIA_NORM'],
                             int(r['n_sucursales']), r['valor_mediana'], r['valor_promedio']))
            # cadena
            _pcd = (_pp.groupby([_freq,'prod','cadena'])
                      .agg(valor_mediana=('precio_med','median'),
                           valor_promedio=('precio_prom', _pmean),
                           n_sucursales=('precio_med','size')).reset_index())
            _pcd = _pcd[_pcd['n_sucursales'] >= ECON_MIN_SUC]
            for _, r in _pcd.iterrows():
                _out.append((_freq, r[_freq], 'producto', r['prod'], 'cadena', r['cadena'],
                             int(r['n_sucursales']), r['valor_mediana'], r['valor_promedio']))
            # nacional (mediana y promedio simples entre sucursales)
            _pnc = (_pp.groupby([_freq,'prod'])
                      .agg(valor_mediana=('precio_med','median'),
                           valor_promedio=('precio_prom', _pmean),
                           n_sucursales=('precio_med','size')).reset_index())
            for _, r in _pnc.iterrows():
                _out.append((_freq, r[_freq], 'producto', r['prod'], 'nacional', 'Nacional',
                             int(r['n_sucursales']), r['valor_mediana'], r['valor_promedio']))
    if not _out: return None
    return pd.DataFrame(_out, columns=['frecuencia','periodo','tipo','clave','nivel','grupo',
                                       'n_sucursales','valor_mediana','valor_promedio'])

# ── Recorrer historia con caché (meses cerrados) + mes en curso fresco ──────────
_econ_key = hashlib.md5(('|'.join(sorted(EANS_ECON)) + '|econv1').encode()).hexdigest()[:8]
_econ_cache = CACHE_DIR / f'econ_{_econ_key}.parquet'
_econ_mes_actual = _meses_disp[-1]

if USE_CACHE and _econ_cache.exists():
    _econ_df = pd.read_parquet(_econ_cache)
    _econ_df = _econ_df[_econ_df['_mes_src'] < _econ_mes_actual].copy()
else:
    _econ_df = pd.DataFrame()

_en_cache = set(_econ_df['_mes_src'].unique()) if len(_econ_df) else set()
_falt = [m for m in _meses_disp if m < _econ_mes_actual and m not in _en_cache]
print(f'Datos econometria: {len(_meses_disp)} meses | en cache: {len(_en_cache)} | a leer: {len(_falt)}')
_nuevos_e = []
for _lbl in _falt:
    _r = _agg_desde_diario(_leer_mes_dia_econ(_lbl))
    if _r is not None:
        _r['_mes_src'] = _lbl
        _nuevos_e.append(_r)
        print(f'  {_lbl}: {len(_r)} filas')
    gc.collect()
if _nuevos_e:
    _econ_df = pd.concat([_econ_df] + _nuevos_e, ignore_index=True)
    if USE_CACHE:
        _econ_df.to_parquet(_econ_cache, compression='snappy', index=False)
        print(f'Cache econometria actualizado: {_econ_cache.name}')

# Mes en curso: siempre fresco
_r_act = _agg_desde_diario(_leer_mes_dia_econ(_econ_mes_actual))
if _r_act is not None:
    _r_act['_mes_src'] = _econ_mes_actual
    print(f'  {_econ_mes_actual}: {len(_r_act)} filas (mes en curso — fresco)')
econ_long = pd.concat([_econ_df] + ([_r_act] if _r_act is not None else []), ignore_index=True)
econ_long = econ_long[econ_long['periodo'].notna()].copy()
# Semanas ISO de BORDE: una misma semana partida entre 2 archivos mensuales genera
# fragmentos duplicados. Se asigna cada semana a su mes "dueño" (el del jueves ISO)
# y se conserva ese fragmento (el más completo). Los meses no tienen este problema.
def _owner_mes(_p):
    _y, _w = _p.split('-W')
    return _dt.date.fromisocalendar(int(_y), int(_w), 4).strftime('%Y-%m')
_wk = econ_long[econ_long['frecuencia'] == 'semanal'].copy()
_ms = econ_long[econ_long['frecuencia'] == 'mensual'].copy()
if len(_wk):
    _wk['_own'] = _wk['periodo'].map(_owner_mes)
    _wk['_is_own'] = (_wk['_own'] == _wk['_mes_src']).astype(int)
    _wk = (_wk.sort_values('_is_own', ascending=False)
             .drop_duplicates(['tipo','clave','nivel','grupo','periodo'], keep='first')
             .drop(columns=['_own','_is_own']))
econ_long = pd.concat([_ms, _wk], ignore_index=True)
econ_long['fecha'] = [ _fecha_de_periodo(f, p) for f, p in zip(econ_long['frecuencia'], econ_long['periodo']) ]
econ_long = econ_long.drop(columns=['_mes_src'])
econ_long = econ_long.sort_values(['tipo','clave','nivel','grupo','frecuencia','fecha']).reset_index(drop=True)
# variación % dentro de cada serie
econ_long['variacion_%'] = (econ_long.groupby(['frecuencia','tipo','clave','nivel','grupo'])['valor_mediana']
                            .pct_change() * 100)
print(f'econ_long: {len(econ_long):,} filas | periodos {econ_long["periodo"].nunique()}')'''))

# ── CELL 23 — DATOS ECONOMETRIA: exportar Excel tidy ────────────────────────────
cells.append(cell_code(r'''# ============================================================
# CELDA 23 — DATOS ECONOMETRIA: exportar Excel tidy (una hoja por tipo x nivel)
# ============================================================
_COLS = ['frecuencia','periodo','fecha','clave','grupo','n_sucursales',
         'valor_mediana','valor_promedio','variacion_%']

def _hoja(_tipo, _nivel):
    _s = econ_long[(econ_long['tipo'] == _tipo) & (econ_long['nivel'] == _nivel)].copy()
    if _nivel == 'nacional':
        _cols = [c for c in _COLS if c != 'grupo']
    else:
        _cols = _COLS
    return _s[_cols].sort_values(['clave','grupo','frecuencia','fecha'] if _nivel != 'nacional'
                                 else ['clave','frecuencia','fecha']).reset_index(drop=True)

_dicc = pd.DataFrame([
    ('frecuencia',    'semanal (ISO) o mensual'),
    ('periodo',       'etiqueta del período: YYYY-Www (semana ISO) o YYYY-MM (mes)'),
    ('fecha',         'fecha real: lunes de la semana ISO, o primer día del mes (para series de tiempo)'),
    ('clave',         'nombre de la canasta (cantidad_01..06) o del producto'),
    ('nivel',         'nacional | provincia | cadena (en el nombre de la hoja)'),
    ('grupo',         'provincia o cadena concreta (Nacional en las hojas nacionales)'),
    ('n_sucursales',  'nº de sucursales que respaldan el valor'),
    ('valor_mediana', 'costo de canasta / precio de producto — MEDIANA entre sucursales'),
    ('valor_promedio','ídem — PROMEDIO con outliers fuera [mediana/4, mediana x4]'),
    ('variacion_%',   'variación % del valor_mediana respecto del período anterior de esa serie'),
    ('nacional (canasta)', 'ponderado por población provincial (mediana provincial -> peso poblacional)'),
    ('nacional (producto)','mediana/promedio simple entre sucursales del país'),
    ('imputación',    'items faltantes en una sucursal se imputan con la referencia nacional del período'),
    ('umbral',        f'se reportan celdas provincia/cadena con n_sucursales >= {ECON_MIN_SUC}'),
], columns=['campo','descripcion'])

with pd.ExcelWriter(ECON_XLSX, engine='openpyxl') as _w:
    _dicc.to_excel(_w, sheet_name='Diccionario', index=False)
    _hoja('canasta','nacional').to_excel(_w, sheet_name='canastas_nacional', index=False)
    _hoja('canasta','provincia').to_excel(_w, sheet_name='canastas_provincia', index=False)
    _hoja('canasta','cadena').to_excel(_w, sheet_name='canastas_cadena', index=False)
    _hoja('producto','nacional').to_excel(_w, sheet_name='productos_nacional', index=False)
    _hoja('producto','provincia').to_excel(_w, sheet_name='productos_provincia', index=False)
    _hoja('producto','cadena').to_excel(_w, sheet_name='productos_cadena', index=False)

print(f'✅ Excel econometria guardado: {ECON_XLSX}')
_res = (econ_long.groupby(['tipo','nivel','frecuencia'])
        .agg(filas=('valor_mediana','size'), series=('clave','nunique'),
             periodos=('periodo','nunique')).reset_index())
print(_res.to_string(index=False))
_prod_con = set(econ_long[econ_long['tipo']=='producto']['clave'])
_prod_sin = [v for v in PRODUCTOS_ECONOMETRIA.values() if v not in _prod_con]
if _prod_sin:
    print(f'⚠️ Productos SIN datos (revisar EAN en PRODUCTOS_ECONOMETRIA): {_prod_sin}')'''))

# ── Write notebook ──────────────────────────────────────────────────────────────
nb = {
    'cells': cells,
    'metadata': {
        'kernelspec': {'display_name':'Python 3','language':'python','name':'python3'},
        'language_info': {'codemirror_mode':{'name':'ipython','version':3},
                          'file_extension':'.py','mimetype':'text/x-python',
                          'name':'python','nbformat':4,'pygments_lexer':'ipython3','version':'3.10.0'}
    },
    'nbformat': 4,
    'nbformat_minor': 5
}

script_dir = os.path.dirname(os.path.abspath(__file__))
out_path = os.path.join(script_dir, '02_evolucion_canasta_representativa.ipynb')
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump(nb, f, ensure_ascii=False, indent=1)

print(f'Written: {out_path}')
print(f'Cells: {len(cells)}')
