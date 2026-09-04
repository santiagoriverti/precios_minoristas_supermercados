# ============================================================
# CARGAR LA CANASTA FEMENINA EN "Productos unicos"  (cantidad_06)
# ------------------------------------------------------------
# Canasta de consumo femenino (gestion menstrual + depilacion + cuidado
# personal). 11 productos, todos con cobertura >=4 cadenas Y con historia de
# precios desde 2024-01 en la hoja "Productos unicos" (clave para que el indice
# vs-IPC no tenga saltos de base). Solo escribe la columna cantidad_06: NO toca
# cantidad_01..05 (asi conservas las 5 canastas que ya cargaste).
#
# Uso en Colab:
#   1. Pega TODO este bloque en una celda y ejecutalo.
#   2. Cuando te pida el archivo, subi el Excel que YA tiene cargadas las
#      canastas 01..05 (por ej. canasta_representativa_2026-08_con_canastas.xlsx).
#   3. Descarga el archivo *_con_femenina.xlsx y subilo al Drive a
#      output_canasta/ renombrado canasta_representativa_2026-08.xlsx
#      (o el que lea el notebook 07).
# ============================================================
import openpyxl, os
from google.colab import files

uploaded = files.upload()
archivo_excel = next(iter(uploaded.keys()))

SHEET = "Productos unicos"
COL   = "cantidad_06"   # <- canasta Femenina

# ---- Cantidades por EAN (unidades/mes). EDITABLE. ----
# (cantidad, descripcion de referencia, bloque)
CANTIDADES = {
    # --- Gestion menstrual ---
    '7790010002769': 2,  # Toallas Femeninas con Alas Ultrafina Suave Siempre Libre 8 Un
    '7790770601851': 1,  # Toallas Femeninas Nocturna Max Nosotras 8 Un
    '7891010254773': 2,  # Tampones Medio O.B. 8 Un
    '7790010002899': 1,  # Protectores Diarios Todoslosdias Compact Carefree 80 Un
    # --- Depilacion ---
    '7506460101279': 1,  # Banda de Cera Depilatoria Corporal Piel Sensible Veet
    '8002990292139': 1,  # Crema Depilatoria Piel Sensible Silck & Fresh Veet
    '7702018072392': 1,  # Rasuradora Desechable Simply Venus Gillette 2 Un
    '7500435173100': 1,  # Crema de Afeitar Protectora Suave y Lisa Venus 150 Ml
    # --- Cuidado personal femenino ---
    '7790064001909': 1,  # Algodon Discos Redondos Estrella 80 Un (desmaquillante)
    '7791293049557': 2,  # Desodorante Aerosol Women Nutrivive Rexona 150 Cc
    '7509552924121': 1,  # Tintura Coloracion Permanente Nutrisse (castano)
    # NOTA: se saco el Jabon Intimo Nivea (4005808538553): recien aparece en el
    # SEPA en 2024-07, sin dato en 2024-01, lo que inflaba artificialmente el
    # indice vs-IPC de la canasta. Si queres re-incorporar higiene intima, elegi
    # un producto con historia desde 2024-01.
}

def norm(e):
    s = str(e).strip()
    if s.endswith(".0"): s = s[:-2]
    return s.lstrip("0")

LOOKUP = {norm(k): v for k, v in CANTIDADES.items()}
ORIG   = {norm(k): k for k in CANTIDADES}

wb = openpyxl.load_workbook(archivo_excel)
if SHEET not in wb.sheetnames:
    raise ValueError(f"No existe la hoja '{SHEET}'. Hojas: {wb.sheetnames}")
ws = wb[SHEET]
headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
if "id_producto" not in headers:
    raise ValueError("La hoja no tiene columna 'id_producto'.")
col_id = headers.index("id_producto") + 1

# Crear la columna cantidad_06 si no existe (al final de la fila de headers)
if COL in headers:
    col_c6 = headers.index(COL) + 1
else:
    col_c6 = len(headers) + 1
    ws.cell(row=1, column=col_c6, value=COL)
    print(f"Columna '{COL}' no existia: se creo en la posicion {col_c6}.")

# Limpiar SOLO cantidad_06 (no se tocan 01..05)
for row in ws.iter_rows(min_row=2, min_col=col_c6, max_col=col_c6):
    row[0].value = None

# Escribir cantidad_06
matches = 0; encontrados = set()
for row in ws.iter_rows(min_row=2):
    ev = row[col_id-1].value
    if ev is None: continue
    e = norm(ev)
    if e in LOOKUP:
        encontrados.add(e)
        row[col_c6-1].value = LOOKUP[e]; matches += 1

print(f"\n=== Hoja '{SHEET}' — canasta Femenina (cantidad_06) ===")
print(f"  productos cargados: {matches}/{len(LOOKUP)}")
faltantes = sorted(set(LOOKUP) - encontrados)
if faltantes:
    print(f"\nAVISO: {len(faltantes)} EAN no estan en la hoja:")
    for e in faltantes: print("   ", ORIG[e])
else:
    print(f"\nOK: los {len(LOOKUP)} EAN estan en la hoja.")

base = os.path.splitext(os.path.basename(archivo_excel))[0]
# evita encadenar '_con_femenina' dos veces si re-ejecutas
base = base.replace("_con_femenina", "")
salida = f"{base}_con_femenina.xlsx"
wb.save(salida); print(f"\nGuardado: {salida}"); files.download(salida)
