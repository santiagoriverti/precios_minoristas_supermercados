# ============================================================
# CARGAR LAS 4 CANASTAS EN "Productos unicos"  (script de Colab)
# cantidad_01=Popular | cantidad_02=Media | cantidad_03=Ejecutiva | cantidad_04=Tecnológica
# ------------------------------------------------------------
# Uso: pegar en una celda de Colab, ejecutar, subir el Excel
# canasta_representativa_*.xlsx. Devuelve el Excel con cantidad_01..04
# cargadas en la hoja "Productos unicos".
# Los 90 EAN vienen embebidos (fuente: docs/canastas_alternativas/cantidades_dict.py).
# ============================================================
import openpyxl, os
from google.colab import files

# ---- 1. Subir el Excel canasta_representativa_*.xlsx ----
uploaded = files.upload()
archivo_excel = next(iter(uploaded.keys()))

# ---- 2. Cantidades por EAN (90 productos, todos cad>=4 / cobertura real) ----
CANTIDADES = {
    '7790070012050': {'cantidad_01': 2, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0},  # Aceite de Girasol Cocinero 900 Ml
    '7790272001029': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 2, 'cantidad_04': 0},  # Aceite de Girasol Natura 1.5 Lt
    '7790272001005': {'cantidad_01': 0, 'cantidad_02': 2, 'cantidad_03': 0, 'cantidad_04': 0},  # Aceite de Girasol Natura 900 Ml
    '7790070231833': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 1, 'cantidad_04': 0},  # Aceite de Oliva Extra Virgen Lira 500 Ml
    '7792540250450': {'cantidad_01': 3, 'cantidad_02': 3, 'cantidad_03': 2, 'cantidad_04': 0},  # Azúcar Molida Superior Ledesma 1 Kg
    '7790580132163': {'cantidad_01': 1, 'cantidad_02': 1, 'cantidad_03': 0, 'cantidad_04': 0},  # Mermelada Durazno Arcor 454 Gr
    '7793360131516': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 2, 'cantidad_04': 0},  # Mermelada de Durazno Bc La Campagnola 390 Gr
    '7792798007387': {'cantidad_01': 0, 'cantidad_02': 4, 'cantidad_03': 0, 'cantidad_04': 0},  # Cerveza Rubia Cristal Quilmes 1 Lt
    '7793147118860': {'cantidad_01': 2, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0},  # Cerveza Rubia en Lata Schneider 473 Cc
    '7792798003716': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 6, 'cantidad_04': 0},  # Cerveza en Botella No Retornable Corona 710
    '7791728248265': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 3, 'cantidad_04': 0},  # Vino Tinto Dulce Santa Julia 750 Ml
    '7798074864675': {'cantidad_01': 0, 'cantidad_02': 2, 'cantidad_03': 0, 'cantidad_04': 0},  # Vino Tinto Malbec Portillo 750 Ml
    '7799155000197': {'cantidad_01': 0, 'cantidad_02': 6, 'cantidad_03': 8, 'cantidad_04': 0},  # Agua Mineral sin Gas Pet Villavicencio 2 Lt
    '7792931000039': {'cantidad_01': 4, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0},  # Agua Mineralizada sin Gas Glaciar 2 Lt
    '7790895000997': {'cantidad_01': 0, 'cantidad_02': 5, 'cantidad_03': 6, 'cantidad_04': 0},  # Coca Cola Sabor Original 2.25 Lt
    '7791813888468': {'cantidad_01': 4, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0},  # Gaseosa Cola Pepsi 2 Lt
    '7622201735685': {'cantidad_01': 6, 'cantidad_02': 6, 'cantidad_03': 0, 'cantidad_04': 0},  # Jugo en Polvo Limón Tang 15 Gr
    '7790580131364': {'cantidad_01': 0, 'cantidad_02': 2, 'cantidad_03': 3, 'cantidad_04': 0},  # Atún al Natural sin TACC La Campagnola 170 Gr
    '7790625001171': {'cantidad_01': 2, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0},  # Jamón Cocido Feteado Lario 150 Gr
    '7798013102875': {'cantidad_01': 0, 'cantidad_02': 3, 'cantidad_03': 4, 'cantidad_04': 0},  # Jamón Cocido Genuino Cagnoli 120 Gr
    '7790670052388': {'cantidad_01': 2, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0},  # Salchicha Clásica Patyviena 6 Un
    '7790670052401': {'cantidad_01': 0, 'cantidad_02': 2, 'cantidad_03': 0, 'cantidad_04': 0},  # Salchichas Vienísima 6 Un
    '7794000008557': {'cantidad_01': 1, 'cantidad_02': 1, 'cantidad_03': 0, 'cantidad_04': 0},  # Caldo de Verdura Knorr 12 Un
    '7791866001203': {'cantidad_01': 1, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0},  # Mayonesa Doypack Natura 237 Gr
    '7791866001364': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 0, 'cantidad_04': 0},  # Mayonesa Doypack Natura 500 Ml
    '7794000006072': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 2, 'cantidad_04': 0},  # Mayonesa Regular Doypack Hellmanns 475 Gr
    '7790072001014': {'cantidad_01': 1, 'cantidad_02': 1, 'cantidad_03': 1, 'cantidad_04': 0},  # Sal Fina en Paquete Celusal 500 Gr
    '7790130000058': {'cantidad_01': 1, 'cantidad_02': 1, 'cantidad_03': 0, 'cantidad_04': 0},  # Vinagre de Alcohol Menoyo 1 Lt
    '7790580132392': {'cantidad_01': 2, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0},  # Arvejas Secas Remojadas Arcor 350 Gr
    '7790580132422': {'cantidad_01': 0, 'cantidad_02': 2, 'cantidad_03': 3, 'cantidad_04': 0},  # Choclo Amarillo Desgranado Arcor 300 Gr
    '7790580567903': {'cantidad_01': 3, 'cantidad_02': 3, 'cantidad_03': 4, 'cantidad_04': 0},  # Tomate Pelado Perita en Lata Arcor 400 Gr
    '7509546687292': {'cantidad_01': 1, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0},  # Crema Dental Doble Protección Odol 90 Gr
    '7509546702605': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 2, 'cantidad_04': 0},  # Crema Dental Total Original Mint Colgate 90 Gr
    '7793100111891': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 0, 'cantidad_04': 0},  # Crema Dental Ultra Blanco Colgate 90 Gr
    '7791293043791': {'cantidad_01': 1, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0},  # Desodorante Aerosol Marine Axe 150 Cc
    '7791293049557': {'cantidad_01': 0, 'cantidad_02': 2, 'cantidad_03': 3, 'cantidad_04': 0},  # Desodorante Aerosol Women Rexona 150 Ml
    '7791293051208': {'cantidad_01': 0, 'cantidad_02': 3, 'cantidad_03': 4, 'cantidad_04': 0},  # Jabón de Tocador Original Dove 90 Gr
    '7891024034781': {'cantidad_01': 2, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0},  # Jabón de Tocador con Karité Palmolive 85 Gr
    '7790250015536': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 4, 'cantidad_04': 0},  # Papel Higiénico Doble Hoja Higienol 4 Un
    '7790250016182': {'cantidad_01': 0, 'cantidad_02': 3, 'cantidad_03': 0, 'cantidad_04': 0},  # Papel Higiénico Fresh Hoja Simple 4 Un
    '7791070000696': {'cantidad_01': 2, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0},  # Papel Higiénico Texturado Campanita 4 Un
    '7791293045740': {'cantidad_01': 1, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0},  # Shampoo Crema Balance Sedal 340 Ml
    '7791293047102': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 2, 'cantidad_04': 0},  # Shampoo Reconstrucción Completa Dove 400 Ml
    '7790150079904': {'cantidad_01': 1, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0},  # Café Clásico en Saquitos La Virginia 20 Un
    '8445291082137': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 2, 'cantidad_04': 0},  # Café Instantáneo Suave Origen Dolca 170 Gr
    '7790550000157': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 0, 'cantidad_04': 0},  # Café Molido Torrado Cabrales 250 Gr
    '7790150355084': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 1, 'cantidad_04': 0},  # Té Verde en Saquitos La Virginia 20 Un
    '7790480008261': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 0, 'cantidad_04': 0},  # Té en Saquitos Green Hills 50 Un
    '7792710000175': {'cantidad_01': 2, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0},  # Yerba Mate Amanda 1 Kg
    '7790387013610': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 2, 'cantidad_04': 0},  # Yerba Mate con Palo 4Flex Taragui 1 Kg
    '7793704000928': {'cantidad_01': 0, 'cantidad_02': 2, 'cantidad_03': 0, 'cantidad_04': 0},  # Yerba Suave Playadito 1 Kg
    '7790787153664': {'cantidad_01': 1, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0},  # Dulce de Leche Clásico Ilolay 400 Gr
    '7790742625304': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 0, 'cantidad_04': 0},  # Dulce de Leche Clásico La Serenísima 400 Gr
    '7790742067005': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 2, 'cantidad_04': 0},  # Dulce de Leche Repostero La Serenísima 400 Gr
    '7790742335500': {'cantidad_01': 0, 'cantidad_02': 30, 'cantidad_03': 30, 'cantidad_04': 0},  # Leche Entera 3% Uat La Serenísima 1 Lt
    '7798338290028': {'cantidad_01': 24, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0},  # Leche Entera Larga Vida Tres Niñas 1 Lt
    '7798060850026': {'cantidad_01': 1, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0},  # Manteca Calidad Extra Tonadita 200 Gr
    '7793940054006': {'cantidad_01': 0, 'cantidad_02': 2, 'cantidad_03': 2, 'cantidad_04': 0},  # Manteca La Serenísima 200 Gr
    '7790398100132': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 0, 'cantidad_04': 0},  # Queso Rallado La Paulina 150 Gr
    '7790742222909': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 2, 'cantidad_04': 0},  # Queso Rallado La Serenísima 35 Gr
    '7790742373809': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 3, 'cantidad_04': 0},  # Queso Untable Cremón Light La Serenísima
    '7798060852990': {'cantidad_01': 2, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0},  # Queso Untable Jamón Tonadita 180 Gr
    '7791337061439': {'cantidad_01': 0, 'cantidad_02': 2, 'cantidad_03': 0, 'cantidad_04': 0},  # Queso Untable Light La Serenísima 290 Gr
    '7791337008694': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 8, 'cantidad_04': 0},  # Yogur Bebible Frutilla Danonino 185 Gr
    '7791337009387': {'cantidad_01': 0, 'cantidad_02': 6, 'cantidad_03': 0, 'cantidad_04': 0},  # Yogur Bebible Frutilla Yogurísimo 190 Gr
    '7793913013535': {'cantidad_01': 4, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0},  # Yogur Descremado Vainilla Tregar
    '7790990003039': {'cantidad_01': 0, 'cantidad_02': 2, 'cantidad_03': 0, 'cantidad_04': 0},  # Detergente Limón Magistral 20 Un 500 Ml
    '7791290794054': {'cantidad_01': 1, 'cantidad_02': 0, 'cantidad_03': 2, 'cantidad_04': 0},  # Detergente Bioactive Limón Cif 300 Ml
    '7791290792814': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 2, 'cantidad_04': 0},  # Jabón Líquido Ropa Bio Enzimas Skip 800 Ml
    '7791290792036': {'cantidad_01': 1, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0},  # Jabón Líquido Ropa Eco Lavado Ala
    '7790132098459': {'cantidad_01': 2, 'cantidad_02': 2, 'cantidad_03': 2, 'cantidad_04': 0},  # Lavandina Concentrada Ayudín 1 Lt
    '7791120031557': {'cantidad_01': 3, 'cantidad_02': 0, 'cantidad_03': 3, 'cantidad_04': 0},  # Arroz Largo Fino 00000 Molinos Ala 1 Kg
    '7790070431417': {'cantidad_01': 0, 'cantidad_02': 3, 'cantidad_03': 0, 'cantidad_04': 0},  # Arroz Parboil Gallo Oro 1 Kg
    '7790070336545': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 8, 'cantidad_04': 0},  # Fideos Fetuccini Don Vicente 500 Gr
    '7790070336385': {'cantidad_01': 6, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0},  # Fideos Spaghetti N7 Lucchetti 500 Gr
    '7790070336293': {'cantidad_01': 0, 'cantidad_02': 6, 'cantidad_03': 0, 'cantidad_04': 0},  # Fideos Tirabuzón N28 Matarazzo 500 Gr
    '7790040143524': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 5, 'cantidad_04': 0},  # Galletitas Rumba Bagley 330 Gr
    '7622201735906': {'cantidad_01': 3, 'cantidad_02': 4, 'cantidad_03': 0, 'cantidad_04': 0},  # Galletitas Chips Chocolate Pepitos 119 Gr
    '7792180139320': {'cantidad_01': 3, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0},  # Harina 000 Cañuelas 1 Kg
    '7790070562258': {'cantidad_01': 0, 'cantidad_02': 2, 'cantidad_03': 0, 'cantidad_04': 0},  # Harina 000 Favorita 1 Kg
    '7790070562265': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 2, 'cantidad_04': 0},  # Harina 0000 Favorita 1 Kg
    '7793890258783': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 4, 'cantidad_04': 0},  # Pan con Salvado Lactal 560 Gr
    '7793890258769': {'cantidad_01': 0, 'cantidad_02': 3, 'cantidad_03': 0, 'cantidad_04': 0},  # Pan de Mesa Blanco Lactal 315 Gr
    '7796962987321': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 1},  # [TEC] Aire Acondicionado Philco Split 3200W
    '6932554416638': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 1},  # [TEC] Celular Redmi 14C 4/128 Gb
    '7796885063874': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 1},  # [TEC] Heladera Brs400i1a 397l Bgh
    '7796885495194': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 1},  # [TEC] Microondas 20Lt 700 W Bgh
    '8806094365573': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 1},  # [TEC] Lavarropas Frontal WW65A4000EEU
    '0196802082679': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 1},  # [TEC] Notebook 14 Celn4120 Lenovo
    '7790653062106': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 1},  # [TEC] Televisión 43 Ur8750Psb Smart 4K
}

# ---- 3. Cargar en la hoja "Productos unicos" ----
SHEET = "Productos unicos"
COLS  = ["cantidad_01", "cantidad_02", "cantidad_03", "cantidad_04"]

def norm(e):
    s = str(e).strip()
    if s.endswith(".0"):          # por si el EAN vino como número
        s = s[:-2]
    return s.lstrip("0")

LOOKUP = {norm(k): v for k, v in CANTIDADES.items()}
ORIG   = {norm(k): k for k in CANTIDADES}

wb = openpyxl.load_workbook(archivo_excel)
if SHEET not in wb.sheetnames:
    raise ValueError(f"No existe la hoja '{SHEET}'. Hojas disponibles: {wb.sheetnames}")
ws = wb[SHEET]

headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
if "id_producto" not in headers:
    raise ValueError(f"La hoja '{SHEET}' no tiene columna 'id_producto'.")
col_id  = headers.index("id_producto") + 1
col_idx = {c: headers.index(c) + 1 for c in COLS if c in headers}
faltan  = [c for c in COLS if c not in headers]
if faltan:
    print(f"⚠️ La hoja no tiene estas columnas (se omiten): {faltan}")

# 3a) Limpiar cantidad_01..04 (carga limpia, sin valores viejos)
for ci in col_idx.values():
    for row in ws.iter_rows(min_row=2, min_col=ci, max_col=ci):
        row[0].value = None

# 3b) Escribir las cantidades por EAN (solo > 0)
matches = {c: 0 for c in col_idx}
encontrados = set()
for row in ws.iter_rows(min_row=2):
    ev = row[col_id - 1].value
    if ev is None:
        continue
    e = norm(ev)
    if e in LOOKUP:
        encontrados.add(e)
        for c, cant in LOOKUP[e].items():
            if c in col_idx and cant:
                row[col_idx[c] - 1].value = cant
                matches[c] += 1

# ---- 4. Reporte ----
print(f"\n=== Hoja '{SHEET}' — productos cargados ===")
for c in COLS:
    tot = sum(1 for v in CANTIDADES.values() if v.get(c, 0) > 0)
    print(f"  {c}: {matches.get(c, 0)}/{tot}")
faltantes = sorted(set(LOOKUP) - encontrados)
if faltantes:
    print(f"\n⚠️ {len(faltantes)} EAN de las canastas NO están en la hoja (revisar/reemplazar):")
    for e in faltantes:
        print("   ", ORIG[e])
else:
    print("\n✅ Los 90 EAN de las canastas están en la hoja.")

# ---- 5. Guardar y descargar ----
base = os.path.splitext(os.path.basename(archivo_excel))[0]
archivo_salida = f"{base}_con_canastas.xlsx"
wb.save(archivo_salida)
print(f"\n✅ Guardado: {archivo_salida}")
files.download(archivo_salida)
