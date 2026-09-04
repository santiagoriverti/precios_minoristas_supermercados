# ============================================================
# CARGAR LAS 6 CANASTAS EN "Productos unicos"  (script de Colab)  -- v4
# cantidad_01=Popular 02=Media 03=Ejecutiva 04=Tecnologica 05=Representativa 06=Femenina
# ------------------------------------------------------------
# 196 productos, seleccionados de la hoja real por cobertura (>= 4 cadenas, >= 15 provincias;
# durables >= 3 cadenas). EDITABLE: cambia cantidades abajo. LIMPIA y reescribe cantidad_01..06.
# Uso: pegar en Colab, ejecutar, subir el Excel canasta_representativa_*.xlsx (cualquier version),
#      descargar *_con_canastas.xlsx y subirlo a Drive/output_canasta/ como canasta_representativa_<periodo>.xlsx
# ============================================================
import openpyxl, os
from google.colab import files

uploaded = files.upload()
archivo_excel = next(iter(uploaded.keys()))

SHEET = "Productos unicos"
COLS  = ["cantidad_01","cantidad_02","cantidad_03","cantidad_04","cantidad_05","cantidad_06"]

# ---- Cantidades por EAN (unidades/mes). EDITAR ACA ----
CANTIDADES = {
    '7790070012050': {'cantidad_01': 2, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Almacén | Aceite de Girasol Cocinero 900 Ml
    '7790272001029': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 2, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Almacén | Aceite de Girasol Natura 1.5 Lt
    '7790272001005': {'cantidad_01': 0, 'cantidad_02': 2, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 2, 'cantidad_06': 0},  # Almacén | Aceite de Girasol Natura 900 Ml
    '7790070231833': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Almacén | Aceite de Oliva Extra Virgen Lira 500 Ml
    '7791866000381': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Almacén | Kétchup Natura 250 Gr
    '7791866001203': {'cantidad_01': 1, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Almacén | Mayonesa Doypack Natura 237 Gr
    '7791866001364': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Almacén | Mayonesa Doypack Natura 500 Ml
    '7794000006072': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 2, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Almacén | Mayonesa Regular Doypack Hellmanns 475 Gr
    '7790150540183': {'cantidad_01': 1, 'cantidad_02': 1, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Almacén | Orégano sin TACC Alicante 25 Gr
    '7790072001014': {'cantidad_01': 1, 'cantidad_02': 1, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Almacén | Sal Fina en Paquete Celusal 500 Gr
    '7791866000480': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Almacén | Salsa Golf Natura 250 Cc
    '7790130000058': {'cantidad_01': 1, 'cantidad_02': 1, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Almacén | Vinagre de Alcohol Menoyo 1 Lt
    '7791120031557': {'cantidad_01': 3, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Almacén | Arroz Largo Fino 00000 Molinos Ala 1 Kg
    '7790070431417': {'cantidad_01': 0, 'cantidad_02': 3, 'cantidad_03': 2, 'cantidad_04': 0, 'cantidad_05': 3, 'cantidad_06': 0},  # Almacén | Arroz Parboil Bolsa Gallo Oro 1 Kg
    '7792540250450': {'cantidad_01': 3, 'cantidad_02': 3, 'cantidad_03': 2, 'cantidad_04': 0, 'cantidad_05': 3, 'cantidad_06': 0},  # Almacén | Azúcar Molida Superior Ledesma 1 Kg
    '7790490998231': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Almacén | Edulcorante Líquido Clásico Hileret 250 Ml
    '7791019172101': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Almacén | Aritos de Avena y Miel 3 Arroyos 180 Gr
    '7896004004990': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 2, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Almacén | Cereal sabor Chocolate Müsli Kellog´s 255 Gr
    '7790045001584': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Almacén | Copos de Maíz Naturales Granix 160 Gr
    '7790580132392': {'cantidad_01': 2, 'cantidad_02': 2, 'cantidad_03': 3, 'cantidad_04': 0, 'cantidad_05': 2, 'cantidad_06': 0},  # Almacén | Arvejas Secas Remojadas sin Conservantes Lata 
    '7790580131364': {'cantidad_01': 0, 'cantidad_02': 2, 'cantidad_03': 3, 'cantidad_04': 0, 'cantidad_05': 2, 'cantidad_06': 0},  # Almacén | Atún al Natural Libre de Gluten sin TACC La Ca
    '7790580132422': {'cantidad_01': 0, 'cantidad_02': 2, 'cantidad_03': 2, 'cantidad_04': 0, 'cantidad_05': 2, 'cantidad_06': 0},  # Almacén | Choclo Amarillo Desgranado Arcor 300 Gr
    '7793360132469': {'cantidad_01': 2, 'cantidad_02': 1, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 1.5, 'cantidad_06': 0},  # Almacén | Garbanzos Secos Remojados Lata La Campagnola 3
    '7790360720115': {'cantidad_01': 1, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0.5, 'cantidad_06': 0},  # Almacén | Picadillo de Carne en Lata Swift 90 Gr
    '7790580133153': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Almacén | Salsa Lista para Pizza Arcor 340 Gr
    '7790580567903': {'cantidad_01': 3, 'cantidad_02': 3, 'cantidad_03': 4, 'cantidad_04': 0, 'cantidad_05': 3, 'cantidad_06': 0},  # Almacén | Tomate Pelado Perita en Lata Arcor 400 Gr
    '7795735000335': {'cantidad_01': 1, 'cantidad_02': 1, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Almacén | Bizcochos Agridulces Don Satur 200 Gr
    '7792180006448': {'cantidad_01': 3, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Almacén | Galletitas Crackers Clásicas Paseo 300 Gr
    '7790040143524': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 4, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Almacén | Galletitas Rumba Bagley 330 Gr
    '7622201735906': {'cantidad_01': 4, 'cantidad_02': 4, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 4, 'cantidad_06': 0},  # Almacén | Galletitas con Chips Chocolate Pepitos 119 Gr
    '7790045826095': {'cantidad_01': 0, 'cantidad_02': 3, 'cantidad_03': 3, 'cantidad_04': 0, 'cantidad_05': 3, 'cantidad_06': 0},  # Almacén | Galletitas de Salvado sin Sal Granix 230 Gr
    '7790040133587': {'cantidad_01': 0, 'cantidad_02': 4, 'cantidad_03': 6, 'cantidad_04': 0, 'cantidad_05': 4, 'cantidad_06': 0},  # Almacén | Alfajor Minitorta Clásica Águila 69 Gr
    '0000077991584': {'cantidad_01': 4, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Almacén | Alfajor Triple de Chocolate Fantoche 85 Gr
    '0000077953124': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Almacén | Chocolate con Maní Cofler Block 38 Gr
    '0000077940131': {'cantidad_01': 2, 'cantidad_02': 1, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Almacén | Turrón Arcor 25 Gr
    '7792180139320': {'cantidad_01': 3, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Almacén | Harina 000 Fortificada con Calcio Cañuelas 1 K
    '7790070562258': {'cantidad_01': 0, 'cantidad_02': 2, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 2, 'cantidad_06': 0},  # Almacén | Harina 000 con Vitamina Zinc Favorita 1 Kg
    '7790070562265': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 2, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Almacén | Harina 0000 con Vitamina Zinc Favorita 1 Kg
    '7790070562180': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Almacén | Harina Leudante Blancaflor 1 Kg
    '7790580138721': {'cantidad_01': 2, 'cantidad_02': 1, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Almacén | Polenta Prestopronta 730 Gr
    '8445291121881': {'cantidad_01': 1, 'cantidad_02': 1, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Almacén | Cacao Paquete Nesquik 360 Gr
    '7790070933638': {'cantidad_01': 1, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Almacén | Café Instantáneo Suave Extra Aroma Arlistán 17
    '8445291082137': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Almacén | Café Instantáneo Suave Origen Dolca 170 Gr
    '7790550000157': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Almacén | Café Molido Torrado en Bolsa Cabrales 250 Gr
    '7790150355084': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Almacén | Té Verde en Saquitos La Virginia 20 Un
    '7790480008261': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Almacén | Té en Saquitos Green Hills 50 Un
    '7790710334573': {'cantidad_01': 2, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Almacén | Yerba Hierbas Serranas Cbsé 1 Kg
    '7790387013610': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 2, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Almacén | Yerba Mate con Palo 4Flex Taragui 1 Kg
    '7793704000928': {'cantidad_01': 0, 'cantidad_02': 2, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 2, 'cantidad_06': 0},  # Almacén | Yerba Suave Playadito 1 Kg
    '7790580132163': {'cantidad_01': 1, 'cantidad_02': 1, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Almacén | Mermelada Durazno Arcor 454 Gr
    '7793360131516': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 2, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Almacén | Mermelada de Durazno Bc La Campagnola 390 Gr
    '7790158229547': {'cantidad_01': 0, 'cantidad_02': 0.5, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 0.5, 'cantidad_06': 0},  # Almacén | Miel Squeezze Aleluya 250 Gr
    '7798114470194': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Almacén | Budín de Pan con Ralladura de Limón Flanchello
    '7790070433169': {'cantidad_01': 1, 'cantidad_02': 1, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Almacén | Pan Rallado Fortificado Preferido 500 Gr
    '7793890258783': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 4, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Almacén | Pan con Salvado Lactal 560 Gr
    '7793890258769': {'cantidad_01': 1, 'cantidad_02': 3, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 2, 'cantidad_06': 0},  # Almacén | Pan de Mesa Blanco Chico Lactal 315 Gr
    '7790070336545': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 3, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Almacén | Fideos Fetuccini Don Vicente 500 Gr
    '7790070336149': {'cantidad_01': 0, 'cantidad_02': 3, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 3, 'cantidad_06': 0},  # Almacén | Fideos Mostachol N51 Lucchetti 500 Gr
    '7790070336316': {'cantidad_01': 0, 'cantidad_02': 3, 'cantidad_03': 3, 'cantidad_04': 0, 'cantidad_05': 3, 'cantidad_06': 0},  # Almacén | Fideos Tallarines N5 Matarazzo 500 Gr
    '7790070336118': {'cantidad_01': 3, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Almacén | Fideos Tallarín N5 Lucchetti 500 Gr
    '7790070337085': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 2, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Almacén | Fideos al Huevo Tirabuzón Lucchetti 500 Gr
    '7790310985465': {'cantidad_01': 1, 'cantidad_02': 2, 'cantidad_03': 2, 'cantidad_04': 0, 'cantidad_05': 2, 'cantidad_06': 0},  # Almacén | Papas Fritas Clásicas Lays 134 Gr
    '7794000008557': {'cantidad_01': 1, 'cantidad_02': 1, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Almacén | Caldo de Verdura Knorr 12 Un
    '8445290993021': {'cantidad_01': 1, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0.5, 'cantidad_06': 0},  # Almacén | Puré de Papas Cremoso Original Maggi 125 Gr
    '7794000005884': {'cantidad_01': 1, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0.5, 'cantidad_06': 0},  # Almacén | Sopa Crema de Zapallo con Nuez Moscada Knorr 7
    '7799155000203': {'cantidad_01': 2, 'cantidad_02': 2, 'cantidad_03': 2, 'cantidad_04': 0, 'cantidad_05': 2, 'cantidad_06': 0},  # Bebidas | Agua Mineral con Gas Pet Villavicencio 500 Ml
    '7799155000197': {'cantidad_01': 0, 'cantidad_02': 6, 'cantidad_03': 8, 'cantidad_04': 0, 'cantidad_05': 6, 'cantidad_06': 0},  # Bebidas | Agua Mineral sin Gas Pet Villavicencio 2 Lt
    '7792931000039': {'cantidad_01': 4, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Bebidas | Agua Mineralizada sin Gas Glaciar 2 Lt
    '7790895640483': {'cantidad_01': 0, 'cantidad_02': 4, 'cantidad_03': 4, 'cantidad_04': 0, 'cantidad_05': 3, 'cantidad_06': 0},  # Bebidas | Agua Saborizada Manzana sin Gas Aquarius 1.5 L
    '7790950000160': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Bebidas | Aperitivo Americano Gancia 950 Cc
    '7790290001179': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Bebidas | Fernet Branca 450 Cc
    '7790290101602': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Bebidas | Fernet Branca 750 Ml
    '7791250001345': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Bebidas | Vodka Smirnoff 750 Cc
    '7792798007387': {'cantidad_01': 0, 'cantidad_02': 4, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 4, 'cantidad_06': 0},  # Bebidas | Cerveza Rubia Cristal Quilmes 1 Lt
    '7793147118860': {'cantidad_01': 4, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Bebidas | Cerveza Rubia en Lata Schneider 473 Cc
    '7792798003716': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 6, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Bebidas | Cerveza en Botella No Retornable Corona 710 Cc
    '7790895000997': {'cantidad_01': 0, 'cantidad_02': 5, 'cantidad_03': 6, 'cantidad_04': 0, 'cantidad_05': 5, 'cantidad_06': 0},  # Bebidas | Coca Cola Sabor Original 2.25 Lt
    '7790895067570': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 3, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Bebidas | Coca Cola sin Azúcar 2.25 Lt
    '7791813888468': {'cantidad_01': 4, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Bebidas | Gaseosa Cola Pepsi 2 Lt
    '7790639003468': {'cantidad_01': 2, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Bebidas | Gaseosa Lima Limón Classic Cunnington 2.25 Lt
    '7790895001017': {'cantidad_01': 0, 'cantidad_02': 2, 'cantidad_03': 2, 'cantidad_04': 0, 'cantidad_05': 2, 'cantidad_06': 0},  # Bebidas | Gaseosa Naranja Fanta 2.25 Lt
    '7790895001000': {'cantidad_01': 0, 'cantidad_02': 2, 'cantidad_03': 3, 'cantidad_04': 0, 'cantidad_05': 2, 'cantidad_06': 0},  # Bebidas | Gaseosa Sprite 2.25 Lt
    '7790895640025': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 4, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Bebidas | Bebida Isotónica Mountain Blast Powerade 500 M
    '7790895000737': {'cantidad_01': 0, 'cantidad_02': 3, 'cantidad_03': 4, 'cantidad_04': 0, 'cantidad_05': 3, 'cantidad_06': 0},  # Bebidas | Jugo Listo Multifruta Cepita 200 Ml
    '7622201735685': {'cantidad_01': 6, 'cantidad_02': 6, 'cantidad_03': 2, 'cantidad_04': 0, 'cantidad_05': 6, 'cantidad_06': 0},  # Bebidas | Jugo en Polvo Sabor Limón Dulce Tang 15 Gr
    '7790240002010': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 3, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Bebidas | Vino Tinto Cabernet Sauvignon Fond de Cave 750
    '7798074864675': {'cantidad_01': 0, 'cantidad_02': 2, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 2, 'cantidad_06': 0},  # Bebidas | Vino Tinto Malbec Portillo 750 Ml
    '7791540053351': {'cantidad_01': 2, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Bebidas | Vino Tinto Tradicional Brick Termidor 1 Lt
    '7500435228763': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Bebés y mamás | Pañal Talle XG Baby-dry Hipoalergénico Pampers
    '7794626011023': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Bebés y mamás | Toallitas Húmedas Clásico y Cotidiano Huggies 
    '7796962987321': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 1, 'cantidad_05': 0, 'cantidad_06': 0},  # CLIMATIZACIÓN | Aire Acondicionado Philco Split 3200W Frio Cal
    '7798115593090': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 1, 'cantidad_05': 0, 'cantidad_06': 0},  # Climatización | Caloventor Fh801 2000 W Blanco Protalia 1 Un
    '7502271278278': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 1, 'cantidad_05': 0, 'cantidad_06': 0},  # Cocinas y Hornos | Cocina Multigas Cd5602Ab0 Drean 1 Un
    '7796885495194': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 1, 'cantidad_05': 0, 'cantidad_06': 0},  # Cocinas y Hornos | Hornos Microondas 20Lt Blanco 700 W Bgh 1 Un
    '7790670050650': {'cantidad_01': 1, 'cantidad_02': 2, 'cantidad_03': 3, 'cantidad_04': 0, 'cantidad_05': 2, 'cantidad_06': 0},  # Congelados | Hamburguesas de Carne Clásicas sin TACC Paty 4
    '7790174001349': {'cantidad_01': 1, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Congelados | Milanesa de Soja Tradicional Vegetalex 4 Un
    '7790070034410': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Congelados | Nuggets de Pollo Bolsa Granja Del Sol 400 Gr
    '7790070036278': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 2, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Congelados | Espinaca Congelada Granja Del Sol 500 Gr
    '7791451071000': {'cantidad_01': 1, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0.5, 'cantidad_06': 0},  # Frescos | Dulce de Batata con Vainilla Esnaola 500 Gr
    '7790079022593': {'cantidad_01': 2, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Frescos | Jamon Cocido Fetas Finas Paladini 200 Gr
    '7798013102875': {'cantidad_01': 0, 'cantidad_02': 3, 'cantidad_03': 4, 'cantidad_04': 0, 'cantidad_05': 3, 'cantidad_06': 0},  # Frescos | Jamón Cocido Genuino Cagnoli 120 Gr
    '7790079001031': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Frescos | Mortadela Mini Paladini 300 Gr
    '7794990880225': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 2, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Frescos | Queso Dambo en Fetas La Paulina 180 Gr
    '7790742373304': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 3, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Frescos | Queso Light Finlandia 290 Gr
    '7790398100132': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Frescos | Queso Rallado La Paulina 150 Gr
    '7790742222909': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 2, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Frescos | Queso Rallado La Serenisima 35 Gr
    '7794990880232': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Frescos | Queso Tybo en Feta La Paulina 180 Gr
    '7798060852990': {'cantidad_01': 2, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Frescos | Queso Untable Jamón Tonadita 180 Gr
    '7791337061439': {'cantidad_01': 0, 'cantidad_02': 2, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 2, 'cantidad_06': 0},  # Frescos | Queso Untable Light La Serenisima 290 Gr
    '7798013100697': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 2, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Frescos | Salame Tandilero Tipo Fuet en Pieza Cagnoli 15
    '7790670052388': {'cantidad_01': 2, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Frescos | Salchicha Clásica Flow Pack Patyviena 6 Un
    '7790670052418': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Frescos | Salchichas Vienísima 12 Un
    '7790670052401': {'cantidad_01': 0, 'cantidad_02': 2, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 2, 'cantidad_06': 0},  # Frescos | Salchichas Vienísima 6 Un
    '7622202247316': {'cantidad_01': 2, 'cantidad_02': 1, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Frescos | Caramelos Mentolyptus Halls 25.2 Gr
    '7790742141101': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Frescos | Crema de Leche Uat Tetra Top La Serenísima 330
    '7790787153664': {'cantidad_01': 1, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Frescos | Dulce de Leche Clásico Ilolay 400 Gr
    '7790742625304': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Frescos | Dulce de Leche Clásico La Serenísima 400 Gr
    '7790742067005': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 2, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Frescos | Dulce de Leche Repostero La Serenísima 400 Gr
    '7790742335500': {'cantidad_01': 0, 'cantidad_02': 24, 'cantidad_03': 24, 'cantidad_04': 0, 'cantidad_05': 28, 'cantidad_06': 0},  # Frescos | Leche Entera 3% Uat La Serenísima 1 Lt
    '7798338290028': {'cantidad_01': 24, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Frescos | Leche Entera Larga Vida Tradición Tres Niñas 1
    '7791337007819': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 6, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Frescos | Leche Fermentada Parcialmente Descremada con P
    '7791337007260': {'cantidad_01': 0, 'cantidad_02': 2, 'cantidad_03': 2, 'cantidad_04': 0, 'cantidad_05': 2, 'cantidad_06': 0},  # Frescos | Leche Larga Vida Chocolatada Brick Cindor 200 
    '7790742333605': {'cantidad_01': 0, 'cantidad_02': 6, 'cantidad_03': 10, 'cantidad_04': 0, 'cantidad_05': 6, 'cantidad_06': 0},  # Frescos | Leche UAT Parcialmente Descremada Zero Lactos 
    '7790742436207': {'cantidad_01': 1, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0.5, 'cantidad_06': 0},  # Frescos | Leche en Polvo Descremada Fortificada La Seren
    '7798060850026': {'cantidad_01': 1, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Frescos | Manteca Calidad Extra Tonadita 200 Gr
    '7793940054006': {'cantidad_01': 0, 'cantidad_02': 2, 'cantidad_03': 2, 'cantidad_04': 0, 'cantidad_05': 2, 'cantidad_06': 0},  # Frescos | Manteca La Serenísima 200 Gr
    '7791620187778': {'cantidad_01': 1, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0.5, 'cantidad_06': 0},  # Frescos | Margarina Dorada Dánica 210 Gr
    '7791337008472': {'cantidad_01': 0, 'cantidad_02': 2, 'cantidad_03': 3, 'cantidad_04': 0, 'cantidad_05': 2, 'cantidad_06': 0},  # Frescos | Postre Vainilla con Rocklets Serenito 120 Gr
    '7791337009981': {'cantidad_01': 0, 'cantidad_02': 6, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 6, 'cantidad_06': 0},  # Frescos | Yogur Batido Natural Yogurísimo 190 Gr
    '7791337008694': {'cantidad_01': 0, 'cantidad_02': 4, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 4, 'cantidad_06': 0},  # Frescos | Yogur Bebible Frutilla Danonino 185 Gr
    '7793913013566': {'cantidad_01': 4, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Frescos | Yogur Firme Descremado Vainilla Pote Tregar 12
    '7791337007246': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 8, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Frescos | Yogur Firme Descremado con Colágeno Vainilla P
    '7790070621863': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 2, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Frescos | Ravioles Ricota Mozzarella Danbo Parmesano La 
    '7790070622037': {'cantidad_01': 2, 'cantidad_02': 2, 'cantidad_03': 2, 'cantidad_04': 0, 'cantidad_05': 2, 'cantidad_06': 0},  # Frescos | Tapa para Pascualina Criolla La Salteña 400 Gr
    '7799111685901': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 1, 'cantidad_05': 0, 'cantidad_06': 0},  # Heladeras y Freezers | Heladera Sidebyside Inox 428 Lts Philco Phsb45
    '7798111354763': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 1, 'cantidad_05': 0, 'cantidad_06': 0},  # Informática | Auricular Bt Earbud Aiwa Mod. Twa 70N Negro
    '6932554416638': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 1, 'cantidad_05': 0, 'cantidad_06': 0},  # Informática | Celular 14C 4Gb/128Gb Starry Blue Azul Redmi X
    '0196802082679': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 1, 'cantidad_05': 0, 'cantidad_06': 0},  # Informática | Notebook 14 Celn4120 4/128 Gb 82V Lenovo 1 Un
    '7797026970037': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 1, 'cantidad_05': 0, 'cantidad_06': 0},  # Informática | Parlante Bt Stromberg Boombox Force
    '8806094365573': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 1, 'cantidad_05': 0, 'cantidad_06': 0},  # Lavado | Lavarropas Frontal Inventer WW65A4000EEU Blanc
    '7794440003310': {'cantidad_01': 1, 'cantidad_02': 1, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Limpieza | Esponja Pintada Lisa Virulana 1 Un
    '7793253004231': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Limpieza | Aromatizante Ambiente en Aerosol Frescura de L
    '7790520995285': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Limpieza | Desinfectante de Ambientes y Superficies Origi
    '7790520997623': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 0.5, 'cantidad_06': 0},  # Limpieza | Mata Moscas y Mosquitos Aerosol Raid 370 Ml
    '7790132098459': {'cantidad_01': 2, 'cantidad_02': 2, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 2, 'cantidad_06': 0},  # Limpieza | Lavandina Concentrada Ayudín 1 Lt
    '7793253003722': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 2, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Limpieza | Lavandina Lavanda Triple Poder Ayudín 2 Lt
    '7791290795617': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Limpieza | Limpiador Vidrios y Multiuso Gatillo Expert Ci
    '7790117000200': {'cantidad_01': 1, 'cantidad_02': 1, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Limpieza | Bolsas de Residuos de 45 Cm 60 Cm en Rollo Asu
    '7790990003039': {'cantidad_01': 0, 'cantidad_02': 2, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 2, 'cantidad_06': 0},  # Limpieza | Detergente Limón Magistral 20 Un 500 Ml
    '7791290794054': {'cantidad_01': 1, 'cantidad_02': 2, 'cantidad_03': 2, 'cantidad_04': 0, 'cantidad_05': 2, 'cantidad_06': 0},  # Limpieza | Detergente Líquido Bioactive Limón Cif 300 Ml
    '7791290792814': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 2, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Limpieza | Jabón Líquido para Ropa Bio Encimas Skip 800 M
    '7791290792036': {'cantidad_01': 1, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Limpieza | Jabón Líquido para Ropa Eco Lavado Doypack Ala
    '7791290793460': {'cantidad_01': 1, 'cantidad_02': 1, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Limpieza | Jabón de Lavar en Pan Fragancia Coco Ala 2 Un 
    '7791290792142': {'cantidad_01': 1, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Limpieza | Jabón en Polvo Matic Mañana de Sol Ala 800 Gr
    '7891150000971': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 2, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Limpieza | Suavizante para Ropa Concentrado Cuidado Intes
    '7790250015536': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 4, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Limpieza | Papel Higiénico Doble Hoja Manzanilla Higienol
    '7790250016182': {'cantidad_01': 0, 'cantidad_02': 3, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 3, 'cantidad_06': 0},  # Limpieza | Papel Higiénico Fresh Hoja Simple con Aloe Ver
    '7793344904143': {'cantidad_01': 2, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Limpieza | Papel Higiénico Hoja Simple Blanco Elegante Pa
    '7790250056881': {'cantidad_01': 0, 'cantidad_02': 2, 'cantidad_03': 2, 'cantidad_04': 0, 'cantidad_05': 2, 'cantidad_06': 0},  # Limpieza | Rollos de Cocina Clásico 50 Paños Sussex 3 Un
    '7790250057765': {'cantidad_01': 1, 'cantidad_02': 1, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Limpieza | Servilletas Papel Hoja Simple Clásica 30X30 Su
    '8445290057433': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 0.5, 'cantidad_06': 0},  # Mascotas | Alimento para Gatos Adultos Pescado y Pollo Ca
    '8445290944559': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Mascotas | Alimento para Perro Cachorro Gran Comienzo Dog
    '7613287613431': {'cantidad_01': 1, 'cantidad_02': 1, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Mascotas | Alimento para Perros Adultos Bolsa Dogui 3 Kg
    '7799111033290': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 1, 'cantidad_05': 0, 'cantidad_06': 0},  # Pequeños Electrodomésticos | Pava Eléctrica Blanca con Corte 1.7 Lt Atma 1 
    '0053891142960': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 1, 'cantidad_05': 0, 'cantidad_06': 0},  # Pequeños Electrodomésticos | Plancha A Vapor Oster Gcstbs5905
    '0034264476424': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 1, 'cantidad_05': 0, 'cantidad_06': 0},  # Pequeños Electrodomésticos | Licuadora De Mano Oster Fpsthb2800 354 Negra 8
    '7790064001909': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 1},  # Perfumería | Algodón Discos Redondos Estrella 80 Un
    '7790064000261': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Perfumería | Algodón Estrella Clásico 75 Gr
    '7506460101279': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 1},  # Perfumería | Banda de Cera Depilatoria Corporal Piel Sensib
    '4005808319695': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 1},  # Perfumería | Crema Corporal Milk Nutritiva Piel Extra Seca 
    '8002990292139': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 1},  # Perfumería | Crema Depilatoria Piel Sensible Silck & Fresh 
    '7791293043791': {'cantidad_01': 1, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Perfumería | Desodorante Aerosol Marine Axe 150 Cc
    '7791293049502': {'cantidad_01': 0, 'cantidad_02': 2, 'cantidad_03': 3, 'cantidad_04': 0, 'cantidad_05': 2, 'cantidad_06': 0},  # Perfumería | Desodorante Aerosol Men Rexona 150 Ml
    '7791293049557': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 2},  # Perfumería | Desodorante Aerosol Women Nutrivive Rexona 150
    '7791293051208': {'cantidad_01': 0, 'cantidad_02': 3, 'cantidad_03': 4, 'cantidad_04': 0, 'cantidad_05': 3, 'cantidad_06': 0},  # Perfumería | Jabón de Tocador Original Dove 90 Gr
    '7891024034781': {'cantidad_01': 2, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Perfumería | Jabón de Tocador con Karite Palmolive 85 Gr
    '7500435173100': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 1},  # Perfumería | Crema Afeitar Protectora Suave y Lisa Venus 15
    '7702018072392': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 1},  # Perfumería | Rasuradora Desechable Simply Venus Gillette 2 
    '7500435225366': {'cantidad_01': 1, 'cantidad_02': 1, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Perfumería | Repuesto Máquina Afeitar Carbono Gillette Mach
    '7702018072477': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0.5},  # Perfumería | Repuesto Rasuradora Recargable Venus Gillette 
    '4005900985828': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 1},  # Perfumería | Toallitas Desmaquillantes Micellair 3 en 1 Niv
    '5900273001566': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Perfumería | Cepillo Dental Colgate Extra Clean 1 Un
    '7509546687292': {'cantidad_01': 1, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Perfumería | Crema Dental Doble Protección Odol 90 Gr
    '7509546702605': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 2, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Perfumería | Crema Dental Total Original Mint Colgate 90 Gr
    '7793100111891': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Perfumería | Crema Dental Ultra Blanco Colgate 90 Gr
    '7891010255237': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Perfumería | Enjuague Bucal Cool Mint Listerine 500 Ml
    '7791293047515': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 1},  # Perfumería | Acondicionador Oleo Nutrición Dove 400 Cc
    '7791293045740': {'cantidad_01': 1, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 0},  # Perfumería | Shampoo Crema Balance Sedal 340 Ml
    '7791293047102': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 2, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 1},  # Perfumería | Shampoo Reconstrucción Completa Dove 400 Ml
    '7509552924121': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 1},  # Perfumería | Tintura Coloración Permanente 50 Castaño Nutri
    '7790139101053': {'cantidad_01': 0, 'cantidad_02': 1, 'cantidad_03': 1, 'cantidad_04': 0, 'cantidad_05': 1, 'cantidad_06': 0},  # Perfumería | Alcohol Etílico al 70% Spray Bialcohol 100 Ml
    '7790010002899': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 1},  # Perfumería | Protectores Diarios Todoslosdias Compact Caref
    '7891010254773': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 2},  # Perfumería | Tampones Medio O.B. 8 Un
    '7790770601851': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 0, 'cantidad_05': 0, 'cantidad_06': 1},  # Perfumería | Toallas Femeninas Nocturna Max Nosotras 8 Un
    '7790010002769': {'cantidad_01': 1, 'cantidad_02': 2, 'cantidad_03': 2, 'cantidad_04': 0, 'cantidad_05': 2, 'cantidad_06': 2},  # Perfumería | Toallas Femeninas con Alas Ultrafina Suave Sie
    '7796941329296': {'cantidad_01': 0, 'cantidad_02': 0, 'cantidad_03': 0, 'cantidad_04': 1, 'cantidad_05': 0, 'cantidad_06': 0},  # Tv Audio y Video | Smart Tv Qled 43" Fhd 43s5k Go Tcl 1 Un
}

def norm(e):
    s = str(e).strip()
    if s.endswith(".0"): s = s[:-2]
    return s.lstrip("0")

LOOKUP = {norm(k): v for k, v in CANTIDADES.items()}
ORIG   = {norm(k): k for k in CANTIDADES}

wb = openpyxl.load_workbook(archivo_excel)
if SHEET not in wb.sheetnames:
    raise ValueError(f"No existe la hoja '{SHEET}'. Hojas: {wb.sheetnames}")
ws = wb[SHEET]
headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
if "id_producto" not in headers:
    raise ValueError("La hoja no tiene columna 'id_producto'.")
col_id = headers.index("id_producto") + 1
col_idx = {}
for c in COLS:
    if c in headers:
        col_idx[c] = headers.index(c) + 1
    else:  # crear la columna al final
        ci = len(headers) + 1; ws.cell(row=1, column=ci, value=c); headers.append(c); col_idx[c] = ci
        print(f"Columna '{c}' no existia: se creo en la posicion {ci}.")

# limpiar cantidad_01..06
for ci in col_idx.values():
    for row in ws.iter_rows(min_row=2, min_col=ci, max_col=ci):
        row[0].value = None
# escribir
matches = {c: 0 for c in COLS}; encontrados = set()
for row in ws.iter_rows(min_row=2):
    ev = row[col_id-1].value
    if ev is None: continue
    e = norm(ev)
    if e in LOOKUP:
        encontrados.add(e)
        for c, cant in LOOKUP[e].items():
            if cant: row[col_idx[c]-1].value = cant; matches[c] += 1

print(f"\n=== Hoja '{SHEET}' - productos cargados ===")
for c in COLS:
    tot = sum(1 for v in CANTIDADES.values() if v.get(c,0) > 0)
    print(f"  {c}: {matches[c]}/{tot}")
faltantes = sorted(set(LOOKUP) - encontrados)
if faltantes:
    print(f"\nAVISO: {len(faltantes)} EAN no estan en la hoja:")
    for e in faltantes: print("   ", ORIG[e])
else:
    print("\nOK: los", len(LOOKUP), "EAN estan en la hoja.")

base = os.path.splitext(os.path.basename(archivo_excel))[0]
for suf in ("_con_canastas", "_con_femenina"):
    base = base.replace(suf, "")
salida = f"{base}_con_canastas.xlsx"
wb.save(salida); print(f"\nGuardado: {salida}"); files.download(salida)
